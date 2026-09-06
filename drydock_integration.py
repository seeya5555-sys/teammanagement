"""
drydock_integration.py — Dock Manager × TRMT 통합 shim (프로덕션)
설계 §0b 확정결정(MCP/OAuth 폐기 + SSO 세션신뢰+admin게이팅)의 실배선.
PoC(poc_merged.py 11/11 PASS)에서 검증된 로직을 프로덕션 모듈로 승격.

원칙: drydock 원본 코드(app.py)는 건드리지 않고 import 후 런타임 패치.
      → drydock git 업데이트 pull 시 충돌 최소화(방식① 코드분리 유지).

사용:
    import drydock_integration as ddi
    dd_app = ddi.apply(dd_module, trmt_app)   # 반환: 통합·잠금 완료된 drydock WSGI app
"""
from contextlib import contextmanager
from copy import copy as _copy
from datetime import date as _date, datetime as _datetime, timedelta as _timedelta, timezone as _timezone
import hashlib
import io
import json
import math
import os
import re
import secrets
import sqlite3
import threading
import time
from urllib.parse import urlsplit


DAILY_EVENTS_PATH = "/api/integration/daily-events"
ROSTER_VESSELS_PATH = "/api/integration/roster-vessels"
SVMS_IMPORT_PENDING_PATH = "/api/integration/svms-job-import/pending"
SVMS_IMPORT_COMPLETE_PATH = "/api/integration/svms-job-import/complete"
DAILY_EVENT_SOURCES = (
    "jobs",
    "discussions",
    "class_items",
    "steel_repair",
    "pipe_repair",
    "outfitting",
    "wbt_cot",
    "portable_fan",
    "staging",
    "gas_free",
)
_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")

YARD_JOB_SECTION_MAP = {
    "2": "GENERAL", "3": "GENERAL", "4": "PAINT", "6": "STEEL",
    "5": "DECK", "9": "DECK", "10": "DECK", "11": "DECK",
    "12": "DECK", "18": "DECK", "24": "DECK",
    "7": "ENGINE", "8": "ENGINE", "13": "ENGINE", "14": "ENGINE",
    "15": "ENGINE", "19": "ENGINE", "20": "ENGINE", "21": "ENGINE",
    "22": "ENGINE", "23": "ENGINE", "25": "ENGINE", "40": "ENGINE",
    "16": "ELECTRIC", "17": "ELECTRIC",
}
_YARD_TOTAL_ROW = re.compile(
    r"total price|final discount|after di(?:s)?count|normal total|sub\s*total|소계|합계",
    re.I,
)
_YARD_QUOTE_REQUEST_ROW = re.compile(
    r"^\s*please\s+provide\s+(?:a\s+)?quotation\b|"
    r"\bto\s+be\s+quoted\b[.\s]*$|\bfor\s+quotation\s+only\b[.\s]*$",
    re.I,
)
_YARD_ITEM_NO = re.compile(r"^\d+(?:\.\d+)*$")
_YARD_IMPORT_TTL = 30 * 60
_YARD_IMPORT_MAX = 8
_SVMS_JOB_TAG = re.compile(r"^\s*\[([^\]]+)\]\s*(.*)$")
_SVMS_JOB_BARE_TAG = re.compile(
    r"^\s*((?:ST|S|P|R)\d+(?:-[A-Z0-9]+)?)\b\s*[-:]?\s*(.*)$", re.I
)
_SVMS_JOB_CANONICAL_NUMBER = re.compile(r"^(?:ST|S|P|R)\d+(?:-[A-Z0-9]+)?$", re.I)
_SVMS_IMPORT_TTL_SECONDS = 15 * 60
_JOB_PROGRESS_TEMPLATE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "static", "dock_templates", "job_progress.xlsx"
)
_JOB_PROGRESS_FIRST_ROW = 7
_JOB_PROGRESS_LAST_ROW = 370
_JOB_PROGRESS_TIMELINE_FIRST_COL = 16  # P
_JOB_PROGRESS_TIMELINE_LAST_COL = 53   # BA
_JOB_PROGRESS_HANGUL = re.compile(r"[\uac00-\ud7a3]")


def _text(value):
    """Return a stable, whitespace-normalized text value for event fields."""
    if value is None:
        return ""
    return str(value).strip()


def _bool(value):
    if isinstance(value, bool):
        return value
    return _text(value).lower() in {"1", "true", "yes", "y", "on", "urgent", "critical"}


def _job_progress_date(value):
    raw = _business_date(value)
    return _datetime.strptime(raw, "%Y-%m-%d").date() if raw else None


def _job_progress_remarks(value):
    if not value:
        return ""
    if isinstance(value, str):
        try:
            value = json.loads(value)
        except (TypeError, ValueError):
            return value.strip()
    if not isinstance(value, list):
        return _text(value)
    lines = []
    for item in value:
        if not isinstance(item, dict):
            text = _text(item)
            if text:
                lines.append(text)
            continue
        text = _text(item.get("progress") or item.get("remark") or item.get("text"))
        if not text:
            continue
        prefix = _text(item.get("date"))
        if item.get("important"):
            prefix = (prefix + " !").strip()
        lines.append((prefix + " " + text).strip())
    return "\n".join(lines)


def _job_progress_remarks_en(jobs, translator=None):
    """Return copied job rows with Korean progress remarks translated to English.

    Export must never silently retain Korean when translation is unavailable or
    incomplete; the user can retry without receiving a partly translated report.
    """
    copied = [dict(job) for job in jobs]
    remarks = [_job_progress_remarks(job.get("remarks")) for job in copied]
    targets = [index for index, value in enumerate(remarks) if _JOB_PROGRESS_HANGUL.search(value)]
    if targets:
        try:
            if translator is None:
                from helpers_shared import _translate_texts_en
                translator = _translate_texts_en
            translated = translator([remarks[index] for index in targets])
        except Exception as exc:
            raise RuntimeError("Job Progress 영문 번역에 실패했습니다") from exc
        if not isinstance(translated, (list, tuple)) or len(translated) != len(targets):
            raise RuntimeError("Job Progress 영문 번역에 실패했습니다")
        for index, value in zip(targets, translated):
            value = _text(value)
            if not value or _JOB_PROGRESS_HANGUL.search(value):
                raise RuntimeError("Job Progress 영문 번역에 실패했습니다")
            remarks[index] = value
    for job, value in zip(copied, remarks):
        job["remarks"] = value
    return copied


def _job_progress_number(value):
    try:
        number = float(value or 0)
    except (TypeError, ValueError):
        return 0.0
    return number if math.isfinite(number) else 0.0


def _yard_is_quote_request(description):
    """Identify quotation-request instructions that are not executable Jobs."""
    return bool(_YARD_QUOTE_REQUEST_ROW.search(_text(description)))


def _job_progress_download_name(value):
    name = re.sub(r'[\x00-\x1f\x7f/\\]+', "_", _text(value)).strip(" ._")
    return "%s_DD_JOB_PROGRESS.xlsx" % ((name or "vessel")[:120])


def _job_progress_month_headers(sheet, dates):
    first = _JOB_PROGRESS_TIMELINE_FIRST_COL
    last = _JOB_PROGRESS_TIMELINE_LAST_COL
    month_style = _copy(sheet.cell(5, first)._style)
    month_alignment = _copy(sheet.cell(5, first).alignment)
    for merged in list(sheet.merged_cells.ranges):
        if (merged.min_row == 5 and merged.max_row == 5
                and merged.max_col >= first and merged.min_col <= last):
            sheet.unmerge_cells(str(merged))
    for col in range(first, last + 1):
        sheet.cell(5, col).value = None
        sheet.cell(5, col)._style = _copy(month_style)
        sheet.cell(5, col).alignment = _copy(month_alignment)
    start = first
    while start <= last:
        key = (dates[start - first].year, dates[start - first].month)
        end = start
        while end < last:
            next_date = dates[end + 1 - first]
            if (next_date.year, next_date.month) != key:
                break
            end += 1
        if end > start:
            sheet.merge_cells(start_row=5, start_column=start, end_row=5, end_column=end)
        sheet.cell(5, start).value = dates[start - first].strftime("%b")
        start = end + 1


def _build_job_progress_workbook(vessel, jobs, template_path=_JOB_PROGRESS_TEMPLATE,
                                 translator=None):
    """Fill the owner's Job Progress template without carrying its sample data."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for Job Progress export") from exc

    if len(jobs) > _JOB_PROGRESS_LAST_ROW - _JOB_PROGRESS_FIRST_ROW + 1:
        raise ValueError("Job Progress 템플릿 최대 364개 Job을 초과했습니다")
    workbook = load_workbook(template_path)
    sheet = workbook["Job progress"]
    jobs = _job_progress_remarks_en(jobs, translator=translator)
    for merged in list(sheet.merged_cells.ranges):
        if (merged.min_row >= _JOB_PROGRESS_FIRST_ROW
                and merged.max_row <= _JOB_PROGRESS_LAST_ROW
                and merged.min_col >= 2 and merged.max_col <= 15):
            source = sheet.cell(merged.min_row, merged.min_col)
            targets = [(row, col) for row in range(merged.min_row, merged.max_row + 1)
                       for col in range(merged.min_col, merged.max_col + 1)]
            sheet.unmerge_cells(str(merged))
            for row, col in targets:
                target = sheet.cell(row, col)
                target._style = _copy(source._style)
                target.alignment = _copy(source.alignment)
    sheet["B2"] = "%s DD JOB PROGRESS" % _text(vessel["name"])
    sheet.freeze_panes = "F7"  # Keep Job No. through Vendor visible while scrolling.

    white_fill = PatternFill(fill_type="solid", fgColor="FFFFFFFF")
    for row in range(_JOB_PROGRESS_FIRST_ROW, _JOB_PROGRESS_LAST_ROW + 1):
        for col in range(2, _JOB_PROGRESS_TIMELINE_LAST_COL + 1):
            sheet.cell(row, col).fill = _copy(white_fill)
    for conditional_range in sheet.conditional_formatting:
        if "P7" not in str(conditional_range):
            continue
        for rule in sheet.conditional_formatting[conditional_range]:
            if not rule.formula:
                continue
            rule.formula = [{
                'AND(P$6>$K7,P$6<$L7)': 'AND($K7<>"",$L7<>"",P$6>$K7,P$6<$L7)',
                'AND(P$6=$L7)': 'AND($L7<>"",P$6=$L7)',
                'AND(P$6=$K7)': 'AND($K7<>"",P$6=$K7)',
            }.get(formula, formula) for formula in rule.formula]

    commence = _job_progress_date(vessel["dock_in"])
    starts = [_job_progress_date(job["start_date"]) for job in jobs]
    starts = [value for value in starts if value is not None]
    timeline_anchor = commence or (min(starts) if starts else _date.today())
    sheet["M4"] = commence
    sheet["N4"] = '=IF(M4="","",DATEDIF(M4,L4,"d")+1)'
    timeline_dates = [timeline_anchor - _timedelta(days=10) + _timedelta(days=offset)
                      for offset in range(_JOB_PROGRESS_TIMELINE_LAST_COL - _JOB_PROGRESS_TIMELINE_FIRST_COL + 1)]
    _job_progress_month_headers(sheet, timeline_dates)
    for col, value in enumerate(timeline_dates, _JOB_PROGRESS_TIMELINE_FIRST_COL):
        sheet.cell(6, col).value = value

    for row in range(_JOB_PROGRESS_FIRST_ROW, _JOB_PROGRESS_LAST_ROW + 1):
        for col in (2, 3, 4, 5, 6, 7, 8, 9, 11, 12, 15):
            sheet.cell(row, col).value = None
        sheet.cell(row, 10).value = "=H%d-I%d" % (row, row)
        sheet.cell(row, 13).value = '=IF(OR(K%d="",L%d=""),"",DATEDIF(K%d,L%d,"d"))' % (row, row, row, row)
        sheet.cell(row, 14).value = '=IF(OR(K%d="",L%d=""),"",MAX(0,MIN(DATEDIF(K%d,$L$4,"d"),M%d)))' % (
            row, row, row, row)
        sheet.cell(row, 11).number_format = "yyyy-mm-dd"
        sheet.cell(row, 12).number_format = "yyyy-mm-dd"

    for offset, job in enumerate(jobs):
        row = _JOB_PROGRESS_FIRST_ROW + offset
        completion = max(0.0, min(100.0, _job_progress_number(job["completion"]))) / 100.0
        values = {
            2: _text(job["number"]),
            3: _text(job["category"]),
            4: _text(job["description"]),
            5: _text(job["vendor"]),
            6: _job_progress_remarks(job["remarks"]),
            8: _job_progress_number(job["budget"]),
            9: _job_progress_number(job["consumption"]),
            11: _job_progress_date(job["start_date"]),
            12: _job_progress_date(job["end_date"]),
            15: completion,
        }
        for col, value in values.items():
            sheet.cell(row, col).value = value

    for row in (3, 4):
        sheet.cell(row, 8).value = '=SUMIF(C%d:C%d,G%d,H%d:H%d)' % (
            _JOB_PROGRESS_FIRST_ROW, _JOB_PROGRESS_LAST_ROW, row,
            _JOB_PROGRESS_FIRST_ROW, _JOB_PROGRESS_LAST_ROW)
        sheet.cell(row, 9).value = '=SUMIF(C%d:C%d,G%d,I%d:I%d)' % (
            _JOB_PROGRESS_FIRST_ROW, _JOB_PROGRESS_LAST_ROW, row,
            _JOB_PROGRESS_FIRST_ROW, _JOB_PROGRESS_LAST_ROW)
        sheet.cell(row, 10).value = "=H%d-I%d" % (row, row)

    class_sheet = workbook["Class item"]
    for row in range(4, class_sheet.max_row + 1):
        for col in range(2, 9):
            class_sheet.cell(row, col).value = None
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _business_date(value):
    """Extract a YYYY-MM-DD business date without accepting a date prefix."""
    raw = _text(value)
    if len(raw) < 10 or not _DATE_RE.match(raw[:10]):
        return None
    return raw[:10]


def _timezone_timestamp(value):
    """Make a Dock Manager timestamp explicit and canonical.

    The Dock Manager schema uses SQLite ``datetime('now')`` defaults, which
    are UTC but do not carry an offset.  Treating those known-naive values as
    UTC preserves their meaning while ensuring the integration contract never
    emits an ambiguous timestamp.
    """
    if isinstance(value, _datetime):
        parsed = value
    else:
        raw = _text(value)
        if not raw:
            raise ValueError("source_updated_at is required")
        if raw.endswith("Z"):
            raw = raw[:-1] + "+00:00"
        try:
            parsed = _datetime.fromisoformat(raw)
        except ValueError as exc:
            raise ValueError("invalid source_updated_at") from exc
    if parsed.tzinfo is None or parsed.utcoffset() is None:
        parsed = parsed.replace(tzinfo=_timezone.utc)
    return parsed.isoformat(timespec="seconds")


def _json_array(value, source):
    if value is None or value == "":
        return []
    try:
        parsed = json.loads(value) if isinstance(value, str) else value
    except (TypeError, ValueError) as exc:
        raise ValueError("invalid JSON in %s" % source) from exc
    if not isinstance(parsed, list):
        raise ValueError("%s must be a JSON array" % source)
    if any(not isinstance(item, dict) for item in parsed):
        raise ValueError("%s contains a non-object item" % source)
    return parsed


def _canonical_hash(event):
    content = {key: event[key] for key in event if key != "source_hash"}
    raw = json.dumps(content, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return "sha256:" + hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _section(source_table, values):
    haystack = " ".join(_text(value) for value in values).lower()
    if "egcs" in haystack:
        return "egcs"
    if source_table == "class_items":
        return "survey"
    # Only an explicit vendor/supplier tag or label routes a discussion to
    # Vendor.  Other discussions remain Shipyard instead of being guessed.
    if source_table == "discussions" and re.search(r"(?:^|[\s#\[(:])(?:vendor|supplier)(?:$|[\s#\]\):])", haystack):
        return "vendor"
    return "shipyard"


def _tracking_body(source_table, row):
    labels = {
        "steel_repair": ("description", "location", "status", "priority", "remark"),
        "pipe_repair": ("description", "system_line", "position_tank", "status", "remark"),
        "outfitting": ("description", "location", "status", "priority", "remark"),
        "wbt_cot": ("tank_name", "manhole_status", "remark"),
        "portable_fan": ("location", "qty", "remark"),
        "staging": ("location", "staging_area", "qty", "remark"),
        "gas_free": ("tank", "certificate", "remark"),
    }
    return "; ".join(
        "%s: %s" % (key.replace("_", " "), _text(row[key]))
        for key in labels[source_table]
        if key in row.keys() and _text(row[key])
    )


def _event_base(source_table, row, project_id, source_subkey, event_date,
                source_updated_at, kind, title, body, progress, important,
                values):
    event = {
        "source_project_id": _text(project_id),
        "source_table": source_table,
        "source_id": _text(row["id"]),
        "source_subkey": source_subkey,
        "source_updated_at": _timezone_timestamp(source_updated_at),
        "updated_at": _timezone_timestamp(source_updated_at),
        "date": event_date,
        "kind": kind,
        "title": _text(title),
        "body": _text(body),
        "progress": _text(progress),
        "important": _bool(important),
        "suggested_section": _section(source_table, values),
    }
    event["source_hash"] = _canonical_hash(event)
    return event


def _job_events(row, project_id, wanted):
    out = []
    remarks = [item for item in _json_array(row["remarks"], "jobs.remarks")
               if _business_date(item.get("date")) == wanted]
    prepared = []
    for remark in remarks:
        # There is no child id in the legacy JSON shape.  The business date
        # is stable; same-day duplicates use a content fingerprint, never an
        # array index.
        child = {key: remark[key] for key in sorted(remark) if key != "updated_at"}
        fingerprint = hashlib.sha256(json.dumps(child, ensure_ascii=False, sort_keys=True,
                                                separators=(",", ":")).encode("utf-8")).hexdigest()[:16]
        prepared.append((fingerprint, remark))
    for fingerprint, remark in sorted(prepared, key=lambda item: item[0]):
        progress = remark.get("progress") or remark.get("body") or remark.get("remark")
        # Always include the content fingerprint. If a second same-day remark
        # is inserted later, identities of existing children must not shift.
        subkey = "job:%s:remark:%s:%s" % (row["id"], wanted, fingerprint)
        out.append(_event_base(
            "jobs", row, project_id, subkey, wanted,
            remark.get("updated_at") or row["updated_at"], "job_remark",
            "%s %s" % (_text(row["number"]), _text(row["description"])),
            progress, progress, remark.get("important", False),
            (row["section"], row["category"], row["vendor"], row["description"]),
        ))
    return out


def _discussion_events(row, project_id, wanted):
    out = []
    values = (row["item"], row["description"], row["no"])
    if _business_date(row["date"]) == wanted:
        out.append(_event_base(
            "discussions", row, project_id, "discussion:%s:body:%s" % (row["id"], wanted),
            wanted, row["updated_at"], "discussion", row["item"] or row["no"],
            row["description"], "", row["priority"] in ("Urgent", "Critical"), values,
        ))
    for action in _json_array(row["actions"], "discussions.actions"):
        if _business_date(action.get("date")) != wanted:
            continue
        child = {key: action[key] for key in sorted(action) if key != "updated_at"}
        fingerprint = hashlib.sha256(json.dumps(child, ensure_ascii=False, sort_keys=True,
                                                separators=(",", ":")).encode("utf-8")).hexdigest()[:16]
        progress = action.get("progress") or action.get("body") or action.get("remark")
        out.append(_event_base(
            "discussions", row, project_id,
            "discussion:%s:action:%s:%s" % (row["id"], wanted, fingerprint),
            wanted, action.get("updated_at") or row["updated_at"], "discussion_action",
            row["item"] or row["no"], progress, progress,
            action.get("important", row["priority"] in ("Urgent", "Critical")), values,
        ))
    return out


def _class_events(row, project_id, wanted):
    out = []
    values = (row["finding"], row["description"], row["responsible"], row["priority"])
    for action in _json_array(row["actions"], "class_items.actions"):
        if _business_date(action.get("date")) != wanted:
            continue
        child = {key: action[key] for key in sorted(action) if key != "updated_at"}
        fingerprint = hashlib.sha256(json.dumps(child, ensure_ascii=False, sort_keys=True,
                                                separators=(",", ":")).encode("utf-8")).hexdigest()[:16]
        progress = action.get("progress") or action.get("body") or action.get("remark")
        out.append(_event_base(
            "class_items", row, project_id,
            "class_item:%s:action:%s:%s" % (row["id"], wanted, fingerprint),
            wanted, action.get("updated_at") or row["updated_at"], "class_item_action",
            row["finding"] or row["no"], progress, progress,
            action.get("important", row["priority"] in ("Urgent", "Critical")), values,
        ))
    return out


_TRACKING_DATES = {
    "steel_repair": ("start_date", "completion_date", "last_updated"),
    "pipe_repair": ("start_date", "completion_date", "last_updated"),
    "outfitting": ("start_date", "completion_date", "last_updated"),
    "wbt_cot": ("open_date", "close_date", "bottom_plug_open", "bottom_plug_close", "updated_at"),
    "portable_fan": ("start_date", "stop_date", "updated_at"),
    "staging": ("updated_at",),
    "gas_free": ("date", "updated_at"),
}


def _tracking_events(source_table, row, project_id, wanted):
    date_fields = _TRACKING_DATES[source_table]
    if not any(_business_date(row[field]) == wanted for field in date_fields if field in row.keys()):
        return []
    updated = row["last_updated"] if "last_updated" in row.keys() else row["updated_at"]
    description = row["description"] if "description" in row.keys() else ""
    title = "%s %s" % (_text(row["no"]), _text(description))
    body = _tracking_body(source_table, row)
    important = _text(row["priority"]).lower() in {"urgent", "critical"} if "priority" in row.keys() else False
    values = tuple(row[field] for field in row.keys())
    return [_event_base(
        source_table, row, project_id, "%s:%s:row:%s" % (source_table, row["id"], wanted),
        wanted, updated, "%s_update" % source_table, title, body,
        row["status"] if "status" in row.keys() else "", important, values,
    )]


def _database_path(dd, dd_app):
    candidate = dd_app.config.get("DATABASE")
    if not candidate:
        candidate = getattr(dd, "DATABASE", None) or getattr(dd, "DB_PATH", None)
    if not candidate or candidate == ":memory:":
        return None
    candidate = os.fspath(candidate)
    if candidate.startswith("file:"):
        return candidate
    if not os.path.isabs(candidate):
        candidate = os.path.join(dd_app.instance_path, candidate)
    return candidate


@contextmanager
def _read_connection(dd, dd_app):
    """Open the Dock DB read-only; the fallback exists only for test doubles."""
    path = _database_path(dd, dd_app)
    if path:
        if path.startswith("file:"):
            uri = path + ("&" if "?" in path else "?") + "mode=ro"
        else:
            uri = "file:%s?mode=ro" % path
        conn = sqlite3.connect(uri, uri=True)
        conn.row_factory = sqlite3.Row
        try:
            yield conn
        finally:
            conn.close()
        return
    # A real Dock Manager app has DATABASE.  Test doubles may expose only a
    # connection factory; do not add writes or PRAGMAs to that connection.
    conn = getattr(dd, "_trmt_original_get_db", dd.get_db)()
    try:
        yield conn
    finally:
        close = getattr(conn, "close", None)
        if callable(close):
            close()


def _daily_events_payload(dd, dd_app, project_ids, wanted):
    events = []
    complete_sources = []
    with _read_connection(dd, dd_app) as conn:
        # Check the complete source set before returning anything.  A missing
        # table is an unavailable source, not an empty daily report.
        table_rows = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name IN "
            "(?,?,?,?,?,?,?,?,?,?)", DAILY_EVENT_SOURCES,
        ).fetchall()
        available = {row["name"] for row in table_rows}
        if available != set(DAILY_EVENT_SOURCES):
            missing = [name for name in DAILY_EVENT_SOURCES if name not in available]
            raise RuntimeError("Dock Manager source unavailable: " + ",".join(missing))
        for source_table in DAILY_EVENT_SOURCES:
            allowed = set(project_ids)
            # Keep each identifier literal: source_table is a server-owned
            # allowlist, but this also keeps the read boundary auditable by
            # the repository SQL-construction contract.  Project filtering is
            # intentionally performed after the read; no user value reaches
            # SQL text and the connection is opened mode=ro.
            if source_table == "jobs":
                rows = conn.execute("SELECT * FROM jobs ORDER BY vessel_id, id").fetchall()
            elif source_table == "discussions":
                rows = conn.execute("SELECT * FROM discussions ORDER BY vessel_id, id").fetchall()
            elif source_table == "class_items":
                rows = conn.execute("SELECT * FROM class_items ORDER BY vessel_id, id").fetchall()
            elif source_table == "steel_repair":
                rows = conn.execute("SELECT * FROM steel_repair ORDER BY vessel_id, id").fetchall()
            elif source_table == "pipe_repair":
                rows = conn.execute("SELECT * FROM pipe_repair ORDER BY vessel_id, id").fetchall()
            elif source_table == "outfitting":
                rows = conn.execute("SELECT * FROM outfitting ORDER BY vessel_id, id").fetchall()
            elif source_table == "wbt_cot":
                rows = conn.execute("SELECT * FROM wbt_cot ORDER BY vessel_id, id").fetchall()
            elif source_table == "portable_fan":
                rows = conn.execute("SELECT * FROM portable_fan ORDER BY vessel_id, id").fetchall()
            elif source_table == "staging":
                rows = conn.execute("SELECT * FROM staging ORDER BY vessel_id, id").fetchall()
            else:
                rows = conn.execute("SELECT * FROM gas_free ORDER BY vessel_id, id").fetchall()
            for row in rows:
                project_id = row["vessel_id"]
                if project_id not in allowed:
                    continue
                if source_table == "jobs":
                    events.extend(_job_events(row, project_id, wanted))
                elif source_table == "discussions":
                    events.extend(_discussion_events(row, project_id, wanted))
                elif source_table == "class_items":
                    events.extend(_class_events(row, project_id, wanted))
                else:
                    events.extend(_tracking_events(source_table, row, project_id, wanted))
            complete_sources.append(source_table)
    events.sort(key=lambda item: (item["source_table"], item["source_id"], item["source_subkey"]))
    return {
        "complete": True,
        "partial": False,
        "complete_sources": complete_sources,
        "requested_project_ids": project_ids,
        "date": wanted,
        "events": events,
    }


def _install_daily_events_endpoint(dd, trmt_app, dd_app):
    """Install the one read-only integration route on the mounted app."""
    from flask import jsonify, request

    def check_trmt_api_key():
        # app.py exposes the helper in its module namespace in production;
        # importing the shared symbol is the compatibility path for the Flask
        # object passed by wsgi.py and for lightweight test apps.
        checker = getattr(trmt_app, "_check_api_key", None)
        if not callable(checker):
            from helpers_shared import _check_api_key as checker
        return bool(checker())

    dd_app._trmt_check_daily_api_key = check_trmt_api_key

    if getattr(dd_app, "_trmt_daily_events_installed", False):
        return

    @dd_app.route(DAILY_EVENTS_PATH, methods=["GET"])
    def _trmt_daily_events():
        raw_projects = request.args.get("project_ids", "")
        project_ids = []
        try:
            decoded = json.loads(raw_projects)
            raw_values = decoded if isinstance(decoded, list) else raw_projects.split(",")
        except (TypeError, ValueError):
            raw_values = raw_projects.split(",")
        for value in raw_values:
            value = _text(value)
            if value and value not in project_ids:
                if not re.fullmatch(r"v_[A-Za-z0-9][A-Za-z0-9_.:-]*", value):
                    return jsonify({"error": "project_ids must contain Dock Manager v_ ids"}), 400
                project_ids.append(value)
        wanted = request.args.get("date", "")
        try:
            parsed_date = _date.fromisoformat(wanted)
        except (TypeError, ValueError):
            parsed_date = None
        if not project_ids or not parsed_date or parsed_date.isoformat() != wanted or not _DATE_RE.match(wanted):
            return jsonify({"error": "project_ids and date=YYYY-MM-DD are required"}), 400
        try:
            return jsonify(_daily_events_payload(dd, dd_app, project_ids, wanted))
        except (OSError, sqlite3.Error, ValueError, RuntimeError) as exc:
            # No partial events are ever returned.  Keep source details out of
            # the response because this endpoint is also reachable by a key.
            dd_app.logger.warning("daily-events unavailable: %s", exc)
            return jsonify({"error": "daily event sources unavailable", "complete": False}), 503

    dd_app._trmt_daily_events_installed = True


def _install_roster_vessels_endpoints(dd, trmt_app, dd_app):
    """Bridge the signed-in supervisor's canonical TRMT roster into Dock Manager."""
    from flask import jsonify, request, session

    if getattr(dd_app, "_trmt_roster_vessels_installed", False):
        return

    def _trmt_db():
        path = trmt_app.config.get("DATABASE")
        if not path:
            raise RuntimeError("TRMT database is not configured")
        db = sqlite3.connect("file:%s?mode=ro" % os.path.abspath(path), uri=True)
        db.row_factory = sqlite3.Row
        return db

    def _assigned_vessel(vessel_id=None):
        supervisor_id = session.get("supervisor_id")
        if supervisor_id is None:
            return [] if vessel_id is None else None
        params = [supervisor_id]
        where = "sv.supervisor_id=? AND v.active=1"
        if vessel_id is not None:
            where += " AND v.id=?"
            params.append(vessel_id)
        sql = """
            SELECT v.id, v.name, v.vessel_type, v.imo, v.class_society,
                   COALESCE(NULLIF(TRIM(v.gross_tonnage), ''),
                     (SELECT d.gross_tonnage FROM dock_reports d
                       WHERE d.vessel_id=v.id AND TRIM(COALESCE(d.gross_tonnage,''))<>''
                       ORDER BY d.updated_at DESC, d.id DESC LIMIT 1)) AS gross_tonnage,
                   COALESCE(NULLIF(TRIM(v.dead_weight), ''),
                     (SELECT d.dead_weight FROM dock_reports d
                       WHERE d.vessel_id=v.id AND TRIM(COALESCE(d.dead_weight,''))<>''
                       ORDER BY d.updated_at DESC, d.id DESC LIMIT 1)) AS dead_weight
              FROM vessels v
              JOIN supervisor_vessels sv ON sv.vessel_id=v.id
             WHERE %s
             ORDER BY v.name
        """ % where
        db = _trmt_db()
        try:
            result = db.execute(sql, params).fetchall()
        finally:
            db.close()
        if vessel_id is not None:
            return result[0] if result else None
        return result

    def _grt_dwt(vessel):
        parts = []
        if _text(vessel["gross_tonnage"]):
            parts.append("GRT %s" % _text(vessel["gross_tonnage"]))
        if _text(vessel["dead_weight"]):
            parts.append("DWT %s" % _text(vessel["dead_weight"]))
        return " / ".join(parts)

    def _payload(vessel):
        return {
            "id": vessel["id"],
            "name": _text(vessel["name"]),
            "type": _text(vessel["vessel_type"]),
            "imo": _text(vessel["imo"]),
            "classSociety": _text(vessel["class_society"]),
            "grtDwt": _grt_dwt(vessel),
            "grossTonnage": _text(vessel["gross_tonnage"]),
            "deadWeight": _text(vessel["dead_weight"]),
        }

    @dd_app.route(ROSTER_VESSELS_PATH, methods=["GET"])
    def _trmt_roster_vessels():
        return jsonify([_payload(vessel) for vessel in _assigned_vessel()])

    @dd_app.route(ROSTER_VESSELS_PATH, methods=["POST"])
    def _trmt_create_roster_vessel():
        body = request.get_json(silent=True) or {}
        try:
            vessel_id = int(body.get("trmtVesselId"))
        except (TypeError, ValueError):
            return jsonify({"error": "담당선박을 선택하세요"}), 400
        vessel = _assigned_vessel(vessel_id)
        if vessel is None:
            return jsonify({"error": "선택한 선박은 내 TRMT 담당선박이 아닙니다"}), 403

        canonical = _payload(vessel)
        dock_id = "v_%s" % secrets.token_hex(4)
        db = dd.get_db()
        db.execute(
            "INSERT INTO vessels(id,name,type,imo,shipyard,class_society,berthing_date,"
            "dock_in,dock_out,departure_date,duration,grt) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
            (dock_id, canonical["name"], canonical["type"], canonical["imo"],
             _text(body.get("shipyard")), canonical["classSociety"],
             body.get("berthingDate") or None, body.get("dockIn") or None,
             body.get("dockOut") or None, body.get("departureDate") or None,
             body.get("duration") or None, canonical["grtDwt"]),
        )
        db.commit()
        return jsonify(dd.to_vessel(dd.row("SELECT * FROM vessels WHERE id=?", dock_id))), 201

    dd_app._trmt_roster_vessels_installed = True


def _yard_cell_text(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return format(value, ".15g")
    return re.sub(r"\s+", " ", str(value)).strip()


def _svms_number(value):
    try:
        number = float(str(value).replace(",", ""))
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number):
        return None
    return round(number, 2)


def _svms_job_category(subject, source_key):
    """Map an SVMS Dock sub-tab row to the Dock Manager category/section."""
    if source_key == "P_RS_DP":
        return "Paint", "PAINT"
    if source_key == "P_RS_SR":
        return "Shore Repair", "SHORE REPAIR"
    upper = subject.upper()
    local_tail = upper.split("LOCAL SUPPLY", 1)[1] if "LOCAL SUPPLY" in upper else ""
    if re.search(r"\bSTORES?\b", local_tail):
        return "Store", "STORE"
    if re.search(r"\bSPARE\b", local_tail):
        return "Spare", "SPARE"
    tag, _description = _svms_subject_parts(subject)
    tag = tag.upper()
    if re.fullmatch(r"ST\d+[A-Z0-9-]*", tag):
        return "Store", "STORE"
    # [S1-A]/[S2]/[S3] and every unclassified Spare/Store row default to Spare.
    return "Spare", "SPARE"


def _svms_subject_parts(subject):
    bracketed = _SVMS_JOB_TAG.match(subject)
    if bracketed:
        return _text(bracketed.group(1)), _text(bracketed.group(2))
    bare = _SVMS_JOB_BARE_TAG.match(subject)
    if bare:
        return _text(bare.group(1)), _text(bare.group(2))
    return "", _text(subject)


def _svms_job_budget(row, warnings, label):
    """Use SVMS's own USD conversion, never a locally invented exchange rate."""
    amount = _svms_number(row.get("AMT"))
    usd = _svms_number(row.get("AMT_USD"))
    currency = _text(row.get("CUR_CD")).upper()
    if amount is None and usd is None:
        amount = _svms_number(row.get("R_AMT"))
        usd = _svms_number(row.get("R_AMT_USD"))
        currency = _text(row.get("R_CUR_CD")).upper()
    if amount is None:
        if usd is not None:
            return usd
        return 0.0
    if usd is not None:
        return usd
    if not currency or currency == "USD":
        return amount
    warnings.append("%s: %s 금액에 SVMS USD 환산값이 없어 Budget 0으로 보류" % (label, currency))
    return 0.0


def _svms_job_identity(job):
    """Stable re-import identity.

    Canonical P/S/ST/R numbers are globally identified by Job No. because their prefix
    encodes the source. Non-canonical bracket tags keep Category in the key so an old
    manual/foreign tag cannot absorb an unrelated SVMS row. Numberless rows need
    Category + title.
    """
    def value(key):
        try:
            return job[key]
        except (KeyError, IndexError, TypeError):
            return ""

    number = _text(value("number")).upper()
    if _SVMS_JOB_CANONICAL_NUMBER.fullmatch(number):
        return ("number", number)
    if number:
        return ("category_number", _text(value("category")).lower(), number)
    return ("title", _text(value("category")).lower(),
            _text(value("description")).upper())


def _normalize_svms_dock_jobs(payload):
    """Minimal SP_GET_DOCK_INFO payload -> deterministic Dock Manager preview."""
    if not isinstance(payload, dict):
        raise ValueError("SVMS 응답 형식이 올바르지 않습니다")
    jobs = []
    seen = set()
    warnings = []
    for source_key in ("P_RS_DP", "P_RS_SS", "P_RS_SR"):
        source_rows = payload.get(source_key) or []
        if not isinstance(source_rows, list):
            raise ValueError("%s 응답 형식이 올바르지 않습니다" % source_key)
        for row in source_rows:
            if not isinstance(row, dict):
                continue
            subject = _text(row.get("SUBJ"))
            number, description = _svms_subject_parts(subject)
            if not description:
                description = subject or "SVMS Dock Job"
            if not number:
                warnings.append("%s: Job No. 없음 → 제목 기준으로 재가져오기 매칭" % description)
            category, section = _svms_job_category(subject, source_key)
            label = number or description
            candidate = {"number": number, "category": category, "description": description}
            identity = _svms_job_identity(candidate)
            if identity in seen:
                warnings.append("%s: 같은 Job 식별자가 중복되어 첫 행만 사용" % label)
                continue
            seen.add(identity)
            jobs.append({
                "number": number,
                "section": section,
                "category": category,
                "description": description,
                "vendor": _text(row.get("VNDR_NM")),
                "budget": _svms_job_budget(row, warnings, label),
            })
    return {
        "jobs": jobs,
        "job_count": len(jobs),
        "total_budget": round(sum(job["budget"] for job in jobs), 2),
        "counts": {category: sum(job["category"] == category for job in jobs)
                   for category in ("Paint", "Spare", "Store", "Shore Repair")},
        "warnings": list(dict.fromkeys(warnings)),
    }


def _apply_svms_job_import(db, vessel_id, parsed):
    if not db.execute("SELECT 1 FROM vessels WHERE id=?", (vessel_id,)).fetchone():
        raise ValueError("선박을 찾을 수 없습니다")
    existing = {}
    for row in db.execute("SELECT * FROM jobs WHERE vessel_id=? ORDER BY id", (vessel_id,)).fetchall():
        key = _svms_job_identity(row)
        existing.setdefault(key, row)
    inserted = updated = unchanged = 0
    for job in parsed["jobs"]:
        key = _svms_job_identity(job)
        current = existing.get(key)
        values = (job["section"], job["category"], job["description"],
                  job["vendor"], job["budget"])
        if current is None:
            cursor = db.execute(
                "INSERT INTO jobs(vessel_id,number,section,category,description,vendor,budget,"
                "consumption,start_date,end_date,completion,remarks) VALUES(?,?,?,?,?,?,?,0,NULL,NULL,0,'[]')",
                (vessel_id, job["number"]) + values,
            )
            inserted += 1
            existing[key] = db.execute("SELECT * FROM jobs WHERE id=?", (cursor.lastrowid,)).fetchone()
            continue
        incoming_budget = round(float(job["budget"] or 0), 2)
        current_budget = round(float(current["budget"] or 0), 2)
        if current_budget == incoming_budget:
            unchanged += 1
            continue
        # Existing Job identity is authoritative in TRMT. A re-import may only refresh
        # the SVMS USD Budget; preserve title/vendor/classification and every live field.
        db.execute("UPDATE jobs SET budget=?,updated_at=datetime('now') WHERE id=?",
                   (incoming_budget, current["id"]))
        updated += 1
    return {"inserted": inserted, "updated": updated, "unchanged": unchanged,
            "job_count": parsed["job_count"], "total_budget": parsed["total_budget"]}


def _yard_number(value):
    raw = _yard_cell_text(value).replace(" ", "")
    return raw if _YARD_ITEM_NO.fullmatch(raw) else ""


def _yard_numeric_after(row, start):
    for value in row[start + 1:]:
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return float(value)
    return None


def _yard_find_sheet_and_columns(workbook):
    aliases = {
        "item": ("itemno", "itemnumber"),
        "description": ("workdescription", "description"),
        "net_total": ("nettotal", "netamount"),
    }
    preferred = sorted(workbook.worksheets, key=lambda ws: ws.title.lower() != "quotation")
    for sheet in preferred:
        for row_no, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(30, sheet.max_row), values_only=True), 1):
            normalized = [re.sub(r"[^a-z]", "", _yard_cell_text(value).lower()) for value in row]
            found = {}
            for key, names in aliases.items():
                found[key] = next((idx for idx, value in enumerate(normalized) if value in names), None)
            if all(value is not None for value in found.values()):
                return sheet, row_no, found
    raise ValueError("Item No. / Work Description / Net Total 헤더를 찾지 못했습니다")


def _yard_scan_quote_totals(workbook, quote_sheet):
    discount = gross = net = None
    for sheet in (quote_sheet,):
        for row in sheet.iter_rows(values_only=True):
            for idx, value in enumerate(row):
                label = _yard_cell_text(value).lower()
                if not label:
                    continue
                number = _yard_numeric_after(row, idx)
                if number is None:
                    continue
                if "final discount" in label:
                    discount = number * 100 if 0 <= number <= 1 else number
                elif "normal total" in label:
                    gross = number
                elif "total price after" in label or "after discount" in label or "after dicount" in label:
                    net = number
    if discount is None:
        for sheet in workbook.worksheets:
            if sheet is quote_sheet:
                continue
            for row in sheet.iter_rows(values_only=True):
                for idx, value in enumerate(row):
                    if "final discount" not in _yard_cell_text(value).lower():
                        continue
                    number = _yard_numeric_after(row, idx)
                    if number is not None:
                        discount = number * 100 if 0 <= number <= 1 else number
                        break
                if discount is not None:
                    break
            if discount is not None:
                break
    return discount, gross, net


def _parse_yard_job_workbook(raw_bytes):
    """YiuLian-style quotation workbook -> Dock Manager Job Progress preview.

    Every numbered line remains visible as a Job, matching the Kuwait project.
    Money below a third-level line rolls up to its second-level parent; a
    stand-alone top-level job (for example item 24) owns its own amount.
    """
    try:
        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError("엑셀 파일을 읽지 못했습니다") from exc

    sheet, header_row, columns = _yard_find_sheet_and_columns(workbook)
    source_rows = list(sheet.iter_rows(min_row=header_row + 1, values_only=True))
    numbered = []
    roots_with_children = set()
    for row in source_rows:
        number = _yard_number(row[columns["item"]] if columns["item"] < len(row) else None)
        if number:
            numbered.append(number)
            if "." in number:
                roots_with_children.add(number.split(".", 1)[0])
    if not numbered:
        raise ValueError("파싱 가능한 Job 번호가 없습니다")

    jobs = []
    jobs_by_number = {}
    budget_owner = None
    warnings = []
    for row in source_rows:
        number = _yard_number(row[columns["item"]] if columns["item"] < len(row) else None)
        description = _yard_cell_text(
            row[columns["description"]] if columns["description"] < len(row) else None
        )
        net_value = row[columns["net_total"]] if columns["net_total"] < len(row) else None
        if number:
            if number in jobs_by_number:
                raise ValueError("중복 Job 번호가 있습니다: %s" % number)
            root = number.split(".", 1)[0]
            has_direct_price = (isinstance(net_value, (int, float))
                                and not isinstance(net_value, bool) and bool(net_value))
            if _yard_is_quote_request(description) and not has_direct_price:
                # A third-level quotation instruction belongs to its visible
                # second-level repair item but must not become a Job itself.
                # Preserve that parent as the budget owner for following
                # unnumbered price rows.
                if number.count(".") < 2:
                    budget_owner = None
                continue
            section = YARD_JOB_SECTION_MAP.get(root, "ADD")
            if section == "ADD" and root not in {"1"}:
                warnings.append("미지정 섹션 %s → ADD로 분류" % root)
            job = {
                "number": number,
                "section": section,
                "category": "Shipyard",
                "description": description or ("Item %s" % number),
                "budget": 0.0,
            }
            jobs.append(job)
            jobs_by_number[number] = job
            depth = number.count(".") + 1
            if depth == 1:
                budget_owner = None if root in roots_with_children else job
            elif depth == 2:
                budget_owner = job
            elif number.rsplit(".", 1)[0] not in jobs_by_number:
                # If a filtered quotation-only row was the second-level
                # parent, retain a real priced child as its own budget owner.
                budget_owner = job
            # Third-level headings stay visible but their money belongs to the
            # second-level repair item, as in Kuwait Prosperity.

        row_text = " ".join(_yard_cell_text(value) for value in row if isinstance(value, str))
        if (budget_owner is not None and isinstance(net_value, (int, float))
                and not isinstance(net_value, bool) and net_value
                and not _YARD_TOTAL_ROW.search(row_text)):
            budget_owner["budget"] += float(net_value)

    for job in jobs:
        job["budget"] = round(job["budget"], 2)
    gross_total = round(sum(job["budget"] for job in jobs), 2)
    discount, quoted_gross, quoted_net = _yard_scan_quote_totals(workbook, sheet)
    discount = round(discount, 4) if discount is not None and 0 <= discount <= 100 else None
    after_discount = round(gross_total * (1 - (discount or 0) / 100), 2)
    if quoted_gross is not None and abs(gross_total - quoted_gross) > 0.05:
        warnings.append("견적 Gross Total과 파싱 합계가 일치하지 않습니다")
    if quoted_net is not None and discount is not None and abs(after_discount - quoted_net) > 0.05:
        warnings.append("견적 D/C 후 합계와 파싱 합계가 일치하지 않습니다")
    return {
        "sheet": sheet.title,
        "jobs": jobs,
        "job_count": len(jobs),
        "priced_count": sum(bool(job["budget"]) for job in jobs),
        "gross_total": gross_total,
        "discount_rate": discount,
        "after_discount": after_discount,
        "quoted_gross": round(quoted_gross, 2) if quoted_gross is not None else None,
        "quoted_net": round(quoted_net, 2) if quoted_net is not None else None,
        "warnings": list(dict.fromkeys(warnings)),
    }


def _apply_yard_job_import(db, vessel_id, parsed):
    vessel = db.execute("SELECT id FROM vessels WHERE id=?", (vessel_id,)).fetchone()
    if not vessel:
        raise ValueError("선박을 찾을 수 없습니다")
    existing_rows = db.execute(
        "SELECT id,number,category,description,budget FROM jobs WHERE vessel_id=? ORDER BY id",
        (vessel_id,),
    ).fetchall()
    existing = {}
    for row in existing_rows:
        number = row["number"] if isinstance(row, sqlite3.Row) else row[1]
        if number and number not in existing:
            existing[number] = row
    inserted = updated = preserved_manual = unchanged = 0
    for job in parsed["jobs"]:
        current = existing.get(job["number"])
        if current is None:
            db.execute(
                "INSERT INTO jobs(vessel_id,number,section,category,description,vendor,budget,"
                "consumption,start_date,end_date,completion,remarks) VALUES(?,?,?,?,?,'',?,0,NULL,NULL,0,'[]')",
                (vessel_id, job["number"], job["section"], "Shipyard", job["description"], job["budget"]),
            )
            inserted += 1
            continue
        category = current["category"] if isinstance(current, sqlite3.Row) else current[2]
        if category != "Shipyard":
            preserved_manual += 1
            continue
        old_description = current["description"] if isinstance(current, sqlite3.Row) else current[3]
        old_budget = current["budget"] if isinstance(current, sqlite3.Row) else current[4]
        description = old_description or job["description"]
        if abs(float(old_budget or 0) - job["budget"]) <= 0.005 and description == old_description:
            unchanged += 1
            continue
        # Deliberately preserve section/vendor/consumption/dates/progress/remarks:
        # those fields may contain an owner's cancellation or live progress.
        db.execute("UPDATE jobs SET description=?,budget=?,updated_at=datetime('now') WHERE id=?",
                   (description, job["budget"], current["id"] if isinstance(current, sqlite3.Row) else current[0]))
        updated += 1
    if parsed.get("discount_rate") is not None:
        db.execute("UPDATE vessels SET dc_rate=? WHERE id=?", (parsed["discount_rate"], vessel_id))
    return {
        "inserted": inserted,
        "updated": updated,
        "preserved_manual": preserved_manual,
        "unchanged": unchanged,
        "discount_rate": parsed.get("discount_rate"),
    }


def _install_yard_job_import(dd, dd_app):
    if getattr(dd_app, "_trmt_yard_import_installed", False):
        return
    from flask import jsonify, request

    cache = {}
    cache_lock = threading.Lock()

    def prune(now):
        expired = [token for token, entry in cache.items() if now - entry["created"] > _YARD_IMPORT_TTL]
        for token in expired:
            cache.pop(token, None)
        while len(cache) >= _YARD_IMPORT_MAX:
            oldest = min(cache, key=lambda token: cache[token]["created"])
            cache.pop(oldest, None)

    @dd_app.route("/api/vessels/<vid>/jobs/xlsx/preview", methods=["POST"])
    def _trmt_yard_xlsx_preview(vid):
        upload = request.files.get("file")
        filename = (upload.filename or "") if upload else ""
        if not upload or not filename.lower().endswith((".xlsx", ".xlsm")):
            return jsonify({"error": ".xlsx 또는 .xlsm 조선소 견적서가 필요합니다"}), 400
        raw = upload.read()
        if not raw.startswith(b"PK\x03\x04"):
            return jsonify({"error": "올바른 Excel OOXML 파일이 아닙니다"}), 400
        try:
            parsed = _parse_yard_job_workbook(raw)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        db = dd.get_db()
        if not db.execute("SELECT 1 FROM vessels WHERE id=?", (vid,)).fetchone():
            return jsonify({"error": "선박을 찾을 수 없습니다"}), 404
        token = secrets.token_urlsafe(24)
        now = time.time()
        with cache_lock:
            prune(now)
            cache[token] = {"created": now, "vessel_id": vid, "parsed": parsed}
        preview = dict(parsed)
        preview["jobs"] = parsed["jobs"][:12]
        preview["preview_token"] = token
        return jsonify(preview)

    @dd_app.route("/api/vessels/<vid>/jobs/xlsx/apply", methods=["POST"])
    def _trmt_yard_xlsx_apply(vid):
        token = _yard_cell_text((request.get_json(silent=True) or {}).get("preview_token"))
        now = time.time()
        with cache_lock:
            prune(now)
            entry = cache.pop(token, None)
        if not entry or entry["vessel_id"] != vid:
            return jsonify({"error": "미리보기가 만료됐습니다. 견적서를 다시 선택하세요"}), 409
        db = dd.get_db()
        try:
            result = _apply_yard_job_import(db, vid, entry["parsed"])
            db.commit()
        except (sqlite3.Error, ValueError) as exc:
            db.rollback()
            dd_app.logger.warning("yard job import failed: %s", exc)
            return jsonify({"error": "견적서 반영에 실패했습니다"}), 500
        result["job_count"] = entry["parsed"]["job_count"]
        result["gross_total"] = entry["parsed"]["gross_total"]
        result["after_discount"] = entry["parsed"]["after_discount"]
        return jsonify(result)

    @dd_app.after_request
    def _trmt_yard_import_asset(response):
        if (request.method == "GET" and request.path == "/" and response.status_code == 200
                and response.mimetype == "text/html" and not response.headers.get("Content-Encoding")):
            body = response.get_data(as_text=True)
            asset = '<script src="/static/js/drydock-yard-import.js?v=20260906-1"></script>'
            if asset not in body and "</body>" in body:
                response.set_data(body.replace("</body>", asset + "</body>"))
                response.headers["Content-Length"] = len(response.get_data())
        return response

    dd_app._trmt_yard_import_installed = True


def _install_job_progress_export(dd, dd_app):
    if getattr(dd_app, "_trmt_job_progress_export_installed", False):
        return
    from flask import jsonify, send_file

    @dd_app.route("/api/vessels/<vid>/jobs/progress.xlsx", methods=["GET"])
    def _trmt_job_progress_export(vid):
        if not dd.is_admin():
            return jsonify({"error": "Unauthorized (admin only)"}), 401
        db = dd.get_db()
        vessel = db.execute(
            "SELECT id,name,dock_in FROM vessels WHERE id=?", (vid,)
        ).fetchone()
        if not vessel:
            return jsonify({"error": "선박을 찾을 수 없습니다"}), 404
        jobs = db.execute(
            "SELECT number,category,description,vendor,budget,consumption,start_date,end_date,"
            "completion,remarks FROM jobs WHERE vessel_id=? ORDER BY id", (vid,)
        ).fetchall()
        try:
            output = _build_job_progress_workbook(vessel, jobs)
        except ValueError as exc:
            dd_app.logger.warning("job progress export failed vessel=%s: %s", vid, exc)
            return jsonify({"error": str(exc)}), 409
        except (OSError, RuntimeError) as exc:
            dd_app.logger.warning("job progress export failed vessel=%s: %s", vid, exc)
            return jsonify({"error": "Job Progress 엑셀 생성에 실패했습니다"}), 500
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=_job_progress_download_name(vessel["name"]),
        )

    dd_app._trmt_job_progress_export_installed = True


def _pending_svms_import_request(db, vessel_id):
    return db.execute(
        "SELECT token,status FROM trmt_svms_job_import WHERE vessel_id=? "
        "AND status='pending' ORDER BY requested_at DESC LIMIT 1",
        (vessel_id,),
    ).fetchone()


def _install_svms_job_import(dd, trmt_app, dd_app):
    """Queue a live read on the credential-holding Mac, then preview/apply it here."""
    if getattr(dd_app, "_trmt_svms_job_import_installed", False):
        return
    from flask import jsonify, request

    def _ensure_table(db):
        try:
            db.execute("SELECT 1 FROM trmt_svms_job_import LIMIT 0")
            return
        except sqlite3.OperationalError as exc:
            if "no such table" not in str(exc).lower():
                raise
        db.execute("""CREATE TABLE IF NOT EXISTS trmt_svms_job_import (
            token TEXT PRIMARY KEY,
            vessel_id TEXT NOT NULL,
            vsl_cd TEXT NOT NULL,
            dk_cd TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'pending',
            payload_json TEXT,
            error TEXT,
            requested_at TEXT NOT NULL DEFAULT (datetime('now')),
            completed_at TEXT,
            applied_at TEXT
        )""")
        db.execute("CREATE INDEX IF NOT EXISTS idx_trmt_svms_job_import_status "
                   "ON trmt_svms_job_import(status,requested_at)")
        db.commit()

    # apply() runs while the parent TRMT app is importing, outside the mounted
    # Dock Manager app context. Enter the owning app explicitly for its g-bound DB.
    with dd_app.app_context():
        _ensure_table(dd.get_db())

    def _trmt_db():
        path = trmt_app.config.get("DATABASE")
        if not path:
            raise RuntimeError("TRMT database is not configured")
        conn = sqlite3.connect("file:%s?mode=ro" % os.path.abspath(path), uri=True)
        conn.row_factory = sqlite3.Row
        return conn

    def _queue_db():
        """Open the Dock DB explicitly for API-key worker routes.

        The parent TRMT API-key checker uses Flask ``g.db`` too. On the mounted
        app it can therefore populate that shared slot with the TRMT database
        before our route runs. Worker routes must never trust the shared slot.
        """
        conn = sqlite3.connect(os.path.abspath(dd_app.config["DATABASE"]), timeout=5)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA busy_timeout=5000")
        return conn

    def _target(vid):
        vessel = dd.get_db().execute("SELECT name,imo FROM vessels WHERE id=?", (vid,)).fetchone()
        if not vessel:
            return None
        conn = _trmt_db()
        try:
            return conn.execute("""
                SELECT v.vsl_cd, dpv.dk_cd
                  FROM vessels v
                  JOIN dock_procure_vessel dpv
                    ON UPPER(TRIM(dpv.vsl_nm))=UPPER(TRIM(v.name))
                 WHERE v.active=1 AND TRIM(COALESCE(v.vsl_cd,''))<>''
                   AND TRIM(COALESCE(dpv.dk_cd,''))<>''
                   AND (UPPER(TRIM(v.name))=UPPER(TRIM(?))
                        OR (TRIM(COALESCE(?,''))<>'' AND TRIM(v.imo)=TRIM(?)))
                 ORDER BY dpv.updated_at DESC LIMIT 1
            """, (_text(vessel["name"]), _text(vessel["imo"]), _text(vessel["imo"]))).fetchone()
        finally:
            conn.close()

    @dd_app.route("/api/vessels/<vid>/jobs/svms-import/request", methods=["POST"])
    def _trmt_svms_import_request(vid):
        _ensure_table(dd.get_db())
        target = _target(vid)
        if target is None:
            return jsonify({"error": "이 선박에 연결된 SVMS 입거수리 Draft(Dock No)가 없습니다"}), 409
        now = time.time()
        dd.get_db().execute(
            "UPDATE trmt_svms_job_import SET status='failed',error='가져오기 요청 만료',"
            "completed_at=datetime('now') WHERE vessel_id=? AND status='pending' "
            "AND requested_at < datetime('now', ?)",
            (vid, "-%d seconds" % _SVMS_IMPORT_TTL_SECONDS),
        )
        dd.get_db().commit()
        # Coalesce only an in-flight read. A ready preview is deliberately not
        # reused: the user can update the SVMS draft and immediately press the
        # import button again, in which case a fresh Mac-side read is required.
        # The old ready token remains independently applicable until its TTL.
        recent = _pending_svms_import_request(dd.get_db(), vid)
        if recent:
            return jsonify({"preview_token": recent["token"], "status": recent["status"]}), 202
        token = secrets.token_urlsafe(24)
        dd.get_db().execute(
            "INSERT INTO trmt_svms_job_import(token,vessel_id,vsl_cd,dk_cd,status,requested_at) "
            "VALUES(?,?,?,?, 'pending', datetime('now'))",
            (token, vid, target["vsl_cd"], target["dk_cd"]),
        )
        dd.get_db().commit()
        return jsonify({"preview_token": token, "status": "pending", "requested_at": now}), 202

    @dd_app.route("/api/vessels/<vid>/jobs/svms-import/status/<token>", methods=["GET"])
    def _trmt_svms_import_status(vid, token):
        _ensure_table(dd.get_db())
        row = dd.get_db().execute(
            "SELECT status,payload_json,error,completed_at FROM trmt_svms_job_import "
            "WHERE token=? AND vessel_id=?", (token, vid),
        ).fetchone()
        if not row:
            return jsonify({"error": "가져오기 요청을 찾을 수 없습니다"}), 404
        result = {"status": row["status"], "preview_token": token,
                  "error": row["error"], "completed_at": row["completed_at"]}
        if row["status"] == "ready":
            result.update(json.loads(row["payload_json"] or "{}"))
            result["jobs"] = (result.get("jobs") or [])[:20]
        return jsonify(result)

    @dd_app.route("/api/vessels/<vid>/jobs/svms-import/apply", methods=["POST"])
    def _trmt_svms_import_apply(vid):
        _ensure_table(dd.get_db())
        token = _text((request.get_json(silent=True) or {}).get("preview_token"))
        row = dd.get_db().execute(
            "SELECT status,payload_json FROM trmt_svms_job_import WHERE token=? AND vessel_id=? "
            "AND completed_at >= datetime('now', ?)",
            (token, vid, "-%d seconds" % _SVMS_IMPORT_TTL_SECONDS),
        ).fetchone()
        if not row or row["status"] != "ready":
            return jsonify({"error": "적용 가능한 SVMS 미리보기가 없습니다"}), 409
        try:
            parsed = json.loads(row["payload_json"] or "{}")
            result = _apply_svms_job_import(dd.get_db(), vid, parsed)
            dd.get_db().execute("UPDATE trmt_svms_job_import SET status='applied',"
                                "applied_at=datetime('now') WHERE token=? AND status='ready'", (token,))
            dd.get_db().commit()
        except (ValueError, sqlite3.Error, TypeError) as exc:
            dd.get_db().rollback()
            dd_app.logger.warning("SVMS job import apply failed: %s", exc)
            return jsonify({"error": "SVMS Job 반영에 실패했습니다"}), 500
        return jsonify(result)

    @dd_app.route(SVMS_IMPORT_PENDING_PATH, methods=["GET"])
    def _trmt_svms_import_pending():
        db = _queue_db()
        try:
            _ensure_table(db)
            row = db.execute(
                "SELECT token,vsl_cd,dk_cd FROM trmt_svms_job_import WHERE status='pending' "
                "ORDER BY requested_at LIMIT 1"
            ).fetchone()
        finally:
            db.close()
        return jsonify(dict(row) if row else {})

    @dd_app.route(SVMS_IMPORT_COMPLETE_PATH, methods=["POST"])
    def _trmt_svms_import_complete():
        body = request.get_json(silent=True) or {}
        token = _text(body.get("token"))
        db = _queue_db()
        try:
            _ensure_table(db)
            row = db.execute(
                "SELECT status FROM trmt_svms_job_import WHERE token=?", (token,),
            ).fetchone()
            if not row or row["status"] != "pending":
                return jsonify({"error": "pending 요청을 찾을 수 없습니다"}), 409
            if body.get("error"):
                db.execute(
                    "UPDATE trmt_svms_job_import SET status='failed',error=?,completed_at=datetime('now') "
                    "WHERE token=? AND status='pending'", (_text(body.get("error"))[:300], token),
                )
            else:
                try:
                    parsed = _normalize_svms_dock_jobs(body.get("payload"))
                except ValueError as exc:
                    return jsonify({"error": str(exc)}), 400
                db.execute(
                    "UPDATE trmt_svms_job_import SET status='ready',payload_json=?,error=NULL,"
                    "completed_at=datetime('now') WHERE token=? AND status='pending'",
                    (json.dumps(parsed, ensure_ascii=False, separators=(",", ":")), token),
                )
            db.commit()
        finally:
            db.close()
        return jsonify({"ok": True})

    dd_app._trmt_svms_job_import_installed = True


def apply(dd, trmt_app):
    """dd = import된 drydock app.py 모듈, trmt_app = TRMT Flask app.
    반환 = 통합 적용된 dd.app (서브마운트용)."""
    from flask import session, request, jsonify, redirect

    dd_app = dd.app

    # ── 0. 템플릿 자동리로드 (drydock UI 패치 후 gunicorn 재시작 없이 반영) ──
    dd_app.config["TEMPLATES_AUTO_RELOAD"] = True
    dd_app.jinja_env.auto_reload = True

    # ── 1. SECRET_KEY 공유 (결정2: 하드코딩 fallback 제거 대체) ──────────
    # TRMT의 서명키를 그대로 써야 세션 쿠키 상호 인식(SSO). 값 불일치 시 즉시 실패.
    if not trmt_app.secret_key:
        raise RuntimeError("TRMT secret_key 미설정 — SSO 불가, 기동 거부")
    dd_app.secret_key = trmt_app.secret_key

    def _live_trmt_admin():
        """Resolve Dock Manager authorization from the TRMT DB, not stale cookie role."""
        username = session.get("username")
        if not username:
            return False
        database = trmt_app.config.get('DATABASE')
        # Minimal integration test doubles do not own a TRMT database. Production
        # always does; preserve the old gate only for those isolated harnesses.
        try:
            if not database:
                return bool(trmt_app.config.get('TESTING') and session.get('role') == 'admin')
            db = sqlite3.connect(f"file:{database}?mode=ro", uri=True)
            db.row_factory = sqlite3.Row
            account = db.execute(
                "SELECT role, active, app_scope FROM users WHERE username=?", (username,)
            ).fetchone()
        except (OSError, sqlite3.Error):
            return False
        finally:
            if 'db' in locals(): db.close()
        if not account or account['active'] != 1 or account['role'] != 'admin':
            return False
        if session.get('role') != account['role']:
            session['role'] = account['role']
        return account['app_scope'] != 'family'

    # ── 2. get_current_user: 세션 신뢰로 패치 (결정2-b, C1 해소) ─────────
    #    drydock users 테이블 조회 X. TRMT 세션(username/role)만 신뢰.
    def _session_user():
        if not session.get("username"):
            return None
        return {
            "username": session.get("username"),
            "role": session.get("role", "editor"),
            "vessels": "[]",          # admin은 전 선박 허용이라 무관
        }
    dd.get_current_user = _session_user
    dd.is_admin = lambda: (_session_user() or {}).get("role") == "admin"
    _install_daily_events_endpoint(dd, trmt_app, dd_app)
    _install_roster_vessels_endpoints(dd, trmt_app, dd_app)
    _install_yard_job_import(dd, dd_app)
    _install_job_progress_export(dd, dd_app)
    _install_svms_job_import(dd, trmt_app, dd_app)
    # is_viewer/can_access_vessel도 세션 유저 기준으로 동작(원본이 get_current_user 참조)

    # ── 3. 전 route admin 게이팅 + login/logout 차단 (결정2, C2·C3 해소) ─
    @dd_app.before_request
    def _sso_admin_gate():
        # The read-only daily adapter is the sole API-key exception.  Keep the
        # comparison exact (path and method) so a trailing-slash, case, or
        # neighbouring Dock Manager API cannot inherit TRMT key access.
        if ((request.path == DAILY_EVENTS_PATH and request.method == "GET")
                or (request.path == SVMS_IMPORT_PENDING_PATH and request.method == "GET")
                or (request.path == SVMS_IMPORT_COMPLETE_PATH and request.method == "POST")):
            if not dd_app._trmt_check_daily_api_key():
                return jsonify({"error": "Unauthorized"}), 401
            return None
        # 경로 정규화(올마이트: 대소문자·중복슬래시·trailing 우회 차단)
        raw = request.path
        norm = "/" + "/".join(seg for seg in raw.split("/") if seg)   # 중복슬래시 제거
        low = norm.lower().rstrip("/")
        # 3-a. drydock 자체 login/logout 비활성(공유쿠키 session.clear 차단). 모든 method.
        if low in ("/login", "/logout"):
            return ("Dock Manager 로그인은 TRMT 통합 로그인으로 대체됨", 404)
        # 3-b. MCP/OAuth 폐기 (결정1, C5). 접두 변형 포함.
        if low == "/oauth" or low.startswith(("/oauth/", "/mcp", "/.well-known")):
            return ("Not Found", 404)
        # 3-c. 정적파일만 게이트 통과(정확히 /static/ 하위. traversal은 Flask가 차단)
        if low == "/static" or low.startswith("/static/"):
            return None
        # 3-d. admin만 통과 (username+role=admin 동시 요구)
        if _live_trmt_admin():
            # Transition gate: reject an explicit foreign/malformed Origin while
            # temporarily allowing a missing header for installed pre-migration iOS.
            if request.method in ("POST", "PUT", "PATCH", "DELETE"):
                claimed = request.headers.get("Origin")
                public_origin = trmt_app.config.get(
                    'PUBLIC_ORIGIN', os.environ.get('TRMT_PUBLIC_ORIGIN', 'https://vslmanager.duckdns.org'))
                def origin_key(value):
                    try:
                        parsed = urlsplit(value)
                        if (parsed.scheme not in ('http', 'https') or not parsed.hostname
                                or parsed.username or parsed.password or parsed.path
                                or parsed.query or parsed.fragment):
                            return None
                        port = parsed.port or (443 if parsed.scheme == 'https' else 80)
                        return (parsed.scheme, parsed.hostname.lower(), port)
                    except (TypeError, ValueError):
                        return None
                if claimed and (origin_key(claimed) is None or origin_key(claimed) != origin_key(public_origin)):
                    return jsonify({"error": "cross-origin write blocked"}), 403
            # login_required 호환 shim — 이미 있으면 재대입 안 함(Set-Cookie 반복 방지, 올마이트)
            if not session.get("logged_in"):
                session["logged_in"] = True
            # 신규 생성은 TRMT 담당선박 정본을 강제하는 integration route만 허용한다.
            # 구버전 자유입력 POST를 남겨두면 직접 호출로 roster scope를 우회할 수 있다.
            if request.path == "/api/vessels" and request.method == "POST":
                return jsonify({"error": "TRMT 담당선박 선택 경로를 사용하세요"}), 410
            return None
        # 3-e. 비인가 → API는 401 JSON, 그 외는 TRMT(루트) 로그인으로
        if low.startswith("/api"):
            return jsonify({"error": "Unauthorized (admin only)"}), 401
        return redirect("/login")

    # ── 4. sqlite 안정화 (C6): busy_timeout + WAL ──────────────────────
    #    A1은 gunicorn worker=1이라 경합 낮지만, threads=8 대비 방어.
    _orig_get_db = dd.get_db
    # The daily adapter must remain strictly read-only; retain a connection
    # factory that predates the general Dock Manager hardening wrapper so its
    # fallback test-double path never executes WAL/PRAGMA writes.
    dd._trmt_original_get_db = _orig_get_db
    def _get_db_hardened():
        db = _orig_get_db()
        try:
            db.execute("PRAGMA busy_timeout=5000")
            db.execute("PRAGMA journal_mode=WAL")
        except Exception:
            pass
        return db
    dd.get_db = _get_db_hardened

    return dd_app
