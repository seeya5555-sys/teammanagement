#   sync 가 부활시키고(삭제는 하드 DELETE), 옛 입거 잔상까지 끌어올 수 있어서다.
_DOCKPROC_ORPHAN_KEY = 'dockproc_orphans'
_DOCKPROC_ORPHAN_MAX = 50                                # 선박당 보관 상한


def _dockproc_orphans_all():
    row = query("SELECT v FROM api_settings WHERE k=?", (_DOCKPROC_ORPHAN_KEY,), one=True)
    if not row or not row['v']:
        return {}
    try:
        d = json.loads(row['v'])
        return d if isinstance(d, dict) else {}
    except Exception:                                    # 값이 깨졌으면 비어있는 것으로 본다(화면만 영향)
        return {}


def _dockproc_orphans_save(items, orphans):
    """이번 sync payload 가 다룬 선박의 목록만 갈아치운다 — 폴러는 선박별로 호출하므로
    남의 선박 목록을 지우면 배너가 사라진다. 0건이 된 선박은 키를 없애 배너도 사라진다."""
    seen = {(it.get('vsl_cd') or '').strip().upper() for it in items}
    seen.discard('')
    if not seen:
        return
    cur = _dockproc_orphans_all()
    for vc in seen:
        cur.pop(vc, None)
    for o in orphans:
        vc = (o.get('vsl_cd') or '').strip().upper()
        if vc and len(cur.setdefault(vc, [])) < _DOCKPROC_ORPHAN_MAX:
            cur[vc].append(o)
    execute("INSERT OR REPLACE INTO api_settings (k,v) VALUES (?,?)",
            (_DOCKPROC_ORPHAN_KEY, json.dumps(cur, ensure_ascii=False)))


def _dockproc_orphans_of(vsl_cd, have=()):
    """화면용 — 그 선박의 미적재 목록. 이미 행이 생긴 번호는 즉시 빠진다(다음 sync 를 안 기다림)."""
    vc = (vsl_cd or '').strip().upper()
    if not vc:
        return []
    hv = {(h or '').strip().upper() for h in have}
    out = []
    for o in (_dockproc_orphans_all().get(vc) or []):
        rq = (o.get('req_no') or '').strip().upper()
        if rq and rq not in hv:
            out.append(o)
    return out


def _dockproc_cell(ws, coord):
    v = ws[coord].value
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return v


def _dockproc_parse_index(stream):
    """INDEX 시트 → (vessel_meta, [line...]). 빈 슬롯(equipment·subject 모두 없음) 제외.
    R/S/ST 만 자동수집(P/SY=메일견적은 사이트서 수동추가)."""
    import re as _re
    from openpyxl import load_workbook
    wb = load_workbook(stream, data_only=True, read_only=True)
    if 'INDEX' not in wb.sheetnames:
        raise ValueError('INDEX 시트가 없음')
    ws = wb['INDEX']
    meta = {'vsl_nm': None, 'owner_co': None, 'vtype': None,
            'survey': None, 'shipyard': None, 'due_date': None}
    label_map = [('VESSEL NAME', 'vsl_nm'), ('OWNER', 'owner_co'),
                 ('TYPE OF VESSEL', 'vtype'), ('KIND OF SURVEY', 'survey'),
                 ('SHIPYARD', 'shipyard'), ('DUE DATE', 'due_date')]
    for row in ws.iter_rows(min_row=1, max_row=8, max_col=8, values_only=True):
        for i, v in enumerate(row):
            if not isinstance(v, str):
                continue
            u = v.strip().upper()
            for lbl, key in label_map:
                if u == lbl and meta[key] is None:
                    for w in row[i + 1:]:
                        if w is not None and (not isinstance(w, str) or w.strip()):
                            meta[key] = w.strip() if isinstance(w, str) else w
                            break
    if meta['due_date'] is not None and not isinstance(meta['due_date'], str):
        try:
            meta['due_date'] = meta['due_date'].strftime('%Y-%m-%d')
        except Exception:
            app.logger.exception('dockproc-parse-index')
            meta['due_date'] = str(meta['due_date'])
    # 헤더행 탐색(REQ. NUMBER / CATEGORY 포함)
    hdr_row = None
    for r in range(1, 12):
        vals = [str(_dockproc_cell(ws, f'{c}{r}') or '').upper() for c in 'ABCDEFGH']
        if any(('REQ' in x and 'NUMBER' in x) for x in vals) or 'CATEGORY' in vals:
            hdr_row = r
            break
    if hdr_row is None:
        hdr_row = 5
    lines = []
    for r in range(hdr_row + 1, ws.max_row + 1):
        no = _dockproc_cell(ws, f'A{r}')
        reqb = _dockproc_cell(ws, f'B{r}')         # REQ.NUMBER (수기 오타 가능)
        cat = _dockproc_cell(ws, f'C{r}')
        equip = _dockproc_cell(ws, f'D{r}')
        subj = _dockproc_cell(ws, f'E{r}')
        prep = _dockproc_cell(ws, f'F{r}')
        link = _dockproc_cell(ws, f'G{r}')         # LINK = 실제 시트 ID(유니크) → dedup 키 우선
        rmk = _dockproc_cell(ws, f'H{r}')
        # 정규 req_no: LINK(G) 우선(시트탭과 1:1, 유니크), 없으면 REQ.NUMBER(B) fallback
        req = None
        for cand in (link, reqb):
            if cand and _re.match(r'^(SY|ST|R|S|P)\d+$', str(cand).strip().upper()):
                req = str(cand).strip().upper()
                break
        if not req:
            continue
        if not equip and not subj:                       # grey 빈 슬롯 제외
            continue
        code = _dockproc_cat_code(req)
        prep_v = (str(prep).strip().upper() if prep else None)
        lines.append({
            'req_no': req, 'cat_code': code,
            'category': _DOCKPROC_CAT_NM.get(code, (cat or None)),
            'equipment': equip, 'subject': subj,
            'prepared_by': prep_v,
            'source': _dockproc_source(code, prep_v),
            'remark': rmk,
            'sort_no': (int(no) if isinstance(no, (int, float)) else None),
            'content_hash': _dockproc_hash(equip, subj),
        })
    return meta, lines


_DOCKPROC_ORDER = ("ORDER BY CASE cat_code WHEN 'R' THEN 0 WHEN 'S' THEN 1 "
                   "WHEN 'ST' THEN 2 WHEN 'P' THEN 3 WHEN 'SY' THEN 4 ELSE 5 END, "
                   "COALESCE(sort_no, 999999), id")


@app.route('/dock_procure')
@login_required
def dock_procure_page():
    return render_template('dock_procure.html')


@app.route('/api/dock_procure/lines')
@login_required
def api_dockproc_lines():
    vsl = request.args.get('vsl_nm')
    vessels = [dict(r) for r in query(
        "SELECT * FROM dock_procure_vessel ORDER BY updated_at DESC")]
    # 선박별 집계(카드 선택기용): 총건수 + 발주완료 건수
    agg = {r['vsl_nm']: r for r in query(
        "SELECT vsl_nm, COUNT(*) tot, COALESCE(SUM(stg_order),0) done FROM dock_procure GROUP BY vsl_nm")}
    for v in vessels:
        a = agg.get(v['vsl_nm'])
        v['total'] = (a['tot'] if a else 0)
        v['done'] = (a['done'] if a else 0)
    if not vsl and vessels:
        vsl = vessels[0]['vsl_nm']
    rows = []
    orphans = []
    if vsl:
        rows = [dict(r) for r in query(
            "SELECT * FROM dock_procure WHERE vsl_nm=? " + _DOCKPROC_ORDER, (vsl,))]
        ves = next((v for v in vessels if v['vsl_nm'] == vsl), None)
        prefix = _reqgen_vsl_prefix((ves or {}).get('vtype'))
        vcode = (ves or {}).get('vsl_cd')
        disk = _dockatt_disk_map()                       # 디스크 1회 스캔 — 행마다 listdir 하면 목록이 느려진다
        # 각 R/S/ST 행에 SVMS 정규 제목(수동작성 시 복사용 = reqgen 자동건과 동일 포맷) 생성
        for r in rows:
            # 실제로 열 수 있는 견적서 idx = 디스크에 **지문까지 일치하는** 캐시가 있는 자리만
            r['att_cached'] = (_dockatt_cached_idx(_dockproc_files_of(r['att_files']), disk.get(r['id']))
                               if r['att_files'] else [])
            # 견적요청 버튼이 쓸 문서종류/키 — **어느 컬럼인지 판단은 서버가 한다**(웹·앱이 각자
            # `cat_code`→컬럼 매핑을 들고 있으면 한쪽만 고쳐졌을 때 엉뚱한 번호로 요청이 나간다).
            r['inq_doc'], r['inq_key'] = _dockproc_inq_target(r)
            # 버튼을 미리 회색처리할 사유(없으면 None) — 라벨만 보는 순수 판정이라 쿼리가 늘지 않는다.
            r['inq_block'] = _dockproc_inq_stage_block(r['inq_doc'], r.get('svms_status'))
            # 라벨이 '이미 견적요청 이후'인지도 **서버가** 판정해 내려준다 — 화면이 부분일치 리스트를
            # 각자 복사해 들고 있으면 `_DOCK_INQ_PRE` 예외가 빠져 실패 이력이 초록으로 뒤집힌다.
            r['inq_posted'] = _dockproc_inq_posted(r['inq_doc'], r.get('svms_status'))
            # 상신 쪽도 같은 규약 — 화면은 이 값만 읽는다(`_DOCK_SUBMIT_PRE` 예외가 서버에만 있다).
            r['sbm_posted'] = _dockproc_sbm_posted(r.get('svms_status'))
            vc = r.get('vsl_cd') or vcode
            if r.get('cat_code') in ('R', 'S', 'ST') and vc:
                r['svms_subj'] = _reqgen_build_subj(vc, r['req_no'], r['vsl_nm'], prefix, r.get('subject'))
            else:
                r['svms_subj'] = None
        # SVMS 엔 있는데 이 목록에 행이 없는 청구 — 화면 배너에서 [적재] 로 끌어올린다.
        orphans = _dockproc_orphans_of(vcode, [r['req_no'] for r in rows])
    return jsonify({'vessels': vessels, 'current': vsl, 'lines': rows,
                    'orphans': orphans if vsl else []})


@app.route('/api/dock_procure/vessel_code', methods=['POST'])
@login_required
def api_dockproc_vessel_code():
    """선박 SVMS 코드(예: SAPS) 설정 — 정규 제목 생성·Phase2 역추적 매칭용. 선박헤더+모든 행에 반영."""
    d = request.get_json(silent=True) or {}
    vsl_nm = (d.get('vsl_nm') or '').strip()
    vsl_cd = (d.get('vsl_cd') or '').strip().upper() or None
    if not vsl_nm:
        return jsonify({'error': 'vsl_nm 필수'}), 400
    execute("UPDATE dock_procure_vessel SET vsl_cd=?, updated_at=datetime('now','localtime') WHERE vsl_nm=?",
            (vsl_cd, vsl_nm))
    execute("UPDATE dock_procure SET vsl_cd=?, updated_at=datetime('now','localtime') WHERE vsl_nm=?",
            (vsl_cd, vsl_nm))
    return jsonify({'vsl_nm': vsl_nm, 'vsl_cd': vsl_cd})


@app.route('/api/dock_procure/vessel', methods=['POST'])
@login_required
def api_dockproc_vessel_create():
    """새 입거선박 등록 — INDEX 엑셀 없이 빈 선박을 직접 생성(여러 선박 동시 진행용).
    라인은 이후 '＋ 라인 추가(P/SY)'·조선소 견적 업로드·INDEX 엑셀로 채운다."""
    d = request.get_json(silent=True)
    if not isinstance(d, dict):                        # array/scalar/None 바디 → 400(500 방지)
        return jsonify({'error': 'JSON object 필요'}), 400
    vsl_nm = str(d.get('vsl_nm') or '').strip()        # 비문자 입력도 안전 강제
    if not vsl_nm:
        return jsonify({'error': '선박명(vsl_nm) 필수'}), 400
    if len(vsl_nm) > 120:
        return jsonify({'error': '선박명이 너무 김(최대 120자)'}), 400
    if query("SELECT vsl_nm FROM dock_procure_vessel WHERE vsl_nm=?", (vsl_nm,), one=True):
        return jsonify({'error': f'"{vsl_nm}" 이미 등록됨'}), 409
    vsl_cd = (str(d.get('vsl_cd') or '').strip().upper()[:20]) or None
    vtype = (str(d.get('vtype') or '').strip()[:60]) or None
    try:                                               # PK(vsl_nm) race → IntegrityError 를 409 로(pre-check TOCTOU 보강)
        execute(
            "INSERT INTO dock_procure_vessel (vsl_nm, vsl_cd, vtype, updated_at) "
            "VALUES (?,?,?,datetime('now','localtime'))",
            (vsl_nm, vsl_cd, vtype))
    except sqlite3.IntegrityError:
        return jsonify({'error': f'"{vsl_nm}" 이미 등록됨'}), 409
    return jsonify({'vsl_nm': vsl_nm, 'vsl_cd': vsl_cd, 'vtype': vtype}), 201


@app.route('/api/dock_procure/vessel', methods=['DELETE'])
@login_required
def api_dockproc_vessel_delete():
    """입거선박 삭제 — 선박 레코드 + 해당 선박의 모든 라인(dock_procure)·조선소(dock_yard) 데이터 일괄 삭제.
    되돌릴 수 없음(UI confirm 게이트). Dry Dock 보고서(dock_reports 계열)는 별개 기능이라 건드리지 않음."""
    d = request.get_json(silent=True)
    if not isinstance(d, dict):
        return jsonify({'error': 'JSON object 필요'}), 400
    vsl_nm = str(d.get('vsl_nm') or '').strip()
    if not vsl_nm:
        return jsonify({'error': '선박명(vsl_nm) 필수'}), 400
    if not query("SELECT vsl_nm FROM dock_procure_vessel WHERE vsl_nm=?", (vsl_nm,), one=True):
        return jsonify({'error': f'"{vsl_nm}" 없음'}), 404
    db = get_db()
    lines = db.execute("SELECT COUNT(*) c FROM dock_procure WHERE vsl_nm=?", (vsl_nm,)).fetchone()['c']
    yard = db.execute("SELECT COUNT(*) c FROM dock_yard WHERE vsl_nm=?", (vsl_nm,)).fetchone()['c']
    # 3개 테이블 원자적 삭제 — 단일 트랜잭션(중간 실패 시 자동 rollback, 부분삭제 방지)
    with db:
        db.execute("DELETE FROM dock_procure WHERE vsl_nm=?", (vsl_nm,))
        db.execute("DELETE FROM dock_yard WHERE vsl_nm=?", (vsl_nm,))
        db.execute("DELETE FROM dock_procure_vessel WHERE vsl_nm=?", (vsl_nm,))
    return jsonify({'ok': True, 'vsl_nm': vsl_nm, 'deleted_lines': lines, 'deleted_yard': yard})


@app.route('/api/dock_procure/upload', methods=['POST'])
@login_required
def api_dockproc_upload():
    """INDEX 엑셀 업로드 → 라인 큐 증분생성. dedup=(vsl_nm, req_no). 기존건은 skip(진행 보존)."""
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'error': '엑셀 파일(file) 필요'}), 400
    if not f.filename.lower().endswith(('.xlsx', '.xlsm')):
        return jsonify({'error': '.xlsx 파일만 가능'}), 400
    try:
        import io as _io
        meta, lines = _dockproc_parse_index(_io.BytesIO(f.read()))
    except Exception as e:
        app.logger.exception('dockproc-upload')
        return jsonify({'error': f'파싱 실패: {e}'}), 400
    vsl_nm = meta.get('vsl_nm')
    if not vsl_nm:
        return jsonify({'error': 'INDEX 에서 VESSEL NAME 을 못 찾음'}), 400
    if not lines:
        return jsonify({'error': 'INDEX 에 유효한 항목(R/S/ST)이 없음'}), 400
    vsl_cd = (request.form.get('vsl_cd') or '').strip().upper() or None
    execute(
        "INSERT INTO dock_procure_vessel (vsl_nm, vsl_cd, owner_co, vtype, survey, shipyard, due_date, updated_at) "
        "VALUES (?,?,?,?,?,?,?,datetime('now','localtime')) "
        "ON CONFLICT(vsl_nm) DO UPDATE SET "
        "  vsl_cd=COALESCE(excluded.vsl_cd, dock_procure_vessel.vsl_cd), "
        "  owner_co=excluded.owner_co, vtype=excluded.vtype, survey=excluded.survey, "
        "  shipyard=excluded.shipyard, due_date=excluded.due_date, updated_at=excluded.updated_at",
        (vsl_nm, vsl_cd, meta.get('owner_co'), meta.get('vtype'), meta.get('survey'),
         meta.get('shipyard'), meta.get('due_date')))
    batch = uuid.uuid4().hex[:12]
    added, skipped, updated = 0, 0, 0
    added_reqs = []
    for ln in lines:
        ex = query("SELECT id, content_hash FROM dock_procure WHERE vsl_nm=? AND req_no=?",
                   (vsl_nm, ln['req_no']), one=True)
        if ex:
            if ex['content_hash'] != ln['content_hash']:
                # 내용 변경 — 진행 체크박스는 보존, 서술필드만 갱신
                execute("UPDATE dock_procure SET equipment=?, subject=?, category=?, prepared_by=?, "
                        "remark=?, content_hash=?, sort_no=?, updated_at=datetime('now','localtime') WHERE id=?",
                        (ln['equipment'], ln['subject'], ln['category'], ln['prepared_by'],
                         ln['remark'], ln['content_hash'], ln['sort_no'], ex['id']))
                updated += 1
            else:
                skipped += 1
            continue
        execute(
            "INSERT INTO dock_procure (vsl_nm, vsl_cd, req_no, cat_code, category, equipment, subject, "
            "prepared_by, source, content_hash, remark, sort_no, rev_batch) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (vsl_nm, vsl_cd, ln['req_no'], ln['cat_code'], ln['category'], ln['equipment'], ln['subject'],
             ln['prepared_by'], ln['source'], ln['content_hash'], ln['remark'], ln['sort_no'], batch))
        added += 1
        added_reqs.append(ln['req_no'])
    return jsonify({'vsl_nm': vsl_nm, 'vsl_cd': vsl_cd, 'batch': batch,
                    'added': added, 'skipped': skipped, 'updated': updated,
                    'added_reqs': added_reqs, 'total': len(lines)}), 201


# ===== 입거 requisition 템플릿 다운로드 (예시용 / 작성용) =====
#   예시용 = 손유석이 검토한 실제 채움본(Owner reviewed) 원본 그대로.
#   작성용 = 같은 워크북에서 선박별 입력 내용만 제거(구조·라벨·수식·하이퍼링크·슬롯 보존).
_DOCKPROC_TMPL = os.path.join(app.root_path, 'static', 'dock_templates', 'docking_requisition.xlsx')


def _dockproc_blank_workbook(wb):
    """Docking Requisition 워크북을 작성용(빈) 버전으로 변환(in-place).
    - INDEX: 선박별 헤더(VESSEL/TYPE/SURVEY/SHIPYARD/DUE) + 슬롯 EQUIPMENT/SUBJECT/REMARK 제거.
             OWNER 기본값·No.·REQ.NUMBER·CATEGORY·PREPARED BY·LINK(하이퍼링크)는 보존.
    - R*/S*/ST*: 헤더 입력값·ITEM LIST 본문 제거. OWNER/VESSEL 수식·REQ.NO·라벨·No. 보존.
    - _TEMPLATE(빈 마스터)·HOW TO USE(설명)는 그대로.
    """
    import re
    from openpyxl.cell.cell import MergedCell

    def _clr(ws, coord):
        c = ws[coord]
        if not isinstance(c, MergedCell):
            c.value = None

    for ws in wb.worksheets:
        name = ws.title
        if name in ('HOW TO USE', '_TEMPLATE'):
            continue
        if name == 'INDEX':
            for coord in ('G2', 'C3', 'G3', 'C4', 'G4'):
                _clr(ws, coord)
            for r in range(6, ws.max_row + 1):
                for col in ('D', 'E', 'H'):
                    _clr(ws, f'{col}{r}')
            continue
        if re.fullmatch(r'(R|S|ST)\d+', name):
            for coord in ('G3', 'C5', 'C6', 'C7', 'G5', 'G6'):
                _clr(ws, coord)
            for r in range(11, ws.max_row + 1):
                for c in range(2, 10):  # B..I
                    cell = ws.cell(row=r, column=c)
                    if not isinstance(cell, MergedCell):
                        cell.value = None
    return wb


@app.route('/dock_procure/template/example')
@login_required
def dockproc_tmpl_example():
    from flask import send_file
    if not os.path.exists(_DOCKPROC_TMPL):
        abort(404)
    return send_file(_DOCKPROC_TMPL, as_attachment=True,
                     download_name='Docking_Requisition_예시용.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/dock_procure/template/blank')
@login_required
def dockproc_tmpl_blank():
    from flask import send_file
    import io as _io, openpyxl
    if not os.path.exists(_DOCKPROC_TMPL):
        abort(404)
    try:
        wb = openpyxl.load_workbook(_DOCKPROC_TMPL)
        _dockproc_blank_workbook(wb)
        bio = _io.BytesIO()
        wb.save(bio)
        bio.seek(0)
    except Exception as e:
        app.logger.exception('dockproc-blank-template')
        return jsonify({'error': f'작성용 템플릿 생성 실패: {e}'}), 500
    return send_file(bio, as_attachment=True,
                     download_name='Docking_Requisition_작성용.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/dock_procure/template/<kind>')
@login_required
def api_dockproc_tmpl(kind):
    """iOS 앱용 alias.

    위 두 라우트는 `/api/` 밖이라 `_bearer_auth` before_request 의 Bearer→세션 투명주입이 걸리지 않아
    앱에서 호출하면 로그인 페이지로 튄다. 파일 생성 로직은 웹 라우트를 그대로 호출해 재사용한다
    (템플릿 생성 코드를 복제하면 두 경로가 갈라짐).
    """
    if kind == 'example':
        return dockproc_tmpl_example()
    if kind == 'blank':
        return dockproc_tmpl_blank()
    return jsonify({'error': 'kind must be example or blank'}), 404


@app.route('/api/dock_procure/<int:lid>/stage', methods=['POST'])
@login_required
def api_dockproc_stage(lid):
    """4단계 체크 토글 + 종속 cascade(상위체크→하위완료, 하위해제→상위해제)."""
    d = request.get_json(silent=True) or {}
    stage = d.get('stage')
    val = 1 if d.get('value') else 0
    if stage not in ('quote', 'vendor', 'confirm', 'order'):
        return jsonify({'error': 'stage must be quote/vendor/confirm/order'}), 400
    row = query("SELECT * FROM dock_procure WHERE id=?", (lid,), one=True)
    if not row:
        return jsonify({'error': 'not found'}), 404
    q, v, f, o = row['stg_quote'], row['stg_vendor'], row['stg_confirm'], row['stg_order']
    if stage == 'quote':
        q = val
        if not val:
            v = f = o = 0
    elif stage == 'vendor':
        v = val
        if val:
            q = 1
        else:
            f = o = 0
    elif stage == 'confirm':
        f = val
        if val:
            q = v = 1
        else:
            o = 0
    else:  # order
        o = val
        if val:
            q = v = f = 1
    # 🔴 SVMS 정본 행은 사람이 단계를 **내리지 못한다**(2026-08-05 실사고 BGBB S13 = id 148).
    #   실측: SVMS 'Ordered'(발주서 `BGBBES2608A21A` 발급) 행을 sync 가 08-04 21:50 에 [●●●●] 로
    #   세팅해 뒀는데, 08-05 13:38 이 엔드포인트 호출 1건이 '벤더제출' 을 끄면서 cascade 로
    #   v=f=o=0 → **발주완료 건이 '견적작성' 버킷으로 떨어졌다**(형 제보 캡처). 이 엔드포인트를
    #   자동 호출하는 코드는 웹·iOS·워커 어디에도 없다. ⚠️단, 서버에 access log 가 없어
    #   (gunicorn `--error-logfile -` 만) **호출 주체는 특정 못 했다** — 화면 탭이 가장 그럴듯하지만
    #   직접 API 호출·스크립트도 배제 못 한다(올마이트 지적 수용). 확정된 건 "이 엔드포인트를 탄 쓰기
    #   1건" 까지다.
    #   다음 정기 sync(최대 1시간)가 되돌리지만, **되돌아온다는 사실 자체가 이 토글이 무의미**하다는
    #   뜻이고 그 사이 화면·필터·상태배지는 명백히 틀린 값을 보여준다.
    #   기준선 = min(SVMS 라벨 rank, 지금 켜져 있는 rank).
    #     · SVMS rank = 정본이 "여기까지 왔다"고 말한 단계는 사람이 못 내린다.
    #     · 현재 rank 로 한 번 더 낮추는 이유 = 발주근거 fail-closed 게이트가 rank4 를 3 으로 눌러둔
    #       행(`ordered_evidence` False/None)에서 화면에 없는 단계까지 요구하면 정상 행이 통째로
    #       잠긴다. 사람이 sync 보다 앞서 켜 둔 단계를 스스로 되돌리는 것도 이 min() 덕에 계속 된다.
    #   ⚠️rank 0 은 **fail-open**(종전대로 자유 해제)이다. rank 0 = "SVMS 미연결"이 아니라
    #     "이 라벨로는 단계를 주장할 수 없음"이다 — 빈 라벨(수동관리 행. 2026-08-03 실측 73행 중
    #     50행이 사람이 켠 단계만 보유) 뿐 아니라 `HQ Received`·`VSL Approved` 같은 **견적의뢰 이전
    #     SVMS 라벨**과 미등재·오탈자 라벨도 전부 여기 떨어진다(올마이트 지적 — 옛 주석은 rank0 을
    #     '미연결'로 뭉뚱그렸는데 틀렸다). 처음 보는 라벨 하나로 사람 조작을 잠그지 않는 쪽이 맞다.
    _sv_rank = _dockproc_status_rank(row['svms_status'])
    if _sv_rank >= 1:
        _cur_rank = (4 if row['stg_order'] else 3 if row['stg_confirm']
                     else 2 if row['stg_vendor'] else 1 if row['stg_quote'] else 0)
        _new_rank = 4 if o else 3 if f else 2 if v else 1 if q else 0
        if _new_rank < min(_sv_rank, _cur_rank):
            return jsonify({'error': "SVMS 상태가 '%s' 라 이 단계는 해제할 수 없음 — 해제해도 다음 "
                                     "동기화에서 되돌아옵니다. SVMS 에서 먼저 되돌려 주세요."
                                     % (row['svms_status'] or '')}), 409
    # 🔴 낙관적 락 — SELECT 와 UPDATE 사이에 sync 가 끼어들면 위 게이트가 **stale 스냅샷**으로 통과한
    #   뒤 옛 단계값을 그대로 덮어써 같은 사고가 재현된다(올마이트 지적 수용). 스냅샷 4개를 WHERE 에
    #   실어 그 사이 값이 바뀌었으면 0행 = 409 로 돌린다. 클라이언트는 새로고침 후 다시 누르면 된다.
    #   (`rowcount` 가 필요해 `execute` 대신 `execute_rc`.)
    # 🔴 사람이 확정한 단계는 **floor** 로 굳힌다(2026-08-07 형 지시, 실사고 BGBB S5).
    #   실측: 형이 '발주완료' 를 켜고 remark 에 "이메일 발주 : 오션어스" 를 적어둔 행을 다음 정기
    #   sync 가 SVMS 라벨(rank 2 = 벤더제출)로 되돌렸다. SVMS 로 발주하지 않은 건(메일·직접 발주)은
    #   SVMS 라벨이 영원히 올라오지 않으므로, 이대로면 사람 입력이 매시간 지워진다.
    #   ⇒ 이 엔드포인트로 세운 rank 를 기록하고, sync 는 그 아래로 내리지 못한다(올리는 건 계속 허용).
    #   해제 경로는 살아 있다 — 사람이 다시 내리면 floor 도 같이 내려간다(위 SVMS 하한 게이트까지).
    _man = 4 if o else 3 if f else 2 if v else 1 if q else 0
    rc = execute_rc(
        "UPDATE dock_procure SET stg_quote=?, stg_vendor=?, stg_confirm=?, stg_order=?, stg_manual=?, "
        "updated_at=datetime('now','localtime') WHERE id=? AND stg_quote=? AND stg_vendor=? "
        "AND stg_confirm=? AND stg_order=?",
        (q, v, f, o, _man, lid, row['stg_quote'], row['stg_vendor'], row['stg_confirm'], row['stg_order']))
    if not rc:
        return jsonify({'error': '동기화가 방금 이 항목을 갱신했습니다. 새로고침 후 다시 시도하세요.'}), 409
    return jsonify({'id': lid, 'stg_quote': q, 'stg_vendor': v,
                    'stg_confirm': f, 'stg_order': o, 'stg_manual': _man})


@app.route('/api/dock_procure/add', methods=['POST'])
@login_required
def api_dockproc_add():
    """라인 수동추가(주로 페인트 P/조선소 SY 메일견적)."""
    d = request.get_json(silent=True) or {}
    vsl_nm = (d.get('vsl_nm') or '').strip()
    req_no = (d.get('req_no') or '').strip().upper()
    if not vsl_nm or not req_no:
        return jsonify({'error': 'vsl_nm, req_no 필수'}), 400
    code = _dockproc_cat_code(req_no)
    if not code:
        return jsonify({'error': 'req_no 는 R/S/ST/P/SY + 숫자 형식'}), 400
    if query("SELECT id FROM dock_procure WHERE vsl_nm=? AND req_no=?", (vsl_nm, req_no), one=True):
        return jsonify({'error': f'{req_no} 이미 존재'}), 409
    equip = (d.get('equipment') or '').strip() or None
    subj = (d.get('subject') or '').strip() or None
    prep = (d.get('prepared_by') or 'MANAGER').strip().upper()
    lid = execute(
        "INSERT INTO dock_procure (vsl_nm, vsl_cd, req_no, cat_code, category, equipment, subject, "
        "prepared_by, source, content_hash, remark) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
        (vsl_nm, (d.get('vsl_cd') or None), req_no, code, _DOCKPROC_CAT_NM.get(code),
         equip, subj, prep,
         _dockproc_source(code, prep), _dockproc_hash(equip, subj),
         (d.get('remark') or None)))
    return jsonify({'id': lid, 'req_no': req_no}), 201


@app.route('/api/dock_procure/adopt', methods=['POST'])
@login_required
def api_dockproc_adopt():
    """미적재 배너의 [적재] — SVMS 에만 있는 입거 청구를 발주현황 행으로 끌어온다.

    🔒 입력은 **폴러가 남긴 미적재 목록 안에서만** 고른다(클라가 보낸 제목·키를 그대로 쓰지 않음) —
       화면에서 임의 행을 만들어 넣는 경로가 되면 안 되고, 제목·문서키는 SVMS 실측값이어야 한다.
    자동생성이 아니라 사람 클릭인 이유: 삭제가 하드 DELETE 라 자동적재는 형이 지운 행을 되살리고,
    옛 입거 잔상까지 끌어올 수 있다(태그 번호는 입거마다 재사용됨).
    """
    d = request.get_json(silent=True)
    # 문자열 아닌 값(dict·list·숫자)이 오면 `.strip()` 에서 500 이 난다 — 계약 위반은 400 으로 떨어뜨린다.
    def _s(v):
        return v.strip().upper() if isinstance(v, str) else ''
    if not isinstance(d, dict):
        return jsonify({'error': 'JSON object 필요'}), 400
    vc, rq = _s(d.get('vsl_cd')), _s(d.get('req_no'))
    if not vc or not rq:
        return jsonify({'error': 'vsl_cd, req_no 필수'}), 400
    ves = query("SELECT vsl_nm, vsl_cd FROM dock_procure_vessel WHERE UPPER(vsl_cd)=? "
                "ORDER BY updated_at DESC", (vc,), one=True)
    if not ves:
        return jsonify({'error': f'{vc} 가 입거선박 목록에 없음'}), 404
    src = next((o for o in (_dockproc_orphans_all().get(vc) or [])
                if (o.get('req_no') or '').strip().upper() == rq), None)
    # 🔴 이미 행이 있으면 목록 밖이어도 성공(200)으로 돌려준다. sync 가 목록에서 뺀 뒤 같은 요청이
    #    오는 건 정상 경로(두 탭·더블클릭·동시 적재)이고, 이때 409 를 주면 형이 '실패'로 읽는다.
    #    행 **생성**만 목록 안으로 제한한다 — 임의 행 생성 경로 차단이라는 목적은 그대로다.
    ex = query("SELECT id FROM dock_procure WHERE vsl_nm=? AND UPPER(req_no)=?",
               (ves['vsl_nm'], rq), one=True)
    if not ex and not src:
        return jsonify({'error': f'{rq} 는 동기화가 남긴 미적재 목록에 없음 — 새로고침(동기화) 후 다시'}), 409
    src = src or {}
    try:
        lid, created, key_state = _dockproc_adopt_svms(
            ves['vsl_nm'], ves['vsl_cd'] or vc, rq, _dockproc_subject_from_svms(src.get('subject')),
            None, src.get('doc'), src.get('key'))
    except sqlite3.IntegrityError:
        # UNIQUE(vsl_nm, req_no) — 동시 적재로 남이 먼저 만든 경우. 실패가 아니라 이미 된 것이다.
        row = query("SELECT id FROM dock_procure WHERE vsl_nm=? AND UPPER(req_no)=?",
                    (ves['vsl_nm'], rq), one=True)
        if not row:
            raise
        lid, created, key_state = row['id'], False, 'none'
    if not lid:
        return jsonify({'error': f'{rq} 는 적재 대상 아님 — SVMS 청구(R/S/ST)만 가능'}), 400
    return jsonify({'id': lid, 'req_no': rq, 'vsl_nm': ves['vsl_nm'],
                    'created': bool(created), 'key_state': key_state}), (201 if created else 200)


@app.route('/api/dock_procure/<int:lid>/prep', methods=['POST'])
@login_required
def api_dockproc_prep(lid):
    """담당(OWNER↔MANAGER) 토글 — 견적출처 자동 동기화(MANAGER→AOR / OWNER→SVMS, P·SY=MAIL 고정)."""
    row = query("SELECT * FROM dock_procure WHERE id=?", (lid,), one=True)
    if not row:
        return jsonify({'error': 'not found'}), 404
    d = request.get_json(silent=True) or {}
    nv = (d.get('prepared_by') or '').strip().upper()
    if nv not in ('OWNER', 'MANAGER'):                 # 값 없으면 토글
        nv = 'MANAGER' if (row['prepared_by'] or '').upper() == 'OWNER' else 'OWNER'
    src = _dockproc_source(row['cat_code'], nv)
    execute("UPDATE dock_procure SET prepared_by=?, source=?, updated_at=datetime('now','localtime') WHERE id=?",
            (nv, src, lid))
    return jsonify({'id': lid, 'prepared_by': nv, 'source': src})


@app.route('/api/dock_procure/<int:lid>', methods=['PATCH'])
@login_required
def api_dockproc_patch(lid):
    d = request.get_json(silent=True) or {}
    # 검증 먼저 전부 통과시킨 뒤 단일 UPDATE — partial update 방지(올마이트 검토 반영)
    sets, params = [], []
    if 'remark' in d:
        sets.append('remark=?'); params.append(d.get('remark'))
    if 'vendor' in d:                                   # 페인트(P) 수동 업체명(SVMS Dock Paint 02 VNDR_NM 소스)
        v = d.get('vendor')
        if v is not None and not isinstance(v, str):    # 타입 엄격(조용한 null overwrite 방지)
            return jsonify({'error': 'vendor must be a string or null'}), 400
        _row = query("SELECT cat_code FROM dock_procure WHERE id=?", (lid,), one=True)
        if not _row or _row['cat_code'] != 'P':         # 서버단 P라인 강제(UI 게이팅 우회 차단)
            return jsonify({'error': 'vendor is only editable on Paint(P) lines'}), 400
        vv = (v.strip()[:200] or None) if isinstance(v, str) else None   # trim + 200자 상한
        sets.append('vendor=?'); params.append(vv)
    if 'quote_amt' in d:                                # 발주업체 확정 견적금액(수정가능, SVMS 연동 소스)
        raw = d.get('quote_amt')
        if raw in (None, ''):
            amt = None
        else:
            try:
                amt = float(str(raw).replace(',', ''))
            except (TypeError, ValueError):
                return jsonify({'error': 'quote_amt must be numeric'}), 400
            if not math.isfinite(amt) or amt < 0:      # nan/inf/음수 차단(금액 도메인)
                return jsonify({'error': 'quote_amt must be a finite non-negative number'}), 400
        sets.append('quote_amt=?'); params.append(amt)
    if 'quote_cur' in d:
        cur = (d.get('quote_cur') or '').strip().upper()
        if not re.fullmatch(r'[A-Z]{3}', cur):         # 3자 통화코드 strict(silent truncation 금지)
            return jsonify({'error': 'quote_cur must be a 3-letter code'}), 400
        sets.append('quote_cur=?'); params.append(cur)
    if 'quote_amt' in d or 'quote_cur' in d:            # 사용자 직접수정 → manual 잠금(폴러 자동덮어쓰기 차단)
        # 금액만 비우면(clear) 자동입력 재개, 그 외(값 입력/통화 변경)는 manual 잠금
        if 'quote_amt' in d and d.get('quote_amt') in (None, ''):
            sets.append('quote_src=?'); params.append('auto')   # 금액 clear = 자동입력 재개(통화 동반 무관)
        else:
            sets.append('quote_src=?'); params.append('manual')
    if sets:
        sets.append("updated_at=datetime('now','localtime')")
        params.append(lid)
        execute(f"UPDATE dock_procure SET {', '.join(sets)} WHERE id=?", tuple(params))
    return jsonify({'ok': True})


@app.route('/api/dock_procure/<int:lid>', methods=['DELETE'])
@login_required
def api_dockproc_delete(lid):
    execute("DELETE FROM dock_procure WHERE id=?", (lid,))
    return jsonify({'ok': True})


@app.route('/api/dock_procure/<int:lid>/link', methods=['POST'])
@login_required
def api_dockproc_link(lid):
    """Tier 3 — 제목규칙 안 지킨 수동 SVMS건을 Inq No 직접입력으로 연결(이후 폴러가 자동추적)."""
    d = request.get_json(silent=True) or {}
    inq = (d.get('svms_req_no') or '').strip() or None
    execute("UPDATE dock_procure SET svms_req_no=?, updated_at=datetime('now','localtime') WHERE id=?",
            (inq, lid))
    return jsonify({'id': lid, 'svms_req_no': inq})


@app.route('/api/ext/dock_procure/vessels')
@api_key_required
def api_ext_dockproc_vessels():
    """맥 폴러용 — SVMS코드(vsl_cd) 설정된 입거선박 목록(역동기화 대상)."""
    rows = query("SELECT vsl_nm, vsl_cd FROM dock_procure_vessel WHERE vsl_cd IS NOT NULL AND vsl_cd<>'' "
                 "ORDER BY updated_at DESC")
    return jsonify({'vessels': [dict(r) for r in rows]})


@app.route('/api/ext/dock_procure/quotes')
@api_key_required
def api_ext_dockproc_quotes():
    """SVMS Dock draft 봉투 조립용 — 발주완료(stg_order=1)+견적금액 있는 R/S/ST 라인.
    cat_code R=Shore Repair(ITEM_CD 04) · S/ST=Spare/Store(03). 조립·환산은 맥 조립기가 수행."""
    vc = (request.args.get('vsl_cd') or '').strip().upper()
    if not vc:
        return jsonify({'error': 'vsl_cd required'}), 400
    rows = query(
        "SELECT d.vsl_nm, d.vsl_cd, d.req_no, d.cat_code, d.category, d.subject, d.equipment, "
        "d.quote_amt, d.quote_cur, d.quote_src, d.svms_req_no "
        "FROM dock_procure d "
        "WHERE d.quote_amt IS NOT NULL AND d.stg_order=1 AND d.cat_code IN ('R','S','ST') "
        "AND (UPPER(d.vsl_cd)=? OR d.vsl_nm IN (SELECT vsl_nm FROM dock_procure_vessel WHERE UPPER(vsl_cd)=?)) "
        "ORDER BY d.cat_code, d.req_no",
        (vc, vc))
    return jsonify({'vsl_cd': vc, 'quotes': [dict(r) for r in rows]})


@app.route('/api/ext/dock/push_data')
@api_key_required
def api_ext_dock_push_data():
    """④ SVMS Dock draft 조립기(맥 build_envelope.py DRY)용 통합 소스.
    vessel(조선소 벤더) + yard 7카테고리 + paint(P) + repair(R) + spare/store(S/ST) 계획금액.
    ⚠️ 읽기전용. 조립·환산·BATCH_FLAG diff·SP_SET 저장은 전부 맥 조립기+형 컨펌(안전커널)."""
    vc = (request.args.get('vsl_cd') or '').strip().upper()
    if not vc:
        return jsonify({'error': 'vsl_cd required'}), 400
    ves = query("SELECT vsl_nm, vsl_cd, shipyard, shipyard_vndr_cd, shipyard_vndr_nm "
                "FROM dock_procure_vessel WHERE UPPER(vsl_cd)=? ORDER BY updated_at DESC", (vc,), one=True)
    if not ves:
        return jsonify({'error': 'unknown vsl_cd (dock_procure_vessel에 vsl_cd 매칭 없음)'}), 404
    vsl_nm = ves['vsl_nm']
    yard = query("SELECT category, amount, cur, remark, src, sort_no FROM dock_yard "
                 "WHERE vsl_nm=? ORDER BY sort_no, category", (vsl_nm,))
    # 🔴 `quote_amt IS NOT NULL` 만 걸면 **분할발주(업체 2곳)가 경고도 없이 통째로 빠진다** —
    #   통화가 섞이면 합계를 만들지 않으므로 `quote_amt=NULL` 이 정상이다(실측 [BGBB S1] KRW+USD).
    #   그래서 `ord_vendors` 가 있는 행도 내려보내고, 라인 조립·보류 판단은 조립기(build_envelope)에
    #   맡긴다. 조립기는 금액을 못 만들면 **warn 을 남기고 omit** 하므로 조용한 누락이 사라진다.
    lines = query(
        "SELECT req_no, cat_code, category, subject, equipment, quote_amt, quote_cur, quote_src, "
        "vendor, svms_req_no, stg_order, ord_vendors FROM dock_procure "
        "WHERE vsl_nm=? AND (quote_amt IS NOT NULL "
        "  OR (ord_vendors IS NOT NULL AND TRIM(ord_vendors) NOT IN ('', '[]'))) "
        "ORDER BY cat_code, req_no", (vsl_nm,))
    def bycat(*codes):
        return [dict(r) for r in lines if r['cat_code'] in codes]
    return jsonify({
        'vessel': {'vsl_nm': vsl_nm, 'vsl_cd': ves['vsl_cd'],
                   'shipyard': ves['shipyard'],
                   'shipyard_vndr_cd': ves['shipyard_vndr_cd'],
                   'shipyard_vndr_nm': ves['shipyard_vndr_nm']},
        'yard': [dict(r) for r in yard],       # dock_yard 7카테고리 → P_IC_YR
        'paint': bycat('P'),                   # → P_IC_DP(02)
        'repair': bycat('R'),                  # → P_IC_SR(04)
        'spare': bycat('S', 'ST'),             # → P_IC_SS(03)
    })


# ===== 조선소(Yard) 견적 → SVMS Yard Repair 7카테고리 (dock_yard) =====
YARD_CATEGORIES = ["General", "Paint", "Steel", "Deck", "Engine", "Electric", "Discount"]
_YARD_TOTAL_ROW = re.compile(r'total price|final discount|after dicount|after discount|normal total|sub ?total|소계|합계', re.I)

# General/Paint는 "항상 고정 형식"(손유석 지시) — AI가 형식을 못 지키면 빈 스켈레톤으로 강제(값은 형 수동입력).
_YARD_GEN_SKELETON = "입거 예상일정 : 일, 상가일정 : "
_YARD_PAINT_SKELETON = "Top : SA %, SA %, The other area :  (m2)"
# full-shape 검증(lead token만 아니라 구조 토큰 전부 존재해야 통과 — 올마이트 반영)
_YARD_GEN_RE = re.compile(r'^입거 예상일정 : .*상가일정 : ', re.S)
_YARD_PAINT_RE = re.compile(r'^Top : .*The other area : .*m2', re.S)


def _yard_norm_remark(cat, remark):
    """General/Paint remark를 고정 형식으로 보장(구조 토큰 전부 있어야 AI 원문 유지, 아니면 빈 스켈레톤). 나머지 카테고리는 AI 원문."""
    r = (remark or '').strip()
    if cat == "General":
        return r if _YARD_GEN_RE.match(r) else _YARD_GEN_SKELETON
    if cat == "Paint":
        return r if _YARD_PAINT_RE.match(r) else _YARD_PAINT_SKELETON
    return r or None

_YARD_AI_PROMPT = """너는 선박 입거수리(dry dock) 견적 분석가다. 조선소 견적서를 SVMS Yard Repair
7카테고리로 집계하고 카테고리별 작업요약(remark)을 작성한다.

카테고리 배정 기준:
- General : 일반서비스·입거비 (general service, docking)
- Paint   : 선체도장 (hull painting)
- Steel   : 강재수리 (structural steelwork)
- Deck    : 갑판부 (seachest, rudder, propeller, windlass, anchor, cargo pump, life boat, fire wire)
- Engine  : 기관부 (valve, tank cleaning, main/aux engine, boiler, pump, pipe/WBT, IGS, cooler, ER crane)
- Electric: 전기 (alternator, electric motor)
- Discount: 최종할인 (final discount) — 반드시 음수 금액

규칙:
- 각 라인의 Net Total(할인 반영된 라인 금액)만 합산한다. 소계/총계행(Total, Sub-total, Normal Total, discount 라벨)은 합산에서 제외.
- EGCS/스크러버(scrubber) 등 별도 스페셜 프로젝트 시트는 제외한다.
- remark(Steel/Deck/Engine/Electric) = 해당 카테고리에서 **금액이 큰 작업 위주로** 영문 1줄 요약(고액 항목을 앞에, 소액은 "etc."로 묶음). 예 Engine: "E/R pipe fabrication, Valves, Aux Boiler & Donkey boiler, IG Scrubber etc."
- ⚠️ General·Paint remark는 **반드시 아래 고정 형식 그대로** 출력한다(형식 문구·구두점 유지). 각 값은 견적서에서 **확실히 찾은 경우에만** 채우고, 없거나 불확실하면 그 자리는 **공란으로 비워둔다**(절대 추정·창작 금지 — 사람이 수동입력):
    General  형식: "입거 예상일정 : {N}일, 상가일정 : {상가 날짜범위}"      (예 "입거 예상일정 : 48일, 상가일정 : 4/25-30")
    Paint    형식: "Top : SA{등급} {비율}%, SA{등급} {비율}%, The other area : {처리방식} ({면적}m2)"   (예 "Top : SA2.0 20%, SA1.0 10%, The other area : full blasting (28,899m2)")
  값을 못 찾으면 예: General="입거 예상일정 : 일, 상가일정 : " / Paint="Top : SA %, SA %, The other area :  (m2)" 처럼 숫자만 비운 채 형식은 유지.
- currency는 견적 표기 그대로. ⚠️ 견적서에 없는 금액·작업을 지어내지 마라.
- quote_total = 견적서에 명시된 최종 총액(할인 후). 없으면 카테고리 합.

- ⚠️ categories 배열에는 7개 카테고리(General,Paint,Steel,Deck,Engine,Electric,Discount)를
  빠짐없이 모두 포함하고, 각 항목의 remark를 반드시 작성한다(해당 작업이 없으면 remark="").

출력은 JSON만:
{"currency":"USD","quote_total":873184.25,
 "categories":[{"cat":"General","amount":449244,"remark":"..."}, ... 7개 전부]}"""


def _yard_xlsx_to_text(raw_bytes, max_rows=3000):
    """조선소 견적 xlsx → 텍스트(전체 시트, Net Total 잘림 방지 위해 행제한 넉넉히)."""
    import io as _io
    from openpyxl import load_workbook
    wb = load_workbook(_io.BytesIO(raw_bytes), read_only=True, data_only=True)
    out = []
    n = 0
    for ws in wb.worksheets:
        out.append(f"### SHEET: {ws.title}")
        for r in ws.iter_rows(values_only=True):
            cells = ['' if c is None else str(c).strip() for c in r]
            while cells and cells[-1] == '':
                cells.pop()
            if not cells:
                continue
            out.append(' | '.join(cells))
            n += 1
            if n >= max_rows:
                return '\n'.join(out)
    return '\n'.join(out)


def _yard_ai_extract(raw_bytes):
    """Gemini Flash로 견적 → 7카테고리 금액+remark+총액. 실패/키없음 시 None."""
    if not GEMINI_API_KEY:
        return None
    try:
        text = _yard_xlsx_to_text(raw_bytes)
        res = _gemini_call_json([{'text': _YARD_AI_PROMPT + "\n\n[견적서]\n" + text}])
    except Exception:
        app.logger.exception('yard-ai-extract')
        return None
    if not isinstance(res, dict) or res.get('error') or not res.get('categories'):
        return None
    return res


def _yard_profiles_dir():
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), 'yard_profiles')


def _list_yard_profiles():
    d = _yard_profiles_dir()
    out = []
    if os.path.isdir(d):
        for fn in sorted(os.listdir(d)):
            if not fn.endswith('.json'):
                continue
            try:
                with open(os.path.join(d, fn), encoding='utf-8') as f:
                    p = json.load(f)
                out.append({'file': fn, 'yard_name': p.get('yard_name', fn),
                            'vndr_cd': p.get('vndr_cd')})   # 조선소 벤더(SVMS) 매칭용
            except Exception:
                app.logger.debug('yard-profile load skip: %s', fn, exc_info=True)
    return out


def _find_yard_profile_by_vndr(vndr_cd):
    """선택된 조선소 벤더코드로 파싱 프로파일 파일명 찾기(없으면 None → AI 폴백)."""
    if not vndr_cd:
        return None
    for p in _list_yard_profiles():
        if (p.get('vndr_cd') or '').strip().upper() == vndr_cd.strip().upper():
            return p['file']
    return None


def _load_yard_profile(name):
    fn = name if name.endswith('.json') else name + '.json'
    path = os.path.join(_yard_profiles_dir(), os.path.basename(fn))   # basename=경로탈출 방지
    with open(path, encoding='utf-8') as f:
        return json.load(f)


def _yard_parse_quote(fileobj, profile):
    """조선소 견적 xlsx → 7카테고리 소계. 총계행(텍스트) 제외 + Item No 첫정수=섹션. (yard_parse.py 검증본 이식)"""
    import openpyxl
    c = profile["cols"]
    ci, cd, cq, cn = c["item_no"], c["desc"], c["qty"], c["net_total"]
    smap = profile["section_map"]
    wb = openpyxl.load_workbook(fileobj, data_only=True, read_only=True)
    ws = wb[profile.get("sheet", "Quotation")]
    sect = {}
    cur_sec = None
    for r in ws.iter_rows(values_only=True):
        def cell(i):
            return r[i] if i < len(r) else None
        itm, desc, qty, nt = cell(ci), cell(cd), cell(cq), cell(cn)
        s = str(itm).strip() if itm is not None else ""
        m = re.match(r'^(\d+)', s)
        if m:
            cur_sec = m.group(1)
        rowtext = " ".join(str(x) for x in r if isinstance(x, str))
        if not isinstance(nt, (int, float)) or not nt or not cur_sec:
            continue
        if _YARD_TOTAL_ROW.search(rowtext):              # 총계/소계행 제외
            continue
        if isinstance(qty, str):
            try:
                float(qty.replace(',', ''))             # 숫자문자열 qty("1")는 라인 허용
            except (TypeError, ValueError):
                continue                                 # 진짜 텍스트(총계 라벨) = 제외
        sect[cur_sec] = sect.get(cur_sec, 0.0) + nt
    cat = {k: 0.0 for k in YARD_CATEGORIES}
    unmapped = {}
    for sec, amt in sect.items():
        c2 = smap.get(sec)
        if c2 in cat:
            cat[c2] += amt
        else:
            unmapped[sec] = round(unmapped.get(sec, 0.0) + amt, 2)
    line_total = sum(cat.values())
    cat["Discount"] = round(-line_total * profile.get("discount_rate", 0.0), 2)
    cat = {k: round(v, 2) for k, v in cat.items()}
    return {"categories": cat, "line_total": round(line_total, 2),
            "final_total": round(sum(cat.values()), 2), "unmapped": unmapped,
            "yard_name": profile.get("yard_name")}


@app.route('/api/dock_yard/profiles')
@login_required
def api_dock_yard_profiles():
    return jsonify({'profiles': _list_yard_profiles()})


@app.route('/api/dock_yard/shipyards')
@login_required
def api_dock_yard_shipyards():
    """조선소 드롭다운 소스 — SVMS 벤더마스터(SYD_YN=Y) 캐시 + 로컬 프로파일 vndr_cd 매칭 표시."""
    rows = query("SELECT vndr_cd, vndr_nm, vndr_nm_eng FROM yard_vendor ORDER BY COALESCE(NULLIF(vndr_nm_eng,''),vndr_nm)")
    profs = {(p.get('vndr_cd') or '').strip().upper() for p in _list_yard_profiles() if (p.get('vndr_cd') or '').strip()}
    out = [dict(r, has_profile=((r['vndr_cd'] or '').strip().upper() in profs)) for r in [dict(x) for x in rows]]
    return jsonify({'shipyards': out, 'synced': bool(rows)})


@app.route('/api/ext/dock_yard/shipyards', methods=['POST'])
@api_key_required
def api_ext_dock_yard_shipyards():
    """맥 yard_vendors_sync.py 가 SVMS 조선소 벤더 목록 적재(full-replace)."""
    d = request.get_json(silent=True) or {}
    ships = d.get('shipyards') or []
    if not isinstance(ships, list) or not ships:
        return jsonify({'error': 'shipyards[] 필요'}), 400
    dedup = {}                                                # vndr_cd 중복 제거(마지막 값 채택)
    for s in ships:
        if not isinstance(s, dict):
            continue
        cd = (s.get('vndr_cd') or '').strip()
        if not cd:
            continue
        dedup[cd] = (cd, (s.get('vndr_nm') or '').strip()[:200], (s.get('vndr_nm_eng') or '').strip()[:200])
    if not dedup:
        return jsonify({'error': '유효 vndr_cd 없음'}), 400
    rows = [(cd, nm, en) for (cd, nm, en) in dedup.values()]
    db = get_db()                                             # 원자적 full-replace(DELETE+INSERT 단일 트랜잭션, 부분상태 방지)
    try:
        db.execute("DELETE FROM yard_vendor")
        db.executemany("INSERT OR REPLACE INTO yard_vendor (vndr_cd, vndr_nm, vndr_nm_eng, updated_at) "
                       "VALUES (?,?,?,datetime('now','localtime'))", rows)
        db.commit()
    except Exception:
        db.rollback()
        app.logger.exception('yard-vendor replace')
        return jsonify({'error': '적재 실패(rollback)'}), 500
    return jsonify({'ok': True, 'count': len(rows)})


@app.route('/api/dock_procure/shipyard', methods=['POST'])
@login_required
def api_dockproc_set_shipyard():
    """선박의 조선소 벤더 선택 저장(드롭다운) → dock 봉투 DR_CD/VNDR_CD/VNDR_NM 소스."""
    d = request.get_json(silent=True) or {}
    vsl_nm = (d.get('vsl_nm') or '').strip()
    vndr_cd = (d.get('vndr_cd') or '').strip() or None
    if not vsl_nm:
        return jsonify({'error': 'vsl_nm 필요'}), 400
    vndr_nm = None
    if vndr_cd:
        row = query("SELECT vndr_nm FROM yard_vendor WHERE vndr_cd=?", (vndr_cd,), one=True)
        if not row:
            return jsonify({'error': '알 수 없는 조선소 벤더코드'}), 400
        vndr_nm = row['vndr_nm']
    rc = execute_rc("UPDATE dock_procure_vessel SET shipyard_vndr_cd=?, shipyard_vndr_nm=?, "
                    "updated_at=datetime('now','localtime') WHERE vsl_nm=?", (vndr_cd, vndr_nm, vsl_nm))
    if not rc:                                                # 없는 선박 → 404(조용한 ok 방지)
        return jsonify({'error': 'unknown vsl_nm'}), 404
    return jsonify({'ok': True, 'vndr_cd': vndr_cd, 'vndr_nm': vndr_nm})


@app.route('/api/dock_yard')
@login_required
def api_dock_yard_lines():
    vsl = request.args.get('vsl_nm')
    rows = query("SELECT * FROM dock_yard WHERE vsl_nm=? ORDER BY sort_no, category", (vsl,)) if vsl else []
    return jsonify({'lines': [dict(r) for r in rows]})


@app.route('/api/dock_yard/upload', methods=['POST'])
@login_required
def api_dock_yard_upload():
    """조선소 견적 xlsx 업로드 → 7카테고리 파싱 → dock_yard upsert(manual 잠금은 금액 보존)."""
    f = request.files.get('file')
    vsl_nm = (request.form.get('vsl_nm') or '').strip()
    prof_name = (request.form.get('profile') or '').strip()
    if not f or not f.filename or not f.filename.lower().endswith(('.xlsx', '.xlsm')):
        return jsonify({'error': '.xlsx 견적 파일 필요'}), 400
    if not vsl_nm:
        return jsonify({'error': 'vsl_nm 필요'}), 400
    data = f.read()
    import io as _io
    # 프로파일 해석: 명시된 profile 우선, 없으면 선택된 조선소 벤더(vndr_cd)로 자동매칭
    if not prof_name:
        _v = query("SELECT shipyard_vndr_cd FROM dock_procure_vessel WHERE vsl_nm=?", (vsl_nm,), one=True)
        if _v and _v['shipyard_vndr_cd']:
            prof_name = _find_yard_profile_by_vndr(_v['shipyard_vndr_cd']) or ''
    # 프로파일(선택) — 규칙파서(결정적 금액). 없으면 AI 폴백(비결정 경고).
    profile = None
    if prof_name:
        try:
            profile = _load_yard_profile(prof_name)
        except Exception:
            profile = None
    # 하이브리드: 금액=규칙파서(결정적) 우선, Remark=Gemini(AI). 프로파일 없으면 AI 금액 폴백(비결정 경고).
    ai = _yard_ai_extract(data)                       # Remark(+프로파일 없을때 금액 폴백)
    ai_remarks = {}
    if ai and ai.get('categories'):
        for c in ai['categories']:
            if c.get('cat') in YARD_CATEGORIES:
                ai_remarks[c['cat']] = (c.get('remark') or None)
    rule = None
    if profile:
        try:
            rule = _yard_parse_quote(_io.BytesIO(data), profile)
        except Exception:
            app.logger.exception('yard-rule')
            rule = None

    warns = []
    yard_nm = (profile or {}).get('yard_name')
    if rule:                                           # ✅ 금액=규칙(결정), Remark=AI
        source = 'rule+ai'
        cur_default = 'USD'
        catmap = {c: {'amount': round(rule['categories'][c], 2), 'remark': _yard_norm_remark(c, ai_remarks.get(c))}
                  for c in YARD_CATEGORIES}
        if rule.get('unmapped'):
            warns.append('⚠️ 미매핑 섹션: ' + ','.join(rule['unmapped'].keys()) + ' — 프로파일 보강 필요')
        if not ai:
            warns.append('Remark 생성 실패(Gemini) — 금액만 반영')
    elif ai and ai.get('categories'):                  # 프로파일 없음 → AI 금액(비결정 경고)
        source = 'ai'
        cur_default = (ai.get('currency') or 'USD').strip().upper()[:3] or 'USD'
        catmap = {}
        for c in ai['categories']:
            cn = c.get('cat')
            if cn not in YARD_CATEGORIES:
                continue
            try:
                amt = round(float(str(c.get('amount') or 0).replace(',', '')), 2)
            except (TypeError, ValueError):
                amt = 0.0
            if not math.isfinite(amt):
                amt = 0.0
            if cn == 'Discount' and amt > 0:
                amt = -amt
            catmap[cn] = {'amount': amt, 'remark': _yard_norm_remark(cn, c.get('remark'))}
        _missing = [x for x in YARD_CATEGORIES if x not in catmap]
        if _missing:
            warns.append('⚠️ AI 누락 카테고리: ' + ','.join(_missing))
        warns.append('⚠️ 프로파일 없음 — AI 금액(같은 견적도 값 변동 가능). 반드시 확인, 프로파일 요청 권장')
    else:
        return jsonify({'error': 'AI 파싱 실패 + 규칙 폴백 없음 — 조선소 프로파일 선택 또는 Gemini 키 확인'}), 400

    vsl_cd = (request.form.get('vsl_cd') or '').strip().upper() or None
    added = updated = skipped = 0
    for i, catn in enumerate(YARD_CATEGORIES):
        c = catmap.get(catn) or {'amount': 0.0, 'remark': None}
        amt, rmk = c['amount'], c.get('remark')
        ex = query("SELECT id, src FROM dock_yard WHERE vsl_nm=? AND category=?", (vsl_nm, catn), one=True)
        if ex and (ex['src'] or 'auto') == 'manual':   # 수동수정건: 금액/통화/remark 보존, metadata만 갱신
            execute("UPDATE dock_yard SET yard_name=?, vsl_cd=COALESCE(?,vsl_cd), sort_no=?, "
                    "updated_at=datetime('now','localtime') WHERE id=?", (yard_nm, vsl_cd, i, ex['id']))
            skipped += 1
            continue
        if ex:
            execute("UPDATE dock_yard SET amount=?, cur=?, remark=?, src='auto', "
                    "yard_name=?, vsl_cd=COALESCE(?,vsl_cd), sort_no=?, updated_at=datetime('now','localtime') WHERE id=?",
                    (amt, cur_default, rmk, yard_nm, vsl_cd, i, ex['id']))
            updated += 1
        else:
            execute("INSERT INTO dock_yard (vsl_nm, vsl_cd, category, amount, cur, remark, src, yard_name, sort_no) "
                    "VALUES (?,?,?,?,?,?,'auto',?,?)",
                    (vsl_nm, vsl_cd, catn, amt, cur_default, rmk, yard_nm, i))
            added += 1
    final = round(sum(c['amount'] for c in catmap.values()), 2)
    verified = not any('⚠️' in w for w in warns)
    return jsonify({'ok': True, 'source': source, 'verified': verified, 'warns': warns,
                    'added': added, 'updated': updated, 'skipped_manual': skipped,
                    'final_total': final})


@app.route('/api/dock_yard/<int:lid>', methods=['PATCH'])
@login_required
def api_dock_yard_patch(lid):
    if not query("SELECT id FROM dock_yard WHERE id=?", (lid,), one=True):
        return jsonify({'error': 'not found'}), 404
    d = request.get_json(silent=True) or {}
    sets, params = [], []
    if 'amount' in d:
        raw = d.get('amount')
        if raw in (None, ''):
            amt = None
        else:
            try:
                amt = float(str(raw).replace(',', ''))
            except (TypeError, ValueError):
                return jsonify({'error': 'amount must be numeric'}), 400
            if not math.isfinite(amt):
                return jsonify({'error': 'amount must be finite'}), 400
        sets.append('amount=?'); params.append(amt)
        sets.append("src='manual'")
    if 'cur' in d:
        cur = (d.get('cur') or '').strip().upper()
        if not re.fullmatch(r'[A-Z]{3}', cur):
            return jsonify({'error': 'cur must be 3-letter'}), 400
        sets.append('cur=?'); params.append(cur)
        sets.append("src='manual'")
    if 'remark' in d:
        sets.append('remark=?'); params.append(d.get('remark'))
    if d.get('src') == 'auto':                          # 🔒 언락 — 재업로드 시 덮어씀
        sets.append("src=?"); params.append('auto')
    if sets:
        sets.append("updated_at=datetime('now','localtime')")
        params.append(lid)
        execute(f"UPDATE dock_yard SET {', '.join(sets)} WHERE id=?", tuple(params))
    return jsonify({'ok': True})


@app.route('/api/ext/dock_procure/links')
@api_key_required
def api_ext_dockproc_links():
    """진단/폴러용 — 수동연결(svms_req_no 설정된) dock 행 목록."""
    vc = (request.args.get('vsl_cd') or '').strip().upper()
    rows = query(
        "SELECT d.req_no, d.svms_req_no, d.cat_code, d.stg_quote, d.stg_vendor, "
        "d.stg_confirm, d.stg_order, d.vsl_nm "
        "FROM dock_procure d WHERE d.svms_req_no IS NOT NULL AND d.svms_req_no<>'' "
        + ("AND (UPPER(d.vsl_cd)=? OR d.vsl_nm IN (SELECT vsl_nm FROM dock_procure_vessel WHERE UPPER(vsl_cd)=?))" if vc else ""),
        ((vc, vc) if vc else ()))
    return jsonify({'links': [dict(r) for r in rows]})


@app.route('/api/ext/dock_procure/sync', methods=['POST'])
@api_key_required
def api_ext_dockproc_sync():
    """Phase 2 역동기화 — 맥 폴러가 SVMS 수리/구매 목록을 보내면 Status→체크박스 자동전진 + 발주완료시 Vendor→Remark.
    매칭: ① 저장된 svms_req_no(=Inq No) ② Subject 태그 [VSL_CD REQ_NO]. HQ Canceled 무시. dry=true면 미리보기.
    item 옵션 `quotes`=벤더 제출견적 스냅샷(표시전용, 발주금액과 별개) — 키 미전송 시 기존값 유지.
    item 옵션 `orders`=발주서(ODR_NO)별 업체·금액(분할발주 표시용, `ord_vendors`) — 같은 3상태 계약."""
    import re as _re
    d = request.get_json(silent=True) or {}
    items = d.get('items') or []
    dry = bool(d.get('dry'))
    # 🔴 `partial` = 델타 폴(`dock_sync.py --fast`)이 **변화 후보 행만** 보낸 payload 라는 선언.
    #   미적재(orphan) 배너는 `_dockproc_orphans_save` 가 "payload 에 등장한 선박의 목록을 통째로
    #   갈아치우는" 구조라, 부분 payload 로 그걸 돌리면 **대기 중인 배너가 통째로 사라진다**
    #   (2026-08-06 설계 시 실코드 확인). 배너 갱신 책임은 전량 sync 한쪽에만 둔다.
    partial = bool(d.get('partial'))
    # 단계전이 → 푸시는 `push_outbox` 에 적재하고(행 UPDATE 보다 먼저) 루프 끝에서 비운다.
    TAG = _re.compile(r'\[([A-Z]{2,6})\s+((?:SY|ST|R|S|P)\d+)\]')
    canceled = 0
    unmatched = 0
    misses = []
    orphans = []                                         # 태그는 맞는데 발주현황에 행이 없는 청구(배너용)
    linked = []                                          # rank 0(견적의뢰 이전) 연결만 채운 행
    pc_keys = []                                         # 구매 REQ_NO(`svms_pc_req_no`) 를 새로 채운 행
    plan = {}                                            # row_id -> (rank, status, vendor, inq, row)
    for it in items:
        status = (it.get('status') or '').strip()
        if 'CANCEL' in status.upper():                   # HQ Canceled = 완전 무시
            canceled += 1
            continue
        rank = _dockproc_status_rank(status)
        # ⚠️rank 0 은 행을 아예 안 건드리므로 `svms_status` 라벨도 옛 값으로 남는다(올마이트 2026-08-01).
        #   실측(2026-08-01 전선박): 미등재 상태 = 'VSL Approved' 26 · 'Approved' 2 · 'HQ Received' 1 —
        #   ⛔ 이 뒤에 있던 "상신된 건이 여기로 되돌아오는 경로가 아니다 / 반려는 'RE'(rank 2)로 돌아오니
        #      갱신 경로에 걸린다"는 전제는 **2026-08-03 실측으로 반증됐다.** SVMS 에서 견적요청을
        #      **회수**하면 헤더는 'RE' 가 아니라 'AP'(HQ Received, rank 0)로 돌아온다 → 옛 로직에선
        #      `stg_vendor=1` 이 영구히 남아 게이트가 영영 잠겼다(BGBBME26073116). 그래서 아래
        #      `_DOCKPROC_PRE_INQUIRY` allowlist 로 **확인된 pre-inquiry 라벨은 되돌림**을 허용한다.
        #   미지의 상태로 굳는 경우는 종전대로 '게이트 닫힘'(=재컨펌 불가)으로 남고, 오상신은 워커의
        #   pre-read `STATUS=='RE'` 게이트가 최종 차단한다 — 안전한 방향으로 실패한다.
        # 🔴 2026-08-03 보강: rank 0 을 완전 스킵하면 **SVMS 연결(`svms_req_no`)조차 안 채워진다**.
        #    실사고: 형이 캡쳐한 `BGBBME26073116`([BGBB R22], HQ Received)이 dock_procure 129 행과
        #    태그로 매칭되는데도 `svms_req_no=NULL` 이라 견적요청 버튼이 쓸 REP_CD 가 없었다(BGBB R 7행 동일).
        #    그래서 rank 0 은 **연결 전용 경로**로 내린다 — 단계(stg_*)·금액·remark 는 손대지 않는다.
        # 🔴 확인된 pre-inquiry 라벨은 link_only 에서 빼서 정상 경로로 보낸다 = 단계(stg_*)를 0 으로
        #    되돌려 회수를 반영한다. 금액·remark 는 정상 경로에서도 `o`(발주완료)=0 이면 손대지 않으므로
        #    되돌아가는 건 단계와 라벨뿐이다. 빈 라벨('')은 allowlist 에 없으니 종전대로 link_only 다.
        pre_inq = (rank == 0 and (status or '').strip().upper() in _DOCKPROC_PRE_INQUIRY)
        link_only = (rank == 0 and not pre_inq)

        evidence = it.get('ordered_evidence')            # True/False/None(=근거 미확정) — 행 매칭 후 rank 게이트에 씀
        inq = (it.get('inq_no') or '').strip() or None
        inq_alt = (it.get('inq_alt') or '').strip() or None   # 구매 INQ_NO(REQ_NO와 별개) — 둘 다 매칭키
        subj = it.get('subject') or ''
        row = None
        tag = None
        cand = [c for c in (inq, inq_alt) if c]
        if cand:                                              # 저장된 svms_req_no가 REQ_NO/INQ_NO 어느 쪽이든 매칭
            qm = ",".join("?" * len(cand))
            row = query(f"SELECT * FROM dock_procure WHERE svms_req_no IN ({qm})", tuple(cand), one=True)
        if not row:
            m = TAG.search(subj)
            if m:
                vc, rq = m.group(1).upper(), m.group(2).upper()
                tag = (vc, rq)
                row = query(
                    "SELECT * FROM dock_procure WHERE UPPER(req_no)=? AND (UPPER(vsl_cd)=? "
                    "OR vsl_nm IN (SELECT vsl_nm FROM dock_procure_vessel WHERE UPPER(vsl_cd)=?))",
                    (rq, vc, vc), one=True)
        if not row:
            if not link_only:                            # rank 0 은 원래 통째로 스킵이던 구간 — 지표 왜곡 방지
                unmatched += 1
                if len(misses) < 20:
                    misses.append({'inq': inq, 'subject': subj[:70]})
            # 🔴 태그가 있는데 붙을 행이 없는 건 = 'INDEX 엑셀에 없는 시트번호로 나간 청구'다.
            #    종전엔 카운터로만 세고 버려서 형 화면에 흔적이 0 이었다(2026-08-05 BGBB 7건).
            #    여기서 자동으로 행을 만들지는 않는다 — 배너에 띄우고 사람이 [적재] 를 누른다
            #    (자동생성 = 하드 DELETE 로 지운 행 부활 + 옛 입거 잔상 유입 위험).
            #    `link_only`(rank 0) 여부와 무관하게 남긴다 — 단계가 아직 0 이어도 없는 행은 없는 행이다.
            if tag:
                orphans.append({'vsl_cd': tag[0], 'req_no': tag[1], 'subject': subj[:160],
                                'status': status, 'doc': (it.get('doc') or '').strip().upper(),
                                # 적재 시 채울 SVMS 키 — 구매는 REQ_NO(`inq_alt`), 수리는 REP_CD(`inq_no`)
                                'key': (inq_alt if (it.get('doc') or '') == 'PC' else inq)})
            continue
        # 🔴 구매 견적요청 키(REQ_NO)는 **별도 칸**(`svms_pc_req_no`)에 적는다. `svms_req_no` 에 섞으면
        #    Phase ③ 상신·제출견적·첨부가 그 칸을 INQ_NO 로 읽으므로 깨진다. 폴러가 `inq_alt`(=REQ_NO)를
        #    매 sync 마다 이미 실어보내므로 추가 SVMS 조회·선명 정규화 조인이 필요 없다.
        #    ⚠️`rank`·`link_only` 와 **무관하게** 채운다 — 'VSL Approved'(rank 0) 행도 나중에 본사확인이
        #      떨어지면 바로 요청 가능해야 하고, 키를 채워도 `_dock_inq_blocked` 의 라벨 게이트가 막는다.
        #    ⚠️컬럼 존재 여부를 확인하고 쓴다 — 마이그레이션 전 DB 에서 KeyError 로 sync 전체(수리 포함)가
        #      500 나면 안 된다(닫힘 쪽 실패: 키만 안 채워지고 버튼이 안 열린다).
        if (it.get('doc') or '') == 'PC' and inq_alt:
            _pcq = ((row['svms_pc_req_no'] if 'svms_pc_req_no' in row.keys() else None) or '').strip()
            if _pcq != inq_alt:
                pc_keys.append({'id': row['id'], 'req_no': row['req_no'], 'vsl_nm': row['vsl_nm'],
                                'pc_req_no': inq_alt, 'was': _pcq or None})
                if not dry:
                    execute("UPDATE dock_procure SET svms_pc_req_no=?, "
                            "updated_at=datetime('now','localtime') WHERE id=?", (inq_alt, row['id']))
        # 🔴 되돌림 fail-closed(2026-08-03, 올마이트 지적 수용): **발주 흔적이 있는 행은 라벨 하나로
        #    퇴행시키지 않는다.** stale·순서역전 sync 나 미지의 SVMS lifecycle 로 발주완료 행이 rank 0
        #    라벨로 보이면, 되돌림이 `stg_order`·발주금액 이력을 조용히 지울 수 있다. 회수 되돌림은
        #    **견적요청 단계의 회수**만 대상이므로 발주근거가 있으면 종전대로 link_only(=닫힘 쪽)로 남긴다.
        #    닫힘 쪽 실패 = 사람이 수동으로 체크를 풀면 되고, 오상신은 워커 pre-read 게이트가 막는다.
        #    ⚠️2026-08-04 정정: `svms_submit` 은 **존재 여부**로 보지 않는다 — 견적요청 직후 `"0/0"`·
        #      `"0/1"` 로 채워지므로 모든 회수건이 이 가드에 걸려 되돌림이 영구 무력화됐다. 제출수>0
        #      (또는 미지 형식)일 때만 흔적으로 인정한다(`_dockproc_submit_has_quotes`).
        if pre_inq and (row['stg_order'] or row['quote_amt'] is not None
                        or _dockproc_submit_has_quotes(row['svms_submit'])):
            link_only = True
        if link_only:
            # 연결만 채운다. 라벨은 **아직 어느 단계도 안 켜진 행**에만 쓴다 —
            # 상신 이후 라벨을 rank 0 라벨로 되돌리면 재컨펌/재상신 게이트가 열릴 수 있어서다
            # (`_dock_submit_prior` 는 라벨이 상신 이후인지로 판정). 단계가 0 이면 상신 이력도 없다.
            fresh = not (row['stg_vendor'] or row['stg_confirm'] or row['stg_order'])
            need_inq = bool(inq) and not (row['svms_req_no'] or '').strip()
            need_lbl = fresh and status and (row['svms_status'] or '') != status
            if need_inq or need_lbl:
                linked.append({'id': row['id'], 'req_no': row['req_no'], 'vsl_nm': row['vsl_nm'],
                               'inq_no': inq if need_inq else None,
                               'status': status if need_lbl else None})
                if not dry:
                    execute(
                        "UPDATE dock_procure SET svms_req_no=COALESCE(NULLIF(svms_req_no,''),?), "
                        "svms_status=CASE WHEN ? THEN ? ELSE svms_status END, "
                        "svms_synced_at=datetime('now','localtime'), updated_at=datetime('now','localtime') "
                        "WHERE id=?",
                        (inq, 1 if need_lbl else 0, status, row['id']))
            continue
        # 🔴 발주완료 fail-closed 게이트(2026-07-31): 헤더 상태 allowlist 만으로 rank4 를 켜지 않는다.
        #   근거(evidence) = 수리 `VNDR_STATS=='Ordered'` 또는 `ODR_YN=='Y'` / 구매 발주서번호 `ODR_NO` 존재.
        #   False = 근거 없음 → 벤더컨펌(rank3)까지만 인정.
        #   None  = 근거 미확정(SVMS 상세조회 실패·구버전 폴러) → **이미 발주완료인 행만 유지**하고
        #           신규 승격은 막는다. 이렇게 안 하면 조회 한 번 실패했을 때 근거 없이 발주완료가 켜짐(올마이트 지적).
        if rank >= 4 and (evidence is False or (evidence is None and not row['stg_order'])):
            rank = 3
        # 🔴 분할발주 부분완료 게이트(2026-08-05 형 확인 기준 = **전부 발주돼야 완료**): 한 청구를 업체
        #   2곳으로 나눠 발주하면 발주서가 2장이고 PC_PRO 는 장마다 행을 준다. 한 장만 승인돼도 그 행의
        #   rank 는 4 라서, 아래 '최고 rank 채택' 규칙 때문에 **절반만 발주된 건이 발주완료로 켜졌다**
        #   (실측 [BGBB S1]: 에버런스 발주완료 + 딘텍 결재중 → 카드는 발주완료·에버런스 금액만).
        #   ⇒ 하나라도 미승인이면 벤더컨펌(rank 3)까지만 인정하고, 부분완료 사실은 화면이 `ord_vendors`
        #     로 "발주 1/2" 처럼 그대로 말한다.
        #   ⚠️**이미 발주완료인 행은 내리지 않는다**(`row['stg_order']` 가드) — stale·순서역전 sync 나
        #     처음 보는 lifecycle 로 완료 이력이 조용히 되돌아가면 안 된다(위 `evidence is None` 가드와
        #     같은 방향 = 닫힘 쪽 실패).
        #   ⚠️`len>1` 로 좁힌 이유: 발주서 1장인 보통 건은 위 `evidence` 게이트가 이미 같은 판정을 한다.
        #     여기서 1장까지 보면 규칙이 두 곳으로 갈려 판정이 어긋날 수 있다(동작 변화 0 을 보장).
        #   ⚠️키 미전송(`orders` 없음)일 때는 **저장된 스냅샷으로 판정한다** — 폴러가 한 번 못 실어보냈다고
        #     게이트가 열려 stg_order 가 켜지면, 화면엔 같은 카드에 '발주완료' 와 '발주 1/2' 가 같이 뜬다
        #     (모순 = 형이 어느 쪽을 믿을지 알 수 없음). 스냅샷도 없으면 종전 `evidence` 규칙만 적용된다.
        #   ⚠️판정 입력은 **정규화 결과**다(올마이트 2026-08-05 지적 수용). raw 를 세면 무효·중복 ODR_NO 까지
        #     세어져 저장값(canonical 1건)과 판정(2건)이 갈린다 = 화면은 1곳인데 게이트만 닫히는 불일치.
        _ordraw = it.get('orders')
        if not isinstance(_ordraw, list):
            _orders = False                              # 미전송 → 기존값 유지
        elif not _ordraw:
            _orders = None                               # 발주 0건 확정 → clear
        else:
            _orders = _dockproc_norm_orders(_ordraw) or False   # 전부 무효 = 계약위반 → 유지
        if isinstance(_orders, str):
            _gate = _dockproc_orders_of(_orders)
        elif _orders is None:
            _gate = []                                   # 발주 0건 = 분할발주 아님(게이트 대상 아님)
        else:
            _gate = _dockproc_orders_of(row['ord_vendors'] if 'ord_vendors' in row.keys() else None)
        if (rank >= 4 and not row['stg_order'] and len(_gate) > 1
                and not all(o.get('ordered') is True for o in _gate)):
            rank = 3
        prev = plan.get(row['id'])
        if not prev or rank > prev[0]:                   # 같은 행 여러건이면 최고 rank만(취소 제외 후)
            _amt = it.get('amt')
            try:
                _amt = None if _amt in (None, '') else float(str(_amt).replace(',', ''))
            except (TypeError, ValueError):
                _amt = None                              # 파싱 실패=자동입력 안 함(0 저장 방지)
            # 제출견적 스냅샷 3상태 — 값이 조회실패로 사라지는 경로를 전부 막는다(올마이트 지적).
            #   키 없음/리스트 아님   → False = 미전송 → 기존 유지
            #   빈 리스트             → None  = '제출 0건' 확정 → clear
            #   내용 있지만 전부 쓰레기 → False = 계약 위반 패킷으로 보고 기존 유지(clear 아님)
            _raw_q = it.get('quotes')
            if not isinstance(_raw_q, list):
                _quotes = False
            elif not _raw_q:
                _quotes = None
            else:
                _quotes = _dockproc_norm_quotes(_raw_q) or False
            # 견적서 첨부 목록도 같은 3상태 계약(키 없음=기존 유지 / [] =첨부 0건 확정 / 내용=교체).
            _raw_f = it.get('files')
            if not isinstance(_raw_f, list):
                _files = False
            elif not _raw_f:
                _files = None
            else:
                _files = _dockproc_norm_files(_raw_f) or False
            # 분할발주 스냅샷(`_orders`)과 그 정규화 결과(`_gate`)는 위 게이트에서 이미 만들었다 —
            # 여기서 다시 파싱하면 판정과 저장이 갈릴 수 있어 **같은 값을 그대로** 싣는다.
            plan[row['id']] = (rank, status, (it.get('vendor') or '').strip() or None,
                               inq, row, (it.get('submit') or '').strip() or None,
                               _amt, (it.get('cur') or '').strip().upper() or None, _quotes, _files,
                               _orders, _gate)
    changes = []
    for rid, (rank, status, vendor, inq, row, submit, amt, cur, quotes, files, orders, gate) in plan.items():
        # 🔴 사람이 확정한 단계 floor(2026-08-07 형 지시). SVMS 밖 발주(메일·직접)는 SVMS 라벨이 절대
        #   따라오지 않으므로, 사람이 켠 단계를 SVMS rank 로 되돌리면 매시간 입력이 지워진다.
        #   ⚠️floor 는 **단계 표시(stg_*)에만** 적용한다. 발주금액 자동입력·푸시 판정은 아래 `svms_o`
        #     (=SVMS 가 실제로 발주완료라고 말한 경우)로 계속 게이트한다 — 사람 체크 하나로 돈 경로
        #     자동입력이나 '발주완료' 푸시가 열리면 안 된다.
        #   ⚠️컬럼 없는 구버전 DB 에서 KeyError 로 sync 전체가 죽지 않게 keys() 확인 후 읽는다.
        svms_o = 1 if rank >= 4 else 0
        _mfloor = (row['stg_manual'] or 0) if 'stg_manual' in row.keys() else 0
        rank_eff = max(rank, _mfloor)
        q, v, f, o = ((1 if rank_eff >= 1 else 0), (1 if rank_eff >= 2 else 0),
                      (1 if rank_eff >= 3 else 0), (1 if rank_eff >= 4 else 0))
        new_remark = row['remark']
        # 옵션 b: 발주완료 시 Vendor명을 Remark에 기입. 단 신규완료/빈Remark일 때만(매폴 수동메모 덮어쓰기 방지)
        if svms_o and vendor and (not row['stg_order'] or not (row['remark'] or '').strip()):
            new_remark = vendor
        # 🔴 분할발주(업체 2곳 이상) 발주금액은 **다시 계산한다**(올마이트 2026-08-05 지적 수용).
        #   폴러가 주는 `amt` 는 INQ 단위 합산인데 **통화가 섞이면 첫 건만 남긴 값**이라, 그대로 쓰면
        #   한 업체 금액이 전체 발주금액으로 박힌다(형이 전에 잡은 "이게 최종비용이잖아" 와 같은 부류).
        #   ⇒ 발주서 스냅샷으로 **전원 금액확정 + 단일통화일 때만** 합계를 쓰고, 하나라도 미확정이거나
        #     통화가 섞였으면 칸을 **비워둔다** — 업체별 줄이 진실을 말한다(거짓 대표금액 금지).
        if len(gate) > 1:
            _gamts = [g.get('amt') for g in gate]
            _gcurs = {(g.get('cur') or '') for g in gate}
            if all(isinstance(x, (int, float)) for x in _gamts) and len(_gcurs) == 1:
                amt, cur = float(sum(_gamts)), (next(iter(_gcurs)) or None)
            else:
                amt = None
        # 발주금액 자동입력: 발주완료(o)·금액있음·manual아님 일 때만(사용자 수정 우선)
        set_q = (svms_o == 1 and amt is not None and (row['quote_src'] or 'auto') != 'manual')
        new_qamt = amt if set_q else row['quote_amt']
        new_qcur = ((cur if (cur and _re.fullmatch(r'[A-Z]{3}', cur)) else 'USD')
                    if set_q else row['quote_cur'])      # SVMS CUR_CD 이상값 방어
        new_qsrc = 'auto' if set_q else (row['quote_src'] or 'auto')
        new_subq = row['sub_quotes'] if quotes is False else quotes
        new_att = row['att_files'] if files is False else files
        # 컬럼 존재 확인 후 읽는다 — 마이그레이션 전 DB 에서 KeyError 로 sync 전체(수리 포함)가 500 나면
        # 안 된다(닫힘 쪽 실패: 스냅샷만 안 채워지고 나머지 동기화는 정상).
        _cur_ords = row['ord_vendors'] if 'ord_vendors' in row.keys() else None
        new_ords = _cur_ords if orders is False else orders
        # 🔴 `svms_status` 를 비교대상에 포함해야 한다(2026-08-01 실측). 빠뜨리면 **단계가 같은 라벨
        #    전이**(예: 'Quotation Inquiry'→'Submit', 둘 다 rank 2)가 '변경 없음'으로 판정돼 라벨이
        #    영영 갱신되지 않는다. 실사고: BGBBME26073108 은 SVMS 가 'Submit' 인데 DB 는 하루 넘게
        #    'Quotation Inquiry' 였음. 표시만의 문제가 아니라 **재컨펌 게이트가 이 라벨을 읽으므로**,
        #    SVMS 에서 반려돼 라벨이 되돌아가도 sync 가 못 써서 게이트가 영구 잠기는 경로가 됨.
        # 관측 전용(동작 변경 없음) — 구매행에 **다른** INQ_NO 가 들어오면 아래 `COALESCE` 가 그걸 버린다.
        #   버리는 건 의도다: 한 REQ 에 INQ 가 여럿 공존하는 게 실측됐고(2026-08-04 B61 껍데기 + B62 정상)
        #   덮으면 Phase ③ 상신이 쓰는 rep_cd 가 sync 마다 흔들린다(돈경로 키 flapping).
        #   다만 '회수→재요청' 이 sync 사이에 다 끝나면 죽은 번호가 남는 잔여 창이라, 실제로 일어나는지
        #   **로그로만** 남긴다. 실측되면 그때 교체 규칙(제출이력·발주흔적 fail-closed 포함)을 설계한다.
        if ((inq or '').strip() and (row['svms_req_no'] or '').strip()
                and inq.strip() != row['svms_req_no'].strip()
                and _DOCK_INQ_DOC.get((row['cat_code'] or '').strip().upper()) == 'PCRQ'):
            app.logger.warning('dock sync: 구매 INQ_NO 불일치(유지) rid=%s 보관=%s 수신=%s 라벨=%s',
                               rid, row['svms_req_no'], inq, status)
        before = (row['stg_quote'], row['stg_vendor'], row['stg_confirm'], row['stg_order'], row['remark'],
                  row['svms_req_no'], row['svms_status'], row['svms_submit'],
                  row['quote_amt'], row['quote_cur'], row['quote_src'],
                  row['sub_quotes'], row['att_files'], _cur_ords)
        after = (q, v, f, o, new_remark, row['svms_req_no'] or inq, status, submit,
                 new_qamt, new_qcur, new_qsrc, new_subq, new_att, new_ords)   # COALESCE(기존,신규)=멱등
        if before != after:
            changes.append({'id': rid, 'req_no': row['req_no'], 'vsl_nm': row['vsl_nm'],
                            'status': status, 'stages': [q, v, f, o],
                            'remark': new_remark, 'inq_no': inq, 'submit': submit,
                            'quote_amt': new_qamt, 'quote_cur': new_qcur, 'quote_src': new_qsrc,
                            'sub_quotes': new_subq, 'att_files': new_att, 'ord_vendors': new_ords})
            # 🔴 푸시는 여기서 **판정 + 대기함 적재**만 하고 발송은 루프 끝에서 한다. APNs 왕복(수백 ms)
            #   을 이 안에 넣으면 행 수만큼 폴 시간이 늘고, 발송 예외 하나가 남은 행의 동기화까지 날린다.
            #   ⚠️적재는 반드시 **아래 UPDATE 보다 먼저** — 순서가 뒤집히면 그 틈에 죽었을 때
            #     "행은 갱신됐는데 알림은 영영 없는" 미탐이 남는다(올마이트 블로커 지적).
            if not dry:
                try:
                    for _ev in _dockproc_push_events(row, status, svms_o, submit, vendor,
                                                     new_qamt, new_qcur):
                        _push_outbox_add(_ev['kind'], _ev['event_key'], _ev['title'], _ev['body'],
                                         link='trmt://dock', collapse_id=_ev['collapse'])
                except Exception:
                    app.logger.exception('dock sync push 판정 실패 rid=%s', rid)
            if not dry:
                execute(
                    "UPDATE dock_procure SET stg_quote=?, stg_vendor=?, stg_confirm=?, stg_order=?, remark=?, "
                    "svms_req_no=COALESCE(svms_req_no,?), svms_status=?, svms_submit=?, "
                    "quote_amt=?, quote_cur=?, quote_src=?, sub_quotes=?, att_files=?, ord_vendors=?, "
                    "svms_synced_at=datetime('now','localtime'), updated_at=datetime('now','localtime') WHERE id=?",
                    (q, v, f, o, new_remark, inq, status, submit, new_qamt, new_qcur, new_qsrc,
                     new_subq, new_att, new_ords, rid))
                if row['att_files'] != new_att:           # 목록이 바뀌면 안 쓰는 캐시 정리(용량 회수)
                    _dockatt_gc(rid, _dockproc_files_of(new_att))
                # 🔴 회수 되돌림이면 **직전 견적요청 이력을 무효화**한다(2026-08-03 형 지시 "재적재 시
                #    이전에 견적요청된 인포가 남아있는데 초기화해줘").
                #    이유: 웹/앱은 `status=='submitted'` 를 '견적요청됨 ✓' 으로 그린다. 단계는 되돌아가도
                #    이 draft 가 남아 있어서 카드에 초록 체크·"견적요청됨" 이 계속 붙었다 = 실제와 불일치.
                #    지우지 않고 `recalled` 로 **보존**한다(append-only 이력). 어느 버킷에도 안 잡혀
                #    표시가 사라지고, 앱은 재빌드 없이도 즉시 정상화된다(버킷 매칭이 클라이언트 판정이라).
                #    ⚠️활성 큐(`approved`/`submitting`)는 **절대 건드리지 않는다** — 워커가 소유한 행이고
                #      전송 중일 수 있다. 그건 `_dock_inq_blocked` 가 '이미 큐에 있음'으로 계속 막는 게 맞다.
                #      이 행이 나중에 `submitted` 로 완료돼 표시가 되살아나는 건 **버그가 아니다** —
                #      그건 형이 회수 후 새로 올린 요청이 실제로 SVMS 로 나간 것이라 표시가 맞다
                #      (회수된 옛 요청의 draft 는 회수 시점에 이미 `submitted` = 여기서 전이된다).
                #    조건은 `rank==0` 같은 간접지표가 아니라 **①allowlist 라벨 ②직전 단계 켜짐
                #      ③이번 계산 결과 단계 전부 꺼짐** 3개를 직접 확인한다(올마이트 지적 수용) —
                #      나중에 rank 0 라벨이 늘어도 과잉 전이되지 않게.
                #    🔴 2026-08-03 2차(형 제보 BGBB S14): **구매 회수는 라벨이 'HQ Confirmed'(rank 1)로
                #      돌아온다** — 단계가 통째로 0 이 되지 않고 `견적작성`(q=1)만 남는다. 그래서 위
                #      "단계 전부 꺼짐" 조건에 걸리지 않아 draft 가 살아남았다. 회수의 본질은
                #      **견적요청(=벤더제출) 이후 단계가 되돌려진 것**이므로 그 갈래를 따로 인정한다.
                #      라벨 allowlist 도 문서종류별 확정 pre-요청 라벨(`_DOCK_INQ_PRE`)을 합집합으로 본다.
                _inq_doc, _inq_key = _dockproc_inq_target(row)
                _pre_lbls = _DOCKPROC_PRE_INQUIRY | set(_DOCK_INQ_PRE.get(_inq_doc, ()))
                _had_post = bool(row['stg_vendor'] or row['stg_confirm'] or row['stg_order'])
                # 🔴 갈래 ② fail-closed(올마이트 2026-08-03 지적 수용): **발주 흔적이 있는 행은 라벨
                #    하나로 이력을 무효화하지 않는다.** 위 `pre_inq` 되돌림에 이미 걸어둔 것과 같은 기준.
                #    stale·순서역전 sync 로 발주완료 행이 'HQ Confirmed' 로 보이면 갈래 ② 는 단계 flag
                #    만 보고 전이해버린다 → 옛 요청 이력이 조용히 사라져 사람이 확인할 근거를 잃는다.
                #    닫힘 쪽 실패 = 표시가 남는 것뿐이고, 사람이 '완료건 지우기'로 치울 수 있다.
                #    ⚠️2026-08-04 정정: `svms_submit` 존재 여부 → **제출수>0** 으로 좁힌다(위 `pre_inq`
                #      가드와 동일 사유·동일 함수). `"0/0"`·`"0/1"` 은 회수건이 항상 들고 있는 값이라
                #      존재 여부로 보면 갈래 ② 전이가 영구히 죽는다.
                _no_odr = not (row['stg_order'] or row['quote_amt'] is not None
                               or _dockproc_submit_has_quotes(row['svms_submit']))
                if _had_post and not _no_odr:
                    _had_post = False
                _recalled = ((status or '').strip().upper() in _pre_lbls
                             and ((  # ① 단계가 통째로 되돌아감 (수리 'HQ Received' 실측 경로)
                                     (row['stg_quote'] or row['stg_vendor']
                                      or row['stg_confirm'] or row['stg_order'])
                                     and not (q or v or f or o))
                                  # ② 견적요청 이후 단계만 되돌아감 (구매 'HQ Confirmed' 실측 경로)
                                  or (_had_post and not (v or f or o))))
                if _recalled:
                    # 🔴 키는 문서종류별로 다르다 — 수리 `svms_req_no`(REP_CD) / 구매 `svms_pc_req_no`(REQ_NO).
                    #    `svms_req_no` 만 보면 구매 draft 의 `rep_cd` 와 절대 안 맞아 전이가 no-op 이 된다
                    #    (S14: `svms_req_no=NULL`, draft.rep_cd='BGBBES2607B4'=`svms_pc_req_no`).
                    _rep = (_inq_key or row['svms_req_no'] or inq or '').strip()
                    if _rep:
                        # `result` 가 비어 있을 때 선행 ' · ' 가 붙지 않게 분기(올마이트 지적).
                        execute(
                            "UPDATE dock_inquiry_draft SET status='recalled', "
                            "result=CASE WHEN COALESCE(result,'')='' THEN '' ELSE result || ' · ' END "
                            "|| ? || datetime('now','localtime') || ')' "
                            "WHERE rep_cd=? AND status IN ('submitted','failed')",
                            ('SVMS 회수로 무효화(%s, ' % status, _rep))
                    # 🔴 구매 회수는 SVMS 에서 **INQ_NO 자체가 삭제**된다(2026-08-04 실측: `SP_SET_INQ_RTN`
                    #    으로 BGBBES2607B41 회수 → INQ 목록 12→10행). 그런데 `svms_req_no` 는 위 본문
                    #    UPDATE 가 `COALESCE(svms_req_no,?)` 라 **한 번 박히면 영영 안 덮인다** → 죽은
                    #    번호를 계속 들고 있고, 재요청으로 새 INQ_NO 가 나와도 그 자리에 못 들어온다.
                    #    Phase ③ 상신(`api_dock_submit_*`)이 이 칸을 rep_cd 로 그대로 쓰므로
                    #    **삭제된 문서번호로 상신이 나가는 경로**다 → 회수 시점에 비운다.
                    #    비우면 폴러가 다음 sync 에서 현재 INQ_NO 를 COALESCE 로 정상 적재한다.
                    #    ⚠️구매(PCRQ)만. 수리의 `svms_req_no` 는 REP_CD = 견적요청 키 그 자체라 비우면
                    #      버튼이 죽는다(fail-open). 문서종류 미상('')·페인트도 손대지 않는다.
                    #    ⚠️이번 sync 가 INQ_NO 를 실어왔으면(=SVMS 에 아직 살아있음) 건드리지 않는다.
                    #    ⚠️발주 흔적(`_no_odr`)은 **중복 방어**다 — 지금은 상류에서 이미 걸러진다
                    #      (rank0 라벨은 위 `pre_inq` 가드가 link_only 로 빼고, rank1 갈래 ② 는
                    #      `_had_post` 가 꺼진다). 그래서 mutation 으로 지워도 red 가 안 난다.
                    #      그래도 남긴다: 이 칸이 Phase ③ 상신 rep_cd 라 돈경로에 붙어 있고,
                    #      상류 가드 2개가 바뀌면 조용히 열리는 자리이기 때문.
                    #      닫힘 쪽 실패 = 낡은 번호가 남는 것뿐이고 사람이 '연결'로 고칠 수 있다.
                    #    ⚠️남는 창(미수정, 의도): 회수와 재요청 사이에 sync 가 **한 번도** 안 돌면 이
                    #      전이 자체가 안 일어나 옛 번호가 남는다. 실무상 회수 후 버튼 게이트가
                    #      직전 라벨로 잠겨 있어 재요청 전에 sync 가 먼저 도는 게 정상 경로다.
                    if (_inq_doc == 'PCRQ' and _no_odr
                            and not (inq or '').strip() and (row['svms_req_no'] or '').strip()):
                        execute("UPDATE dock_procure SET svms_req_no=NULL, "
                                "updated_at=datetime('now','localtime') WHERE id=?", (rid,))
                        app.logger.info('dock sync: 회수로 svms_req_no 비움 rid=%s old=%s label=%s',
                                        rid, row['svms_req_no'], status)
                        # 응답에도 남긴다 — 본문 UPDATE 와 별개 write 라 안 적으면 `changes` 만 보고는
                        # 번호가 사라진 걸 알 수 없다(올마이트 지적). `changes[-1]` = 바로 위에서
                        # 이 행이 append 한 항목이다(같은 반복, 사이에 다른 append 없음).
                        changes[-1]['req_no_cleared'] = row['svms_req_no']
    if not dry and not partial:                          # 부분 payload 로 배너를 갈아치우지 않는다
        _dockproc_orphans_save(items, orphans)
    # 대기함 비우기 — 이번에 적재한 것 + **지난 폴에서 발송 실패한 것**을 같이 보낸다.
    # 🔴 발송 실패는 로그만 남기고 sync 결과를 뒤집지 않는다(동기화는 이미 성공했고, 못 보낸 알림은
    #   대기함에 남아 다음 폴이 이어받는다).
    pushed = []
    if not dry:
        try:
            pushed = _push_outbox_drain()
        except Exception:
            app.logger.exception('dock sync push 대기함 발송 실패')
    return jsonify({'dry': dry, 'partial': partial, 'pushed': pushed,
                    'matched': len(plan), 'updated': len(changes),
                    'unmatched': unmatched, 'canceled_skipped': canceled,
                    'changes': changes, 'misses': misses,
                    'linked': linked, 'linked_n': len(linked),
                    'pc_keys': pc_keys, 'pc_keys_n': len(pc_keys),
                    'orphans': orphans, 'orphans_n': len(orphans)})


# ---- 벤더 견적서(SVMS MAOE 첨부) 원본 확인 ----
#   흐름: 폴러 sync 가 목록(att_files)을 적재 → 폴러가 pending 을 물어 **없는 파일만** NAS 에서 받아
#   업로드 → 웹/앱이 조회 라우트로 열어본다. 읽기전용이고 금전효과 없음(SVMS 로 나가는 write 아님).
_DOCKATT_MAX_IDX = _DOCKPROC_ATT_MAX - 1


@app.route('/api/ext/dock_procure/attachments/pending')
@api_key_required
def api_ext_dockproc_att_pending():
    """폴러용 — 목록에는 있는데 preview cache 에 없는 견적서. 이미 받은 건 다시 안 받는다(콜·용량 절약).
    `svms_req_no` 가 있어야 폴러가 SVMS 에서 그 건을 다시 찾을 수 있으므로 그 행만 준다."""
    try:
        limit = max(1, min(500, int(request.args.get('limit') or 60)))
    except (TypeError, ValueError):
        limit = 60
    vc = (request.args.get('vsl_cd') or '').strip().upper()
    rows = query(
        "SELECT id, vsl_nm, vsl_cd, req_no, svms_req_no, att_files FROM dock_procure "
        "WHERE att_files IS NOT NULL AND att_files<>'' AND svms_req_no IS NOT NULL AND svms_req_no<>'' "
        + ("AND (UPPER(vsl_cd)=? OR vsl_nm IN (SELECT vsl_nm FROM dock_procure_vessel WHERE UPPER(vsl_cd)=?)) " if vc else "")
        + "ORDER BY id",
        ((vc, vc) if vc else ()))
    disk = _dockatt_disk_map()
    out = []
    for r in rows:
        have = disk.get(r['id']) or {}
        for idx, f in enumerate(_dockproc_files_of(r['att_files'])):
            fp = _dockatt_fp(f)
            if (idx, fp) in have or idx > _DOCKATT_MAX_IDX:
                continue
            # `fp` 는 폴러가 업로드할 때 그대로 되돌려줘야 하는 토큰이다. 그 사이 목록이 바뀌면
            # 지문이 달라져 서버가 409 로 거절한다 → 옛 파일이 새 첨부 자리에 저장되는 race 차단.
            out.append({'id': r['id'], 'vsl_nm': r['vsl_nm'], 'vsl_cd': r['vsl_cd'],
                        'req_no': r['req_no'], 'svms_req_no': r['svms_req_no'], 'idx': idx,
                        'fp': fp, 'nm': f.get('nm'), 'sv': f.get('sv'), 'kb': f.get('kb')})
            if len(out) >= limit:
                return jsonify({'pending': out, 'truncated': True})
    return jsonify({'pending': out, 'truncated': False})


@app.route('/api/ext/dock_procure/<int:rid>/attachments/<int:idx>', methods=['POST'])
@api_key_required
def api_ext_dockproc_att_upload(rid, idx):
    """맥 폴러가 NAS 에서 받은 견적서 원본을 preview cache 로 적재. body = 파일 바이트 그대로.
    확장자는 ?ext= → 저장된 파일명 순으로 정한다(allowlist 밖이면 거부).

    🔴 `?fp=` 필수 = pending 이 준 **그 첨부**가 맞는지 대조(올마이트 지적 반영). 폴러가 pending 을
      받아 NAS 에서 받아오는 동안 sync 로 목록이 바뀌면 같은 idx 가 다른 파일을 가리키게 되는데,
      그때 지문이 어긋나 409 로 거절된다 → 다음 폴에서 새 지문으로 다시 받아간다.
      (예전엔 'idx 가 목록 범위 안인지'만 봐서, 옛 파일이 새 첨부 자리에 저장될 수 있었다.)"""
    if idx < 0 or idx > _DOCKATT_MAX_IDX:
        abort(404)
    if request.content_length and request.content_length > _FUNDREQ_ATT_MAX:
        return jsonify({'error': 'too large'}), 413
    row = query("SELECT id, att_files FROM dock_procure WHERE id=?", (rid,), one=True)
    if not row:
        abort(404)
    files = _dockproc_files_of(row['att_files'])
    if idx >= len(files):                                # 목록에 없는 자리에 파일을 꽂으면 이름↔내용이 어긋난다
        return jsonify({'error': 'idx out of list'}), 409
    fp = _dockatt_fp(files[idx])
    if (request.args.get('fp') or '').strip() != fp:     # fail-closed — 지문 없거나 다르면 저장 안 함
        return jsonify({'error': 'fingerprint mismatch', 'expect': fp}), 409
    ext = _fundreq_att_ext(request.args.get('ext')) or _fundreq_att_ext(files[idx].get('nm'))
    if not ext:
        return jsonify({'error': 'unsupported type'}), 400
    data = request.get_data()
    if not data:
        return jsonify({'error': 'empty'}), 400
    if len(data) > _FUNDREQ_ATT_MAX:
        return jsonify({'error': 'too large'}), 413
    if not _fundreq_att_sniff_ok(ext, data):             # 확장자 위장 방지(inline 서빙되는 경로라 필수)
        return jsonify({'error': 'content/ext mismatch'}), 400
    final = _dockatt_path(rid, idx, fp, ext)
    tmp = final + '.' + uuid.uuid4().hex + '.tmp'
    try:
        with open(tmp, 'wb') as fh:
            fh.write(data)
        os.replace(tmp, final)                           # 원자적 교체 — 반쯤 쓰인 파일이 노출되지 않게
    except OSError:
        try:
            os.remove(tmp)
        except OSError:
            pass
        app.logger.exception('dockatt-upload')
        return jsonify({'error': 'write failed'}), 500
    _dockatt_gc(rid, files)                              # 새 파일이 안착한 뒤 안 쓰는 잔재만 정리(용량)
    return jsonify({'id': rid, 'index': idx, 'fp': fp, 'ext': ext, 'stored': True, 'bytes': len(data)})


@app.route('/api/dock_procure/<int:rid>/attachments/<int:idx>')
@login_required
def api_dockproc_att(rid, idx):
    """견적서 원본 미리보기(읽기전용). `/api/` 경로라 앱 Bearer 도 세션 투명주입으로 그대로 열린다.
    권한은 이 기능의 다른 dock_procure API 와 동일한 `@login_required` (탭 전체가 같은 정책).

    🔴 **현재 목록에 있는 자리 + 지문 일치**일 때만 연다(올마이트 지적 반영). 목록이 비워졌거나
      바뀐 뒤 GC 가 실패해 옛 파일이 남아 있어도, URL 을 직접 쳐서 열 수 없다."""
    if idx < 0 or idx > _DOCKATT_MAX_IDX:
        abort(404)
    row = query("SELECT id, att_files FROM dock_procure WHERE id=?", (rid,), one=True)
    if not row:
        abort(404)
    files = _dockproc_files_of(row['att_files'])
    if idx >= len(files):                                # 목록에서 사라진 첨부는 캐시가 남아도 서빙 안 함
        abort(404)
    # 🔴 호출자가 '자기가 보고 있던 첨부의 신원'을 같이 보내면 그것까지 확인한다(올마이트 2026-07-31 지적).
    #   지문 검증만으로는 **서버 기준 현재 목록**과의 일치만 보장한다 — 화면이 열린 채 목록이 바뀌고
    #   캐시까지 새 파일로 채워지면, 같은 idx 가 이제 다른 업체 파일이라 '칩 이름은 A, 열리는 건 B' 가 된다.
    #   `sv`(SVMS 저장명, 정렬 1순위 키)를 되돌려받아 대조하면 그 창이 닫힌다. 안 보내면 기존 동작 유지.
    want_sv = (request.args.get('sv') or '').strip()
    if want_sv and want_sv != (files[idx].get('sv') or ''):
        abort(404)
    p, ext = _dockatt_find(rid, idx, _dockatt_fp(files[idx]))
    if not p:
        abort(404)
    nm = files[idx].get('nm') or 'quotation_%d_%d.%s' % (rid, idx, ext)
    nm = os.path.basename(str(nm).replace('\\', '/'))[:160] or 'quotation.%s' % ext
    resp = send_file(p, mimetype=_FUNDREQ_ATT_MIME[ext],
                     as_attachment=(ext not in _FUNDREQ_ATT_INLINE),
                     download_name=nm, conditional=True)
    resp.headers['X-Content-Type-Options'] = 'nosniff'
    return resp


# ═════════════════════════════════════════════════════════════════
#  Phase ③ — 수리·구매 견적 상신 큐 (웹/앱 컨펌 → 맥 워커가 SVMS write)
# ═════════════════════════════════════════════════════════════════
#  🔴 이 블록이 다루는 것은 **돈이 나가는 경로**다. 설계 정본 = docs/svms/phase3-submit-design.md,
#     봉투 실측 = docs/svms/repair-submit-envelope.md(§최종 조립 규격).
#  안전 원칙 4개 — 코드로 강제되며 편의를 위해 완화하지 않는다:
#   ① **생성은 세션 로그인 admin 버튼 하나뿐.** ext(api_key) 생성 라우트를 만들지 않는다
#      — 만들면 맥/스크립트가 사람 없이 상신을 큐잉할 수 있다(app.py `ext enqueue` 403 선례와 같은 취지).
#   ② **클라이언트는 rep_cd·금액을 못 보낸다.** 서버가 rid 로 DB 에서 읽어 채운다 — 안 그러면
#      화면에 A 를 띄우고 봉투엔 B 를 넣는 조작이 가능하다.
#   ③ 벤더는 그 건의 `sub_quotes` 에 **`cd` 로 실재하고 제출상태**여야 한다(구버전 폴러가 실은
#      cd 없는 견적은 자동 제외 = fail-closed).
#   ④ 이미 발주완료(stg_order)거나 이미 큐에 있으면 거부(부분 유니크 인덱스와 이중방어).
_DOCK_SUBMIT_APP_NO_RE = re.compile(r'^[0-9A-Z]{1,10}$')
_DOCK_SUBMIT_ACTIVE = ('approved', 'submitting')
_DOCK_SUBMIT_DONE = ('submitted', 'failed', 'canceled')
# 🔴 상신 **이후** 단계 라벨 = 재컨펌 차단 대상. allowlist(상신 전 라벨) 가 아니라
#    denylist 인 이유(올마이트 2026-08-01): 반려가 어떤 라벨로 돌아오는지 전수 확인이 안 된다.
#    allowlist 로 짜면 처음 보는 라벨·NULL 이 **영구 차단**이 되어 업무가 멈춘다.
#    여기서 놓쳐도 이중 상신은 안 난다 — 최후 방어선은 맥 워커의 pre-read 다
#    (SVMS 실헤더를 읽어 `RE` 가 아니면 중단. 오늘 12:31 에 실제로 그게 막았다).
#    실측 라벨 분포(수리 R): HQ Ordered 25 · NULL 24 · Quotation Inquiry 18 · HQ Confirmed 2 · Submit 1.
_DOCK_SUBMIT_POST = ('submit', 'approval', 'progressing', 'confirmed', 'ordered')
_DOCK_SUBMIT_CATS = ('R', 'S', 'ST')
# 🔴 상신 **전**이 rank 맵으로 확정된 라벨 = 위 denylist 부분일치보다 먼저 본다(정확일치).
#   'HQ Confirmed' 는 `_DOCKPROC_STATUS_RANK` 에서 rank **1(견적작성)** 인데 denylist 의 'confirmed' 에
#   걸린다. 견적요청 쪽은 2026-08-03 실사고 후 `_DOCK_INQ_PRE` 로 같은 함정을 막았는데 상신 쪽은
#   안 옮겨져 있었다 → 실패 이력이 있는 행이 초록 '✓ 상신완료' 로 뒤집혀 보였다(표시 결함).
_DOCK_SUBMIT_PRE = ('HQ CONFIRMED',)


def _dockproc_sbm_posted(svms_status):
    """이 헤더 라벨이 '이미 상신 이후'인가 — 순수함수. `_dockproc_inq_posted` 와 같은 규약.

    🔴 이 판정을 **서버만** 한다. 웹 `sbmPosted()` 와 iOS `isPosted()` 가 `_DOCK_SUBMIT_POST`
       부분일치 리스트를 각자 복사해 들고 있었고, 둘 다 `_DOCK_SUBMIT_PRE` 예외가 없어서
       'HQ Confirmed'(=상신 전 rank1) 행의 **실패 이력이 초록 '✓ 상신완료' 로 뒤집혔다.**
       ⇒ 목록 API 가 `sbm_posted` 로 내려주고 두 화면은 그 값만 읽는다.

    ⚠️ 실제 재상신 **차단** 게이트(`_dock_submit_prior`)는 이 예외를 쓰지 않는다 — 돈경로를
       넓히는 방향이라 형 컨펌 없이 완화하지 않는다. 여기서 고치는 건 **표시**뿐이다.
    """
    raw = str(svms_status or '').strip()
    if not raw:
        return False
    if raw.upper() in _DOCK_SUBMIT_PRE:
        return False
    st = raw.lower()
    return any(k in st for k in _DOCK_SUBMIT_POST)


def _dock_submit_prior(rep_cd, svms_status):
    """🔴 같은 건 재컨펌 차단 — 이미 상신된 건을 다시 큐에 올리지 못하게 한다.

    실사고 2026-08-01: `BGBBME26073108` 이 12:03 에 실제 상신됐는데 화면은 아직
    `Quotation Inquiry` 라 형이 같은 건을 다시 컨펌했다. 워커 pre-read 가 SVMS 헤더를
    읽고 막아서 이중 상신은 안 났지만(fail-closed 정상), **버튼 단계에서 막았어야 한다.**

    판정은 시각 비교가 아니라 **헤더 라벨**로 한다 — 오늘 서버 TZ 를 GMT→Asia/Seoul 로 바꿔서
    기존 행은 UTC, 신규는 KST 스탬프라 `done_at` vs `svms_synced_at` 비교는 9시간 오판이 난다.
    상신 성공 시 `/result` 가 이 행의 `svms_status` 를 'Submit' 로 즉시 갱신하고, 이후엔
    sync 가 진실을 덮는다. SVMS 에서 반려돼 라벨이 되돌아오면 게이트가 저절로 열린다."""
    sub = query("SELECT id, done_at FROM dock_submit_draft WHERE rep_cd=? AND status='submitted' "
                "ORDER BY id DESC", (rep_cd,), one=True)
    if not sub:
        return None
    st = str(svms_status or '').strip().lower()
    if not any(k in st for k in _DOCK_SUBMIT_POST):
        return None                       # 상신 이후 라벨이 아님 = 반려됐거나 되돌려짐 → 다시 열어준다
    return '이미 상신됨 (#%d · %s) — SVMS 에서 반려되면 동기화 후 다시 열림' % (
        sub['id'], sub['done_at'] or '')


def _dock_submit_quote_pick(row, vndr_cd):
    """그 수리/구매건의 제출견적 중 `cd` 가 일치하는 1건을 고른다."""
    try:
        quotes = json.loads(row['sub_quotes'] or '[]')
    except (TypeError, ValueError):
        return None, '제출견적 스냅샷 손상 — 폴러 재동기화 필요'
    if not isinstance(quotes, list) or not quotes:
        return None, '제출견적 없음 — 벤더 제출 후 다시 시도'
    hit = [q for q in quotes if isinstance(q, dict) and (q.get('cd') or '') == vndr_cd]
    if not hit:
        # cd 가 하나도 없으면 구버전 폴러가 적재한 스냅샷 — 코드 없이는 SELETED_VDR 를 만들 수 없다.
        if not any((q.get('cd') if isinstance(q, dict) else None) for q in quotes):
            return None, '업체코드(VNDR_CD) 미적재 — 폴러 재동기화 후 다시 시도'
        return None, '선택한 업체가 이 건의 제출견적에 없음'
    q = hit[0]
    st = str(q.get('st') or '')
    if 'submit' not in st.lower():
        # SVMS 는 견적을 낸(Submitted) 업체만 발주 대상으로 삼는다. 미제출 업체로 상신하면
        # 금액 없는 발주가 되므로 여기서 막고, 실제 반례가 관측되면 근거를 보고 완화한다.
        return None, "제출상태가 아님(st=%s) — 발주 대상 아님" % (st or '없음')
    return q, None


_DOCK_DRAFT_COLS = {}


def _dock_draft_cols(table):
    """목록 조회용 컬럼 목록 = `envelope_json` 만 뺀 전부.

    🔴 왜: 목록 API 는 `SELECT *` 로 봉투 스냅샷 원문(행당 수 KB)까지 읽어와 직렬화 직전에
       `pop` 으로 버렸다. 200행이면 그만큼을 헛읽는다. 컬럼 이름을 손으로 나열하지 않고
       **스키마에서** 뽑는 이유는, 하드코딩하면 나중에 추가된 컬럼이 목록에서 조용히 빠져
       화면이 이유 없이 빈칸을 보이기 때문이다(테이블명은 우리 리터럴만 들어온다).
    """
    cols = _DOCK_DRAFT_COLS.get(table)
    if not cols:
        names = [r['name'] for r in query('PRAGMA table_info(%s)' % table)]
        cols = ', '.join(n for n in names if n != 'envelope_json') or '*'
        _DOCK_DRAFT_COLS[table] = cols
    return cols


def _dock_submit_row_json(r):
    d = dict(r)
    d.pop('envelope_json', None)                          # 목록엔 스냅샷 원문 안 실음(길다)
    d['dismissable'] = _dock_submit_dismissable(r)        # 모달 [실패기록 지우기] 문구 분기용
    return d


# 🔴 SVMS write 를 **던지기 전에** 끝난 실패 사유들(워커 `submit_watch.py` 문자열과 1:1).
#    `wrote` 컬럼이 없는 구행 판정에만 쓰는 폴백이다 — 새 행은 워커가 보낸 0/1 이 정본.
#    prefix 매칭인 이유: 사유 뒤에 상세가 붙는다("구매 pre-read 검증 실패: 업체 …").
_DOCK_SUBMIT_NOWRITE_PREFIX = (
    '구매 pre-read 검증 실패', 'pre-read 검증 실패',
    '구매 봉투 조립 실패', '봉투 조립 실패',
    '승인 스냅샷 불일치', '지원하지 않는/누락된 cat_code',
)


def _dock_submit_dismissable(r):
    """이 실패기록을 **자동으로** 화면에서 내려도 되는가.

    🔴 기준은 "SVMS 에 write 가 0 건인가" 하나다. pre-read/조립/스냅샷 단계 실패는 SVMS 가
    전혀 안 바뀌었으니(형 2026-08-04: "실패는 trmt 앱에서 실패인거지 svms는 바뀐게 없잖아")
    다음 시도 때 남아 있을 이유가 없다. 반대로 Confirm 이 이미 나간 뒤의 실패는 SVMS 가
    **반쪽 상태**일 수 있어 조용히 사라지면 안 된다(실사고 #13 신호 유실) — 그건 형이
    모달에서 직접 지워야 한다(force).
    """
    d = dict(r)
    if (d.get('status') or '') != 'failed':
        return False
    w = d.get('wrote')
    if w is not None:
        return int(w) == 0
    return str(d.get('result') or '').startswith(_DOCK_SUBMIT_NOWRITE_PREFIX)


def _dock_submit_dismiss(rid=None, vsl_nm=None, force=False, who='web', ids=None):
    """실패기록 소거 — `dismissed_at` 스탬프만 찍는다(행·사유 보존, 감사 추적 유지).

    범위: `rid` = 그 행의 실패기록 전부 · `vsl_nm` = 그 선박 · 둘 다 없으면 전체.
    🔴 개별 draft id 로 소거하지 않는 이유(올마이트 2026-08-04 지적, 실측 확증): 같은 rid 에
    실패가 여러 건 쌓인다(라이브 rid 148 = #16·#17 둘 다 failed). 최신 1건만 내리면 형이
    삭제를 눌렀는데 **그 아래 있던 옛 실패가 배지로 올라온다.** 그래서 최소 단위가 rid 다.

    force=True 는 모달의 수동 삭제 버튼 전용(형이 사유를 읽고 경고를 확인한 뒤 누르는 자리) —
    그때만 SVMS write 가 이미 나간 실패도 내려간다. 전체 범위에는 force 를 허용하지 않는다."""
    sql = "SELECT * FROM dock_submit_draft WHERE status='failed' AND dismissed_at IS NULL"
    args = []
    if ids is not None:
        ids = [int(x) for x in ids]
        if not ids:
            return [], []
        sql += ' AND id IN (%s)' % ','.join('?' for _ in ids)
        args.extend(ids)
    elif rid is not None:
        sql += ' AND rid=?'
        args.append(rid)
    elif vsl_nm:
        sql += ' AND vsl_nm=?'
        args.append(vsl_nm)
    done, kept = [], []
    for r in query(sql, tuple(args)):
        if not (force or _dock_submit_dismissable(r)):
            kept.append(r['id'])
            continue
        # 조건부 UPDATE 의 실제 반영 건수로 판정한다 — 동시에 두 번 눌리면 뒤엣것은 0건이고,
        # 그걸 성공으로 세면 "지웠다"는 오보가 된다(올마이트 2026-08-04).
        rc = execute_rc("UPDATE dock_submit_draft SET dismissed_at=datetime('now','localtime'), "
                        "result=COALESCE(result,'')||' [dismissed by '||?||']' "
                        "WHERE id=? AND status='failed' AND dismissed_at IS NULL", (who, r['id']))
        (done if rc else kept).append(r['id'])
    return done, kept


@app.route('/api/dock_submit/app_lines')
@login_required
def api_dock_submit_app_lines():
    """결재라인 드롭다운 소스 — 맥이 밀어준 캐시. 표시 전용(봉투는 워커가 재조회해 만든다)."""
    rows = query('SELECT app_no, app_nm, user_id, approvers, updated_at FROM svms_app_line ORDER BY app_no')
    out = []
    for r in rows:
        d = dict(r)
        try:
            d['approvers'] = json.loads(d.get('approvers') or '[]')
        except (TypeError, ValueError):
            d['approvers'] = []
        out.append(d)
    return jsonify({'lines': out})


@app.route('/api/ext/svms/app_lines', methods=['POST'])
@api_key_required
def api_ext_svms_app_lines():
    """맥이 SP_GET_USER_APP(+_D) 를 읽어 캐시를 올린다. 읽기 결과 적재라 SVMS write 0.
    전량 교체(delete+insert)로 SVMS 에서 삭제된 라인이 드롭다운에 남지 않게 한다."""
    d = request.get_json(silent=True) or {}
    lines = d.get('lines')
    if not isinstance(lines, list):
        return jsonify({'error': 'lines 배열 필요'}), 400
    keep = []
    for ln in lines[:50]:
        if not isinstance(ln, dict):
            continue
        app_no = str(ln.get('app_no') or '').strip().upper()
        if not _DOCK_SUBMIT_APP_NO_RE.match(app_no):
            continue
        appr = []
        for a in (ln.get('approvers') or [])[:20]:
            if isinstance(a, dict):
                appr.append({'seq': a.get('seq'),
                             'id': str(a.get('id') or '')[:20],
                             'nm': str(a.get('nm') or '')[:60]})
        keep.append((app_no, str(ln.get('app_nm') or '')[:80],
                     str(ln.get('user_id') or '')[:20],
                     json.dumps(appr, ensure_ascii=False, sort_keys=True, separators=(',', ':'))))
    if not keep:
        return jsonify({'error': '유효한 라인 0건 — 캐시 유지'}), 400   # 빈 푸시로 드롭다운을 비우지 않음
    execute('DELETE FROM svms_app_line')
    for k in keep:
        execute("INSERT INTO svms_app_line (app_no, app_nm, user_id, approvers, updated_at) "
                "VALUES (?,?,?,?,datetime('now','localtime'))", k)
    return jsonify({'ok': True, 'count': len(keep)})


@app.route('/api/dock_submit/drafts')
@login_required
def api_dock_submit_list():
    # 🔴 소거된 실패기록은 안 내린다 — 이 한 줄이 웹·iOS 양쪽 배지를 같이 끈다(앱 업데이트 없이도).
    #    행은 DB 에 그대로 남아 있고(`?all=1` 로 조회 가능), 성공/취소 기록은 소거 대상이 아니다.
    keep_all = request.args.get('all') in ('1', 'true')
    flt = '' if keep_all else ' AND dismissed_at IS NULL'
    rid = request.args.get('rid')
    cols = _dock_draft_cols('dock_submit_draft')     # 봉투 원문 제외 — 아래서 pop 할 값을 읽지 않는다
    if rid:
        try:
            rows = query('SELECT %s FROM dock_submit_draft WHERE rid=?%s ORDER BY id DESC' % (cols, flt),
                         (int(rid),))
        except (TypeError, ValueError):
            return jsonify({'error': 'bad rid'}), 400
    else:
        rows = query('SELECT %s FROM dock_submit_draft WHERE 1=1%s ORDER BY id DESC LIMIT 200' % (cols, flt))
    return jsonify({'drafts': [_dock_submit_row_json(r) for r in rows]})


@app.route('/api/dock_submit/drafts/dismiss', methods=['POST'])
@admin_required
def api_dock_submit_dismiss():
    """직전 상신 **실패기록**을 화면에서 내린다 — SVMS 는 건드리지 않는다(read/write 0).

    형 지시 2026-08-04: "trmt 앱/웹에서 실패한거는 리프래시나 다시 상신 시도할때 자동으로
    이전 failure 기록을 없애는걸로 해 (아니면 모달내에 수동으로 이전 실패기록 삭제 버튼을…)".

    · `vsl_nm`      = 새로고침(웹 동기화 버튼·앱 pull-to-refresh) → 그 **선박**의 write 0 실패 소거
    · 인자 없음      = 전체 (범위 지정 못 하는 호출자용 폴백)
    · `rid` + `force` = 모달 수동 삭제(Confirm 이 나간 뒤의 실패도 형이 사유를 보고 지울 수 있게)
    행을 DELETE 하지 않는 이유 = 사후 추적. `dismissed_at` 만 찍고 목록에서 빠진다."""
    d = request.get_json(silent=True) or {}
    try:
        rid = int(d['rid']) if d.get('rid') is not None else None
    except (TypeError, ValueError):
        return jsonify({'error': 'bad rid'}), 400
    # 모달은 표시 중인 실패 draft의 id를 보낸다. 실패가 같은 rid에 여러 건 쌓일 수
    # 있으므로 id를 그 행의 rid로 해석해 해당 건의 실패 이력을 함께 내린다.
    if d.get('id') is not None:
        try:
            did = int(d['id'])
        except (TypeError, ValueError):
            return jsonify({'error': 'bad id'}), 400
        picked = query("SELECT rid, status FROM dock_submit_draft WHERE id=?",
                       (did,), one=True)
        if not picked:
            return jsonify({'error': '실패 기록을 찾을 수 없음'}), 404
        if picked['status'] != 'failed':
            return jsonify({'ok': True, 'dismissed': [], 'kept': [],
                            'note': '실패 기록만 삭제할 수 있음'}), 200
        rid = picked['rid']
    # force 는 rid 지정 시에만 — 선박/전체 범위에 붙으면 반쪽 상태 실패까지 싹 지운다.
    force = bool(d.get('force')) and rid is not None
    done, kept = _dock_submit_dismiss(rid=rid, vsl_nm=(d.get('vsl_nm') or '').strip() or None,
                                      force=force, who=session.get('username') or 'web')
    return jsonify({'ok': True, 'dismissed': done, 'kept': kept,
                    'note': ('SVMS write 가 이미 나간 실패는 남겨둠 — 모달에서 직접 삭제'
                             if kept else '')})


@app.route('/api/dock_submit/preview')
@login_required
def api_dock_submit_preview():
    """컨펌 모달용 초안 요약 — 이 건에 상신 가능한 벤더 후보 + 결재라인. **write 0.**
    여기서 거절 사유를 미리 보여줘서, 형이 버튼을 누른 뒤에야 실패하는 일을 줄인다."""
    try:
        rid = int(request.args.get('rid') or 0)
    except (TypeError, ValueError):
        return jsonify({'error': 'bad rid'}), 400
    row = query('SELECT * FROM dock_procure WHERE id=?', (rid,), one=True)
    if not row:
        return jsonify({'error': 'not found'}), 404
    try:
        quotes = json.loads(row['sub_quotes'] or '[]')
    except (TypeError, ValueError):
        quotes = []
    cands = []
    for q in (quotes if isinstance(quotes, list) else []):
        if not isinstance(q, dict):
            continue
        why = None
        if not q.get('cd'):
            why = '업체코드 미적재'
        elif 'submit' not in str(q.get('st') or '').lower():
            why = '제출상태 아님'
        cands.append({'cd': q.get('cd'), 'nm': q.get('nm'), 'amt': q.get('amt'),
                      'gross_amt': q.get('gross_amt'), 'dc_rate': q.get('dc_rate'),
                      'final_amt': q.get('final_amt'), 'final_usd': q.get('final_usd'),
                      'cur': q.get('cur'), 'usd': q.get('usd'), 'st': q.get('st'),
                      'att': q.get('att'), 'best': q.get('best'), 'ok': why is None, 'why': why,
                      # 🔴 품목 견적 결함(견적 미제출·단가 0…) = **알림만, 선택은 막지 않는다.**
                      #   판정은 맥 워커의 상신 게이트와 같은 `dock_items.item_gaps` 가 했고 이 값은
                      #   마지막 동기화 시점의 스냅샷이다. 여기서 라디오를 비활성하면 형이 SVMS 에서
                      #   고친 직후에도 다음 동기화까지 상신을 못 한다 — 최종 차단은 워커가 라이브로 한다.
                      #   `hard_n` 만 "이대로면 실패" 다. 단가 0(soft)은 형 지시로 통과시킨다.
                      #   저장된 스냅샷을 그대로 읽으므로 hard/soft 분리 이전 행이 섞일 수 있다 —
                      #   `_dockproc_hard_n` 이 그 경우 "모름 = 전부 차단" 으로 되돌린다(적재 경로와 같은 규칙).
                      'gap_n': q.get('gap_n'), 'hard_n': _dockproc_hard_n(q),
                      'gaps': q.get('gaps')})
    blocked = None
    if (row['cat_code'] or '') not in _DOCK_SUBMIT_CATS:
        blocked = '서비스(R)·자재(S)·스토어(ST) 건만 상신 가능'
    elif not (row['svms_req_no'] or '').strip():
        blocked = 'SVMS 문서번호(Inq No) 연결 안 됨'
    elif row['stg_order']:
        blocked = '이미 발주완료'
    else:
        act = query("SELECT id, status FROM dock_submit_draft WHERE rep_cd=? AND status IN (?,?)",
                    ((row['svms_req_no'] or '').strip(), *_DOCK_SUBMIT_ACTIVE), one=True)
        if act:
            blocked = '이미 상신 큐에 있음(#%d %s)' % (act['id'], act['status'])
        else:
            blocked = _dock_submit_prior((row['svms_req_no'] or '').strip(), row['svms_status'])
    return jsonify({'rid': rid, 'req_no': row['req_no'], 'vsl_nm': row['vsl_nm'],
                    'rep_cd': (row['svms_req_no'] or '').strip() or None,
                    'subject': row['subject'], 'blocked': blocked, 'candidates': cands})


@app.route('/api/dock_submit/drafts', methods=['POST'])
@admin_required
def api_dock_submit_create():
    """🔴 형이 컨펌 버튼을 누르는 자리 = Phase ③ 의 **유일한 승인 게이트.**
    여기서 만들어진 approved 1행이 맥 워커의 SVMS write를 부른다.
    수리(R)는 SP_SET_ODR_INFO→SP_SET_SBM, 구매(S/ST)는 SP_SET_ODR 이다.
    받는 값은 `rid`/`vndr_cd`/`app_no` **3개뿐** — 문서번호·금액은 서버가 DB 에서 읽는다(위 원칙 ②)."""
    d = request.get_json(silent=True) or {}
    try:
        rid = int(d.get('rid') or 0)
    except (TypeError, ValueError):
        return jsonify({'error': 'bad rid'}), 400
    vndr_cd = str(d.get('vndr_cd') or '').strip().upper()
    app_no = str(d.get('app_no') or '').strip().upper()
    if not re.fullmatch(r'[A-Z0-9]{1,20}', vndr_cd):
        return jsonify({'error': '업체코드(vndr_cd) 형식 오류', 'field': 'vndr_cd'}), 400
    if not _DOCK_SUBMIT_APP_NO_RE.match(app_no):
        return jsonify({'error': '결재라인(app_no) 형식 오류', 'field': 'app_no'}), 400
    row = query('SELECT * FROM dock_procure WHERE id=?', (rid,), one=True)
    if not row:
        return jsonify({'error': 'not found'}), 404
    if (row['cat_code'] or '') not in _DOCK_SUBMIT_CATS:
        return jsonify({'error': '서비스(R)·자재(S)·스토어(ST) 건만 상신 가능 — 이 행은 %s' %
                               (row['cat_code'] or '?')}), 400
    rep_cd = (row['svms_req_no'] or '').strip()
    if not rep_cd:
        return jsonify({'error': 'SVMS 문서번호(Inq No) 연결 안 됨 — 먼저 연결하세요'}), 400
    if row['stg_order']:
        return jsonify({'error': '이미 발주완료된 건'}), 409
    prior = _dock_submit_prior(rep_cd, row['svms_status'])     # 🔴 재컨펌 차단(서버가 정본 게이트)
    if prior:
        return jsonify({'error': prior}), 409
    q, why = _dock_submit_quote_pick(row, vndr_cd)
    if not q:
        return jsonify({'error': why, 'field': 'vndr_cd'}), 400
    line = query('SELECT * FROM svms_app_line WHERE app_no=?', (app_no,), one=True)
    if not line:
        return jsonify({'error': '결재라인 캐시에 없음 — 맥 워커 동기화 필요', 'field': 'app_no'}), 400
    # 새 시도와 이전 실패를 분리한다. INSERT 뒤에 rid 전체를 dismiss하면 워커가
    # 새 행을 아주 빠르게 failed로 만든 경우 새 실패까지 이전 기록으로 오인할 수 있다.
    prior_failure_ids = [r['id'] for r in query(
        "SELECT id FROM dock_submit_draft WHERE rid=? AND status='failed' AND dismissed_at IS NULL",
        (rid,))]
    try:
        approvers = json.loads(line['approvers'] or '[]')
    except (TypeError, ValueError):
        approvers = []
    # 사람이 화면에서 본 내용 그대로를 스냅샷으로 남긴다(사후 감사·분쟁 대비).
    # ⚠️ 이건 기록이고 봉투가 아니다 — 실제 봉투는 워커가 상신 시점에 SVMS 를 다시 읽어 만든다.
    envelope = {'rep_cd': rep_cd, 'vsl_cd': row['vsl_cd'], 'req_no': row['req_no'],
                'subject': row['subject'], 'vndr_cd': vndr_cd, 'vndr_nm': q.get('nm'),
                'amt': q.get('amt'), 'cur': q.get('cur'), 'usd': q.get('usd'), 'st': q.get('st'),
                'app_no': app_no, 'app_nm': line['app_nm'], 'approvers': approvers}
    who = session.get('username') or 'web'
    try:
        did = execute(
            "INSERT INTO dock_submit_draft (rid, vsl_nm, vsl_cd, req_no, rep_cd, vndr_cd, vndr_nm, "
            "amt, cur, app_no, app_nm, envelope_json, status, decided_at, decided_by) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,'approved',datetime('now','localtime'),?)",
            (rid, row['vsl_nm'], row['vsl_cd'], row['req_no'], rep_cd, vndr_cd, q.get('nm'),
             q.get('amt'), q.get('cur'), app_no, line['app_nm'],
             json.dumps(envelope, ensure_ascii=False, sort_keys=True, separators=(',', ':')), who))
    except sqlite3.IntegrityError:                         # 부분 유니크 = 같은 건 이중 큐잉
        act = query("SELECT id, status FROM dock_submit_draft WHERE rep_cd=? AND status IN (?,?)",
                    (rep_cd, *_DOCK_SUBMIT_ACTIVE), one=True)
        return jsonify({'error': '이미 상신 큐에 있음', 'id': act['id'] if act else None,
                        'status': act['status'] if act else None}), 409
    # 🔴 재시도 = 이전 실패기록의 수명 끝(형 지시 2026-08-04). 새 시도가 큐에 올라간 **뒤에**
    #    소거해서, 위에서 400/409 로 막힌 경우엔 실패 사유가 화면에 남아 있게 한다.
    #    소거 대상은 SVMS write 0 건인 실패뿐 — 반쪽 상태 실패는 그대로 보인다.
    #    소거는 **부가 기능**이라 여기서 터져도 큐잉(201)을 깨면 안 된다 — 형이 알아야 하는 건
    #    "상신이 큐에 올라갔나"이고, 오래된 실패 배지가 남는 건 그 다음 문제다.
    try:
        dismissed, _kept = _dock_submit_dismiss(ids=prior_failure_ids, who=who)
    except Exception as e:                                 # noqa: BLE001 — 큐잉 성공을 지킨다
        app.logger.warning('dock_submit 재시도 소거 실패 rid=%s: %s', rid, e)
        dismissed = []
    return jsonify({'id': did, 'status': 'approved', 'rep_cd': rep_cd,
                    'vndr_cd': vndr_cd, 'app_no': app_no, 'dismissed': dismissed}), 201


@app.route('/api/dock_submit/drafts/<int:did>/cancel', methods=['POST'])
@admin_required
def api_dock_submit_cancel(did):
    """워커가 집어가기 전(approved)만 취소. submitting 이후는 SVMS 에 이미 나갔을 수 있어 못 되돌린다."""
    rc = execute_rc("UPDATE dock_submit_draft SET status='canceled', done_at=datetime('now','localtime'), "
                    "result=COALESCE(result,'')||' [canceled by '||?||']' "
                    "WHERE id=? AND status='approved'", (session.get('username') or 'web', did))
    if not rc:
        cur = query('SELECT status FROM dock_submit_draft WHERE id=?', (did,), one=True)
        if not cur:
            return jsonify({'error': 'not found'}), 404
        return jsonify({'error': '대기(approved) 상태만 취소 가능', 'status': cur['status']}), 409
    return jsonify({'id': did, 'status': 'canceled'})


@app.route('/api/dock_submit/drafts/<int:did>/push', methods=['POST'])
@admin_required
def api_dock_submit_push(did):
    """🔴 [지금 전송] — 형이 누른 그 순간에만 1건이 SVMS 로 나간다. **스케줄러 없음.**

    이 서버(OCI)는 SVMS 에 못 붙는다. 그래서 맥이 `ssh -R` 역터널로 loopback 포트를
    열어두고, 이 라우트가 그 포트를 호출한다. 맥이 claim(`?id=`)·상신·readback 까지
    끝낸 **실제 결과**를 그대로 돌려주므로 모달에 결과가 즉시 뜬다.

    fail-closed: 터널이 없으면 큐에 남겨두고 조용히 기다리지 않고 **503 으로 명확히 실패**한다
    (형의 컨펌이 어디로 갔는지 모르는 상태가 최악)."""
    cur = query('SELECT status FROM dock_submit_draft WHERE id=?', (did,), one=True)
    if not cur:
        return jsonify({'error': 'not found'}), 404
    if cur['status'] != 'approved':
        return jsonify({'error': '대기(approved) 상태만 전송 가능', 'status': cur['status']}), 409

    url = (os.environ.get('DOCK_PUSH_URL') or '').strip()
    token = (os.environ.get('DOCK_PUSH_TOKEN') or '').strip()
    if not url or not token:
        return jsonify({'error': '푸시 경로 미설정 (DOCK_PUSH_URL/TOKEN)'}), 503
    req = urllib.request.Request(
        url, method='POST',
        data=json.dumps({'draft_id': did}).encode(),
        headers={'Content-Type': 'application/json', 'X-Push-Token': token})
    try:
        # 45s — gunicorn `--timeout 60` 보다 짧아야 한다. 넘기면 워커(-w 1)가 통째로 죽어
        # 다른 요청까지 끊긴다. 실측 상신 1건은 Save+Submit+readback 합쳐 20s 안쪽.
        with urllib.request.urlopen(req, timeout=45) as r:
            body = json.loads(r.read().decode('utf-8', 'replace') or '{}')
            code = r.status
    except urllib.error.HTTPError as e:                       # 맥이 사유를 담아 거절한 경우
        try:
            body = json.loads(e.read().decode('utf-8', 'replace') or '{}')
        except Exception:
            body = {'msg': f'맥 응답 오류 {e.code}'}
        code = e.code
    except (TimeoutError, socket.timeout):
        # 🔴 타임아웃은 '안 나갔다'가 아니다 — 맥이 SVMS 로 이미 보냈을 수 있다(올마이트 2026-08-01).
        #    '맥 미연결' 로 뭉뚱그리면 형이 다시 누를 수 있으므로 불확실을 그대로 말한다.
        #    (행은 맥이 claim 해 submitting 이므로 재전송은 아래 상태가드에서 409 로 막힌다.)
        st = query('SELECT status FROM dock_submit_draft WHERE id=?', (did,), one=True)
        return jsonify({'error': '응답 시간 초과 — 전송 여부 불확실. 다시 누르지 말고 SVMS 에서 확인하세요',
                        'status': st['status'] if st else None, 'ambiguous': True}), 504
    except Exception as e:                                    # 터널 끊김·연결거부
        if isinstance(getattr(e, 'reason', None), (TimeoutError, socket.timeout)):
            st = query('SELECT status FROM dock_submit_draft WHERE id=?', (did,), one=True)
            return jsonify({'error': '응답 시간 초과 — 전송 여부 불확실. 다시 누르지 말고 SVMS 에서 확인하세요',
                            'status': st['status'] if st else None, 'ambiguous': True}), 504
        return jsonify({'error': f'맥 미연결 — 전송 못 함 ({type(e).__name__})'}), 503
    st = query('SELECT status, result FROM dock_submit_draft WHERE id=?', (did,), one=True)
    # 🔴 성공 판정은 맥의 `ok` 가 아니라 **DB 최종상태**로 한다. 맥은 상신 후 `/result` 로
    #    submitted 를 기록하고 돌아온다 — 그 기록이 없으면 화면에 완료라고 쓰지 않는다.
    done = bool(st and st['status'] == 'submitted')
    return jsonify({'id': did, 'ok': bool(body.get('ok')) and done, 'msg': body.get('msg') or '',
                    'status': st['status'] if st else None,
                    'result': st['result'] if st else None}), (200 if code == 200 else code)


@app.route('/api/dock_submit/drafts/decided', methods=['DELETE'])
@admin_required
def api_dock_submit_clear_decided():
    """처리완료 정리 — 진행중(approved/submitting)은 보존."""
    n = execute_rc("DELETE FROM dock_submit_draft WHERE status IN (?,?,?)", _DOCK_SUBMIT_DONE)
    return jsonify({'ok': True, 'deleted': n})


# ---- ext (맥 submit_watch) ----
@app.route('/api/ext/dock_submit/approved')
@api_key_required
def api_ext_dock_submit_approved():
    """맥 워커가 상신할 approved 건 → CAS claim 으로 submitting 락.
    fundreq `/approved` 규약 준용:
      · 이번 호출에서 **새로 claim 성공한 행만** 반환(기존 submitting 재서빙 안 함 = 중복 상신 방지)
      · `submitting` 6h 초과 → `failed`. **자동 재큐 안 함** — 절반 성공한 상신의 이중 실행이 최악이다.
      · `?peek=1` = 락 없이 조회(DRY 검증용)
      · `decided_by` 가 빈 행은 claim 대상이 아니다(사람 승인 흔적 없는 행 = 상신 금지)
      · 🔴 `?limit=N` (기본 1) — **claim 은 워커가 이번에 실제로 처리할 만큼만.**
        올마이트 2026-08-01 P0 지적: 예전엔 approved 전부를 submitting 으로 잠갔는데 워커는
        `--max 1` 만 처리해서, 나머지가 아무 일도 안 당한 채 6h 뒤 failed 로 떨어졌다
        (재큐도 안 하므로 형이 다시 컨펌해야 함 = 조용한 승인 유실).
      · 🔴 `?id=N` — [지금 전송] 버튼이 지목한 **그 행만** claim. 가드는 위와 완전히 동일하다
        (같은 CAS·같은 승인흔적 조건). 버튼 경로가 별도 우회로가 되면 안 되므로 코드도 한 곳."""
    cols = ("id, rid, vsl_nm, vsl_cd, req_no, rep_cd, vndr_cd, vndr_nm, amt, cur, app_no, app_nm, "
            "envelope_json, (SELECT cat_code FROM dock_procure WHERE id=dock_submit_draft.rid) AS cat_code")
    if request.args.get('peek'):
        rows = query(f"SELECT {cols} FROM dock_submit_draft WHERE status='approved' ORDER BY id ASC")
        return jsonify({'count': len(rows), 'drafts': [dict(r) for r in rows], 'peek': True})
    try:
        limit = max(1, min(20, int(request.args.get('limit') or 1)))
    except (TypeError, ValueError):
        limit = 1
    execute("UPDATE dock_submit_draft SET status='failed', "
            "result=COALESCE(result,'')||' [auto:6h+ submitting→failed, 사람 재검토]' "
            "WHERE status='submitting' AND done_at IS NOT NULL "
            "AND done_at < datetime('now','localtime','-6 hours')")
    where, params = '', ()
    if 'id' in request.args:
        # 🔴 빈 `?id=` 를 falsy 로 흘리면 bulk 경로로 떨어져 **지목하지 않은 다른 행**이 claim 된다
        #    (올마이트 2026-08-01). 존재 여부로 분기하고, 못 읽으면 claim 하지 말고 거절.
        try:
            where, params, limit = ' AND id=?', (int(request.args['id']),), 1
        except (TypeError, ValueError):
            return jsonify({'count': 0, 'drafts': [], 'error': 'id 형식 오류'}), 400
    out = []
    for r in query(f"SELECT {cols} FROM dock_submit_draft WHERE status='approved' "
                   "AND decided_at IS NOT NULL AND COALESCE(decided_by,'')<>''"
                   + where + " ORDER BY id ASC", params):
        if len(out) >= limit:
            break
        if execute_rc("UPDATE dock_submit_draft SET status='submitting', done_at=datetime('now','localtime') "
                      "WHERE id=? AND status='approved' AND decided_at IS NOT NULL "
                      "AND COALESCE(decided_by,'')<>''", (r['id'],)):
            out.append(dict(r))
    return jsonify({'count': len(out), 'drafts': out, 'limit': limit})


@app.route('/api/ext/dock_submit/drafts/<int:did>/result', methods=['POST'])
@api_key_required
def api_ext_dock_submit_result(did):
    """상신 결과 — ok=True → submitted, else failed. **판정 근거는 워커의 readback**
    (`SP_GET_REP_INFO` 재조회로 상태 전이 확인). 응답 성공키를 몰라도 되는 이유가 이것."""
    d = request.get_json(silent=True) or {}
    ok = bool(d.get('ok'))
    # `wrote` = 이 결과가 SVMS write 를 던진 **뒤**인지(1) 던지기 전인지(0). 실패기록 자동소거 판정용.
    # 키가 없으면 NULL 로 남긴다 — 구 워커 행은 사유 문자열로 보수 판정(`_dock_submit_dismissable`).
    try:
        wrote = None if d.get('wrote') is None else int(bool(int(d['wrote'])))
    except (TypeError, ValueError):
        wrote = 1                                          # 이상값 = 안전한 쪽(write 있었다고 본다)
    rc = execute_rc("UPDATE dock_submit_draft SET status=?, done_at=datetime('now','localtime'), "
                    "result=?, wrote=? WHERE id=? AND status='submitting'",
                    ('submitted' if ok else 'failed', (d.get('result') or '')[:2000], wrote, did))
    if rc and ok:
        # 🔴 상신 성공을 화면에 **즉시** 반영한다. 다음 SVMS sync 까지 기다리면 그 사이 화면은
        #    아직 'Quotation Inquiry' 라 형이 같은 건을 또 컨펌하게 된다(2026-08-01 실사고).
        #    ⚠️ `stg_order` 는 건드리지 않는다 — **Submit 은 발주가 아니다**(rank 2). 이 값은
        #    다음 sync 가 SVMS 진실로 덮으므로, 여기 쓰는 건 그 사이를 메우는 임시 표시다.
        #    갱신 키는 `rid`(그 draft 가 가리키는 행) — `svms_req_no` 로 덮으면 같은 문서번호가
        #    여러 행에 붙었을 때 남의 행까지 건드린다(올마이트 2026-08-01). 현재 라이브 중복은 0건이지만
        #    정확한 키가 있는데 굳이 넓게 쓸 이유가 없다. `svms_synced_at` 은 표시 전용이다
        #    (조회 커서로 쓰이는 곳 0곳 — sync 를 건너뛰게 만들지 않음. 실측 확인).
        dr = query('SELECT d.rid, p.cat_code FROM dock_submit_draft d '
                   'LEFT JOIN dock_procure p ON p.id=d.rid WHERE d.id=?', (did,), one=True)
        if dr and dr['rid']:
            interim = 'Approval(Procssing)' if dr['cat_code'] in ('S', 'ST') else 'Submit'
            execute("UPDATE dock_procure SET stg_quote=1, stg_vendor=1, stg_confirm=1, "
                    "svms_status=?, svms_synced_at=datetime('now','localtime'), "
                    "updated_at=datetime('now','localtime') WHERE id=?", (interim, dr['rid']))
    return jsonify({'id': did, 'ok': ok, 'applied': bool(rc)})


# ══════════════════════════════════════════════════════════════════════════════
# 수리 견적요청 (견적작성 → 벤더제출) — SVMS Confirm + Vendor Submit
#   봉투 정본 = docs/svms/repair-inquiry-envelope.md (Nuxt 실코드 추출, 추측 0)
#     STEP1 `PKG_MA_REP.SP_SET_REP`     PARAM=헤더 51키 verbatim + STATUS='RC'
#     STEP2 `PKG_MA_REP.SP_SET_REP_DTL` PARAM={REP_CD,REP_YN:'Y',USE_YN:'Y',REF_TP:'MARP'}
#                                       CURSOR.P_IC_VNDR=선택벤더행 verbatim(≤5)
#   🔴 Phase ③(dock_submit_draft, 발주벤더확정+결재상신)와 **다른 단계**다. 표·라우트·워커를 섞지 말 것.
#   🔴 벤더 목록은 캐시할 수 없다 — 빈 검색 파라미터로 `SP_GET_VNDR` 를 부르면 SVMS 가 409 로 거절하고,
#      모달이 넘기는 MARP 파라미터가 '이 수리건에 붙을 수 있는 업체'만 좁혀준다(실측 131→4).
#      그래서 검색은 맥을 경유한 **라이브 read** 다. 터널이 없으면 검색도 전송도 못 한다(fail-closed).
# ══════════════════════════════════════════════════════════════════════════════
_DOCK_INQ_MAX_VNDR = 5                                    # SVMS 모달 상한 ("Can't select more than 5.")
# 벤더제출 **이후** 라벨 = 재요청 차단 대상. Phase ③ 와 같은 이유로 denylist 다
# (반려가 어떤 라벨로 돌아오는지 전수 확인이 안 됨 → allowlist 면 처음 보는 라벨이 영구 차단).
_DOCK_INQ_POST = ('quotation', 'submit', 'approval', 'progressing', 'confirmed', 'ordered')
# 🔴 문서종류별 '견적요청 전'이 **확정된** 라벨 = denylist 보다 먼저 보고 재개방한다.
#   구매의 요청 전 라벨은 'HQ Confirmed' 인데 이 문자열이 denylist 의 'confirmed' 에 걸린다 —
#   그대로 두면 회수 후 라벨이 되돌아와도 영구 차단이 되어 2026-08-03 수리 실사고와 같은 부류가 된다.
#   구매가 견적요청 후 이 라벨로 남지 않는 근거(실측 2026-08-03, BGBB 24행 + SAPS 21행 전수):
#   요청된 건은 전원 `INQ_NO` 발급 + PC_PRO `STATUS_NM='Quotation Inquiry'` 이상이고,
#   'HQ Confirmed' 는 `INQ_NO=None` 인 미요청 건에만 나타났다.
_DOCK_INQ_PRE = {'PCRQ': ('HQ CONFIRMED',)}


def _dockproc_inq_posted(doc, svms_status):
    """이 라벨이 '이미 견적요청이 나간 이후'인가 — `_dock_inq_prior` 와 **같은 판정**의 순수함수.

    🔴 이 판정을 **서버만** 한다(2026-08-03 형 제보 실사고). 웹 `inqPosted()`·iOS `isInquiryPosted()`
       가 `_DOCK_INQ_POST` 부분일치 리스트를 각자 복사해 들고 있었는데 `_DOCK_INQ_PRE` 예외가 빠져서,
       구매의 **요청 전** 라벨 'HQ Confirmed' 가 'confirmed' 에 걸려 **직전 실패 이력을
       '견적요청됨 ✓'(초록)으로 뒤집어** 그렸다. 게다가 그 초록이 `inq_block` 회색처리까지 눌러서
       회수·재적재된 행에 옛 요청 정보가 계속 붙어 있었다(BGBB S14 = `BGBBES2607B4`).
       ⇒ 목록 API 가 `inq_posted` 로 내려주고 두 화면은 그 값만 읽는다(`inq_key`/`inq_block` 규약과 동일).
    """
    raw = str(svms_status or '').strip()
    if not raw:
        return False
    if raw.upper() in _DOCK_INQ_PRE.get(doc, ()):
        return False                       # 확정된 pre-요청 라벨 — denylist 부분일치보다 우선
    st = raw.lower()
    return any(k in st for k in _DOCK_INQ_POST)


def _dock_push_sibling(path):
    """맥 리스너의 형제 경로 URL — `DOCK_PUSH_URL`(=…/push) 에서 유도한다.
    별도 env 를 새로 두지 않는 이유: /etc/trmt.env 편집·터널 추가 없이 같은 포트를 쓰기 위함."""
    url = (os.environ.get('DOCK_PUSH_URL') or '').strip()
    token = (os.environ.get('DOCK_PUSH_TOKEN') or '').strip()
    if not url or not token:
        return None, None
    p = urllib.parse.urlsplit(url)
    base = p.path.rsplit('/', 1)[0]
    return urllib.parse.urlunsplit((p.scheme, p.netloc, base + path, '', '')), token


def _dock_mac_call(path, payload, timeout=20):
    """맥 리스너 호출 공통 — (body, http_code, err). err 가 있으면 실패."""
    url, token = _dock_push_sibling(path)
    if not url:
        return None, 503, '푸시 경로 미설정 (DOCK_PUSH_URL/TOKEN)'
    req = urllib.request.Request(
        url, method='POST', data=json.dumps(payload).encode(),
        headers={'Content-Type': 'application/json', 'X-Push-Token': token})
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            return json.loads(r.read().decode('utf-8', 'replace') or '{}'), r.status, None
    except urllib.error.HTTPError as e:
        try:
            body = json.loads(e.read().decode('utf-8', 'replace') or '{}')
        except Exception:
            body = {}
        return body, e.code, body.get('msg') or f'맥 응답 오류 {e.code}'
    except (TimeoutError, socket.timeout):
        return None, 504, '응답 시간 초과'
    except Exception as e:
        if isinstance(getattr(e, 'reason', None), (TimeoutError, socket.timeout)):
            return None, 504, '응답 시간 초과'
        return None, 503, f'맥 미연결 ({type(e).__name__})'


def _dock_inq_prior(doc, rep_cd, svms_status):
    """이미 견적요청이 나간 건의 재요청 차단(Phase ③ `_dock_submit_prior` 와 같은 판정 방식).

    🔴 `_DOCK_INQ_PRE[doc]` 를 denylist(`_DOCK_INQ_POST`) 보다 **먼저** 본다 — 구매의 확정된
       '견적요청 전' 라벨('HQ Confirmed')이 denylist 의 'confirmed' 부분일치에 걸려 영구
       재차단되는 걸 막는다(2026-08-03 수리 회수 사고와 같은 부류를 미리 차단)."""
    sub = query("SELECT id, done_at FROM dock_inquiry_draft WHERE rep_cd=? AND status='submitted' "
                "ORDER BY id DESC", (rep_cd,), one=True)
    if not sub:
        return None
    if not _dockproc_inq_posted(doc, svms_status):
        return None                        # pre-요청 라벨이거나 라벨이 되돌아왔다(반려/취소) → 다시 열어준다
    return '이미 견적요청됨 (#%d · %s) — SVMS 에서 되돌려지면 동기화 후 다시 열림' % (
        sub['id'], sub['done_at'] or '')


def _dock_inq_blocked(row):
    """견적요청 불가 사유(없으면 None). 생성 라우트와 preview 가 **같은 함수**를 쓴다.

    🔴 수리(MARP)/구매(PCRQ) 공통 게이트 — 문서종류·키는 `_dockproc_inq_target()` 이 `cat_code`
       기준으로 판정한다(수리=`svms_req_no`, 구매=`svms_pc_req_no`. 섞으면 Phase③ 조회가 깨진다)."""
    doc, rep_cd = _dockproc_inq_target(row)
    if not doc:
        return '수리(R)·구매(S/ST) 건만 견적요청 가능 — 그 외 분류는 봉투 구조가 없음'
    if not rep_cd:
        return 'SVMS 문서번호 연결 안 됨 — 동기화 후 다시 시도'
    if row['stg_vendor'] or row['stg_confirm'] or row['stg_order']:
        return '이미 벤더제출 이후 단계'
    act = query("SELECT id, status FROM dock_inquiry_draft WHERE rep_cd=? AND status IN (?,?)",
                (rep_cd, *_DOCK_SUBMIT_ACTIVE), one=True)
    if act:
        return '이미 견적요청 큐에 있음(#%d %s)' % (act['id'], act['status'])
    prior = _dock_inq_prior(doc, rep_cd, row['svms_status'])
    if prior:
        return prior
    # 단계 라벨 게이트 — 워커 pre-read 가 어차피 거부할 건을 여기서 먼저 끊는다(SVMS 호출 0).
    # 라벨이 stale 이라 막혔다면 동기화 후 다시 열린다(`_dock_inq_prior` 와 같은 규약).
    return _dockproc_inq_stage_block(doc, row['svms_status'])


@app.route('/api/dock_inquiry/vendor_search', methods=['POST'])
@admin_required
def api_dock_inquiry_vendor_search():
    """벤더 검색 — 맥 경유 **read-only**(`PKG_CM_VNDR.SP_GET_VNDR`, SVMS write 0).
    `rep_cd` 는 클라이언트 값을 쓰지 않고 rid 로 DB 에서 유도한다(다른 건의 벤더를 끌어오지 못하게)."""
    d = request.get_json(silent=True) or {}
    try:
        rid = int(d.get('rid') or 0)
    except (TypeError, ValueError):
        return jsonify({'error': 'bad rid'}), 400
    row = query('SELECT * FROM dock_procure WHERE id=?', (rid,), one=True)
    if not row:
        return jsonify({'error': 'not found'}), 404
    doc, rep_cd = _dockproc_inq_target(row)
    if not doc:
        return jsonify({'error': '수리(R)·구매(S/ST) 건만 벤더 검색 가능'}), 400
    if not rep_cd:
        return jsonify({'error': 'SVMS 문서번호 연결 안 됨'}), 400
    q = str(d.get('q') or '').strip()[:60]
    body, code, err = _dock_mac_call('/vendor_search', {'rep_cd': rep_cd, 'q': q, 'doc': doc}, timeout=25)
    if err:
        return jsonify({'error': err}), (code if code in (503, 504) else 502)
    return jsonify({'rep_cd': rep_cd, 'q': q, 'vendors': (body or {}).get('vendors') or [],
                    'truncated': bool((body or {}).get('truncated'))})


@app.route('/api/dock_inquiry/drafts')
@login_required
def api_dock_inquiry_list():
    rows = query('SELECT * FROM dock_inquiry_draft ORDER BY id DESC LIMIT 200')
    out = []
    for r in rows:
        dd = dict(r)
        dd.pop('envelope_json', None)
        out.append(dd)
    return jsonify({'drafts': out})


@app.route('/api/dock_inquiry/preview')
@login_required
def api_dock_inquiry_preview():
    """컨펌 모달용 요약 — **write 0.** 벤더 목록은 별도 검색 라우트에서 라이브로 받는다."""
    try:
        rid = int(request.args.get('rid') or 0)
    except (TypeError, ValueError):
        return jsonify({'error': 'bad rid'}), 400
    row = query('SELECT * FROM dock_procure WHERE id=?', (rid,), one=True)
    if not row:
        return jsonify({'error': 'not found'}), 404
    doc, rep_cd = _dockproc_inq_target(row)
    return jsonify({'rid': rid, 'req_no': row['req_no'], 'vsl_nm': row['vsl_nm'],
                    'doc_type': doc or None, 'rep_cd': rep_cd or None,
                    'svms_status': row['svms_status'], 'subject': row['subject'],
                    'max_vendor': _DOCK_INQ_MAX_VNDR, 'blocked': _dock_inq_blocked(row)})


@app.route('/api/dock_inquiry/drafts', methods=['POST'])
@admin_required
def api_dock_inquiry_create():
    """🔴 형이 컨펌하는 자리 = 이 단계의 **유일한 승인 게이트.** 여기 만들어진 approved 1행이
    [지금 전송] 을 누를 때 맥 워커의 SVMS write 2회(Confirm + Vendor Submit)를 부른다.

    받는 값은 `rid` + `vndr_cds`(업체코드 ≤5) **뿐이다.** 벤더 행 원문(=봉투 `P_IC_VNDR`)은
    받지 않는다 — 워커가 전송 시점에 `SP_GET_VNDR` 를 다시 읽어 그 코드에 해당하는 행을
    verbatim 으로 만든다(브라우저가 봉투 내용을 정할 수 없게). 이름은 표시용 스냅샷이다."""
    d = request.get_json(silent=True) or {}
    try:
        rid = int(d.get('rid') or 0)
    except (TypeError, ValueError):
        return jsonify({'error': 'bad rid'}), 400
    raw = d.get('vndr_cds')
    if not isinstance(raw, list) or not raw:
        return jsonify({'error': '업체를 1개 이상 선택하세요', 'field': 'vndr_cds'}), 400
    cds, seen = [], set()
    for c in raw:
        cd = str(c or '').strip().upper()
        if not re.fullmatch(r'[A-Z0-9]{1,20}', cd):
            return jsonify({'error': '업체코드 형식 오류: %s' % (cd or '(빈값)'), 'field': 'vndr_cds'}), 400
        if cd not in seen:
            seen.add(cd)
            cds.append(cd)
    if len(cds) > _DOCK_INQ_MAX_VNDR:
        return jsonify({'error': 'SVMS 상한은 %d개' % _DOCK_INQ_MAX_VNDR, 'field': 'vndr_cds'}), 400
    nms = d.get('vndr_nms') if isinstance(d.get('vndr_nms'), dict) else {}
    row = query('SELECT * FROM dock_procure WHERE id=?', (rid,), one=True)
    if not row:
        return jsonify({'error': 'not found'}), 404
    blocked = _dock_inq_blocked(row)                       # 🔴 preview 와 같은 게이트를 서버가 다시 검사
    if blocked:
        return jsonify({'error': blocked}), 409
    doc, rep_cd = _dockproc_inq_target(row)
    picks = [{'cd': cd, 'nm': str(nms.get(cd) or '')[:80]} for cd in cds]
    names = ', '.join([p['nm'] or p['cd'] for p in picks])[:200]
    if doc == 'PCRQ':
        # 🔴 구매는 단일 write(SP_SET_INQ_INFO, 문서 §3) — 수리처럼 별도 Confirm 단계가 없다.
        #    PARAM 은 워커가 SP_GET_REQ_INFO 로 재조회한 헤더 verbatim + STATUS='E'/INQ_NO='' 뿐이라
        #    여기서 지어내지 않는다(워커의 build_inquiry_purchase 가 유일한 조립처).
        envelope = {'step1': {'PACKAGE': 'PKG_PC_INQ', 'PROCEDURE': 'SP_SET_INQ_INFO',
                              'PARAM': '헤더(SP_GET_REQ_INFO) verbatim + STATUS=E/INQ_NO="" '
                                       '(워커가 전송 시점에 재조회·재조립)',
                              'CURSOR': {'P_IC_VNDR': '선택 %d개 (워커가 SP_GET_VNDR 재조회로 verbatim 구성)'
                                                      % len(picks),
                                         'P_IC/P_IC_CS': '부품그리드 verbatim echo(DM_YN 으로 분기)'}},
                    'rep_cd': rep_cd, 'req_no': row['req_no'], 'vsl_cd': row['vsl_cd'],
                    'subject': row['subject'], 'vendors': picks,
                    'svms_status_at_approval': row['svms_status']}
    else:
        envelope = {'step1': {'PACKAGE': 'PKG_MA_REP', 'PROCEDURE': 'SP_SET_REP',
                              'PARAM': '헤더 verbatim + STATUS=RC (워커가 전송 시점에 재조회)'},
                    'step2': {'PACKAGE': 'PKG_MA_REP', 'PROCEDURE': 'SP_SET_REP_DTL',
                              'PARAM': {'REP_CD': rep_cd, 'REP_YN': 'Y', 'USE_YN': 'Y', 'REF_TP': 'MARP'},
                              'CURSOR': {'P_IC_VNDR': '선택 %d개 (워커가 SP_GET_VNDR 재조회로 verbatim 구성)'
                                                      % len(picks)}},
                    'rep_cd': rep_cd, 'req_no': row['req_no'], 'vsl_cd': row['vsl_cd'],
                    'subject': row['subject'], 'vendors': picks,
                    'svms_status_at_approval': row['svms_status']}
    who = session.get('username') or 'web'
    try:
        did = execute(
            "INSERT INTO dock_inquiry_draft (rid, vsl_nm, vsl_cd, req_no, rep_cd, doc_type, vndr_json, "
            "vndr_names, envelope_json, status, decided_at, decided_by) "
            "VALUES (?,?,?,?,?,?,?,?,?,'approved',datetime('now','localtime'),?)",
            (rid, row['vsl_nm'], row['vsl_cd'], row['req_no'], rep_cd, doc or 'MARP',
             json.dumps(picks, ensure_ascii=False, sort_keys=True, separators=(',', ':')), names,
             json.dumps(envelope, ensure_ascii=False, sort_keys=True, separators=(',', ':')), who))
    except sqlite3.IntegrityError:                         # 부분 유니크 = 같은 건 이중 큐잉
        act = query("SELECT id, status FROM dock_inquiry_draft WHERE rep_cd=? AND status IN (?,?)",
                    (rep_cd, *_DOCK_SUBMIT_ACTIVE), one=True)
        return jsonify({'error': '이미 견적요청 큐에 있음', 'id': act['id'] if act else None,
                        'status': act['status'] if act else None}), 409
    return jsonify({'id': did, 'status': 'approved', 'rep_cd': rep_cd,
                    'vndr_cds': cds, 'vndr_names': names}), 201


@app.route('/api/dock_inquiry/drafts/<int:did>/cancel', methods=['POST'])
@admin_required
def api_dock_inquiry_cancel(did):
    rc = execute_rc("UPDATE dock_inquiry_draft SET status='canceled', done_at=datetime('now','localtime'), "
                    "result=COALESCE(result,'')||' [canceled by '||?||']' "
                    "WHERE id=? AND status='approved'", (session.get('username') or 'web', did))
    if not rc:
        cur = query('SELECT status FROM dock_inquiry_draft WHERE id=?', (did,), one=True)
        if not cur:
            return jsonify({'error': 'not found'}), 404
        return jsonify({'error': '대기(approved) 상태만 취소 가능', 'status': cur['status']}), 409
    return jsonify({'id': did, 'status': 'canceled'})


@app.route('/api/dock_inquiry/drafts/<int:did>/push', methods=['POST'])
@admin_required
def api_dock_inquiry_push(did):
    """🔴 [지금 전송] — 누른 그 순간에만 1건이 SVMS 로 나간다. **스케줄러 없음.**
    성공 판정은 맥의 `ok` 가 아니라 **DB 최종상태**(워커가 readback 후 `/result` 로 기록)."""
    cur = query('SELECT status FROM dock_inquiry_draft WHERE id=?', (did,), one=True)
    if not cur:
        return jsonify({'error': 'not found'}), 404
    if cur['status'] != 'approved':
        return jsonify({'error': '대기(approved) 상태만 전송 가능', 'status': cur['status']}), 409
    # 45s — gunicorn `--timeout 60` 보다 짧아야 한다(넘기면 -w 1 워커가 죽어 다른 요청까지 끊긴다).
    body, code, err = _dock_mac_call('/push_inquiry', {'draft_id': did}, timeout=45)
    if err and code == 504:
        # 🔴 타임아웃은 '안 나갔다'가 아니다 — 맥이 SVMS 로 이미 보냈을 수 있다.
        st = query('SELECT status FROM dock_inquiry_draft WHERE id=?', (did,), one=True)
        return jsonify({'error': '응답 시간 초과 — 전송 여부 불확실. 다시 누르지 말고 SVMS 에서 확인하세요',
                        'status': st['status'] if st else None, 'ambiguous': True}), 504
    if err and code == 503:
        return jsonify({'error': err}), 503
    st = query('SELECT status, result FROM dock_inquiry_draft WHERE id=?', (did,), one=True)
    done = bool(st and st['status'] == 'submitted')
    return jsonify({'id': did, 'ok': bool((body or {}).get('ok')) and done,
                    'msg': (body or {}).get('msg') or err or '',
                    'status': st['status'] if st else None,
                    'result': st['result'] if st else None}), (200 if code == 200 else code)


@app.route('/api/dock_inquiry/drafts/decided', methods=['DELETE'])
@admin_required
def api_dock_inquiry_clear_decided():
    # 🔴 `recalled` 도 종료상태다(2026-08-03). 안 넣으면 회수로 무효화된 이력이 '완료건 지우기'로
    #    영원히 안 지워져 큐 목록에 쌓인다(올마이트가 지적한 consumer 갭 — 실측 확인).
    #    Phase ③ 상신 쪽 `_DOCK_SUBMIT_DONE` 의미는 건드리지 않고 이 라우트에서만 확장한다.
    _term = _DOCK_SUBMIT_DONE + ('recalled',)
    n = execute_rc("DELETE FROM dock_inquiry_draft WHERE status IN (%s)" % ','.join('?' * len(_term)),
                   _term)
    return jsonify({'ok': True, 'deleted': n})


# ---- ext (맥 inquiry_watch) ----
@app.route('/api/ext/dock_inquiry/approved')
@api_key_required
def api_ext_dock_inquiry_approved():
    """맥 워커가 처리할 approved 건 → CAS claim 으로 submitting 락. Phase ③ `/approved` 규약 동일:
    새로 claim 한 행만 반환 · `submitting` 6h 초과는 `failed`(**자동 재큐 안 함**) ·
    `?peek=1` 조회전용 · `decided_by` 빈 행은 claim 금지 · `?id=N` 은 그 행만(같은 CAS)."""
    cols = "id, rid, vsl_nm, vsl_cd, req_no, rep_cd, doc_type, vndr_json, vndr_names, envelope_json"
    if request.args.get('peek'):
        rows = query(f"SELECT {cols} FROM dock_inquiry_draft WHERE status='approved' ORDER BY id ASC")
        return jsonify({'count': len(rows), 'drafts': [dict(r) for r in rows], 'peek': True})
    try:
        limit = max(1, min(20, int(request.args.get('limit') or 1)))
    except (TypeError, ValueError):
        limit = 1
    execute("UPDATE dock_inquiry_draft SET status='failed', "
            "result=COALESCE(result,'')||' [auto:6h+ submitting→failed, 사람 재검토]' "
            "WHERE status='submitting' AND done_at IS NOT NULL "
            "AND done_at < datetime('now','localtime','-6 hours')")
    where, params = '', ()
    if 'id' in request.args:
        try:
            where, params, limit = ' AND id=?', (int(request.args['id']),), 1
        except (TypeError, ValueError):
            return jsonify({'count': 0, 'drafts': [], 'error': 'id 형식 오류'}), 400
    out = []
    for r in query(f"SELECT {cols} FROM dock_inquiry_draft WHERE status='approved' "
                   "AND decided_at IS NOT NULL AND COALESCE(decided_by,'')<>''"
                   + where + " ORDER BY id ASC", params):
        if len(out) >= limit:
            break
        if execute_rc("UPDATE dock_inquiry_draft SET status='submitting', done_at=datetime('now','localtime') "
                      "WHERE id=? AND status='approved' AND decided_at IS NOT NULL "
                      "AND COALESCE(decided_by,'')<>''", (r['id'],)):
            out.append(dict(r))
    return jsonify({'count': len(out), 'drafts': out, 'limit': limit})


@app.route('/api/ext/dock_inquiry/drafts/<int:did>/result', methods=['POST'])
@api_key_required
def api_ext_dock_inquiry_result(did):
    """견적요청 결과 — ok=True → submitted, else failed. 판정 근거는 워커의 readback
    (`SP_GET_REP_INFO` 재조회로 헤더가 `RE`(Quotation Inquiry)로 갔는지 확인)."""
    d = request.get_json(silent=True) or {}
    ok = bool(d.get('ok'))
    rc = execute_rc("UPDATE dock_inquiry_draft SET status=?, done_at=datetime('now','localtime'), result=? "
                    "WHERE id=? AND status='submitting'",
                    ('submitted' if ok else 'failed', (d.get('result') or '')[:2000], did))
    if rc and ok:
        # 다음 SVMS sync 까지의 공백을 메우는 임시 표시 — 벤더제출(rank 2)까지만 켠다.
        # 🔴 `stg_confirm`/`stg_order` 는 건드리지 않는다(견적요청은 컨펌·발주가 아니다).
        dr = query('SELECT rid FROM dock_inquiry_draft WHERE id=?', (did,), one=True)
        if dr and dr['rid']:
            execute("UPDATE dock_procure SET stg_quote=1, stg_vendor=1, svms_status='Quotation Inquiry', "
                    "svms_synced_at=datetime('now','localtime'), updated_at=datetime('now','localtime') "
                    "WHERE id=?", (dr['rid'],))
        # 표시칸 `svms_submit`(제출수/요청수)·`svms_req_no`(INQ_NO)·SVMS 실단계 라벨은 위 낙관적
        # UPDATE 가 못 채운다 — **폴러(dock_sync)만** 채우는 칸이다. 자동 폴러는 1시간 간격이라
        # 견적요청 성공 후 최대 1시간 동안 "0/N 없음 · Quotation Inquiry 링크 없음"이 보였다
        # (2026-08-04 형 실관측, S14/BGBBES2607B41). 기존 온디맨드 트리거 flag 를 재사용해
        # ~1분 안에 채워지게 한다(새 인프라·새 스케줄러 0개). 실패해도 결과기록은 유지(격리).
        try:
            _dock_sync_flag_bump()
        except Exception:
            app.logger.warning('dock_sync flag bump 실패(무시) did=%s', did, exc_info=True)
    return jsonify({'id': did, 'ok': ok, 'applied': bool(rc)})


@app.route('/automation')
@admin_required
def automation_page():
    return render_template('automation.html')


@app.route('/api/automation/run', methods=['POST'])
@admin_required
def api_automation_run():
    _ensure_api_table()
    d = request.get_json(silent=True)
    if not isinstance(d, dict):
        return jsonify({'error': 'bad body'}), 400
    task, mode = d.get('task'), (d.get('mode') or 'verify')
    if not isinstance(task, str) or not isinstance(mode, str):   # non-str 방어(500 회피, 올마이트)
        return jsonify({'error': 'bad task/mode'}), 400
    task, mode = task.strip(), mode.strip()
    if task not in automation_tasks() or mode not in AUTOMATION_MODES:
        return jsonify({'error': 'bad task/mode'}), 400
    # 선박별 SOA 검증: params(vsl_cd/vsl_cds 필수, 기간·부서·검증모델 옵션) 검증.
    # live=실기입(체크박스+리젝리마크). 순수 DRY는 카나리/CLI용으로만 유지.
    params = None
    vessel_params = []
    if task == 'soa_vessel':
        p = d.get('params')
        if not isinstance(p, dict):
            p = {}
        try:
            vessel_params = [_soa_vessel_params(p, vsl) for vsl in _soa_vessel_codes_from_params(p)]
        except ValueError as e:
            return jsonify({'error': str(e)}), 400
        params = json.dumps(vessel_params[0], ensure_ascii=False)
    if not _automation_enabled():
        return jsonify({'error': 'killswitch ON — 자동화 정지중. 마스터 스위치 먼저 켜세요.'}), 409
    # lock: 같은 task가 queued/running이면 거부(중복클릭·동시실행 방지)
    busy = query("SELECT 1 FROM automation_run WHERE task=? AND status IN ('queued','running') LIMIT 1",
                 (task,), one=True)
    if busy:
        return jsonify({'error': '이미 실행 대기/진행중입니다.'}), 409
    import uuid
    user = session.get('username', '')
    if task == 'soa_vessel' and len(vessel_params) > 1:
        run_ids = []
        db = get_db()
        try:
            for pp in vessel_params:
                rid = uuid.uuid4().hex[:12]
                db.execute("INSERT INTO automation_run (run_id, task, mode, status, requested_by, params) "
                           "VALUES (?,?,?, 'queued', ?, ?)",
                           (rid, task, mode, user, json.dumps(pp, ensure_ascii=False)))
                run_ids.append({'run_id': rid, 'vsl_cd': pp['vsl_cd']})
            db.commit()
        except sqlite3.Error:
            db.rollback()
            app.logger.exception('soa-vessel multi enqueue failed')
            return jsonify({'error': '선박별 SOA 검증 큐 적재 실패 — 아무 작업도 큐에 넣지 않았습니다.'}), 500
        return jsonify({'ok': True, 'run_ids': run_ids, 'count': len(run_ids)})
    rid = uuid.uuid4().hex[:12]
    execute("INSERT INTO automation_run (run_id, task, mode, status, requested_by, params) "
            "VALUES (?,?,?, 'queued', ?, ?)", (rid, task, mode, user, params))
    return jsonify({'ok': True, 'run_id': rid})


@app.route('/api/automation/runs')
@admin_required
def api_automation_runs():
    rows = query("SELECT run_id,task,mode,status,requested_at,started_at,finished_at,exit_code,summary,progress "
                 "FROM automation_run ORDER BY id DESC LIMIT 40")
    total = query("SELECT COUNT(*) c FROM automation_run", one=True)['c']
    cleared = None
    try:
        r = query("SELECT v FROM api_settings WHERE k='automation_log_cleared'", one=True)
        if r and r['v']:
            cleared = json.loads(r['v'])
    except (sqlite3.Error, ValueError):
        pass
    return jsonify({
        'enabled': _automation_enabled(),
        'tasks': automation_tasks(),
        'runs': [dict(r) for r in rows],
        'total': total,
        'cleared': cleared,
    })


@app.route('/api/automation/runs', methods=['DELETE'])
@admin_required
def api_automation_runs_clear():
    """완료/실패 로그만 삭제(진행중 보존). 삭제 행위 자체는 api_settings 에 기록."""
    _ensure_api_table()
    n = execute_rc("DELETE FROM automation_run WHERE status IN ('done','failed')")
    user = session.get('username', '')
    now = query("SELECT datetime('now','localtime') t", one=True)['t']
    execute("INSERT OR REPLACE INTO api_settings (k, v) VALUES ('automation_log_cleared', ?)",
            (json.dumps({'at': now, 'by': user, 'n': n}, ensure_ascii=False),))
    return jsonify({'ok': True, 'deleted': n})


@app.route('/api/automation/killswitch', methods=['POST'])
@admin_required
def api_automation_killswitch():
    _ensure_api_table()
    d = request.get_json(silent=True) or {}
    on = bool(d.get('enabled'))
    execute("INSERT OR REPLACE INTO api_settings (k, v) VALUES ('automation_enabled', ?)",
            ('1' if on else '0',))
    return jsonify({'ok': True, 'enabled': on})


# ---- ext (맥미니 launchd 폴링) ----
@app.route('/api/ext/automation/enqueue', methods=['POST'])
@api_key_required
def api_ext_automation_enqueue():
    """무인 스케줄러(launchd)가 task 를 큐에 적재. ⚠️ 안전상 verify(읽기전용)만 허용 —
    무인 자동으로는 절대 상신/승인(live)이 안 되게 잠근다. live 는 사람이 허브 버튼으로만."""
    d = request.get_json(silent=True) or {}
    task = (d.get('task') or '').strip()
    mode = (d.get('mode') or 'verify').strip()
    if task not in automation_tasks():
        return jsonify({'error': 'bad task'}), 400
    if mode != 'verify':
        return jsonify({'error': 'ext enqueue 는 verify 만 허용(무인 상신 차단)'}), 403
    if not _automation_enabled():
        return jsonify({'error': 'killswitch ON'}), 409
    busy = query("SELECT 1 FROM automation_run WHERE task=? AND status IN ('queued','running') LIMIT 1",
                 (task,), one=True)
    if busy:
        return jsonify({'skipped': True, 'reason': '이미 대기/진행중'}), 200
    import uuid
    rid = uuid.uuid4().hex[:12]
    execute("INSERT INTO automation_run (run_id, task, mode, status, requested_by) "
            "VALUES (?,?,?, 'queued', 'scheduler')", (rid, task, mode))
    return jsonify({'ok': True, 'run_id': rid})


@app.route('/api/ext/automation/claim', methods=['POST'])
@api_key_required
def api_ext_automation_claim():
    if not _automation_enabled():
        return jsonify({'run': None, 'disabled': True})
    # stuck-running 회수(보수적): 러너 사망(맥 다운 등)으로 6시간 넘게 running 이면 failed 처리.
    # 짧게 잡으면 살아있는 장기 run 을 오판→이중 dispatch(돈경로) 위험이라 길게(6h) —
    # 재큐잉 안 함(사람이 허브에서 재실행). 정상 run 은 수 분 내라 6h 오탐 없음.
    execute("UPDATE automation_run SET status='failed', finished_at=datetime('now','localtime'), "
            "summary=COALESCE(summary,'') || ' [auto-expired: running>6h, 러너 무응답 간주]', progress=NULL "
            "WHERE status='running' AND started_at IS NOT NULL "
            "AND started_at < datetime('now','localtime','-6 hours')")
    # 진행중이 있으면 신규 claim 안 함(스크립트 순차 실행 — SVMS 세션 충돌 방지)
    running = query("SELECT 1 FROM automation_run WHERE status='running' LIMIT 1", one=True)
    if running:
        return jsonify({'run': None, 'busy': True})
    row = query("SELECT id,run_id,task,mode,params FROM automation_run WHERE status='queued' ORDER BY id ASC LIMIT 1",
                one=True)
    if not row:
        return jsonify({'run': None})
    # 조건부 claim — rowcount 0 이면(다른 폴러가 먼저 잡음) dispatch 안 함(이중실행 방지)
    rc = execute_rc("UPDATE automation_run SET status='running', started_at=datetime('now','localtime') "
                    "WHERE id=? AND status='queued'", (row['id'],))
    if not rc:
        return jsonify({'run': None, 'busy': True})
    try:
        _params = json.loads(row['params']) if row['params'] else {}
    except Exception:
        _params = {}
    if not isinstance(_params, dict):
        _params = {}
    # soa_vessel은 vsl_cd 필수 — 무효면 dispatch 안 하고 failed 처리(fail-closed, 올마이트)
    if row['task'] == 'soa_vessel' and not re.match(r'^[A-Z]{4}$', str(_params.get('vsl_cd') or '')):
        execute("UPDATE automation_run SET status='failed', finished_at=datetime('now','localtime'), "
                "summary='params 무효(vsl_cd 없음/형식오류) — dispatch 취소' WHERE id=?", (row['id'],))
        return jsonify({'run': None})
    return jsonify({'run': {'run_id': row['run_id'], 'task': row['task'], 'mode': row['mode'], 'params': _params}})


@app.route('/api/ext/automation/<run_id>/progress', methods=['POST'])
@api_key_required
def api_ext_automation_progress(run_id):
    """러너 중간보고. 화면에서 '굳음'과 '도는중'을 구분하기 위한 표시 전용 필드.
    · status='running' 인 행에만 쓴다 — 끝난 run 에 진행문구가 되살아나면 오히려 오독을 만든다.
    · 제어문자 제거 + 길이 컷: 러너 stdout 이 그대로 화면에 들어오는 경로라 로그 injection 방지.
    · 실패해도 러너는 계속 돈다(호출측이 삼킴) — 여기서 500 을 내도 돈경로에는 영향 없음."""
    d = request.get_json(silent=True) or {}
    raw = str(d.get('progress') or '')[:300]
    text = ''.join(c if c.isprintable() else ' ' for c in raw).strip()
    execute("UPDATE automation_run SET progress=? WHERE run_id=? AND status='running'",
            (text or None, run_id))
    return jsonify({'ok': True})


@app.route('/api/ext/automation/<run_id>/done', methods=['POST'])
@api_key_required
def api_ext_automation_done(run_id):
    d = request.get_json(silent=True) or {}
    status = 'failed' if (d.get('status') == 'failed' or d.get('exit_code')) else 'done'
    summary = (d.get('summary') or '')[:4000]
    # progress=NULL: 끝난 run 에 '준비 3/12 판독중' 같은 중간문구가 남아 있으면 진행중으로 오독된다.
    execute("UPDATE automation_run SET status=?, finished_at=datetime('now','localtime'), "
            "exit_code=?, summary=?, progress=NULL WHERE run_id=?",
            (status, d.get('exit_code'), summary, run_id))
    # Fail-safe: review scripts normally POST their structured result first. If they crash before that,
    # never leave the case permanently locked; the run summary remains visible for manual reconcile.
    if query('SELECT 1 FROM soa_review_case WHERE queued_run_id=?', (run_id,), one=True):
        _soa_review_case_unlock(run_id, result=f'{status}: {summary[:500]}')
    return jsonify({'ok': True})


# ---- 전자결재(jeonja) 검증 결과 적재 / 자동상신 제외 체크 ----
def _jeonja_ref(ref):
    """Canonical safe document ref for preview cache/API."""
    value = (ref or '').strip().upper()
    return value if re.fullmatch(r'[A-Z0-9_-]{6,64}', value) else None


def _jeonja_pdf_path(ref):
    """Hash-backed cache path; raw business ref never becomes a filesystem path."""
    safe = _jeonja_ref(ref)
    if not safe:
        raise ValueError('invalid ref')
    key = hashlib.sha256(safe.encode('utf-8')).hexdigest()
    return os.path.join(JEONJA_PDF_DIR, key + '.pdf')


def _jeonja_pdf_delete(ref):
    """Best-effort preview cache cleanup; never affects the SVMS/NAS original."""
    try:
        p = _jeonja_pdf_path(ref)
        if os.path.exists(p):
            os.remove(p)
            return True
    except Exception:
        app.logger.exception('jeonja-pdf-delete')
    return False


@app.route('/api/automation/jeonja/items/<ref>/pdf')
@admin_required
def api_automation_jeonja_pdf(ref):
    """Reviewed invoice/DN cache preview. Completed/removed items naturally 404."""
    safe = _jeonja_ref(ref)
    if not safe:
        abort(404)
    row = query('SELECT ref FROM jeonja_review_item WHERE ref=?', (safe,), one=True)
    if not row:
        abort(404)
    p = _jeonja_pdf_path(safe)
    if not os.path.exists(p):
        abort(404)
    return send_file(p, mimetype='application/pdf', as_attachment=False,
                     download_name='jeonja_%s.pdf' % safe, conditional=True)


@app.route('/api/ext/jeonja/review/<ref>/pdf', methods=['POST'])
@api_key_required
def api_ext_jeonja_pdf_upload(ref):
    """Review runner uploads only the invoice/DN PDF actually used for judgment."""
    MAX = 25 * 1024 * 1024
    safe = _jeonja_ref(ref)
    if not safe:
        return jsonify({'error': 'invalid ref'}), 400
    if request.content_length and request.content_length > MAX:
        return jsonify({'error': 'too large'}), 413
    row = query('SELECT bucket FROM jeonja_review_item WHERE ref=?', (safe,), one=True)
    if not row:
        return jsonify({'error': 'not found'}), 404
    if row['bucket'] == 'already':
        return jsonify({'error': 'completed'}), 409
    data = request.files['pdf'].read() if request.files.get('pdf') else request.get_data()
    if not data:
        return jsonify({'error': 'empty'}), 400
    if len(data) > MAX:
        return jsonify({'error': 'too large'}), 413
    if data[:5] != b'%PDF-':
        return jsonify({'error': 'not pdf'}), 400
    final = _jeonja_pdf_path(safe)
    tmp = final + '.' + uuid.uuid4().hex + '.tmp'
    try:
        with open(tmp, 'wb') as fh:
            fh.write(data)
        os.replace(tmp, final)
    finally:
        if os.path.exists(tmp):
            os.remove(tmp)
    return jsonify({'ref': safe, 'stored': True, 'bytes': len(data)})


@app.route('/api/ext/jeonja/review/<ref>/complete', methods=['POST'])
@api_key_required
def api_ext_jeonja_pdf_complete(ref):
    """Confirmed submission cleanup. Held rows are never deletable through this path."""
    safe = _jeonja_ref(ref)
    if not safe:
        return jsonify({'error': 'invalid ref'}), 400
    row = query('SELECT excluded FROM jeonja_review_item WHERE ref=?', (safe,), one=True)
    if row and row['excluded'] != 0:
        return jsonify({'error': 'held item', 'ref': safe}), 409
    row_deleted = bool(execute_rc('DELETE FROM jeonja_review_item WHERE ref=? AND excluded=0', (safe,)))
    pdf_deleted = _jeonja_pdf_delete(safe) if (row_deleted or not row) else False
    return jsonify({'ref': safe, 'row_deleted': row_deleted, 'pdf_deleted': pdf_deleted})


@app.route('/api/ext/jeonja/review', methods=['POST'])
@api_key_required
def api_ext_jeonja_review():
    """Store current review set atomically and reset preview cache for fresh upload."""
    _ensure_api_table()
    d = request.get_json(silent=True) or {}
    items = d.get('items') or []
    run_id = (d.get('run_id') or '').strip()
    db = get_db()
    prev_rows = db.execute('SELECT ref, excluded FROM jeonja_review_item').fetchall()
    canon = lambda value: _jeonja_ref(value) or (value or '').strip()
    prev_refs = {canon(r['ref']) for r in prev_rows if canon(r['ref'])}
    prev_excluded = {canon(r['ref']) for r in prev_rows if r['excluded'] == 1 and canon(r['ref'])}
    DEFAULT_HOLD = {'mismatch'}
    current_refs, completed_refs = set(), set()
    n = invalid = 0
    try:
        db.execute('DELETE FROM jeonja_review_item')
        for it in items:
            raw_ref = (it.get('ref') or '').strip()
            if not raw_ref:
                continue
            safe_ref = _jeonja_ref(raw_ref)
            ref = safe_ref or raw_ref
            bucket = (it.get('bucket') or 'flag')
            why = it.get('why')
            if not safe_ref:
                bucket = 'flag'
                why = ('비정규 REF 형식 — 자동상신 보류: ' + raw_ref)[:500]
                invalid += 1
            elif bucket == 'already':
                completed_refs.add(safe_ref)
                continue
            excl = 1 if (not safe_ref or ref in prev_excluded or bucket in DEFAULT_HOLD) else 0
            db.execute("INSERT OR REPLACE INTO jeonja_review_item "
                       "(ref,vsl_cd,subj,fund,cost,dn,bucket,why,excluded,run_id) VALUES (?,?,?,?,?,?,?,?,?,?)",
                       (ref, it.get('vsl_cd'), it.get('subj'), it.get('fund'), it.get('cost'),
                        it.get('dn'), bucket, why, excl, run_id))
            current_refs.add(ref)
            if bucket == 'already': completed_refs.add(ref)
            n += 1
        db.commit()
    except Exception:
        db.rollback()
        raise
    for ref in prev_refs | current_refs | completed_refs:
        if _jeonja_ref(ref): _jeonja_pdf_delete(ref)
    kept = len(prev_excluded & current_refs)
    return jsonify({'ok': True, 'count': n, 'kept_excluded': kept, 'invalid_refs': invalid})


@app.route('/api/automation/jeonja/items')
@admin_required
def api_automation_jeonja_items():
    """Review checklist plus read-only preview availability."""
    rows = query("SELECT ref,vsl_cd,subj,fund,cost,dn,bucket,why,excluded,reviewed_at "
                 "FROM jeonja_review_item ORDER BY CASE bucket "
                 "WHEN 'pass' THEN 0 WHEN 'costslip' THEN 1 WHEN 'mismatch' THEN 2 "
                 "WHEN 'escalate' THEN 3 WHEN 'flag' THEN 4 WHEN 'already' THEN 5 ELSE 6 END, ref") or []
    items = []
    for row in rows:
        item = dict(row)
        try:
            item['has_pdf'] = os.path.exists(_jeonja_pdf_path(item['ref']))
        except ValueError:
            item['has_pdf'] = False
        items.append(item)
    return jsonify({'items': items,
                    'reviewed_at': rows[0]['reviewed_at'] if rows else None})


@app.route('/api/automation/jeonja/exclude', methods=['POST'])
@admin_required
def api_automation_jeonja_exclude():
    """항목별 '자동상신 제외(보류)' 토글. 검증 통과건이어도 excluded=1 이면 live 가 skip."""
    d = request.get_json(silent=True) or {}
    ref = (d.get('ref') or '').strip()
    excluded = 1 if d.get('excluded') else 0
    if not ref:
        return jsonify({'error': 'no ref'}), 400
    rc = execute_rc("UPDATE jeonja_review_item SET excluded=? WHERE ref=?", (excluded, ref))
    return jsonify({'ok': bool(rc), 'ref': ref, 'excluded': bool(excluded)})


@app.route('/api/ext/jeonja/exclusions')
@api_key_required
def api_ext_jeonja_exclusions():
    """맥 live(jeonja_approve) 가 자동상신 직전 호출 — 보류 ref 는 상신에서 제외."""
    rows = query("SELECT ref FROM jeonja_review_item WHERE excluded=1")
    return jsonify({'refs': [r['ref'] for r in rows]})


# mail_card historical rows are retained in SQLite for audit only.
# The UI/API/automation workflow was retired on 2026-07-13.

# ═════════════════════════════════════════════════════════════════
#  Ship-Issue Wiki — 선박별 이슈 지식노트 검토/승격 큐
#   파이프라인: 맥 crawl→librarian→pending → [이 탭: 사람 승격/병합/리젝] → wiki(confirmed)
#   브릿지: push(맥→TRMT 적재) / decided(맥 pull) / result(맥→TRMT 결과). 발송·자동확정 없음.
# ═════════════════════════════════════════════════════════════════
SHIPWIKI_TIERS = ('pending', 'auto', 'confirmed')
SHIPWIKI_DECISIONS = ('promote', 'reject', 'split_flag', 'upgrade')


@app.route('/shipwiki')
@admin_required
def shipwiki_page():
    return render_template('shipwiki.html')


@app.route('/api/shipwiki/cards')
@admin_required
def api_shipwiki_cards():
    """탭 카드 목록 + 선박/tier/상태 통계. 기본 정렬: 미결(open) 우선, tier(pending>auto>confirmed), 신뢰도 낮은 순."""
    ship = (request.args.get('ship') or '').strip()
    where, params = [], []
    if ship:
        where.append('slug=?'); params.append(ship)
    sql = "SELECT * FROM shipwiki_card"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += (" ORDER BY CASE card_status WHEN 'open' THEN 0 WHEN 'decided' THEN 1 "
            "WHEN 'applying' THEN 2 WHEN 'failed' THEN 3 ELSE 4 END, "
            "CASE tier WHEN 'pending' THEN 0 WHEN 'auto' THEN 1 ELSE 2 END, "
            "multi DESC, COALESCE(llm_conf,0) ASC, id DESC")
    rows = [dict(r) for r in query(sql, tuple(params))]
    ships = [dict(r) for r in query(
        "SELECT slug, COALESCE(ship_nm,slug) ship_nm, COUNT(*) n, "
        "SUM(CASE WHEN tier='pending' AND card_status='open' THEN 1 ELSE 0 END) open_pending "
        "FROM shipwiki_card GROUP BY slug ORDER BY ship_nm")]
    stat = query("SELECT "
                 "SUM(CASE WHEN tier='pending' AND card_status='open' THEN 1 ELSE 0 END) pending_open, "
                 "SUM(CASE WHEN tier='auto' THEN 1 ELSE 0 END) auto_n, "
                 "SUM(CASE WHEN tier='confirmed' THEN 1 ELSE 0 END) confirmed_n, "
                 "SUM(CASE WHEN card_status='decided' THEN 1 ELSE 0 END) decided_n "
                 "FROM shipwiki_card", one=True)
    return jsonify({'cards': rows, 'ships': ships, 'stat': dict(stat) if stat else {},
                    'enabled': _automation_enabled()})


@app.route('/api/shipwiki/cards/<int:cid>/decide', methods=['POST'])
@admin_required
def api_shipwiki_decide(cid):
    """사람 결정 기록 → card_status='decided'(맥 apply 대기). 자동적재물 확정 = 100% 여기서만."""
    row = query("SELECT * FROM shipwiki_card WHERE id=?", (cid,), one=True)
    if not row:
        return jsonify({'error': 'not found'}), 404
    if row['card_status'] in ('applying',):
        return jsonify({'error': '맥 적용 진행중 — 잠시 후', 'status': row['card_status']}), 409
    d = request.get_json(silent=True) or {}
    decision = (d.get('decision') or '').strip()
    if decision not in SHIPWIKI_DECISIONS:
        return jsonify({'error': f'bad decision (one of {SHIPWIKI_DECISIONS})'}), 400
    # split_flag = 결정 아님(쪼갤 후보 표시만, materialize 없음) → open 유지
    new_status = 'open' if decision == 'split_flag' else 'decided'
    nt = (d.get('new_title') or '').strip() or row['title']
    nc = (d.get('new_category') or '').strip() or row['category']
    ncf = (d.get('new_conf') or '').strip()
    if decision == 'promote' and ncf not in ('medium', 'high'):
        ncf = 'medium'                                  # 사람 승격은 최소 medium
    if decision == 'upgrade' and ncf not in ('medium', 'high'):
        ncf = 'medium'
    jud = d.get('decided_judgment')
    if jud is not None:
        jud = jud.strip() or None
    mg = (d.get('merge_group') or '').strip() or None
    execute("UPDATE shipwiki_card SET decision=?, new_title=?, new_category=?, new_conf=?, "
            "decided_judgment=?, merge_group=?, card_status=?, decided_by=?, "
            "decided_at=datetime('now','localtime'), result=NULL WHERE id=?",
            (decision, nt, nc, ncf, jud, mg, new_status, session.get('username', ''), cid))
    return jsonify({'id': cid, 'decision': decision, 'card_status': new_status})


@app.route('/api/shipwiki/cards/<int:cid>/reset', methods=['POST'])
@admin_required
def api_shipwiki_reset(cid):
    """결정 취소 → open. 적용완료(applied)/진행중(applying)은 되돌리지 않음(파일 이미 생성)."""
    row = query("SELECT card_status FROM shipwiki_card WHERE id=?", (cid,), one=True)
    if not row:
        return jsonify({'error': 'not found'}), 404
    if row['card_status'] in ('applied', 'applying'):
        return jsonify({'error': '이미 적용됨/진행중 — reset 불가', 'status': row['card_status']}), 409
    execute("UPDATE shipwiki_card SET decision=NULL, new_title=NULL, new_category=NULL, new_conf=NULL, "
            "decided_judgment=NULL, merge_group=NULL, card_status='open', decided_by=NULL, "
            "decided_at=NULL, result=NULL WHERE id=?", (cid,))
    return jsonify({'id': cid, 'card_status': 'open'})


@app.route('/api/shipwiki/cards/<int:cid>', methods=['DELETE'])
@admin_required
def api_shipwiki_delete(cid):
    """카드 1건 삭제(TRMT 목록만 — 맥 파일엔 무영향). 다음 push 때 다시 적재될 수 있음."""
    execute("DELETE FROM shipwiki_card WHERE id=?", (cid,))
    return jsonify({'id': cid, 'deleted': True})


@app.route('/api/shipwiki/cards/applied', methods=['DELETE'])
@admin_required
def api_shipwiki_clear_applied():
    n = execute_rc("DELETE FROM shipwiki_card WHERE card_status='applied'")
    return jsonify({'deleted': n})


