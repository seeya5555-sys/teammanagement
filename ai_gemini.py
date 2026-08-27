"""Vetting / condition-survey APIs — the first boundary converted to a real module.

Until 2026-08-11 this file was executed inside the ``app`` namespace by
``_load_extracted_module``; every dependency was an invisible shared-global.
It is now a normal imported module with a Blueprint: dependencies are the
explicit imports below and nothing else, so a misspelled helper fails at
import time instead of as a request-time ``NameError``.

Import contract (enforced by ``test_converted_modules_are_self_contained``):
only stdlib, Flask, and the layers below it (``app_core``, ``helpers_shared``,
``app``) — never a sibling ``routes_*`` boundary.

Endpoints are prefixed ``ai_gemini.`` by the Blueprint; the URLs themselves are
unchanged and no template or Python code referenced the old endpoint names
(measured 2026-08-11: zero ``url_for``/nav hits).
"""
import json
import os
import re as _re_cls
import uuid
import hashlib
import tempfile
from difflib import SequenceMatcher
from datetime import datetime

from flask import abort, jsonify, request, send_from_directory, session, Response
from flask import Blueprint

from app_core import (
    UPLOAD_DIR, app, execute, get_db, query,
)
from helpers_shared import (
    GEMINI_API_KEY, VETTING_TYPES, _MARITIME_TERMS, _coerce_translation_items,
    _ext_allowed, _findings_workbook, _gemini_call_json, _model_for, _safe_filename,
    _vetting_display_order, _vetting_with_counts, _xlsx_to_text, login_required,
    api_key_required,
)

bp = Blueprint("ai_gemini", __name__)
























def _findings_prompt(kind):
    if kind == 'cs':
        return (
            "다음은 선박 컨디션 서베이(상태검사) 보고서다. 보고서에 적힌 지적/관찰 항목을 "
            "빠짐없이 추출해 지정한 JSON으로만 답하라. 각 항목 필드:\n"
            "- category: 'Defect' 또는 'Observation' (시정이 필요한 지적은 Defect, 권고/관찰사항은 Observation)\n"
            "- item: 짧은 제목 한 줄 (예: 'Main deck 부식')\n"
            "- description: 지적 상세 내용을 원문 그대로 복사한다(영문이면 영문 그대로). 요약·변형 금지.\n"
            "- remark: description의 핵심 지적사항을 한국어로 1~2문장으로 간결하게 요약한다(전체 직역 금지). 문장은 '~함/~됨/~음' 형태의 음슴체(개조식)로 끝맺는다. "
            "기술 명칭·장비명·약어(예: ECDIS, DCP, DRS, smoke detector, high-high level alarm 등)는 번역하지 말고 영문 그대로 둔다." + _MARITIME_TERMS + "\n"
            "없는 내용을 지어내지 말 것. 항목이 하나도 없으면 items를 빈 배열로.\n"
            '형식: {"items":[{"category":"Defect","item":"","description":"","remark":""}]}'
        )
    return (
        "다음은 선박 SIRE 2.0 점검 보고서다. 지적(결함) 사항만 추출한다.\n"
        "■ 포함 대상: 'Observable or detectable deficiency' 또는 'Not as expected'로 표시된 부정적 지적 "
        "(보고서에서 빨간색 글씨로 적힌 항목). 또한 'Photograph' 분류의 지적(예: 'Photo not representative', "
        "'Photograph supplied: ...' 아래 빨간 이탤릭 설명)처럼 사진 증빙이 부적절·불일치하다는 지적도 반드시 포함한다.\n"
        "■ 제외 대상: 'Exceeded normal expectation' 등 칭찬/긍정 평가(초록색 글씨)는 절대 포함하지 마라.\n"
        "각 지적 항목의 필드:\n"
        "- item: 항목 왼쪽에 표시된 분류 라벨을 괄호로 먼저 붙이고, 그 뒤에 굵게 표시된 "
        "지적 제목을 그대로 이어 붙인다. 분류 라벨은 보고서에 나온 그대로 쓴다 — "
        "Hardware · Human · Photograph · Process · Other 등 무엇이든. 예: "
        "'(Hardware)Misc Nautical Equipment – Maintenance deferred, awaiting spares', "
        "'(Human)Senior Engineer Officer – Not as expected', '(Photograph)Photo not representative'.\n"
        "- description: 제목 아래의 상세 본문(설명/이탤릭 문장 포함)을 영어 원문 그대로 복사한다. 요약·변형 금지.\n"
        "- remark: description의 핵심 지적사항을 한국어로 1~2문장으로 간결하게 요약한다(전체 직역 금지). 문장은 '~함/~됨/~음' 형태의 음슴체(개조식)로 끝맺는다. "
        "기술 명칭·장비명·약어(예: ECDIS, DCP, DRS, smoke detector, high-high level alarm, turn table 등)는 번역하지 말고 영문 그대로 둔다." + _MARITIME_TERMS + "\n"
        "없는 내용을 지어내지 말 것. 지적이 하나도 없으면 items를 빈 배열로.\n"
        '형식: {"items":[{"item":"","description":"","remark":""}]}'
    )


def _normalize_findings(parsed, kind):
    out = []
    if isinstance(parsed, list):
        arr = parsed
    elif isinstance(parsed, dict):
        arr = parsed.get('items') or parsed.get('findings') or []
    else:
        arr = []
    for it in (arr or []):
        if not isinstance(it, dict):
            continue
        rec = {
            'item':        (it.get('item') or '').strip(),
            'description': (it.get('description') or '').strip(),
            'remark':      (it.get('remark') or '').strip(),
        }
        if kind == 'cs':
            cat = it.get('category')
            rec['category'] = cat if cat in ('Defect', 'Observation') else 'Observation'
        if rec['item'] or rec['description']:
            out.append(rec)
    return out


def _xlsx_extract(raw_bytes, kind):
    """엑셀: 헤더가 명확하면 직접 매핑(AI 불필요), 자유양식이면 텍스트화 후 Gemini.
    반환: ('items', [...])  또는  ('text', '<탭구분 텍스트>')."""
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append(['' if c is None else str(c).strip() for c in r])
    if not rows:
        return ('items', [])

    KEY = {
        'category':    ['category', '구분', '분류', 'type', 'def/obs'],
        'item':        ['item', '항목', 'title', 'subject', '제목', 'short gen name', 'gen name', 'short name'],
        'description': ['description', 'detail', 'details', '내용', '상세', 'finding', 'observation', 'remarks/finding'],
        'remark':      ['remark', 'remarks', '비고', 'note', 'notes', 'comment', 'action', '조치'],
    }
    header_idx, colmap = None, {}
    for i, row in enumerate(rows[:6]):
        m = {}
        for ci, cell in enumerate(row):
            lc = cell.lower()
            for field, keys in KEY.items():
                if field in m:
                    continue
                if any(k == lc or k in lc for k in keys):
                    m[field] = ci
        if 'description' in m or ('item' in m and len(m) >= 2):
            header_idx, colmap = i, m
            break

    if header_idx is not None:
        items = []
        for row in rows[header_idx + 1:]:
            if not any(row):
                continue
            def g(f):
                ci = colmap.get(f)
                return row[ci] if ci is not None and ci < len(row) else ''
            rec = {'item': g('item'), 'description': g('description'), 'remark': g('remark')}
            if kind == 'cs':
                cat = (g('category') or '').strip().lower()
                rec['category'] = 'Defect' if cat.startswith('def') or '지적' in cat else 'Observation'
            if not rec['description'] and rec['item']:
                rec['description'] = rec['item']
            if rec['item'] or rec['description']:
                items.append(rec)
        return ('items', items)

    # 자유 양식 → 텍스트(TSV)로 변환
    lines = ['\t'.join(r) for r in rows if any(r)]
    return ('text', '\n'.join(lines[:400]))


def _summarize_remarks(items, kind):
    """엑셀 직접매핑 항목들의 remark를, 각 description의 한글 요약으로 채운다(배치 1회 호출).
    GEMINI 키 없거나 실패 시 기존 remark 값을 그대로 유지."""
    if not GEMINI_API_KEY or not items:
        return items
    payload = json.dumps(
        [{'i': idx, 'description': (it.get('description') or '')} for idx, it in enumerate(items)],
        ensure_ascii=False)
    prompt = (
        "아래는 선박 점검 지적 항목들의 description 목록(JSON 배열)이다. 각 항목의 description을 "
        "한국어로 1~2문장으로 간결하게 요약하라(전체 직역 금지). 문장은 '~함/~됨/~음' 형태의 음슴체(개조식)로 끝맺어라. 기술 명칭·장비명·약어"
        "(예: ECDIS, DCP, DRS, smoke detector, high-high level alarm 등)는 번역하지 말고 영문 그대로 둔다." + _MARITIME_TERMS + "\n"
        "입력의 i 값을 그대로 사용해 JSON으로만 답하라.\n"
        '형식: {"summaries":[{"i":0,"remark":"요약문"}]}\n\n[입력]\n' + payload)
    res = _gemini_call_json([{'text': prompt}], model=_model_for('remark'))
    if isinstance(res, dict):
        if res.get('error'):
            return items
        arr = res.get('summaries') or res.get('items') or res.get('translations') or []
    elif isinstance(res, list):
        arr = res
    else:
        arr = []
    by_i = {}
    for s in arr:
        if not isinstance(s, dict):
            continue
        try:
            by_i[int(s.get('i'))] = (s.get('remark') or s.get('en') or '').strip()
        except (TypeError, ValueError):
            pass
    for idx, it in enumerate(items):
        if by_i.get(idx):
            it['remark'] = by_i[idx]
    return items


def _extract_findings_from_upload(f, kind):
    """업로드 FileStorage → 항목 리스트. (items, err) 반환."""
    name = (f.filename or '').lower()
    ext = name.rsplit('.', 1)[-1] if '.' in name else ''
    raw = f.read()
    size_mb = len(raw) / (1024 * 1024)

    if ext in ('xlsx', 'xls'):
        try:
            mode, data = _xlsx_extract(raw, kind)
        except Exception as e:
            app.logger.exception('extract-findings-from-upload')
            return None, {'reason': 'XLSX_PARSE_FAILED', 'message': f'엑셀을 읽지 못했습니다: {e}'}
        if mode == 'items':
            return _summarize_remarks(data, kind), None
        parsed = _gemini_call_json([{'text': _findings_prompt(kind) + '\n\n[보고서 표 내용]\n' + data}], model=_model_for('findings'))
    elif ext == 'pdf':
        if size_mb > 15:
            return None, {'reason': 'TOO_LARGE', 'message': f'PDF가 너무 큽니다({size_mb:.1f}MB). 15MB 이하로 줄이거나 페이지를 나눠 올려주세요.'}
        b64 = __import__('base64').standard_b64encode(raw).decode()
        parsed = _gemini_call_json([
            {'inline_data': {'mime_type': 'application/pdf', 'data': b64}},
            {'text': _findings_prompt(kind)},
        ], model=_model_for('findings'))
    elif ext in ('png', 'jpg', 'jpeg', 'webp', 'gif', 'bmp'):
        if size_mb > 15:
            return None, {'reason': 'TOO_LARGE', 'message': f'이미지가 너무 큽니다({size_mb:.1f}MB).'}
        import mimetypes
        media = mimetypes.guess_type(name)[0] or 'image/jpeg'
        b64 = __import__('base64').standard_b64encode(raw).decode()
        parsed = _gemini_call_json([
            {'inline_data': {'mime_type': media, 'data': b64}},
            {'text': _findings_prompt(kind)},
        ], model=_model_for('findings'))
    else:
        return None, {'reason': 'BAD_TYPE', 'message': 'PDF, 이미지, 엑셀(xlsx) 파일만 지원합니다.'}

    if isinstance(parsed, dict) and parsed.get('error') == 'NO_API_KEY':
        return None, {'reason': 'no_api_key', 'message': 'AI 자동추출이 설정되지 않았습니다(키 미설정).'}
    if isinstance(parsed, dict) and parsed.get('error'):
        return None, {'reason': parsed['error'], 'message': '자동 추출에 실패했습니다.',
                      'detail': parsed.get('detail') or parsed.get('raw')}
    return _normalize_findings(parsed, kind), None


@bp.route('/api/cs/surveys/<int:sid>/extract-report', methods=['POST'])
@login_required
def api_cs_extract_report(sid):
    if not query('SELECT id FROM cs_surveys WHERE id=?', (sid,), one=True):
        abort(404)
    if 'file' not in request.files or not request.files['file'].filename:
        return jsonify({'ok': False, 'message': '파일이 없습니다.'}), 400
    items, err = _extract_findings_from_upload(request.files['file'], 'cs')
    if err:
        return jsonify({'ok': False, **err}), 200
    return jsonify({'ok': True, 'items': items, 'count': len(items)})


@bp.route('/api/cs/surveys/<int:sid>/export')
@login_required
def api_cs_survey_export(sid):
    from flask import send_file
    s = query('''SELECT cs.*, v.name AS vessel_name
                   FROM cs_surveys cs JOIN vessels v ON v.id = cs.vessel_id
                  WHERE cs.id=?''', (sid,), one=True)
    if not s:
        abort(404)
    fr = query('''SELECT category, no, item, description, remark, status
                    FROM cs_findings WHERE survey_id=?
                   ORDER BY CASE category WHEN 'Defect' THEN 0 ELSE 1 END, no, id''', (sid,))
    # RECTIFICATION·PHOTO 2열은 공란으로 출력(현장기입용). REMARK는 export에서 제외.
    rows = [[r['category'], r['no'], r['item'] or '', r['description'] or '',
             '', '', r['status'] or ''] for r in fr]
    vessel = s['vessel_name']
    title = f"Condition Survey — {vessel}  {s['year']} Q{s['quarter']}"
    sub_bits = [f"수검일: {s['inspection_date'] or '-'}", f"Vendor: {s['vendor'] or '-'}",
                f"총 {len(rows)}건 (Defect {sum(1 for r in fr if r['category']=='Defect')} / "
                f"Observation {sum(1 for r in fr if r['category']=='Observation')})"]
    headers = ['Category', 'No.', 'ITEM', 'DESCRIPTION', 'RECTIFICATION', 'PHOTO', 'STATUS']
    bio = _findings_workbook(title, '   │   '.join(sub_bits), headers, rows,
                             wrap_cols={3, 4, 5, 6}, widths=[12, 6, 28, 50, 40, 30, 10])
    fname = f"CS_{_safe_filename(vessel)}_{s['year']}Q{s['quarter']}.xlsx"
    return send_file(bio, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ----- CS 첨부파일 -----

@bp.route('/api/cs/surveys/<int:sid>/attachments', methods=['GET'])
@login_required
def api_cs_attachments_list(sid):
    rows = query(
        'SELECT * FROM cs_attachments WHERE survey_id=? ORDER BY id DESC',
        (sid,),
    )
    return jsonify([dict(r) for r in rows])


@bp.route('/api/cs/surveys/<int:sid>/attachments', methods=['POST'])
@login_required
def api_cs_attachment_upload(sid):
    if not query('SELECT id FROM cs_surveys WHERE id=?', (sid,), one=True):
        abort(404)
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다.'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'error': '파일명이 없습니다.'}), 400
    if not _ext_allowed(f.filename):
        return jsonify({'error': '허용되지 않는 파일 형식입니다.'}), 400

    ext = os.path.splitext(f.filename)[1]
    stored = f"cs_{uuid.uuid4().hex}{ext}"
    save_path = os.path.join(UPLOAD_DIR, stored)
    f.save(save_path)
    size = os.path.getsize(save_path)

    aid = execute("""
        INSERT INTO cs_attachments
            (survey_id, filename, stored_name, file_size, mime_type, uploaded_by)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (sid, f.filename, stored, size, f.mimetype, session.get('username')))
    return jsonify({'id': aid, 'filename': f.filename, 'file_size': size}), 201


@bp.route('/api/cs/attachments/<int:aid>', methods=['GET'])
@login_required
def api_cs_attachment_get(aid):
    a = query('SELECT * FROM cs_attachments WHERE id=?', (aid,), one=True)
    if not a:
        abort(404)
    inline = request.args.get('inline')
    return send_from_directory(
        UPLOAD_DIR, a['stored_name'],
        as_attachment=not inline,
        download_name=a['filename'],
    )


@bp.route('/api/cs/attachments/<int:aid>', methods=['DELETE'])
@login_required
def api_cs_attachment_delete(aid):
    a = query('SELECT * FROM cs_attachments WHERE id=?', (aid,), one=True)
    if not a:
        abort(404)
    p = os.path.join(UPLOAD_DIR, a['stored_name'])
    if os.path.exists(p):
        try: os.remove(p)
        except OSError:
            app.logger.exception('cs-attachment-delete')
    execute('DELETE FROM cs_attachments WHERE id=?', (aid,))
    return jsonify({'ok': True})










# ----- Vettings (vessel별 그룹) -----

@bp.route('/api/vettings', methods=['GET'])
@login_required
def api_vettings_list():
    """선박별 vetting 그룹 응답.
    Query: ?year=2026&supervisor_id=N
    응답: [ { vessel: {...}, vettings: [...with findings...] } ]
    """
    year = request.args.get('year', type=int)
    sup_id = request.args.get('supervisor_id', type=int)

    # 대상 선박: VLCC/AFRAMAX/LR/MR만
    placeholders = ','.join('?' * len(VETTING_TYPES))
    sql = f'SELECT v.* FROM vessels v WHERE v.active=1 AND v.vessel_type IN ({placeholders})'
    params = list(VETTING_TYPES)
    if sup_id:
        sql += ' AND EXISTS (SELECT 1 FROM supervisor_vessels sv WHERE sv.vessel_id=v.id AND sv.supervisor_id=?)'
        params.append(sup_id)
    sql += ' ORDER BY v.name'
    vessels = query(sql, tuple(params))

    # vetting 한번에
    # vetting 필터:
    #  - 검사일이 있는 것은 해당 연도와 일치할 때만
    #  - 검사일이 없는 것 (방금 + 새 Vetting 추가 한 빈 행)은 모든 연도에 항상 표시
    if year:
        vettings = query('SELECT * FROM vettings')
        vettings = [v for v in vettings
                    if (not v['inspection_date'])
                    or (v['inspection_date'].startswith(str(year)))]
    else:
        vettings = query('SELECT * FROM vettings')

    # findings 한번에
    vids = [v['id'] for v in vettings]
    findings_by_vid = {vid: [] for vid in vids}
    if vids:
        ph = ','.join('?' * len(vids))
        all_f = query(
            f'SELECT * FROM vt_findings WHERE vetting_id IN ({ph}) ORDER BY vetting_id, no',
            tuple(vids),
        )
        for f in all_f:
            findings_by_vid[f['vetting_id']].append(_public_vetting_finding(f))

    by_vessel = {}
    for v in vettings:
        d = _vetting_with_counts(v)
        d['findings'] = findings_by_vid.get(v['id'], [])
        by_vessel.setdefault(v['vessel_id'], []).append(d)

    # 표시 순서 = _vetting_display_order 정본 ('Next Plan' 먼저, 그 다음 검사일 내림차순).
    # 웹 상세 테이블 행 순서와 iOS 앱의 대표행(vettings.first) 이 모두 이 순서를 그대로 쓴다.
    for vid in by_vessel:
        by_vessel[vid] = _vetting_display_order(by_vessel[vid])

    # 선박별 담당 감독 ID 매핑 (Daily 이슈 등록 시 필요)
    sv_map = {}
    if vessels:
        v_ids = [v['id'] for v in vessels]
        ph2 = ','.join('?' * len(v_ids))
        rows = query(
            f'SELECT vessel_id, supervisor_id FROM supervisor_vessels WHERE vessel_id IN ({ph2})',
            tuple(v_ids),
        )
        for r in rows:
            sv_map.setdefault(r['vessel_id'], []).append(r['supervisor_id'])

    # 선박별 last_updated (해당 선박의 모든 vettings 중 가장 최근 updated_at)
    last_by_vessel = {}
    for v in vettings:
        u = v['updated_at']
        if u and (v['vessel_id'] not in last_by_vessel or u > last_by_vessel[v['vessel_id']]):
            last_by_vessel[v['vessel_id']] = u

    out = []
    for ves in vessels:
        vd = dict(ves)
        vd['supervisor_ids'] = sv_map.get(ves['id'], [])
        out.append({
            'vessel': vd,
            'vettings': by_vessel.get(ves['id'], []),
            'last_updated': last_by_vessel.get(ves['id']),
        })
    return jsonify(out)


@bp.route('/api/vettings', methods=['POST'])
@login_required
def api_vetting_create():
    """단일 vetting 생성. 선박 ID만 필수, 나머지는 선택."""
    d = request.get_json() or {}
    vid = d.get('vessel_id')
    if not vid:
        return jsonify({'error': 'vessel_id 가 필요합니다.'}), 400
    v = query('SELECT vessel_type FROM vessels WHERE id=?', (vid,), one=True)
    if not v:
        return jsonify({'error': '선박을 찾을 수 없습니다.'}), 404
    if v['vessel_type'] not in VETTING_TYPES:
        return jsonify({'error': f'Vetting은 {", ".join(VETTING_TYPES)} 선박에만 적용됩니다.'}), 400

    st = d.get('sire_type') or None
    if st and st not in ('Idle', 'Bunkering', 'Discharge'):
        st = None
    valid = d.get('valid') or None
    if valid and valid not in ('Next Plan', 'Last Result'):
        valid = None

    new_id = execute("""
        INSERT INTO vettings
            (vessel_id, report_number, inspection_date, inspection_company,
             inspector, port, sire_type, valid, overall_remark, created_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (vid,
          d.get('report_number') or '',
          d.get('inspection_date') or None,
          d.get('inspection_company') or '',
          d.get('inspector') or '',
          d.get('port') or '',
          st,
          valid,
          d.get('overall_remark') or '',
          session.get('username')))
    row = query('SELECT * FROM vettings WHERE id=?', (new_id,), one=True)
    return jsonify(_vetting_with_counts(row)), 201


@bp.route('/api/vettings/<int:vid>', methods=['GET'])
@login_required
def api_vetting_get(vid):
    v = query('SELECT * FROM vettings WHERE id=?', (vid,), one=True)
    if not v:
        abort(404)
    d = _vetting_with_counts(v)
    d['findings'] = [_public_vetting_finding(f) for f in query(
        'SELECT * FROM vt_findings WHERE vetting_id=? ORDER BY no', (vid,))]
    return jsonify(d)


@bp.route('/api/vettings/<int:vid>', methods=['PUT'])
@login_required
def api_vetting_update(vid):
    if not query('SELECT id FROM vettings WHERE id=?', (vid,), one=True):
        abort(404)
    d = request.get_json() or {}
    sets, params = [], []
    for f in ('report_number','inspection_date','inspection_company','inspector',
              'port','sire_type','valid','overall_remark',
              'manual_observation_count','manual_open_count','manual_close_count'):
        if f in d:
            sets.append(f'{f} = ?')
            v = d[f]
            params.append(None if v == '' else v)
    if not sets:
        return jsonify({'ok': True})
    sets.append("updated_at = datetime('now','localtime')")
    execute(f'UPDATE vettings SET {", ".join(sets)} WHERE id=?', tuple(params + [vid]))
    return jsonify({'ok': True})


@bp.route('/api/vettings/<int:vid>', methods=['DELETE'])
@login_required
def api_vetting_delete(vid):
    # 첨부 파일도 같이 삭제 (CASCADE는 DB만, 파일은 직접)
    atts = query('SELECT stored_name FROM vt_attachments WHERE vetting_id=?', (vid,))
    for a in atts:
        p = os.path.join(UPLOAD_DIR, a['stored_name'])
        if os.path.exists(p):
            try: os.remove(p)
            except OSError as e:
                app.logger.warning('vetting-delete: %s', e)
    execute('DELETE FROM vettings WHERE id=?', (vid,))
    return jsonify({'ok': True})


# ----- Findings -----

def _vt_next_no(vid):
    r = query('SELECT COALESCE(MAX(no), 0) + 1 AS next FROM vt_findings WHERE vetting_id=?',
              (vid,), one=True)
    return r['next']


@bp.route('/api/vettings/<int:vid>/findings', methods=['POST'])
@login_required
def api_vt_findings_create(vid):
    """단건 또는 배치(items 배열) 생성."""
    if not query('SELECT id FROM vettings WHERE id=?', (vid,), one=True):
        abort(404)
    d = request.get_json() or {}
    items = d.get('items')
    if items is None:
        items = [{
            'item':        d.get('item'),
            'description': d.get('description'),
            'remark':      d.get('remark'),
            'user_remark': d.get('user_remark'),
            'status':      d.get('status') or 'Open',
        }]

    next_no = _vt_next_no(vid)
    created = []
    for it in items:
        st = it.get('status') or 'Open'
        if st not in ('Open','Closed'): st = 'Open'
        fid = execute("""
            INSERT INTO vt_findings (vetting_id, no, item, description, remark, user_remark, priority, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (vid, next_no,
              it.get('item') or '',
              it.get('description') or '',
              it.get('remark') or '',
              it.get('user_remark') or '',
              1 if it.get('priority') else 0,
              st))
        created.append(fid)
        next_no += 1
    return jsonify({'ids': created, 'count': len(created)}), 201


@bp.route('/api/vt-findings/<int:fid>', methods=['PUT'])
@login_required
def api_vt_finding_update(fid):
    cur = query(
        'SELECT vetting_id,status,user_remark,full_report_remark FROM vt_findings WHERE id=?',
        (fid,), one=True)
    if not cur:
        abort(404)
    d = request.get_json() or {}
    sets, params = [], []
    for f in ('item','description','remark','user_remark','status'):
        if f in d:
            sets.append(f'{f} = ?')
            params.append(d[f] or '')
    if 'user_remark' in d and (d.get('user_remark') or '') != (cur['user_remark'] or ''):
        sets.append("full_report_remark = ''")
    if 'priority' in d:
        sets.append('priority = ?')
        params.append(1 if d.get('priority') else 0)
    if not sets:
        return jsonify({'ok': True})
    sets.append("updated_at = datetime('now','localtime')")
    execute(f'UPDATE vt_findings SET {", ".join(sets)} WHERE id=?', tuple(params + [fid]))

    # status 변경 시 vettings.updated_at 갱신 (선박 헤더의 Last update에 반영)
    if 'status' in d and d['status'] != cur['status']:
        execute(
            "UPDATE vettings SET updated_at = datetime('now','localtime') WHERE id=?",
            (cur['vetting_id'],),
        )
    return jsonify({'ok': True})


@bp.route('/api/vt-findings/<int:fid>', methods=['DELETE'])
@login_required
def api_vt_finding_delete(fid):
    f = query('SELECT vetting_id FROM vt_findings WHERE id=?', (fid,), one=True)
    if not f:
        abort(404)
    vid = f['vetting_id']
    execute('DELETE FROM vt_findings WHERE id=?', (fid,))
    # No 재정렬
    rows = query('SELECT id FROM vt_findings WHERE vetting_id=? ORDER BY no', (vid,))
    for new_no, r in enumerate(rows, start=1):
        execute('UPDATE vt_findings SET no=? WHERE id=?', (new_no, r['id']))
    return jsonify({'ok': True})


# ----- Attachments -----

def _vetting_full_prompt():
    return (
        "다음은 선박 SIRE 2.0 점검(Vetting Inspection) 보고서다. 두 가지를 추출해 지정한 JSON으로만 답하라.\n"
        "■ meta: 보고서 표지/상단의 점검 메타정보. 보고서에 해당 정보가 없으면 반드시 빈 문자열로 둔다(지어내지 말 것).\n"
        "- report_number: Report No / Report # / 보고서 번호\n"
        "- inspection_date: 점검 실시일 (반드시 YYYY-MM-DD 형식. 다른 형식이면 YYYY-MM-DD로 변환)\n"
        "- inspection_company: 점검 주체 / Oil Major / 제출사 (예: VIVA ENERGY, BP, SHELL, TOTAL)\n"
        "- inspector: 점검관(Inspector) 성명\n"
        "- port: 점검 항구명만 추출한다(도시/항구 이름). 국가명·UNLOCODE 코드(예: [SGSIN])·중복 표기는 제거. "
        "예: 'Singapore - Singapore [SGSIN]' → 'Singapore', 'Fujairah - UAE [AEFJR]' → 'Fujairah'.\n"
        "- sire_type: 점검 시 운항 상태. 반드시 'Idle' · 'Bunkering' · 'Discharge' 중 하나로만. 식별 불가 시 빈 문자열.\n"
        "■ items: 지적(결함) 사항만 추출한다.\n"
        "■ 포함: 'Observable or detectable deficiency' / 'Not as expected'로 표시된 부정적 지적(빨간 글씨). "
        "또한 'Photograph' 분류의 지적(예: 'Photo not representative', 'Photograph supplied: ...' 아래 빨간 이탤릭 설명)처럼 "
        "사진 증빙이 부적절·불일치하다는 지적도 반드시 포함한다.\n"
        "■ 제외: 'Exceeded normal expectation' 등 칭찬/긍정 평가(초록 글씨)는 절대 포함하지 마라.\n"
        "- item: 항목 왼쪽에 표시된 분류 라벨을 괄호로 먼저 붙이고, 그 뒤 굵게 표시된 지적 제목을 그대로 이어 붙인다. "
        "분류 라벨은 보고서에 나온 그대로 쓴다 — Hardware · Human · Photograph · Process · Other 등 무엇이든. "
        "예: '(Hardware)Misc Nautical Equipment – Maintenance deferred', '(Human)Senior Engineer Officer – Not as expected', "
        "'(Photograph)Photo not representative'.\n"
        "- description: 제목 아래 상세 본문(설명/이탤릭 문장 포함)을 영어 원문 그대로 복사. 요약·변형 금지.\n"
        "- remark: description의 핵심 지적사항을 한국어 1~2문장으로 간결하게 요약(전체 직역 금지). 문장은 '~함/~됨/~음' 음슴체(개조식). "
        "기술 명칭·장비명·약어(예: ECDIS, DCP, DRS, smoke detector, high-high level alarm 등)는 영문 그대로 둔다." + _MARITIME_TERMS + "\n"
        "없는 내용을 지어내지 말 것. 지적이 하나도 없으면 items를 빈 배열로.\n"
        '형식: {"meta":{"report_number":"","inspection_date":"","inspection_company":"","inspector":"",'
        '"port":"","sire_type":""},"items":[{"item":"","description":"","remark":""}]}'
    )


def _clean_port(p):
    """'Singapore - Singapore [SGSIN]' → 'Singapore'. 국가/코드/중복 제거, 항구명만."""
    s = (p or '').strip()
    if not s:
        return ''
    s = _re_cls.sub(r'\[[^\]]*\]', '', s)      # [SGSIN] 등 코드 제거
    s = s.split(' - ')[0]                       # ' - ' 앞 항구명만
    s = s.split(' / ')[0].split('/')[0]         # '/' 구분도 첫 토큰
    s = _re_cls.sub(r'\s+', ' ', s).strip(' -,')
    return s


def _norm_iso_date(s):
    """LLM 이 준 날짜를 'YYYY-MM-DD' 로만 통과시킨다. 아니면 '' (fail-closed).

    🔴 이 값은 그대로 `vettings.inspection_date` 에 들어가고 두 곳이 문자열 형식을 전제한다:
       ① 연도 필터 `inspection_date.startswith(str(year))` → 형식이 틀리면 그 vetting 이
          모든 연도 화면에서 사라진다.
       ② `date(vt.inspection_date,'+21 days') < date('now')` (routes_tail) → SQLite 는
          비-ISO 문자열에 NULL 을 돌려주므로 **SIRE OBS 3주 경과 알림이 조용히 미탐**된다.
       그래서 '12 Mar 2026' 같은 값은 저장하지 않고 빈 값으로 두어 사람이 채우게 한다.
    """
    s = (s or '').strip()
    if not _re_cls.fullmatch(r'\d{4}-\d{2}-\d{2}', s):
        return ''
    try:
        from datetime import datetime as _dt
        _dt.strptime(s, '%Y-%m-%d')              # 2026-02-31 같은 없는 날짜 배제
    except ValueError:
        return ''
    return s


def _norm_vetting_meta(m):
    m = m if isinstance(m, dict) else {}
    # 텍스트 필드는 길이를 자른다 — LLM 이 문단을 통째로 뱉으면 목록 UI 가 깨진다.
    g = lambda k: (m.get(k) or '').strip()[:200] if isinstance(m.get(k), str) else ''
    sire = g('sire_type')
    return {
        'report_number':      g('report_number'),
        'inspection_date':    _norm_iso_date(g('inspection_date')),
        'inspection_company': g('inspection_company'),
        'inspector':          g('inspector'),
        'port':               _clean_port(g('port')),
        'sire_type':          sire if sire in ('Idle', 'Bunkering', 'Discharge') else '',
        'valid':              '',   # '상태'(Next Plan/Last Result)는 수동 입력 — 보고서에서 추출하지 않음
    }


def _extract_vetting_from_upload(f):
    """SIRE 보고서 업로드 → (items, meta, err). 헤더 메타 + 지적 항목을 한 번에 추출."""
    name = (f.filename or '').lower()
    ext = name.rsplit('.', 1)[-1] if '.' in name else ''
    raw = f.read()
    size_mb = len(raw) / (1024 * 1024)
    prompt = _vetting_full_prompt()

    if ext == 'pdf':
        if size_mb > 15:
            return None, None, {'reason': 'TOO_LARGE', 'message': f'PDF가 너무 큽니다({size_mb:.1f}MB). 15MB 이하로 줄여주세요.'}
        b64 = __import__('base64').standard_b64encode(raw).decode()
        parsed = _gemini_call_json([
            {'inline_data': {'mime_type': 'application/pdf', 'data': b64}},
            {'text': prompt},
        ], model=_model_for('findings'))
    elif ext in ('png', 'jpg', 'jpeg', 'webp', 'gif', 'bmp'):
        if size_mb > 15:
            return None, None, {'reason': 'TOO_LARGE', 'message': f'이미지가 너무 큽니다({size_mb:.1f}MB).'}
        import mimetypes
        media = mimetypes.guess_type(name)[0] or 'image/jpeg'
        b64 = __import__('base64').standard_b64encode(raw).decode()
        parsed = _gemini_call_json([
            {'inline_data': {'mime_type': media, 'data': b64}},
            {'text': prompt},
        ], model=_model_for('findings'))
    elif ext in ('xlsx', 'xls'):
        try:
            txt = _xlsx_to_text(raw)
        except Exception as e:
            app.logger.exception('extract-vetting-from-upload')
            return None, None, {'reason': 'XLSX_PARSE_FAILED', 'message': f'엑셀을 읽지 못했습니다: {e}'}
        parsed = _gemini_call_json([{'text': prompt + '\n\n[보고서 표 내용]\n' + txt}],
                                   model=_model_for('findings'))
    else:
        return None, None, {'reason': 'BAD_TYPE', 'message': 'PDF · 이미지 · 엑셀(xlsx) 파일만 지원합니다.'}

    if isinstance(parsed, dict) and parsed.get('error') == 'NO_API_KEY':
        return None, None, {'reason': 'no_api_key', 'message': 'AI 자동추출이 설정되지 않았습니다(키 미설정).'}
    if isinstance(parsed, dict) and parsed.get('error'):
        return None, None, {'reason': parsed['error'], 'message': '자동 추출에 실패했습니다.',
                            'detail': parsed.get('detail') or parsed.get('raw')}
    items = _normalize_findings(parsed, 'sire')
    meta = _norm_vetting_meta(parsed.get('meta') if isinstance(parsed, dict) else None)
    return items, meta, None


_FULL_REPORT_MARKER = '[SIRE Full Report 자동반영]'
_FULL_REPORT_END_MARKER = '[/SIRE Full Report 자동반영]'


def _legacy_full_report_pattern():
    return (
        r'^[ \t]*' + _re_cls.escape(_FULL_REPORT_MARKER) + r'[ \t]*\r?\n(.*?)^[ \t]*'
        + _re_cls.escape(_FULL_REPORT_END_MARKER) + r'[ \t]*$'
    )


def _public_vetting_finding(row):
    """내부 자동 Remark 추적값은 숨기고 과거 marker도 일반 문장으로 풀어 반환한다."""
    d = dict(row)
    raw = d.get('user_remark') or ''
    blocks = _re_cls.findall(_legacy_full_report_pattern(), raw, flags=_re_cls.S | _re_cls.M)
    if blocks:
        manual = _re_cls.sub(_legacy_full_report_pattern(), '', raw,
                             flags=_re_cls.S | _re_cls.M).strip()
        automatic = _concise_full_report_remark(blocks[-1])
        d['user_remark'] = f'{manual}\n\n{automatic}'.strip() if manual else automatic
    d.pop('full_report_remark', None)
    return d


def _norm_report_number(value):
    """Report 번호 비교용 정규화. 구두점은 보존하고 공백/대소문자만 무시한다."""
    return _re_cls.sub(r'\s+', '', (value or '')).upper()


def _finding_text_similar(left, right):
    """같은 영문 지적의 경미한 표현 변형을 재업로드 중복으로 판정한다."""
    stopwords = {
        'a', 'an', 'and', 'are', 'as', 'at', 'be', 'been', 'by', 'for',
        'from', 'in', 'is', 'it', 'noted', 'of', 'on', 'or', 'that', 'the',
        'this', 'to', 'was', 'were', 'with',
    }

    def tokens(value):
        return [
            token for token in _re_cls.findall(r'[a-z0-9]+', (value or '').lower())
            if token not in stopwords
        ]
    a, b = tokens(left), tokens(right)
    if not a or not b:
        return False
    ac, bc = ''.join(a), ''.join(b)
    if ac == bc:
        return True
    aset, bset = set(a), set(b)
    overlap = len(aset & bset)
    jaccard = overlap / max(1, len(aset | bset))
    containment = overlap / max(1, min(len(aset), len(bset)))
    return (
        (overlap >= 4 and containment >= 0.80)
        or jaccard >= 0.72
        or SequenceMatcher(None, ac, bc).ratio() >= 0.88
    )


def _full_report_prompt(vetting, findings):
    existing = [{
        'finding_id': f['id'],
        'no': f['no'],
        'item': f['item'] or '',
        'description': f['description'] or '',
    } for f in findings]
    return (
        "다음 PDF는 선박 SIRE 2.0 Full Report다. 아래 기존 Observation 각각을 보고서의 동일 지적과 "
        "일대일로 매칭하고, Full report에만 새로 추가된 Observation도 빠짐없이 추출하라. JSON으로만 답하라.\n"
        "- report_type: 표지의 Report Type을 그대로 추출(반드시 Full인지 확인).\n"
        "- report_number: 표지의 Report 번호를 그대로 추출.\n"
        "- items에는 기존 Observation을 합치거나 생략하지 말고 finding_id를 각각 정확히 한 번씩 모두 반환한다. "
        "Full report에서 찾지 못한 기존 항목은 matched=false로 둔다.\n"
        "- new_items에는 Full report의 부정적 지적 중 기존 Observation 어느 것과도 동일하지 않은 신규 항목만 넣는다. "
        "포함 대상은 'Not as expected', 'Observable or detectable deficiency', 'Photo not representative'처럼 "
        "명시적으로 부정 판정된 항목뿐이다. 'As expected', 'Largely as expected', 일반 권고·정보성 코멘트는 제외한다. "
        "기존 항목의 표현 차이·Operator Comments 상세는 신규로 중복 생성하지 않는다.\n"
        "- status는 Open 또는 Closed만 사용한다. Corrective Action이 명시적으로 완료되고 핵심 결함에 "
        "남은 시정·검사·승인·Class closure가 없으면 Closed. 시정/검사/승인/자재/후속 확인이 예정·진행·보류, "
        "임시조치 또는 모니터링 단계면 Open. Preventative Action이 상시/미래형이라는 이유만으로, 이미 완료된 "
        "Corrective Action을 Open으로 두지는 않는다. 적합성 확인으로 시정 불필요가 명확하고 후속조치가 없으면 Closed.\n"
        "- 기존 items의 remark는 Operator Comments를 근거로 현재 조치상태와 남은 핵심 조치만 한국어 음슴체 "
        "한 문장으로 간결하게 요약하고 현재 상태와 남은 핵심 조치만 남긴다. Immediate Cause/Root Cause의 경위 설명은 반복하지 않는다. "
        "Condition of Class, starting valve seating, Cylinder cover, UT/MPI, FIVA, ECDIS 등 기술 명칭·장비명·약어는 "
        "번역하지 말고 보고서의 영문 표기를 그대로 유지한다. "
        "나쁜 예: '메인 엔진 6번 실린더 커버 시동 밸브 시트 수리 후 Class Condition이 발행됨. 수리 부위는 정기 점검 및 모니터링 중이며 차기 연례 검사 시 UT/MPI 재검사가 예정되어 있음.' "
        "좋은 예: 'M/E No.6 Cylinder cover starting valve seating 수리 후 Condition of Class 발행됨, 차기 Annual Survey 시 UT/MPI 재검사 예정.' "
        "없는 내용을 만들지 않는다.\n"
        "- evidence는 status 판정에 직접 사용한 영문 원문 핵심 문장이다.\n"
        "- 보고서에서 동일 지적을 확실히 찾지 못하거나 Open/Closed 판정이 불확실하면 matched=false로 둔다. "
        "추측으로 Closed를 선택하지 않는다.\n"
        "- new_items.item은 보고서 분류 라벨을 괄호로 붙인 제목, description은 지적 원문, translation은 지적의 "
        "한국어 요약, action_remark는 Operator Comments의 현재 조치상태와 남은 핵심 조치만 담은 한국어 음슴체 "
        "한 문장이다. 기술 명칭·장비명·약어는 영문 그대로 유지하고 위 좋은 예의 문체를 따른다.\n"
        '형식: {"report_type":"Full","report_number":"...","items":['
        '{"finding_id":1,"matched":true,"status":"Closed","remark":"...","evidence":"..."}],'
        '"new_items":[{"item":"(Process)Not as expected","description":"영문 지적 원문",'
        '"translation":"한글 지적 요약","status":"Open","action_remark":"한글 조치 결과",'
        '"evidence":"상태 판정 영문 근거"}]}\n\n'
        f"[대상 Vetting]\nreport_number={vetting['report_number'] or ''}\n"
        f"[기존 Observation JSON]\n{json.dumps(existing, ensure_ascii=False)}"
    )


def _replace_full_report_remark(existing, previous_auto, generated):
    """marker 없이 자동 Remark를 멱등 교체하며 수동 입력은 보존한다."""
    current = (existing or '').strip()
    legacy = _re_cls.findall(_legacy_full_report_pattern(), current,
                             flags=_re_cls.S | _re_cls.M)
    if legacy:
        current = _re_cls.sub(_legacy_full_report_pattern(), '', current,
                              flags=_re_cls.S | _re_cls.M).strip()
    else:
        previous = (previous_auto or '').strip()
        if previous and current == previous:
            current = ''
        elif previous and current.endswith('\n\n' + previous):
            current = current[:-(len(previous) + 2)].rstrip()
    automatic = _re_cls.sub(r'\s+', ' ', (generated or '')).strip()
    return f'{current}\n\n{automatic}'.strip() if current else automatic


def _concise_full_report_remark(value, limit=140):
    """AI 조치 Remark를 단일 문장·화면 한두 줄 길이로 강제한다."""
    text = (value or '')
    replacements = (
        (r'메인\s*엔진', 'M/E'),
        (r'(\d+)\s*번\s*실린더\s*커버', r'No.\1 Cylinder cover'),
        (r'실린더\s*커버', 'Cylinder cover'),
        (r'시동\s*밸브\s*(?:시트|시팅)', 'starting valve seating'),
        (r'(?:Class\s*Condition|선급\s*조건)(?:이|가)?', 'Condition of Class'),
        (r'연례\s*검사', 'Annual Survey'),
        (r'재검사가\s*예정되어\s*있음', '재검사 예정'),
    )
    for pattern, replacement in replacements:
        text = _re_cls.sub(pattern, replacement, text, flags=_re_cls.I)
    text = _re_cls.sub(r'\s+', ' ', text).strip()
    if not text:
        return ''
    parts = [part.strip() for part in _re_cls.split(r'(?<=[.!?])\s+', text) if part.strip()]
    chosen = parts[:1]
    if len(parts) > 1 and not _re_cls.search(r'Root\s*Cause|Immediate\s*Cause|원인|경위',
                                             parts[1], flags=_re_cls.I):
        if _re_cls.search(r'예정|검사|Survey|UT/MPI|모니터링|monitor|pending|완료|발행',
                          parts[1], flags=_re_cls.I):
            future = _re_cls.search(r'(차기\s+.+)', parts[1])
            chosen.append(future.group(1) if future else parts[1])
    sentence = ', '.join(part.rstrip(' .!?') for part in chosen).strip()
    if chosen and chosen[-1].endswith(('.', '!', '?')):
        sentence += chosen[-1][-1]
    if len(sentence) <= limit:
        return sentence
    cut = sentence[:limit - 1].rstrip()
    if ' ' in cut:
        cut = cut.rsplit(' ', 1)[0]
    return cut.rstrip(' ,.;:') + '…'


def _summary_full_report_remark(value):
    """종합소견용 Remark: 내부 marker는 숨기고 자동반영 조치문만 간결하게 반환한다."""
    text = (value or '').strip()
    if not text:
        return ''
    blocks = _re_cls.findall(_legacy_full_report_pattern(), text,
                             flags=_re_cls.S | _re_cls.M)
    return _concise_full_report_remark(blocks[-1] if blocks else text)


def _extract_full_report_updates(f, vetting, findings):
    """SIRE Full PDF와 기존 findings를 AI에 함께 주고 원자적 갱신 후보를 검증한다."""
    name = (f.filename or '').lower()
    if not name.endswith('.pdf'):
        return None, {'reason': 'BAD_TYPE', 'message': 'SIRE Full report PDF만 지원합니다.'}
    raw = f.read()
    size_mb = len(raw) / (1024 * 1024)
    if not raw:
        return None, {'reason': 'EMPTY_FILE', 'message': '빈 PDF 파일입니다.'}
    if not raw.lstrip().startswith(b'%PDF-'):
        return None, {'reason': 'BAD_PDF', 'message': '유효한 PDF 파일이 아닙니다.'}
    if size_mb > 15:
        return None, {'reason': 'TOO_LARGE', 'message': f'PDF가 너무 큽니다({size_mb:.1f}MB). 15MB 이하로 줄여주세요.'}
    b64 = __import__('base64').standard_b64encode(raw).decode()
    parsed = _gemini_call_json([
        {'inline_data': {'mime_type': 'application/pdf', 'data': b64}},
        {'text': _full_report_prompt(vetting, findings)},
    ], model=_model_for('findings'))
    if not isinstance(parsed, dict) or parsed.get('error'):
        return None, {'reason': (parsed or {}).get('error', 'PARSE_FAILED') if isinstance(parsed, dict) else 'PARSE_FAILED',
                      'message': 'Full report 자동 분석에 실패했습니다.'}
    if (parsed.get('report_type') or '').strip().lower() != 'full':
        return None, {'reason': 'NOT_FULL_REPORT', 'message': 'Report Type이 Full인 SIRE 보고서가 아닙니다.'}
    expected_report = _norm_report_number(vetting['report_number'])
    actual_report = _norm_report_number(parsed.get('report_number'))
    if not expected_report or actual_report != expected_report:
        return None, {'reason': 'REPORT_MISMATCH',
                      'message': f"Report 번호가 일치하지 않습니다. 기존 {vetting['report_number'] or '-'} / 업로드 {parsed.get('report_number') or '-'}"}

    expected_ids = {int(f['id']) for f in findings}
    updates, unmatched_ids, seen = [], [], set()
    invalid_match_set = False
    for item in parsed.get('items') or []:
        if not isinstance(item, dict):
            continue
        try:
            fid = int(item.get('finding_id'))
        except (TypeError, ValueError):
            continue
        if fid not in expected_ids or fid in seen:
            invalid_match_set = True
            continue
        seen.add(fid)
        if item.get('matched') is not True:
            unmatched_ids.append(fid)
            continue
        status = (item.get('status') or '').strip()
        remark = _concise_full_report_remark(item.get('remark'))
        evidence = (item.get('evidence') or '').strip()
        if status not in ('Open', 'Closed') or not remark or not evidence:
            invalid_match_set = True
            continue
        updates.append({'finding_id': fid, 'status': status, 'remark': remark,
                        'evidence': evidence})
    if invalid_match_set:
        return None, {'reason': 'INVALID_MATCH_SET',
                      'message': '기존 Observation ID 분석 결과가 중복·오류여서 반영하지 않았습니다.'}
    if seen != expected_ids:
        missing = sorted(expected_ids - seen)
        return None, {'reason': 'INCOMPLETE_MATCH',
                      'message': f'기존 Observation 전건의 분석 결과가 없어 반영하지 않았습니다. 누락 {len(missing)}건.'}

    existing_desc = [(f['description'] or '') for f in findings]
    new_items = []
    for item in parsed.get('new_items') or []:
        if not isinstance(item, dict):
            continue
        rec = {
            'item': (item.get('item') or '').strip(),
            'description': (item.get('description') or '').strip(),
            'translation': (item.get('translation') or '').strip(),
            'status': (item.get('status') or '').strip(),
            'action_remark': _concise_full_report_remark(item.get('action_remark')),
            'evidence': (item.get('evidence') or '').strip(),
        }
        label = rec['item'].lower()
        explicitly_negative = any(token in label for token in (
            'not as expected', 'observable or detectable deficiency',
            'photo not representative',
        ))
        if (not explicitly_negative or not rec['item'] or not rec['description'] or not rec['translation']
                or rec['status'] not in ('Open', 'Closed') or not rec['action_remark']
                or not rec['evidence']
                or any(_finding_text_similar(rec['description'], old) for old in existing_desc)):
            continue
        if any(_finding_text_similar(rec['description'], n['description']) for n in new_items):
            continue
        new_items.append(rec)
    return {'updates': updates, 'unmatched_ids': unmatched_ids, 'new_items': new_items}, None


@bp.route('/api/vettings/<int:vid>/extract-report', methods=['POST'])
@login_required
def api_vt_extract_report(vid):
    if not query('SELECT id FROM vettings WHERE id=?', (vid,), one=True):
        abort(404)
    if 'file' not in request.files or not request.files['file'].filename:
        return jsonify({'ok': False, 'message': '파일이 없습니다.'}), 400
    items, meta, err = _extract_vetting_from_upload(request.files['file'])
    if err:
        return jsonify({'ok': False, **err}), 200
    # 헤더 메타 자동 반영: 추출값이 있는 필드만 갱신 (없으면 기존값 유지)
    applied = {}
    sets, params = [], []
    for col in ('report_number', 'inspection_date', 'inspection_company',
                'inspector', 'port', 'sire_type', 'valid'):
        val = (meta or {}).get(col, '')
        if val:
            sets.append(f'{col}=?'); params.append(val); applied[col] = val
    if sets:
        sets.append("updated_at=datetime('now','localtime')")
        params.append(vid)
        execute(f'UPDATE vettings SET {", ".join(sets)} WHERE id=?', tuple(params))
    return jsonify({'ok': True, 'items': items, 'count': len(items),
                    'meta': meta, 'applied': applied})


@bp.route('/api/vettings/<int:vid>/apply-full-report', methods=['POST'])
@login_required
def api_vt_apply_full_report(vid):
    """SIRE Full report로 기존 Observation 갱신 + Full 신규 지적을 원자적 추가한다."""
    vetting = query('SELECT id, report_number FROM vettings WHERE id=?', (vid,), one=True)
    if not vetting:
        abort(404)
    findings = query(
        'SELECT id, no, item, description, user_remark, full_report_remark, status '
        'FROM vt_findings WHERE vetting_id=? ORDER BY no, id', (vid,))
    if not findings:
        message = '먼저 Initial report로 Observation을 생성하세요.'
        return jsonify({'ok': False, 'reason': 'NO_FINDINGS',
                        'message': message, 'error': message}), 422
    if 'file' not in request.files or not request.files['file'].filename:
        message = '파일이 없습니다.'
        return jsonify({'ok': False, 'reason': 'NO_FILE', 'message': message,
                        'error': message}), 400

    uploaded = request.files['file']
    raw_for_hash = uploaded.read()
    uploaded.stream.seek(0)
    file_sha256 = hashlib.sha256(raw_for_hash).hexdigest()
    result, err = _extract_full_report_updates(uploaded, vetting, findings)
    if err:
        return jsonify({'ok': False, **err, 'error': err['message']}), 422
    updates = result['updates']
    new_items = result['new_items']
    unmatched_ids = result['unmatched_ids']
    if not updates and not new_items:
        message = 'Full report에서 반영 가능한 기존 또는 신규 Observation을 찾지 못했습니다.'
        return jsonify({'ok': False, 'reason': 'NO_APPLICABLE_ITEMS',
                        'message': message, 'error': message}), 422

    by_id = {int(f['id']): f for f in findings}
    before = [{'finding_id': int(f['id']), 'status': f['status'],
               'user_remark': f['user_remark'] or ''} for f in findings]
    db = get_db()
    try:
        db.execute('BEGIN IMMEDIATE')
        after = []
        for item in updates:
            fid = item['finding_id']
            old = by_id[fid]
            automatic = _concise_full_report_remark(item['remark'])
            user_remark = _replace_full_report_remark(
                old['user_remark'], old['full_report_remark'], automatic)
            cur = db.execute(
                "UPDATE vt_findings SET status=?, user_remark=?, full_report_remark=?, "
                "updated_at=datetime('now','localtime') "
                'WHERE id=? AND vetting_id=?',
                (item['status'], user_remark, automatic, fid, vid),
            )
            if cur.rowcount != 1:
                raise RuntimeError(f'finding changed during full report apply: {fid}')
            after.append({'finding_id': fid, 'status': item['status'],
                          'user_remark': user_remark, 'full_report_remark': automatic,
                          'evidence': item['evidence']})
        current_desc = [row[0] or '' for row in db.execute(
            'SELECT description FROM vt_findings WHERE vetting_id=?', (vid,))]
        next_no = db.execute(
            'SELECT COALESCE(MAX(no),0)+1 FROM vt_findings WHERE vetting_id=?', (vid,)
        ).fetchone()[0]
        created = []
        for item in new_items:
            if any(_finding_text_similar(item['description'], old) for old in current_desc):
                continue
            automatic = _concise_full_report_remark(item['action_remark'])
            user_remark = _replace_full_report_remark('', '', automatic)
            cur = db.execute(
                "INSERT INTO vt_findings "
                "(vetting_id,no,item,description,remark,user_remark,full_report_remark,priority,status) "
                "VALUES(?,?,?,?,?,?,?,0,?)",
                (vid, next_no, item['item'], item['description'], item['translation'],
                 user_remark, automatic, item['status']),
            )
            fid = cur.lastrowid
            created.append({'finding_id': fid, 'no': next_no, **item})
            after.append({'finding_id': fid, 'created': True, 'status': item['status'],
                          'user_remark': user_remark, 'full_report_remark': automatic,
                          'evidence': item['evidence']})
            current_desc.append(item['description'])
            next_no += 1
        db.execute(
            "UPDATE vettings SET updated_at=datetime('now','localtime') WHERE id=?", (vid,))
        db.execute(
            "INSERT INTO vt_full_report_audit "
            "(vetting_id,report_number,file_sha256,filename,before_json,after_json,applied_by) "
            "VALUES(?,?,?,?,?,?,?)",
            (vid, vetting['report_number'], file_sha256, uploaded.filename,
             json.dumps(before, ensure_ascii=False), json.dumps(after, ensure_ascii=False),
             session.get('username') or session.get('display_name')),
        )
        db.commit()
    except Exception:
        db.rollback()
        app.logger.exception('apply-full-report')
        message = '분석 결과 저장에 실패해 아무 항목도 변경하지 않았습니다.'
        return jsonify({'ok': False, 'reason': 'DB_UPDATE_FAILED',
                        'message': message, 'error': message}), 500

    changed = updates + created
    opened = sum(1 for item in changed if item['status'] == 'Open')
    closed = len(changed) - opened
    return jsonify({'ok': True, 'updated': len(updates), 'created': len(created),
                    'unmatched': len(unmatched_ids), 'open': opened,
                    'closed': closed, 'items': updates, 'new_items': created})


def _md_from_date(d):
    """'2026-04-30' → '4/30'. 파싱 실패 시 원문."""
    try:
        y, m, dd = (d or '').split('-')
        return f'{int(m)}/{int(dd)}'
    except Exception:
        app.logger.exception('md-from-date')
        return (d or '').strip()


def _company_abbr(c):
    """'VIVA ENERGY' → 'VIVA' (첫 토큰 대문자). 빈 값이면 ''."""
    c = (c or '').strip()
    if not c:
        return ''
    return c.split()[0].upper()


def _sire_abbr(s):
    return {'Bunkering': 'BUNKER', 'Discharge': 'DISCHARGE', 'Idle': 'IDLE'}.get(
        (s or '').strip(), (s or '').strip().upper())


def _condense_obs(items):
    """[{i,summary,description,user_remark}] → {i: short}. 선박 약어체 한 줄.
    GEMINI 키 없거나 실패 시 빈 dict (상위에서 번역요약으로 폴백)."""
    out = {}
    if not GEMINI_API_KEY or not items:
        return out
    payload = json.dumps([{'i': it['i'], 'summary': it.get('summary', ''),
                           'description': it.get('description', '')} for it in items],
                         ensure_ascii=False)
    prompt = (
        "아래는 선박 SIRE 점검 지적 항목들이다(JSON 배열). 각 항목의 핵심 결함을 "
        "선박 현업 약어체로 아주 짧게 한 줄로 요약하라.\n"
        "- 장비명은 선박 약어로 대문자 표기: Cargo Oil Tank→COT, Ballast Water Treatment System→BWTS, "
        "Main Engine→M/E, Auxiliary Engine→A/E, pressure→PRESS., No.3 Port→3P, Vapour return manifold→VAP. RETURN MANIFOLD 등.\n"
        "- 결함은 '불량/파손/누설/마모/고장' 등 한 단어로 압축. 군더더기·서술 제거.\n"
        "- 예: 'Cargo tank high level alarm display 결함으로 상시 점등됨' → 'COT HIGH LEVEL ALARM DISPLAY 불량', "
        "'3 Port cargo tank 압력 센서 결함' → '3P COT PRESS. SENSOR 불량'.\n"
        + _MARITIME_TERMS +
        "입력의 i를 그대로 사용해 JSON으로만 답하라.\n"
        '형식: {"items":[{"i":0,"short":"..."}]}\n\n[입력]\n' + payload)
    res = _gemini_call_json([{'text': prompt}], model=_model_for('summary'))
    arr = _coerce_translation_items(res)
    for o in (arr or []):
        if not isinstance(o, dict):
            continue
        try:
            i = int(o.get('i'))
        except (TypeError, ValueError):
            continue
        sh = (o.get('short') or o.get('en') or '').strip()
        if sh:
            out[i] = sh
    return out


@bp.route('/api/vettings/<int:vid>/obs-summary', methods=['POST'])
@login_required
def api_vt_obs_summary(vid):
    """Priority 체크 + Open 항목 기준으로 '지적 상세' 요약을 생성해 overall_remark에 기록."""
    v = query('SELECT * FROM vettings WHERE id=?', (vid,), one=True)
    if not v:
        abort(404)
    findings = query('SELECT * FROM vt_findings WHERE vetting_id=? ORDER BY no, id', (vid,))
    open_f = [f for f in findings if (f['status'] or 'Open') == 'Open']
    def _is_prio(f):
        try:
            return bool(f['priority'])
        except (KeyError, IndexError):
            return False
    prio = [f for f in open_f if _is_prio(f)]
    total_open = len(open_f)
    minor = total_open - len(prio)

    header_bits = [b for b in (_md_from_date(v['inspection_date']),
                               _company_abbr(v['inspection_company']),
                               _sire_abbr(v['sire_type'])) if b]
    header = (' '.join(header_bits) + ' ' if header_bits else '') + \
             f'SIRE OBS 잔여 {total_open}건 조치 중'

    shorts = _condense_obs([
        {'i': i, 'summary': f['remark'] or '', 'description': f['description'] or '',
         'user_remark': f['user_remark'] or ''}
        for i, f in enumerate(prio)
    ])

    lines = [header]
    for i, f in enumerate(prio):
        short = shorts.get(i) or (f['remark'] or f['item'] or '').strip()
        action = _concise_full_report_remark(
            f['full_report_remark'] or _summary_full_report_remark(f['user_remark']))
        lines.append(f'{i + 1}. {short}' + (f' - {action}' if action else ''))
    if minor > 0:
        lines.append(f'그 외 Minor 지적 {minor}건')
    text = '\n'.join(lines)

    execute("UPDATE vettings SET overall_remark=?, updated_at=datetime('now','localtime') WHERE id=?",
            (text, vid))
    return jsonify({'ok': True, 'summary': text,
                    'total_open': total_open, 'priority_open': len(prio), 'minor': minor})


@bp.route('/api/vettings/<int:vid>/export')
@login_required
def api_vt_export(vid):
    from flask import send_file
    v = query('''SELECT vt.*, ve.name AS vessel_name
                   FROM vettings vt JOIN vessels ve ON ve.id = vt.vessel_id
                  WHERE vt.id=?''', (vid,), one=True)
    if not v:
        abort(404)
    fr = query('''SELECT no, item, description, remark, user_remark, status
                    FROM vt_findings WHERE vetting_id=? ORDER BY no, id''', (vid,))
    # RECTIFICATION·PHOTO 2열은 공란으로 출력(현장기입용). 번역요약·Remark는 export에서 제외.
    rows = [[r['no'], r['item'] or '', r['description'] or '',
             '', '', r['status'] or ''] for r in fr]
    vessel = v['vessel_name']
    rno = v['report_number'] or ''
    title = f"SIRE Observation List — {vessel}"
    sub_bits = [f"검사일: {v['inspection_date'] or '-'}", f"Port: {v['port'] or '-'}"]
    if rno:
        sub_bits.append(f"Report: {rno}")
    sub_bits.append(f"총 {len(rows)}건")
    headers = ['No.', 'ITEM', 'DESCRIPTION', 'RECTIFICATION', 'PHOTO', 'STATUS']
    bio = _findings_workbook(title, '   │   '.join(sub_bits), headers, rows,
                             wrap_cols={2, 3, 4, 5}, widths=[6, 26, 46, 40, 30, 10])
    date_tag = (v['inspection_date'] or '').replace('-', '')
    fname = f"SIRE_{_safe_filename(vessel)}_{date_tag or vid}.xlsx"
    return send_file(bio, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/api/vettings/<int:vid>/attachments', methods=['GET'])
@login_required
def api_vt_attachments_list(vid):
    rows = query(
        'SELECT * FROM vt_attachments WHERE vetting_id=? AND inactive_at IS NULL ORDER BY id DESC',
        (vid,),
    )
    return jsonify([dict(r) for r in rows])


@bp.route('/api/vettings/<int:vid>/attachments', methods=['POST'])
@login_required
def api_vt_attachment_upload(vid):
    if not query('SELECT id FROM vettings WHERE id=?', (vid,), one=True):
        abort(404)
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다.'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'error': '파일명이 없습니다.'}), 400
    if not _ext_allowed(f.filename):
        return jsonify({'error': '허용되지 않는 파일 형식입니다.'}), 400

    ext = os.path.splitext(f.filename)[1]
    stored = f"vt_{uuid.uuid4().hex}{ext}"
    save_path = os.path.join(UPLOAD_DIR, stored)
    f.save(save_path)
    size = os.path.getsize(save_path)

    aid = execute("""
        INSERT INTO vt_attachments
            (vetting_id, filename, stored_name, file_size, mime_type, uploaded_by)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (vid, f.filename, stored, size, f.mimetype, session.get('username')))
    return jsonify({'id': aid, 'filename': f.filename, 'file_size': size}), 201


_SVMS_MAX_BYTES = 20 * 1024 * 1024
_SVMS_EXTS = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'jpg', 'jpeg', 'png', 'gif', 'heic', 'heif', 'webp', 'bmp'}

def _svms_norm(value):
    """Comparison normalization only: casefold and retain unicode letters/digits."""
    return ''.join(ch for ch in (value or '').casefold() if ch.isalnum())

def _svms_magic_ok(ext, head):
    if ext == 'pdf': return head.startswith(b'%PDF-')
    if ext in {'jpg', 'jpeg'}: return head.startswith(b'\xff\xd8\xff')
    if ext == 'png': return head.startswith(b'\x89PNG\r\n\x1a\n')
    if ext == 'gif': return head.startswith((b'GIF87a', b'GIF89a'))
    if ext in {'doc', 'xls'}: return head.startswith(b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1')
    if ext in {'docx', 'xlsx'}: return head.startswith(b'PK\x03\x04')
    if ext in {'webp'}: return head.startswith(b'RIFF') and head[8:12] == b'WEBP'
    if ext in {'heic', 'heif'}: return len(head) >= 12 and head[4:8] == b'ftyp'
    if ext == 'bmp': return head.startswith(b'BM')
    return False

def _svms_activate_revision(row_id, external_file_id, vetting_id):
    """Keep revision history while exposing exactly one active source file."""
    execute("UPDATE vt_attachments SET inactive_at=datetime('now','localtime') WHERE source='svms' AND external_file_id=? AND id<>? AND inactive_at IS NULL",
            (external_file_id, row_id))
    execute("UPDATE vt_attachments SET vetting_id=?, inactive_at=NULL, synced_at=datetime('now','localtime') WHERE id=?",
            (vetting_id, row_id))


def _svms_exact_vetting(vessel_name, report_number):
    vessel = _svms_norm(vessel_name); report = _svms_norm(report_number)
    rows = query('''SELECT vt.id, v.name AS vessel_name, vt.report_number
                    FROM vettings vt JOIN vessels v ON v.id=vt.vessel_id
                    WHERE v.name IS NOT NULL AND vt.report_number IS NOT NULL''')
    return [r for r in rows if _svms_norm(r['vessel_name']) == vessel and
                                _svms_norm(r['report_number']) == report]


@bp.route('/api/ext/vettings/svms-status', methods=['POST'])
@api_key_required
def api_ext_vetting_svms_status():
    """Store SVMS header flags for one exact TRMT report; findings stay untouched."""
    d = request.get_json(silent=True) or {}
    uploaded = (d.get('report_uploaded_yn') or 'Y').strip().upper()
    full = (d.get('full_report_yn') or '').strip().upper()
    close = (d.get('close_report_yn') or '').strip().upper()
    if uploaded not in ('Y', 'N'):
        return jsonify({'error': 'report_uploaded_yn must be Y or N'}), 400
    if uploaded == 'Y' and (full not in ('Y', 'N') or close not in ('Y', 'N')):
        return jsonify({'error': 'full_report_yn and close_report_yn must be Y or N'}), 400
    matches = _svms_exact_vetting(d.get('vessel_name'), d.get('report_number'))
    if not matches:
        return jsonify({'error': 'exact vessel/report match not found'}), 409
    if len(matches) != 1:
        return jsonify({'error': 'ambiguous exact vessel/report match'}), 409
    vid = matches[0]['id']
    if uploaded == 'N':
        full = close = None
    execute('''UPDATE vettings SET svms_full_report_yn=?, svms_close_report_yn=?,
               svms_report_uploaded_yn=?, svms_status_synced_at=datetime('now','localtime') WHERE id=?''',
            (full, close, uploaded, vid))
    return jsonify({'ok': True, 'id': vid, 'report_uploaded_yn': uploaded,
                    'full_report_yn': full, 'close_report_yn': close})


@bp.route('/api/ext/vettings/svms-attachment', methods=['POST'])
@api_key_required
def api_ext_vetting_svms_attachment():
    """Idempotent SVMS SIRE attachment ingress; never mutates findings."""
    required = ('vessel_name', 'report_number', 'SIRE_CD', 'FILE_TP', 'external_file_id', 'sha256')
    vals = {k: (request.form.get(k) or '').strip() for k in required}
    if any(not vals[k] for k in required) or 'file' not in request.files:
        return jsonify({'error': 'missing required metadata or file'}), 400
    if vals['FILE_TP'] not in ('SMSR', 'SMSC'):
        return jsonify({'error': 'FILE_TP must be SMSR or SMSC'}), 400
    claimed = vals['sha256'].lower()
    if len(claimed) != 64 or any(c not in '0123456789abcdef' for c in claimed):
        return jsonify({'error': 'invalid sha256'}), 400
    matches = _svms_exact_vetting(vals['vessel_name'], vals['report_number'])
    if not matches:
        return jsonify({'error': 'exact vessel/report match not found'}), 409
    if len(matches) != 1:
        return jsonify({'error': 'ambiguous exact vessel/report match'}), 409
    match = matches[0]
    vid = match['id']; f = request.files['file']; filename = os.path.basename(f.filename or '')
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    if ext not in _SVMS_EXTS: return jsonify({'error': 'unsupported extension'}), 415
    raw = f.read(_SVMS_MAX_BYTES + 1)
    if not raw: return jsonify({'error': 'empty file'}), 400
    if len(raw) > _SVMS_MAX_BYTES: return jsonify({'error': 'file too large'}), 413
    digest = hashlib.sha256(raw).hexdigest()
    if digest != claimed: return jsonify({'error': 'sha256 mismatch'}), 400
    if not _svms_magic_ok(ext, raw[:32]): return jsonify({'error': 'file magic mismatch'}), 415
    old = query('SELECT * FROM vt_attachments WHERE source=\'svms\' AND external_file_id=? AND sha256=?', (vals['external_file_id'], claimed), one=True)
    if old:
        _svms_activate_revision(old['id'], vals['external_file_id'], vid)
        return jsonify({'ok': True, 'id': old['id'], 'deduplicated': True}), 200
    stored = f"vt_svms_{uuid.uuid4().hex}.{ext}"
    stored_path = os.path.join(UPLOAD_DIR, stored)
    committed = False
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    fd, tmp = tempfile.mkstemp(prefix='.svms.', dir=UPLOAD_DIR)
    try:
        with os.fdopen(fd, 'wb') as out:
            out.write(raw); out.flush(); os.fsync(out.fileno())
        os.replace(tmp, stored_path)
        try:
            aid = execute('''INSERT INTO vt_attachments
                (vetting_id,filename,stored_name,file_size,mime_type,uploaded_by,source,source_type,
                 external_sire_cd,external_file_id,sha256,synced_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,datetime('now','localtime'))''',
                (vid, filename, stored, len(raw), f.mimetype, 'svms-sync',
                 'svms', 'close' if vals['FILE_TP'] == 'SMSC' else 'initial', vals['SIRE_CD'],
                 vals['external_file_id'], claimed))
        except Exception:
            race = query('SELECT id FROM vt_attachments WHERE source=\'svms\' AND external_file_id=? AND sha256=?', (vals['external_file_id'], claimed), one=True)
            if race:
                _svms_activate_revision(race['id'], vals['external_file_id'], vid)
                return jsonify({'ok': True, 'id': race['id'], 'deduplicated': True}), 200
            raise
        _svms_activate_revision(aid, vals['external_file_id'], vid)
        committed = True
    finally:
        try: os.unlink(tmp)
        except OSError: pass
        if not committed:
            try: os.unlink(stored_path)
            except OSError: pass
    return jsonify({'ok': True, 'id': aid, 'deduplicated': False}), 201


@bp.route('/api/vt-attachments/<int:aid>', methods=['GET'])
@login_required
def api_vt_attachment_get(aid):
    a = query('SELECT * FROM vt_attachments WHERE id=?', (aid,), one=True)
    if not a:
        abort(404)
    inline = request.args.get('inline')
    return send_from_directory(
        UPLOAD_DIR, a['stored_name'],
        as_attachment=not inline,
        download_name=a['filename'],
    )


@bp.route('/api/vt-attachments/<int:aid>/docx-preview', methods=['GET'])
@login_required
def api_vt_attachment_docx_preview(aid):
    a = query('SELECT * FROM vt_attachments WHERE id=? AND inactive_at IS NULL', (aid,), one=True)
    if not a:
        abort(404)
    if not (a['filename'] or '').lower().endswith('.docx'):
        return jsonify({'error': 'DOCX 파일만 미리보기 가능합니다.'}), 415
    if request.args.get('download'):
        return send_from_directory(UPLOAD_DIR, a['stored_name'], as_attachment=True,
                                   download_name=a['filename'])
    path = os.path.join(UPLOAD_DIR, a['stored_name'])
    try:
        from docx_preview import render_docx_html
        return Response(render_docx_html(path, a['filename']), mimetype='text/html')
    except Exception as exc:
        app.logger.exception('vt-docx-preview aid=%s', aid)
        return jsonify({'error': 'DOCX 미리보기를 생성하지 못했습니다.', 'detail': str(exc)[:180]}), 422


@bp.route('/api/vt-attachments/<int:aid>', methods=['DELETE'])
@login_required
def api_vt_attachment_delete(aid):
    a = query('SELECT * FROM vt_attachments WHERE id=?', (aid,), one=True)
    if not a:
        abort(404)
    if (a['source'] or 'manual') == 'svms':
        return jsonify({'error': 'SVMS synced attachments cannot be deleted'}), 403
    p = os.path.join(UPLOAD_DIR, a['stored_name'])
    if os.path.exists(p):
        try: os.remove(p)
        except OSError:
            app.logger.exception('vt-attachment-delete')
    execute('DELETE FROM vt_attachments WHERE id=?', (aid,))
    return jsonify({'ok': True})
