"""routes_core — converted to a real imported module with Blueprint("routes_core") on 2026-08-11.

Previously executed in the app namespace by ``_load_extracted_module``.
Dependencies are now the explicit imports below and nothing else, and since
2026-08-12 each import names the module that actually *defines* the symbol
(``app_core`` for config/db primitives, ``helpers_shared`` for shared helpers,
``app`` only for what app.py itself defines) instead of routing everything
through ``app``.  Contract enforced by
``test_converted_modules_are_self_contained``: zero unresolved names, and
no sibling boundary imports.
"""
from flask import Blueprint

import json
import os
import re
from datetime import date, timedelta
from flask import abort, g, jsonify, make_response, redirect, render_template, request, session, url_for
from werkzeug.security import check_password_hash, generate_password_hash
from datetime import datetime
from app_core import (
    INSTANCE_DIR, SOA_REVIEW_PDF_DIR, UPLOAD_DIR, app, execute, execute_rc, get_db, query,
)
from token_auth import (
    _DUMMY_PW_HASH, _TOKEN_MAXAGE, _issue_token, _token_note_fail, _token_rate_limited,
    _token_reset_fails,
)
from helpers_shared import (
    VETTING_TYPES, _automation_health_summary, _dashboard_ctx, _ensure_summary_table,
    _fleet_visible_auto_vessels, _issue_to_dict, _issue_write_scope, _run_summary_generate,
    _safe_filename, _soa_review_attachment_path, _translate_rows_en, _vetting_pick, _vkey,
    admin_required, login_required,
)

bp = Blueprint("routes_core", __name__)





# ═════════════════════════════════════════════════════════════════
#  Pages
# ═════════════════════════════════════════════════════════════════
@bp.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        if 'user_id' in session:
            return redirect(url_for('routes_core.dashboard'))
        return render_template('login.html')

    username = (request.form.get('username') or '').strip()
    password = request.form.get('password') or ''
    u = query('SELECT * FROM users WHERE username=? AND active=1',
              (username,), one=True)
    if not u or not check_password_hash(u['password_hash'], password):
        return render_template(
            'login.html',
            error='아이디 또는 비밀번호가 올바르지 않습니다.',
            username=username,
        ), 401

    session.clear()
    session.permanent = True
    session['user_id']       = u['id']
    session['username']      = u['username']
    session['display_name']  = u['display_name'] or u['username']
    session['role']          = u['role']
    session['supervisor_id'] = u['supervisor_id']
    execute('UPDATE users SET last_login_at=datetime("now","localtime") WHERE id=?',
            (u['id'],))

    nxt = request.args.get('next') or url_for('routes_core.dashboard')
    # 외부 URL 리다이렉트 방지 ('//evil.com' 같은 프로토콜-상대 URL 포함)
    # 🔴 백슬래시도 막는다 — 브라우저가 `/\evil.com` 을 `//evil.com` 으로 정규화해서
    #    startswith('//') 검사만으로는 그대로 외부로 나간다(open redirect).
    if (not nxt.startswith('/') or nxt.startswith('//')
            or nxt.startswith('/\\') or '\\' in nxt or '\r' in nxt or '\n' in nxt):
        nxt = url_for('routes_core.dashboard')
    return redirect(nxt)


@bp.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('routes_core.login'))


@bp.route('/api/auth/token', methods=['POST'])
def api_auth_token():
    """네이티브 앱 로그인: username/password → Bearer 토큰."""
    d = request.get_json(silent=True)
    if not isinstance(d, dict):
        return jsonify({'error': 'bad_request'}), 400
    un = d.get('username'); pw = d.get('password')
    username = (un if isinstance(un, str) else '').strip()
    password = pw if isinstance(pw, str) else ''
    # 조회 먼저(username indexed, 값쌈). 존재 유저만 canonical user_id 버킷으로 rate-limit.
    u = query('SELECT * FROM users WHERE username=? AND active=1', (username,), one=True)
    if not u:
        # 비존재/비활성: 버킷 생성·카운트 안 함(임의 username 키증식 차단) → 존재 유저와
        # 동일하게 401. 응답(401)도, 더미 해시로 처리시간도 균일 → status/timing enumeration 방지.
        check_password_hash(_DUMMY_PW_HASH, password)
        return jsonify({'error': 'invalid_credentials'}), 401
    if _token_rate_limited(u['id']):
        return jsonify({'error': 'rate_limited'}), 429
    if not check_password_hash(u['password_hash'], password):
        _token_note_fail(u['id'])
        return jsonify({'error': 'invalid_credentials'}), 401
    _token_reset_fails(u['id'])       # 성공 시 해당 계정 버킷만 초기화
    execute('UPDATE users SET last_login_at=datetime("now","localtime") WHERE id=?', (u['id'],))
    resp = jsonify({
        'token':         _issue_token(u),
        'expires_in':    _TOKEN_MAXAGE,
        'user_id':       u['id'],
        'username':      u['username'],
        'display_name':  u['display_name'] or u['username'],
        'role':          u['role'],
        'supervisor_id': u['supervisor_id'],
    })
    resp.headers['Cache-Control'] = 'no-store'
    return resp


@bp.route('/dashboard')
@login_required
def dashboard():
    """Fleet Map — 지도 기반 대시보드(SVMS noon 선위 + TRMT 현황 조인).
    데이터는 /api/fleet-map/data (감독 스코프). 상단 KPI 스트립은 구 대시보드 집계
    (_dashboard_ctx)를 재사용. 카드형 전체는 /dashboard/classic 백업 경로."""
    embedded = request.args.get('embed') == '1'
    response = make_response(render_template('dashboard.html', embedded=embedded, **_dashboard_ctx()))
    if embedded:
        # `/mobile` is same-origin; reject all external framing of the map surface.
        response.headers['X-Frame-Options'] = 'SAMEORIGIN'
        response.headers['Content-Security-Policy'] = "frame-ancestors 'self'"
    return response


@bp.route('/mobile')
@login_required
def mobile_app():
    """Mobile-first TRMT shell backed by the existing server and session."""
    return render_template('mobile.html', **_dashboard_ctx())




@bp.route('/api/dashboard/cockpit')
@login_required
def api_dashboard_cockpit():
    """대시보드 '오늘의 조종석' 스트립 데이터.
    · due: 45일 내 마감 임박(class_status_items due_date + calendar_events) 병합·정렬 상위6.
    · approvals: 사람 판단대기 큐 카운트(FundReq/AOR/Invoice pending) — admin 전용 큐.
    · automation: automation_health 최신-러너 요약 + worst4.
    담당선박 스코프는 due 에만 적용(_dashboard_ctx 와 동일 vin 패턴). 큐는 전사(admin)."""
    today = date.today()
    today_s = today.isoformat()
    horizon = (today + timedelta(days=45)).isoformat()
    is_admin = (session.get('role') == 'admin')

    sup_id = session.get('supervisor_id')
    scoped = bool(sup_id)
    vessel_ids = []
    if scoped:
        vessel_ids = [r['vessel_id'] for r in
                      query("SELECT vessel_id FROM supervisor_vessels WHERE supervisor_id=?", (sup_id,))]

    def vin(col):
        if not scoped:
            return ("1=1", [])
        if not vessel_ids:
            return ("0=1", [])
        return (f"{col} IN ({','.join('?' * len(vessel_ids))})", list(vessel_ids))

    def _days_left(iso_d):
        try:
            return (date.fromisoformat(iso_d[:10]) - today).days
        except (ValueError, TypeError):
            return None

    # ── due: (1) class_status_items 마감일 ──
    due = []
    cvf, cvp = vin("cs.vessel_id")
    ci = query(
        "SELECT i.due_date, i.description, v.name AS vessel "
        "FROM class_status_items i JOIN class_status cs ON cs.id=i.cs_id "
        "LEFT JOIN vessels v ON v.id=cs.vessel_id "
        "WHERE i.due_date IS NOT NULL AND i.due_date != '' "
        f"AND i.due_date >= ? AND i.due_date <= ? AND {cvf}",
        (today_s, horizon, *cvp))
    for r in ci:
        dl = _days_left(r['due_date'])
        if dl is None:
            continue
        title = (r['description'] or '선급/기국 지적').strip()
        if len(title) > 60:
            title = title[:59] + '…'
        due.append({'days_left': dl, 'vessel': r['vessel'] or '', 'title': title, 'source': 'class'})

    # ── due: (2) calendar_events 45일 내(담당선박/본인/공용) ──
    if scoped:
        evf, evp = vin("vessel_id")
        ce = query(
            "SELECT ce.start_date, ce.title, ce.completed, v.name AS vessel FROM calendar_events ce "
            "LEFT JOIN vessels v ON v.id=ce.vessel_id "
            "WHERE ce.start_date >= ? AND ce.start_date <= ? "
            f"AND (ce.supervisor_id=? OR ce.supervisor_id IS NULL OR {evf})",
            (today_s, horizon, sup_id, *evp))
    else:
        ce = query(
            "SELECT ce.start_date, ce.title, ce.completed, v.name AS vessel FROM calendar_events ce "
            "LEFT JOIN vessels v ON v.id=ce.vessel_id "
            "WHERE ce.start_date >= ? AND ce.start_date <= ?",
            (today_s, horizon))
    for r in ce:
        dl = _days_left(r['start_date'])
        if dl is None:
            continue
        due.append({'days_left': dl, 'vessel': r['vessel'] or '',
                    'title': (r['title'] or '일정').strip(), 'source': 'calendar',
                    'completed': bool(r['completed'])})

    due.sort(key=lambda x: x['days_left'])
    due = due[:6]

    # ── approvals: 사람 판단대기 큐(전사, admin 큐) ──
    approvals = {'fundreq': 0, 'aor': 0, 'invoice': 0, 'oldest': None}
    if is_admin:
        approvals['fundreq'] = (query("SELECT COUNT(*) c FROM fundreq_draft WHERE status='pending'",
                                      one=True) or {'c': 0})['c']
        approvals['aor'] = (query("SELECT COUNT(*) c FROM aor_draft WHERE status='pending'",
                                  one=True) or {'c': 0})['c']
        approvals['invoice'] = (query("SELECT COUNT(*) c FROM invoice_draft WHERE status='pending'",
                                      one=True) or {'c': 0})['c']
        # oldest pending — 3개 큐 중 가장 오래된 created_at
        oldest = None
        for lbl, sql in (
            ('비용청구', "SELECT MIN(created_at) m FROM fundreq_draft WHERE status='pending'"),
            ('AOR',      "SELECT MIN(created_at) m FROM aor_draft WHERE status='pending'"),
            ('인보이스', "SELECT MIN(created_at) m FROM invoice_draft WHERE status='pending'"),
        ):
            row = query(sql, one=True)
            m = row['m'] if row else None
            if not m:
                continue
            dl = _days_left(m)
            age = (0 - dl) if dl is not None else 0
            if oldest is None or age > oldest['days']:
                oldest = {'label': lbl, 'days': age}
        approvals['oldest'] = oldest

    # ── automation: 최신-러너 요약 + worst4 ──
    runners, counts = _automation_health_summary()

    def _ago(iso_d):
        if not iso_d:
            return None
        try:
            delta = datetime.now() - datetime.fromisoformat(iso_d)
        except (ValueError, TypeError):
            return None
        h = delta.total_seconds() / 3600.0
        if h < 1:
            return '방금'
        if h < 48:
            return f'{int(round(h))}시간 전'
        return f'{int(round(h / 24))}일 전'

    worst = [{'label': r['label'], 'status': r['status'],
              'ago': _ago(r['ran_at'] or r['reported_at'])}
             for r in runners if r['status'] in ('fail', 'warn')][:4]
    automation = {'ok': counts['ok'], 'warn': counts['warn'], 'fail': counts['fail'],
                  'total': counts['total'], 'worst': worst}

    return jsonify({'due': due, 'approvals': approvals, 'automation': automation,
                    'is_admin': is_admin})


@bp.route('/')
@login_required
def index():
    return render_template('index.html')


@bp.route('/condition-survey')
@login_required
def condition_survey():
    return render_template('condition_survey.html')


@bp.route('/vetting-status')
@login_required
def vetting_status():
    return render_template('vetting_status.html')


@bp.route('/class-status')
@login_required
def class_status_page():
    return render_template('class_status.html')


@bp.route('/calendar')
@login_required
def calendar_page():
    return render_template('calendar.html')


@bp.route('/dry-dock')
@login_required
def dry_dock_page():
    return render_template('dry_dock.html')


@bp.route('/dry-dock/<int:rid>/edit')
@login_required
def dry_dock_edit_page(rid):
    r = query('SELECT id FROM dock_reports WHERE id=?', (rid,), one=True)
    if not r:
        abort(404)
    return render_template('dry_dock_edit.html', report_id=rid)


@bp.route('/boarding')
@login_required
def boarding_page():
    return render_template('boarding.html')


@bp.route('/boarding/<int:rid>/edit')
@login_required
def boarding_edit_page(rid):
    r = query('SELECT id FROM boarding_reports WHERE id=?', (rid,), one=True)
    if not r:
        abort(404)
    return render_template('boarding_edit.html', report_id=rid)


# ═════════════════════════════════════════════════════════════════
#  API — me / password
# ═════════════════════════════════════════════════════════════════
@bp.route('/api/me')
@login_required
def api_me():
    out = {
        'user_id':       session['user_id'],
        'username':      session['username'],
        'display_name':  session.get('display_name'),
        'role':          session.get('role'),
        'supervisor_id': session.get('supervisor_id'),
    }
    # Bearer 요청에 한해 이 토큰의 만료시각(epoch)을 함께 준다.
    # 네이티브 앱은 오프라인일 때 캐시 프로필로 진입하는데, 만료시각을 모르면 진입을 거부한다
    # (fail-closed — 서버가 준 유효기간을 앱이 임의로 늘려주면 폰 분실·권한 회수가 영영 반영 안 됨).
    # 쿠키 세션(브라우저) 요청엔 발급시각이 없으므로 키 자체를 생략한다.
    issued = getattr(g, '_token_issued_at', None)
    if issued is not None:
        try:
            out['token_expires_at'] = int(issued.timestamp()) + _TOKEN_MAXAGE
        except (AttributeError, OSError, OverflowError, ValueError):
            pass
    return jsonify(out)


@bp.route('/api/drydock/mobile-entry')
@admin_required
def api_drydock_mobile_entry():
    """네이티브 앱 Bearer를 동일 도메인 Dock Manager browser session으로 1회 교환."""
    if not getattr(g, '_token_auth', False):
        return jsonify({'error': 'fresh_bearer_required'}), 401
    g._bearer_session_bridge = True
    session.permanent = False
    session.modified = True
    response = redirect('/drydock/', code=302)
    response.headers['Cache-Control'] = 'no-store'
    return response

@bp.route('/api/me/password', methods=['POST'])
@login_required
def api_me_password():
    d = request.get_json(silent=True) or {}
    old = d.get('old_password') or ''
    new = d.get('new_password') or ''
    if len(new) < 6:
        return jsonify({'error': '신규 비밀번호는 최소 6자 이상이어야 합니다.'}), 400
    u = query('SELECT * FROM users WHERE id=?',
              (session['user_id'],), one=True)
    if not check_password_hash(u['password_hash'], old):
        return jsonify({'error': '기존 비밀번호가 일치하지 않습니다.'}), 400
    execute('UPDATE users SET password_hash=? WHERE id=?',
            (generate_password_hash(new), session['user_id']))
    return jsonify({'ok': True})


# ═════════════════════════════════════════════════════════════════
#  API — supervisors
# ═════════════════════════════════════════════════════════════════
@bp.route('/api/supervisors')
@login_required
def api_supervisors():
    rows = query('''
        SELECT
            s.id, s.name, s.color, s.display_order, s.email,
            (SELECT COUNT(*) FROM issues i WHERE i.supervisor_id = s.id)
                AS total,
            (SELECT COUNT(*) FROM issues i WHERE i.supervisor_id = s.id AND i.status='Open')
                AS open_count,
            (SELECT COUNT(*) FROM issues i WHERE i.supervisor_id = s.id AND i.status='InProgress')
                AS progress_count,
            (SELECT COUNT(*) FROM issues i WHERE i.supervisor_id = s.id AND i.status='Closed')
                AS closed_count,
            (SELECT GROUP_CONCAT(v.name, ', ')
                FROM supervisor_vessels sv
                JOIN vessels v ON v.id = sv.vessel_id
               WHERE sv.supervisor_id = s.id) AS vessels
          FROM supervisors s
         WHERE s.active = 1
         ORDER BY s.display_order, s.id
    ''')
    return jsonify([dict(r) for r in rows])


# ═════════════════════════════════════════════════════════════════
#  API — vessels
# ═════════════════════════════════════════════════════════════════
@bp.route('/api/vessels')
@login_required
def api_vessels():
    sup = request.args.get('supervisor_id', type=int)
    if sup:
        rows = query('''
            SELECT v.* FROM vessels v
              JOIN supervisor_vessels sv ON sv.vessel_id = v.id
             WHERE sv.supervisor_id = ? AND v.active = 1
             ORDER BY v.name
        ''', (sup,))
    else:
        rows = query('SELECT * FROM vessels WHERE active=1 ORDER BY name')
    return jsonify([dict(r) for r in rows])


# Daily 사이드바 선박 커스텀 순서 (유저별, 드래그앤드롭). 빈 배열 = 기본정렬(디펙트순).
@bp.route('/api/vessel-order', methods=['GET', 'POST'])
@login_required
def api_vessel_order():
    uid = session.get('user_id')
    if request.method == 'POST':
        d = request.get_json(silent=True) or {}
        order = d.get('order')
        if not isinstance(order, list) or len(order) > 500:
            return jsonify({'ok': False, 'error': 'invalid order'}), 400
        # 정수 vessel id만 허용
        clean = [int(x) for x in order if str(x).lstrip('-').isdigit()]
        execute("INSERT OR REPLACE INTO user_vessel_order (user_id, order_json, updated_at) "
                "VALUES (?, ?, datetime('now','localtime'))",
                (uid, json.dumps(clean)))
        return jsonify({'ok': True, 'count': len(clean)})
    row = query("SELECT order_json FROM user_vessel_order WHERE user_id=?", (uid,), one=True)
    try:
        order = json.loads(row['order_json']) if row else []
    except (ValueError, TypeError):
        order = []
    return jsonify({'order': order})


# 선박별 활성(Open + InProgress) 이슈 수 — Daily 필터 드롭다운용
#   · 다른 화면 필터(감독, 검색, 우선순위, 선종)는 적용
#   · 선박 필터 자체는 무시 (드롭다운 라벨용이므로)
@bp.route('/api/vessels/active-counts')
@login_required
def api_vessel_active_counts():
    conds = ["i.status IN ('Open', 'InProgress')"]
    params = []

    sup = request.args.get('supervisor_id')
    if sup:
        conds.append('i.supervisor_id = ?')
        params.append(sup)

    q = request.args.get('q')
    if q:
        like = f'%{q}%'
        conds.append('(i.item_topic LIKE ? OR i.description LIKE ? OR i.actions LIKE ?)')
        params += [like, like, like]

    vt = request.args.get('vessel_type')
    if vt:
        conds.append('v.vessel_type = ?')
        params.append(vt)

    pri = request.args.get('priority')
    if pri:
        conds.append('i.priority = ?')
        params.append(pri)

    sql = f'''
        SELECT i.vessel_id, COUNT(*) AS cnt
          FROM issues i
          JOIN vessels v ON v.id = i.vessel_id
         WHERE {' AND '.join(conds)}
         GROUP BY i.vessel_id
    '''
    rows = query(sql, params)
    return jsonify({str(r['vessel_id']): r['cnt'] for r in rows})


# ═════════════════════════════════════════════════════════════════
#  API — issues (list / get / create / update / delete)
# ═════════════════════════════════════════════════════════════════
@bp.route('/api/issues')
@login_required
def api_issue_list():
    conds, params = ['1=1'], []
    for key, col in [('supervisor_id', 'i.supervisor_id'),
                     ('vessel_id',     'i.vessel_id'),
                     ('status',        'i.status'),
                     ('priority',      'i.priority')]:
        val = request.args.get(key)
        if val:
            conds.append(f'{col} = ?')
            params.append(val)

    q = request.args.get('q')
    if q:
        like = f'%{q}%'
        conds.append('(i.item_topic LIKE ? OR i.description LIKE ? OR i.actions LIKE ?)')
        params += [like, like, like]

    # 제목(ITEM) 정확 일치 — 요약 링크에서 해당 이슈만 보기 위함
    item_exact = request.args.get('item_topic')
    if item_exact:
        conds.append('i.item_topic = ?')
        params.append(item_exact)

    # 선종 필터 (vessels.vessel_type JOIN 기준)
    vt = request.args.get('vessel_type')
    if vt:
        conds.append('v.vessel_type = ?')
        params.append(vt)

    sql = f'''
        SELECT i.*,
               s.name       AS supervisor_name,
               s.color      AS supervisor_color,
               v.name       AS vessel_name,
               v.short_name AS vessel_short,
               (SELECT COUNT(*) FROM attachments a WHERE a.issue_id = i.id) AS att_count
          FROM issues i
          JOIN supervisors s ON s.id = i.supervisor_id
          JOIN vessels     v ON v.id = i.vessel_id
         WHERE {' AND '.join(conds)}
         ORDER BY i.issue_date ASC, i.id ASC
    '''
    rows = [_issue_to_dict(r) for r in query(sql, params)]
    return jsonify(rows)




@bp.route('/api/mobile/issues')
@login_required
def api_mobile_issue_list():
    """Mobile Daily card feed with server-enforced supervisor scope.

    기존 `/api/issues`의 범용 목록 권한을 바꾸지 않는다. mobile 화면은
    non-admin에게 현재 로그인 감독의 Open 이슈만 내보내며, 감독 연결이
    없는 member 계정은 빈 목록을 받는다.
    """
    is_admin = session.get('role') == 'admin'
    sup_id = session.get('supervisor_id')
    if not is_admin and not sup_id:
        return jsonify([])

    where = ["i.status != 'Closed'"]
    params = []
    if not is_admin:
        where.append('i.supervisor_id = ?')
        params.append(sup_id)

    sql = f'''
        SELECT i.*, v.name AS vessel_name, v.short_name AS vessel_short
          FROM issues i
          JOIN vessels v ON v.id = i.vessel_id
         WHERE {' AND '.join(where)}
         ORDER BY CASE i.priority
                    WHEN 'Urgent' THEN 0
                    WHEN 'COC & Flag' THEN 1
                    WHEN 'Next DD' THEN 2
                    ELSE 3 END,
                  COALESCE(i.due_date, '9999-12-31'), i.issue_date ASC, i.id ASC
         LIMIT 40
    '''
    return jsonify([_issue_to_dict(r) for r in query(sql, tuple(params))])


# ─────────────────────────────────────────────────────────────────
#  위젯 전용 축약 엔드포인트
#   iOS 위젯 익스텐션은 메모리 30MB·타임라인 생성시간 예산이 매우 짧다. 범용 API 를 그대로
#   쓰면 안 쓰는 필드까지 받아 디코드하느라 예산을 태우고, 그 사이 페이지 전환 탭이 씹힌다.
#   실측 2026-07-29: /api/issues 279KB(전건·i.* 통짜) · /api/fleet-map/data 24KB + 콜드 4초.
#   → 위젯이 **실제로 그리는 필드만** 내보낸다. 집계는 위젯이 그대로 하므로(계약 불변)
#     서버·클라 양쪽에 집계 로직이 이중화되지 않는다.
#   🔴 스코프는 기존 화면과 **같은 규칙**을 쓴다 — 숫자가 웹/앱과 어긋나면 형이 못 믿는다.
# ─────────────────────────────────────────────────────────────────
@bp.route('/api/widget/issues')
@login_required
def api_widget_issues():
    """위젯 Daily 현안 페이지용. 미완 이슈의 5개 필드만(스코프 = /api/mobile/issues 와 동일)."""
    is_admin = session.get('role') == 'admin'
    sup_id = session.get('supervisor_id')
    if not is_admin and not sup_id:
        return jsonify([])
    # COALESCE — 위젯 계약은 `status ?? "" != "Closed"` 라 status 결측을 **미완으로** 센다.
    # `i.status != 'Closed'` 만 쓰면 NULL 행이 SQL 3값논리로 조용히 빠져 숫자가 어긋난다
    # (현재 데이터에 NULL 은 0건이지만 계약을 코드로 못박아 둔다 — 올마이트 지적).
    where = ["COALESCE(i.status, '') != 'Closed'"]
    params = []
    if not is_admin:
        where.append('i.supervisor_id = ?')
        params.append(sup_id)
    rows = query(f'''
        SELECT v.name AS vessel, i.item_topic, i.priority, i.status, i.due_date
          FROM issues i
          JOIN vessels v ON v.id = i.vessel_id
         WHERE {' AND '.join(where)}
    ''', tuple(params))
    return jsonify([dict(r) for r in rows])


# 위젯 선대/선급 페이지가 읽는 필드 — fleet_map.json 이 이미 갖고 있는 값들이다.
WIDGET_FLEET_FIELDS = ('name', 'color', 'cls', 'cls_due_date', 'cls_due_days',
                       'coc', 'urgent', 'issues_open')


@bp.route('/api/widget/fleet')
@login_required
def api_widget_fleet():
    """위젯 선대 현황 · 선급 만기 페이지용.

    🔴 `/api/fleet-map/data` 를 쓰지 않는 이유 = 그 경로는 선위 overlay 를 위해 upstream
       ship-position(33.5MB)에 의존한다. 위젯은 **좌표를 그리지 않으므로** 그 의존을 통째로 뺀다.
       스코프는 `_fleet_visible_auto_vessels()` 를 공유해 Fleet Map 화면과 동일하게 유지한다.
    """
    fleet = [{k: v.get(k) for k in WIDGET_FLEET_FIELDS}
             for v in _fleet_visible_auto_vessels()]
    return jsonify({'fleet': fleet})


@bp.route('/api/widget/vetting')
@login_required
def api_widget_vetting():
    """위젯 SIRE 현황용. **선박당 1행**, 지적 본문은 싣지 않는다.

    🔴 `/api/vettings` 를 쓰지 않는 이유 = 그 응답은 vt_findings 전 컬럼(description·remark·
       user_remark)을 모든 vetting 에 중첩해 실측 약 33KB 이고, 그 중 약 70%가 위젯이 안 그리는
       지적 본문이다. 위젯은 다음 수검일과 건수만 그리므로 본문을 통째로 뺀다.
    🔴 스코프와 상단선정 규칙은 Vetting 탭과 **공유**한다(VETTING_TYPES + supervisor_vessels,
       `_vetting_pick`) — 숫자가 앱/웹과 어긋나면 형이 못 믿는다.
    ⚠️ `_vetting_with_counts()` 의 vetting 당 2쿼리(N+1)를 단일 GROUP BY 로 바꾸지 않았다 —
       manual override(manual_open_count 등)까지 그대로 타야 탭과 숫자가 일치하고, 대상이
       담당선 규모(수십 건)라 실측 부담이 없다. 정합성 > 미세최적화.
    """
    is_admin = session.get('role') == 'admin'
    sup_id = session.get('supervisor_id')
    if not is_admin and not sup_id:
        return jsonify({'vetting': []})     # 감독 미연결 member → 빈 배열(widget/issues 와 동일)

    ph = ','.join('?' * len(VETTING_TYPES))
    sql = (f'SELECT v.id, v.name FROM vessels v '
           f'WHERE v.active=1 AND v.vessel_type IN ({ph})')
    params = list(VETTING_TYPES)
    if not is_admin:
        sql += (' AND EXISTS (SELECT 1 FROM supervisor_vessels sv '
                'WHERE sv.vessel_id=v.id AND sv.supervisor_id=?)')
        params.append(sup_id)
    sql += ' ORDER BY v.name'

    out = []
    for ve in query(sql, tuple(params)):
        latest, _enr = _vetting_pick(ve['id'])
        if not latest:
            continue                        # 수검 이력이 아예 없는 선박은 그릴 게 없다
        out.append({
            'vessel': ve['name'],
            'status': latest.get('valid') or '',                 # 'Next Plan' / 'Last Result' / ''
            'oil_major': latest.get('inspection_company') or '',
            'date': latest.get('inspection_date') or '',         # Next Plan 은 미입력일 수 있음
            'port': latest.get('port') or '',
            'obs_total': latest.get('observation_count') or 0,
            'obs_open': latest.get('open_count') or 0,
            'obs_closed': latest.get('close_count') or 0,
            # obs_* 는 전부 상단행 한 건에서만 나온다(`_vetting_pick` 정본). 아래 두 키는
            # 구 폴백 시절 "수치의 출처"를 따로 알리려고 둔 것이고 지금은 위 oil_major/date 와
            # 같은 값이다. 위젯(WidgetModel)이 아직 읽고 있어 **키는 유지**한다 — 지우면
            # 구버전 앱 화면의 부제가 조용히 빈다.
            'obs_oil_major': latest.get('inspection_company') or '',
            'obs_date': latest.get('inspection_date') or '',
        })
    return jsonify({'vetting': out})


# ─────────────────────────────────────────────────────────────────
#  Daily 업무관리 — Excel 추출 (정형 템플릿)
#   · 화면 구조 그대로 재현: 감독 시트 → 제목 → 컬럼 헤더 →
#     월 그룹 헤더 → 일 그룹 헤더 → 데이터 행
#   · Excel의 행 그룹(outline) 기능으로 월·일 단위 접기/펼치기 가능
#   · 컬럼 헤더 행에 AutoFilter 적용 → 선박명 등 자유롭게 필터
#   · 현재 화면 필터(상태/우선순위/선박/선종/검색어/서브탭) 그대로 반영
# ─────────────────────────────────────────────────────────────────
@bp.route('/api/issues/export')
@login_required
def api_issue_export():
    from io import BytesIO
    from datetime import datetime
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'error': 'openpyxl 미설치 — 서버에 pip install openpyxl 필요'}), 500
    from flask import send_file

    # ── 1) 화면 필터와 동일한 조건 ──────────────────────────────
    conds, params = ['1=1'], []
    for key, col in [('supervisor_id', 'i.supervisor_id'),
                     ('vessel_id',     'i.vessel_id'),
                     ('status',        'i.status'),
                     ('priority',      'i.priority')]:
        val = request.args.get(key)
        if val:
            conds.append(f'{col} = ?')
            params.append(val)

    status_in = request.args.get('status_in')
    if status_in:
        vals = [v.strip() for v in status_in.split(',') if v.strip()]
        if vals:
            placeholders = ','.join('?' for _ in vals)
            conds.append(f'i.status IN ({placeholders})')
            params += vals

    # vessel_ids (복수 선박, 담당자별 추출용)
    vessel_ids = request.args.get('vessel_ids')
    if vessel_ids:
        ids = [v.strip() for v in vessel_ids.split(',') if v.strip().isdigit()]
        if ids:
            ph = ','.join('?' for _ in ids)
            conds.append(f'i.vessel_id IN ({ph})')
            params += ids

    q = request.args.get('q')
    if q:
        like = f'%{q}%'
        conds.append('(i.item_topic LIKE ? OR i.description LIKE ? OR i.actions LIKE ?)')
        params += [like, like, like]

    vt = request.args.get('vessel_type')
    if vt:
        conds.append('v.vessel_type = ?')
        params.append(vt)

    sql = f'''
        SELECT i.*,
               s.id            AS sv_id,
               s.name          AS supervisor_name,
               s.display_order AS sv_order,
               v.name          AS vessel_name,
               v.vessel_type   AS vessel_type
          FROM issues i
          JOIN supervisors s ON s.id = i.supervisor_id
          JOIN vessels     v ON v.id = i.vessel_id
         WHERE {' AND '.join(conds)}
         ORDER BY s.display_order ASC, s.id ASC,
                  i.issue_date ASC, i.id ASC
    '''
    rows = [_issue_to_dict(r) for r in query(sql, params)]

    EN = (request.args.get('lang') == 'en')
    if EN:
        _translate_rows_en(rows)

    # ── 2) 선박별 그룹 (sheet = 선박) ──────────────────────────
    VTYPE_ORDER = ['VLCC', 'LR', 'AFRAMAX', 'MR', 'CNTR']
    def _vrank(t):
        t = (t or '').upper()
        return VTYPE_ORDER.index(t) if t in VTYPE_ORDER else len(VTYPE_ORDER)
    ves_map = {}   # vessel_name -> {'type':, 'rows':[]}
    for r in rows:
        vn = r.get('vessel_name') or ('Unassigned' if EN else '미배정')
        if vn not in ves_map:
            ves_map[vn] = {'type': r.get('vessel_type') or '', 'rows': []}
        ves_map[vn]['rows'].append(r)
    # 시트 순서 = 선종(VLCC→…→CNTR) → 선명
    ves_seq = sorted(ves_map.keys(), key=lambda n: (_vrank(ves_map[n]['type']), n))

    # ── 3) 스타일 / 헤더 ────────────────────────────────────────
    HEADERS = (['No.', 'Issue Date', 'Item', 'Description', 'Action Plan',
                'Priority', 'Status', 'Due Date', 'TSI Comment']
               if EN else
               ['No.', '발생일', '현안업무', '상세 내용', '진행사항 (조치 이력)',
                '우선순위', '상태', '마감일', 'TSI Comment'])
    COL_WIDTHS = [5, 12, 30, 40, 44, 12, 11, 12, 34]
    N_COLS   = len(HEADERS)
    PRI_COL, STAT_COL = 6, 7

    F = 'Malgun Gothic'
    title_font   = Font(name=F, size=14, bold=True, color='FFFFFF')
    sub_font     = Font(name=F, size=10, color='ECF0F1', italic=True)
    title_fill   = PatternFill('solid', start_color='1F3A5F')
    sub_fill     = PatternFill('solid', start_color='2C5282')
    col_hdr_font = Font(name=F, size=10, bold=True, color='FFFFFF')
    col_hdr_fill = PatternFill('solid', start_color='34495E')
    body_font    = Font(name=F, size=10)
    tsi_font     = Font(name=F, size=10, italic=True, color='95A5A6')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    body_align   = Alignment(horizontal='left',   vertical='top',    wrap_text=True)
    cent_top     = Alignment(horizontal='center', vertical='top',    wrap_text=True)

    thin = Side(style='thin',   color='BDC3C7')
    med  = Side(style='medium', color='34495E')
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    PRI_FILL = {
        'COC & Flag': PatternFill('solid', start_color='F8CECC'),
        'Urgent':     PatternFill('solid', start_color='FFE6CC'),
        'Next DD':    PatternFill('solid', start_color='FFF2CC'),
        'Normal':     None,
    }
    PRI_FONT = {
        'COC & Flag': Font(name=F, size=10, bold=True, color='B71C1C'),
        'Urgent':     Font(name=F, size=10, bold=True, color='E65100'),
        'Next DD':    Font(name=F, size=10, bold=True, color='6D4C0F'),
        'Normal':     Font(name=F, size=10, color='5D6D7E'),
    }
    STAT_FILL = {
        'Open':       PatternFill('solid', start_color='E1F5FE'),
        'InProgress': PatternFill('solid', start_color='FFF9C4'),
        'Closed':     PatternFill('solid', start_color='E8F5E9'),
    }
    STAT_FONT = {
        'Open':       Font(name=F, size=10, bold=True, color='0277BD'),
        'InProgress': Font(name=F, size=10, bold=True, color='F57F17'),
        'Closed':     Font(name=F, size=10, bold=True, color='2E7D32'),
    }
    STAT_LABEL = ({'Open': 'Open', 'InProgress': 'In Progress', 'Closed': 'Closed'}
                  if EN else
                  {'Open': 'Open', 'InProgress': '진행중', 'Closed': 'Closed'})

    def _sheet_safe(name):
        bad = '[]:*?/\\'
        out = ''.join('_' if c in bad else c for c in name)
        return (out[:31] or 'Sheet')

    def _fmt_actions(acts):
        if not acts:
            return ''
        lines = []
        for a in acts:
            d = (a.get('date') or '').strip()
            p = (a.get('progress') or '').strip()
            mark = '★ ' if a.get('important') else ''
            if d and p:   lines.append(f'{mark}[{d}] {p}')
            elif d:       lines.append(f'{mark}[{d}]')
            elif p:       lines.append(f'{mark}{p}')
        return '\n'.join(lines)

    # ── 4) Workbook 생성 ────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)
    now = datetime.now()
    today_str = now.strftime('%Y-%m-%d')
    me = session.get('display_name') or session.get('username') or ''

    sub_chips = []
    if status_in:
        sub_chips.append(('Filter: ' if EN else '필터: ') + status_in.replace(',', ' / '))
    elif request.args.get('status'):
        sub_chips.append(('Status: ' if EN else '상태: ') + request.args.get('status'))
    if request.args.get('priority'):
        sub_chips.append(('Priority: ' if EN else '우선순위: ') + request.args.get('priority'))
    if request.args.get('q'):
        sub_chips.append(('Search: ' if EN else '검색: ') + request.args.get('q'))
    sub_text = ' | '.join(sub_chips) if sub_chips else ('All items' if EN else '전체 항목')

    if not ves_seq:
        ws = wb.create_sheet('No Data' if EN else '데이터 없음')
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N_COLS)
        c = ws.cell(row=1, column=1, value=('Daily Work Log — No Data' if EN else 'Daily 업무관리 — 데이터 없음'))
        c.font = title_font; c.fill = title_fill; c.alignment = center_align
        ws.cell(row=3, column=1, value=('No issues match the filter.' if EN else '필터 조건에 해당하는 이슈가 없습니다.')).font = Font(name=F, size=11, italic=True)
        for idx, w in enumerate(COL_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = w
    else:
        for vn in ves_seq:
            info = ves_map[vn]
            ws = wb.create_sheet(_sheet_safe(vn))
            for idx, w in enumerate(COL_WIDTHS, start=1):
                ws.column_dimensions[get_column_letter(idx)].width = w

            # 제목(행1) = 선박명 (+선종),  부제(행2) = 추출 메타
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N_COLS)
            vt = info['type']
            c1 = ws.cell(row=1, column=1, value=(f'{vn}   |   {vt}' if vt else vn))
            c1.font = title_font; c1.fill = title_fill
            c1.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.row_dimensions[1].height = 30

            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N_COLS)
            cnt = len(info['rows'])
            if EN:
                sub_msg = f'Exported: {today_str}    │    Total {cnt}    │    {sub_text}'
                if me: sub_msg += f'    │    By: {me}'
            else:
                sub_msg = f'추출일: {today_str}    │    총 {cnt}건    │    {sub_text}'
                if me: sub_msg += f'    │    출력: {me}'
            c2 = ws.cell(row=2, column=1, value=sub_msg)
            c2.font = sub_font; c2.fill = sub_fill
            c2.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.row_dimensions[2].height = 20
            ws.row_dimensions[3].height = 6

            # 컬럼 헤더(행4)
            HDR_ROW = 4
            for col_idx, h in enumerate(HEADERS, start=1):
                c = ws.cell(row=HDR_ROW, column=col_idx, value=h)
                c.font = col_hdr_font; c.fill = col_hdr_fill
                c.alignment = center_align
                c.border = Border(left=thin, right=thin, top=med, bottom=med)
            ws.row_dimensions[HDR_ROW].height = 26

            # 데이터(행5~) — 날짜 그룹 없이 발생일 오래된순, No.=선박 내 1..N
            cur_row = HDR_ROW + 1
            for no, r in enumerate(sorted(info['rows'],
                                          key=lambda x: ((x.get('issue_date') or ''), x.get('id') or 0)), start=1):
                vals = [
                    no,
                    r.get('issue_date') or '',
                    r.get('item_topic') or '',
                    r.get('description') or '',
                    _fmt_actions(r.get('actions')),
                    r.get('priority') or '',
                    STAT_LABEL.get(r.get('status'), r.get('status') or ''),
                    r.get('due_date') or '',
                    '',                                   # TSI Comment — 수기 기입용 빈 칸
                ]
                for col_idx, v in enumerate(vals, start=1):
                    c = ws.cell(row=cur_row, column=col_idx, value=v)
                    c.font = body_font
                    c.border = border_thin
                    if col_idx in (1, 2, 8):              # No / 발생일 / 마감일
                        c.alignment = cent_top
                    elif col_idx in (PRI_COL, STAT_COL):  # 우선순위 / 상태
                        c.alignment = center_align
                    else:                                 # 현안업무 / 상세 / 진행사항 / TSI
                        c.alignment = body_align
                # 우선순위 / 상태 색
                pri = r.get('priority')
                if PRI_FILL.get(pri): ws.cell(row=cur_row, column=PRI_COL).fill = PRI_FILL[pri]
                if pri in PRI_FONT:   ws.cell(row=cur_row, column=PRI_COL).font = PRI_FONT[pri]
                st = r.get('status')
                if STAT_FILL.get(st): ws.cell(row=cur_row, column=STAT_COL).fill = STAT_FILL[st]
                if st in STAT_FONT:   ws.cell(row=cur_row, column=STAT_COL).font = STAT_FONT[st]
                cur_row += 1

            last_row = cur_row - 1
            if last_row > HDR_ROW:
                ws.auto_filter.ref = f'A{HDR_ROW}:{get_column_letter(N_COLS)}{last_row}'
            ws.freeze_panes = f'A{HDR_ROW + 1}'
            ws.print_options.horizontalCentered = True
            ws.page_setup.orientation = 'landscape'
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.print_title_rows = f'{HDR_ROW}:{HDR_ROW}'

    # ── 5) 파일명 ──
    today = now.strftime('%Y%m%d')
    suffix = '_EN' if EN else ''
    if len(ves_seq) == 1:
        fname = f'TRMT_Daily_{_sheet_safe(ves_seq[0])}_{today}{suffix}.xlsx'
    else:
        fname = f'TRMT_Daily_{today}{suffix}.xlsx'

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=fname,
    )






def _summary_scope():
    sid = request.args.get('supervisor_id')
    return str(sid) if sid else 'all'


@bp.route('/api/issues/summary', methods=['GET'])
@login_required
def api_issue_summary_get():
    _ensure_summary_table()
    row = query('SELECT data, generated_at FROM issue_summaries WHERE scope=?',
                (_summary_scope(),), one=True)
    if not row:
        return jsonify({'rows': [], 'generated_at': None, 'count': 0})
    try:
        rows = json.loads(row['data'])
    except Exception as e:
        app.logger.warning('issue-summary-get: %s', e)
        rows = []
    return jsonify({'rows': rows, 'generated_at': row['generated_at'], 'count': len(rows)})




@bp.route('/api/issues/summary-generate', methods=['POST'])
@login_required
def api_issue_summary_generate():
    sid = request.args.get('supervisor_id') or None
    rows, gen_at, counts = _run_summary_generate(sid)
    return jsonify({'rows': rows, 'generated_at': gen_at, 'counts': counts})


@bp.route('/api/issues/summary-counts', methods=['GET'])
@login_required
def api_issue_summary_counts():
    _ensure_summary_table()
    out = {}
    for r in query('SELECT scope, data FROM issue_summaries'):
        try:
            out[r['scope']] = len(json.loads(r['data']))
        except Exception as e:
            app.logger.warning('issue-summary-counts: %s', e)
            out[r['scope']] = 0
    return jsonify(out)


@bp.route('/api/issues/summary-export')
@login_required
def api_issue_summary_export():
    """현재 탭(대분류)의 저장된 요약(요약 탭 내용)을 엑셀로 추출 — AI 미사용."""
    from io import BytesIO
    from datetime import datetime
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'error': 'openpyxl 미설치'}), 500
    from flask import send_file

    # 저장된 요약(요약 탭 내용)을 그대로 사용
    _ensure_summary_table()
    srow = query('SELECT data FROM issue_summaries WHERE scope=?',
                 (_summary_scope(),), one=True)
    rows = []
    if srow:
        try:
            rows = json.loads(srow['data'])
        except Exception as e:
            app.logger.warning('issue-summary-export: %s', e)
            rows = []

    def build_cell(idx, r):
        return r.get('issue') or ''

    # ── Workbook ──
    wb = Workbook(); ws = wb.active; ws.title = '업무 요약'
    F = 'Malgun Gothic'
    HEADERS = ['No.', 'Vessel Name', '현안업무', 'Priority', 'Status']
    WIDTHS = [6, 24, 85, 13, 12]
    for idx, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    title_fill = PatternFill('solid', start_color='1F3A5F')
    sub_fill   = PatternFill('solid', start_color='2C5282')
    hdr_fill   = PatternFill('solid', start_color='34495E')
    thin = Side(style='thin', color='BBBBBB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    now = datetime.now()
    ws.merge_cells('A1:E1')
    c = ws.cell(row=1, column=1, value='Daily 업무 요약')
    c.font = Font(name=F, size=14, bold=True, color='FFFFFF'); c.fill = title_fill
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:E2')
    me = session.get('display_name') or session.get('username') or ''
    c = ws.cell(row=2, column=1,
                value=f"추출일: {now.strftime('%Y-%m-%d')}    │    총 {len(rows)}건"
                      + (f"    │    {me}" if me else ''))
    c.font = Font(name=F, size=10, italic=True, color='ECF0F1'); c.fill = sub_fill
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6

    HDR = 4
    for ci, h in enumerate(HEADERS, start=1):
        cc = ws.cell(row=HDR, column=ci, value=h)
        cc.font = Font(name=F, size=11, bold=True, color='FFFFFF'); cc.fill = hdr_fill
        cc.alignment = Alignment(horizontal='center', vertical='center')
        cc.border = border
    ws.row_dimensions[HDR].height = 24

    body = Font(name=F, size=10)
    top_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
    center = Alignment(horizontal='center', vertical='center')
    STAT_LABEL = {'Open': 'Open', 'InProgress': '진행중', 'Closed': 'Closed'}
    r_idx = HDR + 1
    for n, r in enumerate(rows, start=1):
        ws.cell(row=r_idx, column=1, value=n).alignment = center
        ws.cell(row=r_idx, column=1).font = body
        ws.cell(row=r_idx, column=2, value=r.get('vessel_name') or '')
        ws.cell(row=r_idx, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.cell(row=r_idx, column=2).font = body
        cell = ws.cell(row=r_idx, column=3, value=build_cell(n - 1, r))
        cell.alignment = top_wrap; cell.font = body
        # D열 Priority, E열 Status
        pc = ws.cell(row=r_idx, column=4, value=r.get('priority') or '')
        pc.alignment = center; pc.font = body
        sc = ws.cell(row=r_idx, column=5,
                     value=STAT_LABEL.get(r.get('status'), r.get('status') or ''))
        sc.alignment = center; sc.font = body
        for ci in range(1, 6):
            ws.cell(row=r_idx, column=ci).border = border
        # 줄 수에 맞춰 행 높이 살짝 키움
        n_lines = (build_cell(n - 1, r).count('\n') + 1)
        ws.row_dimensions[r_idx].height = max(34, 15 * n_lines + 6)
        r_idx += 1

    ws.freeze_panes = f'A{HDR + 1}'
    if r_idx - 1 > HDR:
        ws.auto_filter.ref = f'A{HDR}:E{r_idx - 1}'
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f'{HDR}:{HDR}'

    scope = _summary_scope()
    tag = ''
    if scope != 'all':
        sv = query('SELECT name FROM supervisors WHERE id=?', (scope,), one=True)
        if sv:
            tag = '_' + _safe_filename(sv['name'])
    fname = f"TRMT_업무요약{tag}_{now.strftime('%Y%m%d')}.xlsx"
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    return send_file(
        bio,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=fname)




@bp.route('/api/issues/<int:iid>')
@login_required
def api_issue_get(iid):
    _issue_write_scope(iid)
    r = query('''
        SELECT i.*,
               s.name       AS supervisor_name,
               s.color      AS supervisor_color,
               v.name       AS vessel_name,
               v.short_name AS vessel_short
          FROM issues i
          JOIN supervisors s ON s.id = i.supervisor_id
          JOIN vessels     v ON v.id = i.vessel_id
         WHERE i.id = ?
    ''', (iid,), one=True)
    if not r:
        abort(404)
    out = _issue_to_dict(r)
    out['attachments'] = [dict(a) for a in query(
        'SELECT id, filename, stored_name, file_size, mime_type, uploaded_at '
        'FROM attachments WHERE issue_id=? ORDER BY id', (iid,))]
    return jsonify(out)


@bp.route('/api/issues', methods=['POST'])
@login_required
def api_issue_create():
    d = request.get_json(silent=True) or {}
    _issue_write_scope(payload=d)
    for k in ('supervisor_id', 'vessel_id', 'issue_date', 'item_topic'):
        if not d.get(k):
            return jsonify({'error': f'필수 항목 누락: {k}'}), 400

    actions = d.get('actions') or []
    if not isinstance(actions, list):
        actions = []
    actions_json = json.dumps(actions, ensure_ascii=False)

    iid = execute('''
        INSERT INTO issues
            (supervisor_id, vessel_id, issue_date, due_date,
             item_topic, description, actions,
             priority, status, created_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        d['supervisor_id'], d['vessel_id'], d['issue_date'],
        d.get('due_date') or None,
        d['item_topic'],
        d.get('description') or '',
        actions_json,
        d.get('priority') or 'Normal',
        d.get('status')   or 'Open',
        session.get('username'),
    ))
    return jsonify({'id': iid}), 201


@bp.route('/api/issues/<int:iid>', methods=['PUT'])
@login_required
def api_issue_update(iid):
    current = _issue_write_scope(iid)
    d = request.get_json(silent=True) or {}
    if session.get('role') != 'admin':
        sup_id = session.get('supervisor_id')
        if 'supervisor_id' in d and int(d.get('supervisor_id') or 0) != sup_id:
            abort(403)
        vessel_id = int(d.get('vessel_id') or current['vessel_id'])
        # Preserve a legacy issue's current vessel even if it was later unassigned/inactivated.
        # A changed vessel must still be one of the member's currently assigned vessels.
        if vessel_id != current['vessel_id'] and not query(
                'SELECT 1 FROM supervisor_vessels WHERE supervisor_id=? AND vessel_id=?',
                (sup_id, vessel_id), one=True):
            abort(403)
    fields = ['supervisor_id', 'vessel_id', 'issue_date', 'due_date',
              'item_topic',    'description', 'actions',
              'priority',      'status']
    sets, params = [], []
    for f in fields:
        if f in d:
            val = d[f]
            if f == 'actions':
                if not isinstance(val, list):
                    val = []
                val = json.dumps(val, ensure_ascii=False)
            elif val == '':
                val = None
            sets.append(f'{f} = ?')
            params.append(val)
    if not sets:
        return jsonify({'error': '수정할 필드가 없습니다.'}), 400
    sets.append('updated_at = datetime("now","localtime")')
    params.append(iid)
    execute(f'UPDATE issues SET {", ".join(sets)} WHERE id = ?', params)
    return jsonify({'id': iid})


@bp.route('/api/issues/<int:iid>/actions', methods=['POST'])
@login_required
def api_issue_action_append(iid):
    """진행 경과 1건 **원자적 추가**(앱 오프라인 보관함 전용 경로이자, 일반 추가의 안전판).

    🔴 왜 따로 두는가: 앱은 지금까지 `PUT /api/issues/<id>` 로 actions **배열 전체**를 덮어
       진행을 추가했다. 온라인에서는 직전에 상세를 다시 읽어 위험이 짧지만, 오프라인 보관함은
       읽은 시점과 보내는 시점이 몇 시간~며칠 벌어진다 — 그 사이 웹/다른 기기가 추가한 진행이
       **조용히 사라진다**. 여기서는 서버가 그 순간의 원문을 읽어 append 하므로 유실이 없다.
       (같은 이유로 ext(api_key) 전용이던 append 를 login/Bearer 에도 연다 — 계약은 동일.)
    CAS 로 읽고-쓰는 사이의 변경까지 막고, 밀리면 최신 원문으로 1회 재시도한다
    (append 는 순서만 맞으면 되므로 PATCH 처럼 409 로 사람을 부를 이유가 없다).
    """
    from datetime import date as _date
    _issue_write_scope(iid)
    d = request.get_json(silent=True)
    if not isinstance(d, dict):
        return jsonify({'error': '요청 형식이 올바르지 않습니다.'}), 400
    progress = (d.get('progress') or '').strip()
    if not progress:
        return jsonify({'error': '진행 내용을 입력하세요.'}), 400
    adate = (d.get('date') or '').strip()
    if adate:
        try:
            if adate != datetime.strptime(adate, '%Y-%m-%d').strftime('%Y-%m-%d'):
                raise ValueError(adate)
        except ValueError:
            return jsonify({'error': '발생일 형식이 올바르지 않습니다(YYYY-MM-DD).'}), 400
    else:
        adate = _date.today().isoformat()
    entry = {'date': adate, 'progress': progress, 'important': bool(d.get('important'))}

    for _ in range(2):
        row = query('SELECT actions FROM issues WHERE id=?', (iid,), one=True)
        if not row:
            abort(404)
        raw = row['actions']
        try:
            actions = json.loads(raw) if raw else []
        except Exception as e:
            app.logger.warning('issue-action-append: %s', e)
            actions = []
        if not isinstance(actions, list):
            actions = []
        merged = actions + [entry]
        new_raw = json.dumps(merged, ensure_ascii=False)
        if raw is None:
            rc = execute_rc('UPDATE issues SET actions=?, updated_at=datetime("now","localtime") '
                            'WHERE id=? AND actions IS NULL', (new_raw, iid))
        else:
            rc = execute_rc('UPDATE issues SET actions=?, updated_at=datetime("now","localtime") '
                            'WHERE id=? AND actions=?', (new_raw, iid, raw))
        if rc:
            return jsonify({'id': iid, 'actions': merged})
    # 두 번 연속 밀렸다 = 동시 쓰기가 몰리는 중. 조용히 성공으로 위장하지 않는다.
    return jsonify({'error': '다른 곳에서 동시에 변경 중입니다. 잠시 후 다시 시도하세요.'}), 409


@bp.route('/api/issues/<int:iid>/actions/<int:idx>', methods=['PATCH'])
@login_required
def api_issue_action_patch(iid, idx):
    """진행 경과 1건만 수정 — 내용·발생일·중요표시(모바일 인라인 편집용).

    편집 모달은 `PUT /api/issues/<id>` 로 actions 배열을 통째로 덮지만, 인라인 편집은
    "화면에 보이던 그 줄"만 고쳐야 한다. 배열 전체를 되돌려쓰는 사이 웹/다른 기기가
    진행을 추가하면 그 추가가 조용히 사라지고, 더 나쁘게는 배열이 밀려 **다른 줄**이
    고쳐진다. 그래서 두 겹으로 fail-closed 한다:
      ① `prev`(편집 시작 시점의 date/progress) 대조 — 다르면 409, DB 무변경.
      ② 읽은 원문 JSON 을 조건으로 건 CAS(`WHERE actions=?`) — 읽고 쓰는 사이의
         변경도 409 로 떨어진다(rowcount 0).
    409 응답엔 서버 정본 actions 를 실어 보내 클라이언트가 새로고침 없이 되돌릴 수 있게 한다.
    """
    _issue_write_scope(iid)
    row = query('SELECT actions FROM issues WHERE id=?', (iid,), one=True)
    if not row:
        abort(404)
    raw = row['actions']
    try:
        actions = json.loads(raw) if raw else []
    except Exception as e:
        app.logger.warning('issue-action-patch: %s', e)
        actions = []
    if not isinstance(actions, list):
        actions = []
    if not (0 <= idx < len(actions)) or not isinstance(actions[idx], dict):
        return jsonify({'error': '해당 진행 경과가 없습니다. 새로고침 후 다시 시도하세요.',
                        'actions': actions}), 409

    d = request.get_json(silent=True)
    if not isinstance(d, dict):
        return jsonify({'error': '요청 형식이 올바르지 않습니다.'}), 400
    tgt = actions[idx]

    def _txt(v):
        return '' if v is None else str(v)

    # prev 는 **필수** — 없으면 index 만 믿고 고치게 되어 목록이 밀린 순간 다른 줄이 수정된다(올마이트 지적).
    prev = d.get('prev')
    if not isinstance(prev, dict) or 'date' not in prev or 'progress' not in prev:
        return jsonify({'error': 'prev(date/progress)가 필요합니다.'}), 400
    for k in ('date', 'progress'):
        if not isinstance(prev.get(k), (str, type(None))):
            return jsonify({'error': f'prev.{k} 형식이 올바르지 않습니다.'}), 400
        if _txt(prev.get(k)) != _txt(tgt.get(k)):
            return jsonify({'error': '다른 곳에서 이미 변경됐습니다. 새로고침 후 다시 시도하세요.',
                            'actions': actions}), 409
    # important 도 대조 — 안 보면 다른 기기가 켜 둔 별을 이 요청이 옛 값으로 되돌린다.
    if 'important' in prev and bool(prev.get('important')) != bool(tgt.get('important')):
        return jsonify({'error': '다른 곳에서 이미 변경됐습니다. 새로고침 후 다시 시도하세요.',
                        'actions': actions}), 409

    for k, t in (('progress', str), ('date', str), ('important', bool)):
        if k in d and not isinstance(d[k], t):
            return jsonify({'error': f'{k} 형식이 올바르지 않습니다.'}), 400
    if 'progress' in d:
        p = d['progress'].strip()
        if not p:
            return jsonify({'error': '진행 내용을 입력하세요.'}), 400
        tgt['progress'] = p
    if 'date' in d:
        # 빈 값은 무시(발생일 없는 행으로 떨어뜨리지 않는다). 형식은 화면과 동일하게 YYYY-MM-DD 만.
        nd = d['date'].strip()
        if nd:
            try:
                if nd != datetime.strptime(nd, '%Y-%m-%d').strftime('%Y-%m-%d'):
                    raise ValueError(nd)
            except ValueError:
                return jsonify({'error': '발생일 형식이 올바르지 않습니다(YYYY-MM-DD).'}), 400
            tgt['date'] = nd
    if 'important' in d:
        tgt['important'] = d['important']

    new_raw = json.dumps(actions, ensure_ascii=False)
    if raw is None:
        rc = execute_rc('UPDATE issues SET actions=?, updated_at=datetime("now","localtime") '
                        'WHERE id=? AND actions IS NULL', (new_raw, iid))
    else:
        rc = execute_rc('UPDATE issues SET actions=?, updated_at=datetime("now","localtime") '
                        'WHERE id=? AND actions=?', (new_raw, iid, raw))
    if not rc:
        cur = query('SELECT actions FROM issues WHERE id=?', (iid,), one=True)
        try:
            live = json.loads(cur['actions']) if (cur and cur['actions']) else []
        except Exception:
            live = []
        return jsonify({'error': '다른 곳에서 이미 변경됐습니다. 새로고침 후 다시 시도하세요.',
                        'actions': live if isinstance(live, list) else []}), 409
    return jsonify({'id': iid, 'actions': actions})


@bp.route('/api/issues/<int:iid>', methods=['DELETE'])
@login_required
def api_issue_delete(iid):
    _issue_write_scope(iid)
    atts = query('SELECT stored_name FROM attachments WHERE issue_id=?', (iid,))
    for a in atts:
        p = os.path.join(UPLOAD_DIR, a['stored_name'])
        if os.path.exists(p):
            os.remove(p)
    execute('DELETE FROM issues WHERE id=?', (iid,))
    return jsonify({'ok': True})


# ═════════════════════════════════════════════════════════════════
#  API — admin: supervisors / vessels / users
# ═════════════════════════════════════════════════════════════════

# ----- 감독 (CREATE / UPDATE / DELETE) -----
@bp.route('/api/supervisors', methods=['POST'])
@admin_required
def api_supervisor_create():
    d = request.get_json(silent=True) or {}
    name = (d.get('name') or '').strip()
    if not name:
        return jsonify({'error': '감독명은 필수입니다.'}), 400
    if query('SELECT id FROM supervisors WHERE name=?', (name,), one=True):
        return jsonify({'error': '이미 존재하는 감독명입니다.'}), 400
    max_order = query('SELECT COALESCE(MAX(display_order),0)+1 AS n FROM supervisors',
                      one=True)['n']
    sid = execute('''
        INSERT INTO supervisors (name, color, display_order, email, active)
        VALUES (?, ?, ?, ?, 1)
    ''', (name, d.get('color') or 'blue',
          d.get('display_order') or max_order,
          d.get('email') or ''))
    return jsonify({'id': sid}), 201


@bp.route('/api/supervisors/<int:sid>', methods=['PUT'])
@admin_required
def api_supervisor_update(sid):
    if not query('SELECT id FROM supervisors WHERE id=?', (sid,), one=True):
        abort(404)
    d = request.get_json(silent=True) or {}
    sets, params = [], []
    for f in ('name', 'color', 'display_order', 'email', 'active'):
        if f in d:
            sets.append(f'{f} = ?')
            params.append(d[f])
    if not sets:
        return jsonify({'error': '수정할 필드 없음'}), 400
    params.append(sid)
    execute(f'UPDATE supervisors SET {", ".join(sets)} WHERE id = ?', params)
    return jsonify({'id': sid})


@bp.route('/api/supervisors/<int:sid>', methods=['DELETE'])
@admin_required
def api_supervisor_delete(sid):
    # 이슈 있으면 soft delete 만 수행
    n = query('SELECT COUNT(*) AS n FROM issues WHERE supervisor_id=?',
              (sid,), one=True)['n']
    if n > 0:
        execute('UPDATE supervisors SET active=0 WHERE id=?', (sid,))
        return jsonify({'ok': True, 'soft_delete': True, 'issues': n})
    # Hard delete: FK 해제 먼저
    execute('UPDATE users SET supervisor_id=NULL WHERE supervisor_id=?', (sid,))
    execute('DELETE FROM supervisor_vessels WHERE supervisor_id=?', (sid,))
    execute('DELETE FROM supervisors WHERE id=?', (sid,))
    return jsonify({'ok': True})


# ----- 선박 (CREATE / UPDATE / DELETE / 전체 조회) -----
@bp.route('/api/vessels/all')
@login_required
def api_vessels_all():
    """관리 UI용 — 담당 감독 함께."""
    rows = query('''
        SELECT v.*,
          (SELECT GROUP_CONCAT(s.name, ', ')
             FROM supervisor_vessels sv
             JOIN supervisors s ON s.id = sv.supervisor_id
            WHERE sv.vessel_id = v.id) AS supervisor_names,
          (SELECT GROUP_CONCAT(s.id)
             FROM supervisor_vessels sv
             JOIN supervisors s ON s.id = sv.supervisor_id
            WHERE sv.vessel_id = v.id) AS supervisor_ids_csv
          FROM vessels v
         ORDER BY v.active DESC, v.name
    ''')
    out = []
    for r in rows:
        d = dict(r)
        d['supervisor_ids'] = [int(x) for x in (d.pop('supervisor_ids_csv') or '').split(',') if x]
        out.append(d)
    return jsonify(out)


@bp.route('/api/vessels', methods=['POST'])
@login_required
def api_vessel_create():
    d = request.get_json(silent=True) or {}
    name = (d.get('name') or '').strip()
    if not name:
        return jsonify({'error': '선박명은 필수입니다.'}), 400
    if query('SELECT id FROM vessels WHERE name=?', (name,), one=True):
        return jsonify({'error': '이미 존재하는 선박명입니다.'}), 400

    sids = [int(x) for x in (d.get('supervisor_ids') or [])]

    # 일반 사용자(member) 권한 제약:
    #   - 반드시 본인의 감독 1명에게만 연결 가능
    #   - 다른 감독이나 복수 감독, 미할당은 불가
    if session.get('role') != 'admin':
        my_sup = session.get('supervisor_id')
        if not my_sup:
            return jsonify({'error': '담당 감독이 연결되지 않은 계정입니다. 관리자에게 요청하세요.'}), 403
        if sids != [my_sup]:
            return jsonify({'error': '본인 담당 감독으로만 선박을 추가할 수 있습니다.'}), 403

    vid = execute('''
        INSERT INTO vessels
          (name, short_name, vessel_type, imo, class_society, manager, manager_supervisor, active)
        VALUES (?, ?, ?, ?, ?, ?, ?, 1)
    ''', (name,
          (d.get('short_name') or name[:12]).strip(),
          d.get('vessel_type') or '',
          d.get('imo') or '',
          d.get('class_society') or '',
          d.get('manager') or '',
          str(d.get('manager_supervisor') or '').strip()))
    for sid in sids:
        execute('INSERT OR IGNORE INTO supervisor_vessels (vessel_id, supervisor_id) VALUES (?, ?)',
                (vid, sid))
    return jsonify({'id': vid}), 201


@bp.route('/api/vessels/<int:vid>', methods=['PUT'])
@login_required
def api_vessel_update(vid):
    if not query('SELECT id FROM vessels WHERE id=?', (vid,), one=True):
        abort(404)
    d = request.get_json(silent=True) or {}

    # 일반 사용자(member) 권한 제약:
    #   - 본인 담당 감독에 연결된 선박만 수정 가능
    #   - 담당 감독 변경(supervisor_ids), 비활성화(active) 는 불가
    if session.get('role') != 'admin':
        my_sup = session.get('supervisor_id')
        if not my_sup:
            return jsonify({'error': '담당 감독이 연결되지 않은 계정입니다.'}), 403
        owned = query(
            'SELECT 1 FROM supervisor_vessels WHERE vessel_id=? AND supervisor_id=?',
            (vid, my_sup), one=True,
        )
        if not owned:
            return jsonify({'error': '본인 담당 선박만 수정할 수 있습니다.'}), 403
        # 민감 필드는 서버에서 무시 (이중 방어)
        d.pop('supervisor_ids', None)
        d.pop('active', None)

    sets, params = [], []
    if 'manager_supervisor' in d:
        d['manager_supervisor'] = str(d.get('manager_supervisor') or '').strip()
    for f in ('name', 'short_name', 'vessel_type', 'imo', 'class_society',
              'manager', 'manager_supervisor', 'active'):
        if f in d:
            sets.append(f'{f} = ?')
            params.append(d[f])
    if sets:
        params.append(vid)
        execute(f'UPDATE vessels SET {", ".join(sets)} WHERE id = ?', params)
    # supervisor 매핑 갱신 (admin만 가능 — member는 위에서 pop됨)
    if 'supervisor_ids' in d:
        # 🔴 execute() 는 문장마다 커밋된다 — DELETE 를 먼저 날리고 나서 int(sid) 가
        #    터지면 담당 매핑이 복구 불가로 사라진다. 그래서 검증을 전부 앞에 둔다.
        #    문자열 "12" 가 오면 for 가 문자 단위로 돌던 문제도 여기서 같이 막는다.
        raw = d.get('supervisor_ids') or []
        if not isinstance(raw, (list, tuple)):
            return jsonify({'error': 'supervisor_ids 는 배열이어야 합니다.'}), 400
        try:
            sup_ids = [int(s) for s in raw]
        except (TypeError, ValueError):
            return jsonify({'error': 'supervisor_ids 값이 올바르지 않습니다.'}), 400
        execute('DELETE FROM supervisor_vessels WHERE vessel_id = ?', (vid,))
        for sid in sup_ids:
            execute('INSERT OR IGNORE INTO supervisor_vessels (vessel_id, supervisor_id) VALUES (?, ?)',
                    (vid, sid))
    return jsonify({'id': vid})


# ───── 선박 완전 삭제(purge) ─────────────────────────────────────
# 선박을 지우면 그 선박에 매달린 데이터를 전부 함께 지운다(2026-07-27 운영자 지시).
# soft delete(active=0) 폴백은 폐기 — 잔재가 남아 관리 목록이 지저분해지는 문제 때문.
# 되돌릴 수 없으므로 삭제 직전 전 대상 행을 JSON 으로 덤프해 둔다.
VESSEL_PURGE_BACKUP_DIR = os.path.join(INSTANCE_DIR, 'backups', 'vessel_purge')

# 형이 확정한 2계층 감사/결재 데이터만 명시적으로 purge 한다.
# 스키마 전체에서 vsl_cd 컬럼을 훑으면 의미가 다른 미래 테이블까지 삭제할 수 있어 금지한다.
# soa_review_case는 자식(line/attachment) 백업·삭제 순서 때문에 별도로 처리한다.
_PURGE_VSL_CD_TABLES = (
    'aor_draft', 'invoice_draft', 'fundreq_draft', 'reqgen_draft',
    'jeonja_review_item', 'soa_group_vessel', 'dock_procure',
    # dock_procure의 선박 헤더와 SOA 선주 매핑도 같은 vsl_cd 소유 데이터다.
    'dock_procure_vessel', 'soa_vessel_owner',
    # 조선소 견적 7카테고리. dock_procure와 한 세트라 빠지면 라인만 사라지고
    # 견적 금액이 남는다(UNIQUE(vsl_nm,category) 때문에 동명 선박 재등록 시 stale 값 상속).
    'dock_yard',
)


def _purge_vsl_cd_tables():
    """배포 중인 구스키마와의 호환을 위해 존재하는 manifest 테이블만 반환한다."""
    existing = {r['name'] for r in query(
        "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
    )}
    return [t for t in _PURGE_VSL_CD_TABLES if t in existing]


def _purge_file_allowed(path):
    """DB 값이 손상돼도 허용된 첨부 디렉터리 밖 파일은 절대 건드리지 않게 한다."""
    if not path:
        return False
    real = os.path.realpath(path)
    for root in (UPLOAD_DIR, SOA_REVIEW_PDF_DIR):
        r = os.path.realpath(root)
        if real == r or real.startswith(r + os.sep):
            return True
    return False


def _vessel_own_tokens(v):
    """선박 하나가 주장하는 코드/이름 토큰."""
    codes, names = set(), set()
    # 2계층 결재 데이터는 형이 지정한 대로 vsl_cd로만 연결한다.
    # short_name을 대체 키로 쓰면 선박 별칭과 우연히 겹쳐 다른 선박 감사기록을 지울 수 있다.
    val = ((v['vsl_cd'] if 'vsl_cd' in v.keys() else None) or '').strip()
    if val:
        codes.add(val)
    nm = (v['name'] or '').strip()
    if nm:
        names.add(nm)
    try:
        for a in json.loads((v['aliases'] if 'aliases' in v.keys() else None) or '[]'):
            if isinstance(a, str) and a.strip():
                names.add(a.strip())
    except Exception:
        pass
    return codes, names


def _vessel_purge_codes(v):
    """
    이 선박만 가리키는 코드/이름. 다른 선박도 쓰는 토큰은 제외한다.

    2계층은 vsl_cd만 허용한다. 같은 코드 또는 같은 선명/별칭을 다른 선박도 쓰면
    해당 토큰을 제외해 다른 선박 데이터를 지우지 않는다.
    """
    codes, names = _vessel_own_tokens(v)
    others_c, others_n = set(), set()
    for o in query('SELECT * FROM vessels WHERE id<>?', (v['id'],)):
        oc, on = _vessel_own_tokens(o)
        others_c |= oc
        others_n |= on
    shared = sorted((codes & others_c) | (names & others_n))
    return sorted(codes - others_c), sorted(names - others_n), shared


def _vessel_purge_scan(vid):
    """삭제 대상을 훑어 (건수, 백업용 행, 지울 파일경로) 를 만든다. 쓰기는 하지 않는다."""
    v = query('SELECT * FROM vessels WHERE id=?', (vid,), one=True)
    if not v:
        return None
    codes, names, shared = _vessel_purge_codes(v)
    cph = ','.join('?' * len(codes)) if codes else None
    nph = ','.join('?' * len(names)) if names else None

    counts, backup, files = {}, {'vessel': dict(v)}, []

    def grab(key, sql, params=()):
        rows = [dict(r) for r in query(sql, params)]
        if rows:
            # 같은 키로 두 번 담기면 덮어쓰지 않고 합친다(건수 누락 방지).
            counts[key] = counts.get(key, 0) + len(rows)
            backup.setdefault(key, []).extend(rows)
        return rows

    def grab_children(key, table, fk, parent_ids):
        """CASCADE 로 조용히 사라질 자식까지 백업에 담는다."""
        if not parent_ids:
            return []
        ph = ','.join('?' * len(parent_ids))
        return grab(key, f'SELECT * FROM "{table}" WHERE {fk} IN ({ph})', parent_ids)

    # 1) vessel_id 로 직접 엮인 운영 데이터 + 첨부파일
    issues = grab('issues', 'SELECT * FROM issues WHERE vessel_id=?', (vid,))
    if issues:
        iph = ','.join('?' * len(issues))
        iids = [r['id'] for r in issues]
        for a in grab('attachments',
                      f'SELECT * FROM attachments WHERE issue_id IN ({iph})', iids):
            files.append(os.path.join(UPLOAD_DIR, a['stored_name']))

    surveys = grab('cs_surveys', 'SELECT * FROM cs_surveys WHERE vessel_id=?', (vid,))
    if surveys:
        sph = ','.join('?' * len(surveys))
        sids = [r['id'] for r in surveys]
        grab('cs_findings', f'SELECT * FROM cs_findings WHERE survey_id IN ({sph})', sids)
        for a in grab('cs_attachments',
                      f'SELECT * FROM cs_attachments WHERE survey_id IN ({sph})', sids):
            files.append(os.path.join(UPLOAD_DIR, a['stored_name']))

    vets = grab('vettings', 'SELECT * FROM vettings WHERE vessel_id=?', (vid,))
    if vets:
        vph = ','.join('?' * len(vets))
        vids = [r['id'] for r in vets]
        grab('vt_findings', f'SELECT * FROM vt_findings WHERE vetting_id IN ({vph})', vids)
        for a in grab('vt_attachments',
                      f'SELECT * FROM vt_attachments WHERE vetting_id IN ({vph})', vids):
            files.append(os.path.join(UPLOAD_DIR, a['stored_name']))

    # Dock/Boarding Report 본문 이미지는 'dock-<report_id>-*' / 'brep-<report_id>-*' 규칙
    for key, sql, prefix, sub in (
        ('dock_reports', 'SELECT * FROM dock_reports WHERE vessel_id=?', 'dock', 'dock'),
        ('boarding_reports', 'SELECT * FROM boarding_reports WHERE vessel_id=?', 'brep', 'boarding'),
    ):
        reports = grab(key, sql, (vid,))
        for r in reports:
            d = os.path.join(UPLOAD_DIR, sub)
            if os.path.isdir(d):
                files += [os.path.join(d, f) for f in os.listdir(d)
                          if f.startswith(f'{prefix}-{r["id"]}-')]
        base = key[:-1]  # dock_reports -> dock_report
        secs = grab_children(f'{base}_sections', f'{base}_sections', 'report_id',
                             [r['id'] for r in reports])
        grab_children(f'{base}_blocks', f'{base}_blocks', 'section_id',
                      [s['id'] for s in secs])

    cstat = grab('class_status', 'SELECT * FROM class_status WHERE vessel_id=?', (vid,))
    grab_children('class_status_items', 'class_status_items', 'cs_id', [r['id'] for r in cstat])
    grab('calendar_events', 'SELECT * FROM calendar_events WHERE vessel_id=?', (vid,))
    grab('supervisor_vessels', 'SELECT * FROM supervisor_vessels WHERE vessel_id=?', (vid,))

    # 2) vsl_cd 로 엮인 결재·정산·구매 데이터
    if cph:
        for t in _purge_vsl_cd_tables():
            grab(t, f'SELECT * FROM "{t}" WHERE vsl_cd IN ({cph})', codes)
        cases = grab('soa_review_case', f'SELECT * FROM soa_review_case WHERE vsl_cd IN ({cph})', codes)
        if cases:
            xph = ','.join('?' * len(cases))
            xids = [r['id'] for r in cases]
            grab('soa_review_line', f'SELECT * FROM soa_review_line WHERE case_id IN ({xph})', xids)
            for a in grab('soa_review_attachment',
                          f'SELECT * FROM soa_review_attachment WHERE case_id IN ({xph})', xids):
                p = _soa_review_attachment_path(a.get('stored_name'))
                if p:
                    files.append(p)

    # 3) 이름/키로 엮인 캐시성 데이터
    vkey = _vkey(v['name'])
    for t in ('fleet_eta_override', 'fleet_next_port_override'):
        grab(t, f'SELECT * FROM {t} WHERE vessel_key=?', (vkey,))
    if nph:
        grab('mail_card', f'SELECT * FROM mail_card WHERE issue_vessel IN ({nph})', names)
        grab('shipwiki_card', f'SELECT * FROM shipwiki_card WHERE ship_nm IN ({nph})', names)

    # 4) 사용자별 선박 정렬 순서 — 지우진 않고 항목만 빼지만, 되돌리려면 원본이 필요하다.
    order_rows = [dict(r) for r in query('SELECT * FROM user_vessel_order')]
    touched = []
    for row in order_rows:
        try:
            order = json.loads(row.get('order_json') or '[]')
        except Exception:
            continue
        if any(str(x) == str(vid) for x in order):
            touched.append(row)
    if touched:
        backup['user_vessel_order'] = touched

    return {'vessel': v, 'codes': codes, 'names': names, 'shared': shared, 'vkey': vkey,
            'counts': counts, 'backup': backup,
            # DB 값이 손상되었거나 symlink여도 허용 첨부 디렉터리 밖 파일은 삭제 대상에서 제외한다.
            'files': sorted({f for f in files if _purge_file_allowed(f)})}


@bp.route('/api/vessels/<int:vid>/delete-impact')
@admin_required
def api_vessel_delete_impact(vid):
    """삭제 확인창에 보여줄 '함께 지워질 데이터' 건수."""
    plan = _vessel_purge_scan(vid)
    if not plan:
        abort(404)
    return jsonify({
        'vessel': {'id': plan['vessel']['id'], 'name': plan['vessel']['name']},
        'counts': plan['counts'],
        'files': len(plan['files']),
        'total': sum(plan['counts'].values()),
    })


@bp.route('/api/vessels/<int:vid>', methods=['DELETE'])
@login_required
def api_vessel_delete(vid):
    v = query('SELECT id FROM vessels WHERE id=?', (vid,), one=True)
    if not v:
        abort(404)

    # 일반 사용자(member): 담당 해제만 가능. 관련 데이터를 통째로 지우는 purge 는 admin 전용.
    if session.get('role') != 'admin':
        my_sup = session.get('supervisor_id')
        if not my_sup:
            return jsonify({'error': '담당 감독이 연결되지 않은 계정입니다.'}), 403
        owned = query(
            'SELECT 1 FROM supervisor_vessels WHERE vessel_id=? AND supervisor_id=?',
            (vid, my_sup), one=True,
        )
        if not owned:
            return jsonify({'error': '본인 담당 선박만 삭제할 수 있습니다.'}), 403
        execute('DELETE FROM supervisor_vessels WHERE vessel_id=? AND supervisor_id=?',
                (vid, my_sup))
        return jsonify({'ok': True, 'unassigned_only': True})

    # 조회·백업·모든 DELETE를 같은 SQLite snapshot으로 묶어 중간 실패 시 전부 rollback한다.
    db = get_db()
    db.execute('BEGIN IMMEDIATE')
    g._vessel_purge_transaction = True
    plan = _vessel_purge_scan(vid)
    if not plan:
        db.rollback()
        g.pop('_vessel_purge_transaction', None)
        abort(404)

    # 되돌릴 수 있게 삭제 직전 스냅샷을 남긴다. 덤프 실패 시 삭제하지 않는다.
    backup_path = None
    try:
        os.makedirs(VESSEL_PURGE_BACKUP_DIR, exist_ok=True)
        stamp = datetime.now().strftime('%Y%m%d-%H%M%S')
        safe = re.sub(r'[^A-Za-z0-9_-]+', '_', plan['vessel']['name'] or str(vid))[:40]
        backup_path = os.path.join(VESSEL_PURGE_BACKUP_DIR, f'{stamp}-{vid}-{safe}.json')
        with open(backup_path, 'w', encoding='utf-8') as fp:
            json.dump({'deleted_at': stamp, 'deleted_by': session.get('username'),
                       'codes': plan['codes'], 'names': plan['names'],
                       'counts': plan['counts'], 'files': plan['files'],
                       'rows': plan['backup']}, fp, ensure_ascii=False, default=str)
    except Exception as e:
        db.rollback()
        g.pop('_vessel_purge_transaction', None)
        app.logger.exception('vessel purge 백업 실패 vid=%s', vid)
        return jsonify({'error': f'삭제 전 백업에 실패해 중단했습니다: {e}'}), 500

    codes, names, vkey = plan['codes'], plan['names'], plan['vkey']
    cph = ','.join('?' * len(codes)) if codes else None
    nph = ','.join('?' * len(names)) if names else None

    # 자식 → 부모 순서로 지운다. CASCADE 가 걸린 자식은 부모 삭제로 함께 정리된다.
    execute('DELETE FROM issues WHERE vessel_id=?', (vid,))            # attachments CASCADE
    execute('DELETE FROM cs_surveys WHERE vessel_id=?', (vid,))        # findings/attach CASCADE
    execute('DELETE FROM vettings WHERE vessel_id=?', (vid,))          # findings/attach CASCADE
    execute('DELETE FROM dock_reports WHERE vessel_id=?', (vid,))      # sections/blocks CASCADE
    execute('DELETE FROM boarding_reports WHERE vessel_id=?', (vid,))  # sections/blocks CASCADE
    execute('DELETE FROM class_status WHERE vessel_id=?', (vid,))      # items CASCADE
    execute('DELETE FROM calendar_events WHERE vessel_id=?', (vid,))
    execute('DELETE FROM supervisor_vessels WHERE vessel_id=?', (vid,))
    if cph:
        for t in _purge_vsl_cd_tables():
            execute(f'DELETE FROM "{t}" WHERE vsl_cd IN ({cph})', codes)
        execute(f'DELETE FROM soa_review_case WHERE vsl_cd IN ({cph})', codes)  # line/attach CASCADE
    for t in ('fleet_eta_override', 'fleet_next_port_override'):
        execute(f'DELETE FROM {t} WHERE vessel_key=?', (vkey,))
    if nph:
        execute(f'DELETE FROM mail_card WHERE issue_vessel IN ({nph})', names)
        execute(f'DELETE FROM shipwiki_card WHERE ship_nm IN ({nph})', names)
    execute('DELETE FROM vessels WHERE id=?', (vid,))

    # 사용자별 선박 정렬 순서에서도 제거
    for row in query('SELECT user_id, order_json FROM user_vessel_order'):
        try:
            order = json.loads(row['order_json'] or '[]')
        except Exception:
            continue
        pruned = [x for x in order if str(x) != str(vid)]
        if len(pruned) != len(order):
            execute('UPDATE user_vessel_order SET order_json=?, '
                    'updated_at=datetime("now","localtime") WHERE user_id=?',
                    (json.dumps(pruned), row['user_id']))

    # 이 지점까지의 모든 SQL을 한 번에 확정한다. 그 전 예외는 teardown rollback으로 남는 행이 없다.
    db.commit()
    g.pop('_vessel_purge_transaction', None)

    # DB 정리가 끝난 뒤 파일 삭제 — 파일 실패가 트랜잭션을 되돌리지 않게.
    removed = 0
    for p in plan['files']:
        try:
            if p and os.path.exists(p):
                os.remove(p)
                removed += 1
        except OSError:
            app.logger.warning('vessel purge 파일 삭제 실패 vid=%s %s', vid, p)

    app.logger.info('vessel purge vid=%s name=%s rows=%s files=%s backup=%s',
                    vid, plan['vessel']['name'], sum(plan['counts'].values()),
                    removed, backup_path)
    return jsonify({'ok': True, 'purged': True, 'counts': plan['counts'],
                    'rows': sum(plan['counts'].values()), 'files_removed': removed,
                    'backup': os.path.basename(backup_path) if backup_path else None})


# ----- 사용자 (admin 전용 CRUD) -----
@bp.route('/api/users')
@admin_required
def api_users_list():
    rows = query('''
        SELECT u.id, u.username, u.display_name, u.role, u.supervisor_id, u.active,
               u.created_at, u.last_login_at,
               s.name AS supervisor_name
          FROM users u
          LEFT JOIN supervisors s ON s.id = u.supervisor_id
         ORDER BY u.active DESC, u.role DESC, u.id
    ''')
    return jsonify([dict(r) for r in rows])


@bp.route('/api/users', methods=['POST'])
@admin_required
def api_user_create():
    d = request.get_json(silent=True) or {}
    username = (d.get('username') or '').strip()
    password = d.get('password') or ''
    if not username:
        return jsonify({'error': '사용자명은 필수입니다.'}), 400
    if len(password) < 6:
        return jsonify({'error': '비밀번호는 6자 이상이어야 합니다.'}), 400
    if query('SELECT id FROM users WHERE username=?', (username,), one=True):
        return jsonify({'error': '이미 사용 중인 사용자명입니다.'}), 400
    role = d.get('role') or 'member'
    if role not in ('admin', 'member'):
        role = 'member'
    uid = execute('''
        INSERT INTO users (username, password_hash, display_name, role, supervisor_id, active)
        VALUES (?, ?, ?, ?, ?, 1)
    ''', (username, generate_password_hash(password),
          d.get('display_name') or username,
          role,
          d.get('supervisor_id') or None))
    return jsonify({'id': uid}), 201


@bp.route('/api/users/<int:uid>', methods=['PUT'])
@admin_required
def api_user_update(uid):
    if not query('SELECT id FROM users WHERE id=?', (uid,), one=True):
        abort(404)
    d = request.get_json(silent=True) or {}
    sets, params = [], []
    for f in ('display_name', 'role', 'supervisor_id', 'active'):
        if f in d:
            sets.append(f'{f} = ?')
            params.append(d[f])
    if not sets:
        return jsonify({'error': '수정할 필드 없음'}), 400
    params.append(uid)
    execute(f'UPDATE users SET {", ".join(sets)} WHERE id = ?', params)
    return jsonify({'id': uid})


@bp.route('/api/users/<int:uid>', methods=['DELETE'])
@admin_required
def api_user_delete(uid):
    if uid == session.get('user_id'):
        return jsonify({'error': '자기 자신은 삭제할 수 없습니다.'}), 400
    # admin 계정이 하나만 남을 땐 삭제 금지
    u = query('SELECT role FROM users WHERE id=?', (uid,), one=True)
    if not u:
        abort(404)
    if u['role'] == 'admin':
        n = query("SELECT COUNT(*) AS n FROM users WHERE role='admin' AND active=1 AND id<>?",
                  (uid,), one=True)['n']
        if n == 0:
            return jsonify({'error': '최소 1명의 관리자 계정은 유지되어야 합니다.'}), 400
    execute('UPDATE users SET active=0 WHERE id=?', (uid,))
    return jsonify({'ok': True})


@bp.route('/api/users/<int:uid>/password', methods=['POST'])
@admin_required
def api_user_reset_password(uid):
    d = request.get_json(silent=True) or {}
    new = d.get('new_password') or ''
    if len(new) < 6:
        return jsonify({'error': '비밀번호는 6자 이상이어야 합니다.'}), 400
    if not query('SELECT id FROM users WHERE id=?', (uid,), one=True):
        abort(404)
    execute('UPDATE users SET password_hash=? WHERE id=?',
            (generate_password_hash(new), uid))
    return jsonify({'ok': True})


# ═════════════════════════════════════════════════════════════════
#  API — Condition Survey
# ═════════════════════════════════════════════════════════════════

def _cs_survey_with_counts(s):
    """단일 survey에 카운트 컬럼들 포함시켜 반환 (dict).
    manual_*_count 가 NULL이 아니면 수동 입력값을 우선."""
    sid = s['id']
    rows = query("""
        SELECT category, status, COUNT(*) AS n
          FROM cs_findings
         WHERE survey_id = ?
         GROUP BY category, status
    """, (sid,))
    def_open = def_closed = obs_open = obs_closed = 0
    for r in rows:
        if r['category'] == 'Defect':
            if r['status'] == 'Closed': def_closed = r['n']
            else: def_open = r['n']
        else:
            if r['status'] == 'Closed': obs_closed = r['n']
            else: obs_open = r['n']
    auto_def   = def_open + def_closed
    auto_obs   = obs_open + obs_closed
    auto_close = def_closed + obs_closed

    d = dict(s)
    # 수동 override가 있으면 그 값을, 없으면 자동 카운트
    d['defect_count']      = s['manual_defect_count']      if s['manual_defect_count']      is not None else auto_def
    d['observation_count'] = s['manual_observation_count'] if s['manual_observation_count'] is not None else auto_obs
    d['close_count']       = s['manual_close_count']       if s['manual_close_count']       is not None else auto_close
    d['total_count']       = d['defect_count'] + d['observation_count']
    # Open 카운트는 항상 자동 (전체 - 완료)
    d['open_count']        = max(0, d['total_count'] - d['close_count'])
    # manual flag (UI에서 자동/수동 구분)
    d['defect_manual']      = s['manual_defect_count']      is not None
    d['observation_manual'] = s['manual_observation_count'] is not None
    d['close_manual']       = s['manual_close_count']       is not None
    # 첨부 카운트
    ar = query('SELECT COUNT(*) AS n FROM cs_attachments WHERE survey_id=?',
               (sid,), one=True)
    d['attach_count'] = ar['n'] if ar else 0
    return d


@bp.route('/api/cs/surveys')
@login_required
def api_cs_surveys_list():
    """연도 + (선택)감독별 모든 선박의 분기별 서베이 목록.
    응답 구조: [{vessel: {...}, surveys: {1: {...}, 2: {...}}}]"""
    try:
        year = int(request.args.get('year') or 2026)
    except (TypeError, ValueError):
        year = 2026
    sup_id = request.args.get('supervisor_id')

    # 선박 목록 — 감독 필터 적용
    if sup_id and sup_id != 'all':
        vessels = query("""
            SELECT v.* FROM vessels v
              JOIN supervisor_vessels sv ON sv.vessel_id = v.id
             WHERE v.active = 1 AND sv.supervisor_id = ?
             ORDER BY v.name
        """, (sup_id,))
    else:
        vessels = query('SELECT * FROM vessels WHERE active=1 ORDER BY name')

    # 해당 연도의 모든 서베이 한번에
    surveys = query('SELECT * FROM cs_surveys WHERE year = ?', (year,))

    # 한번에 findings 모두 가져와서 survey_id 별로 매핑 (N+1 회피)
    sids = [s['id'] for s in surveys]
    findings_by_sid = {sid: [] for sid in sids}
    if sids:
        placeholders = ','.join('?' * len(sids))
        all_findings = query(
            f'SELECT * FROM cs_findings WHERE survey_id IN ({placeholders}) ORDER BY survey_id, category, no',
            tuple(sids),
        )
        for f in all_findings:
            findings_by_sid[f['survey_id']].append(dict(f))

    by_vessel = {}
    for s in surveys:
        d = _cs_survey_with_counts(s)
        d['findings'] = findings_by_sid.get(s['id'], [])
        by_vessel.setdefault(s['vessel_id'], {})[s['quarter']] = d

    # 선박별 last_updated (해당 선박의 모든 surveys 중 가장 최근 updated_at)
    last_by_vessel = {}
    for s in surveys:
        u = s['updated_at']
        if u and (s['vessel_id'] not in last_by_vessel or u > last_by_vessel[s['vessel_id']]):
            last_by_vessel[s['vessel_id']] = u

    out = []
    for v in vessels:
        out.append({
            'vessel': dict(v),
            'surveys': by_vessel.get(v['id'], {}),
            'last_updated': last_by_vessel.get(v['id']),
        })
    return jsonify(out)


@bp.route('/api/cs/surveys', methods=['POST'])
@login_required
def api_cs_survey_create():
    """헤더(분기 셀) 생성 또는 upsert."""
    d = request.get_json(silent=True) or {}
    vid = d.get('vessel_id'); year = d.get('year'); q = d.get('quarter')
    if not (vid and year and q in (1,2,3,4)):
        return jsonify({'error': 'vessel_id, year, quarter 필수'}), 400
    if not query('SELECT id FROM vessels WHERE id=?', (vid,), one=True):
        return jsonify({'error': '선박 없음'}), 404

    existing = query(
        'SELECT id FROM cs_surveys WHERE vessel_id=? AND year=? AND quarter=?',
        (vid, year, q), one=True,
    )
    if existing:
        return jsonify({'id': existing['id'], 'existed': True})

    sid = execute("""
        INSERT INTO cs_surveys
            (vessel_id, year, quarter, vendor, management, inspection_date,
             overall_remark, created_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (vid, year, q,
          d.get('vendor') or None,
          d.get('management') or None,
          d.get('inspection_date') or None,
          d.get('overall_remark') or None,
          session.get('username')))
    return jsonify({'id': sid}), 201


@bp.route('/api/cs/surveys/<int:sid>', methods=['GET'])
@login_required
def api_cs_survey_get(sid):
    s = query('SELECT * FROM cs_surveys WHERE id=?', (sid,), one=True)
    if not s: abort(404)
    d = _cs_survey_with_counts(s)
    findings = query(
        "SELECT * FROM cs_findings WHERE survey_id=? ORDER BY category, no",
        (sid,),
    )
    d['findings'] = [dict(f) for f in findings]
    return jsonify(d)


@bp.route('/api/cs/surveys/<int:sid>', methods=['PUT'])
@login_required
def api_cs_survey_update(sid):
    if not query('SELECT id FROM cs_surveys WHERE id=?', (sid,), one=True):
        abort(404)
    d = request.get_json(silent=True) or {}
    sets, params = [], []
    for f in ('vendor','management','inspection_date','overall_remark',
              'manual_defect_count','manual_observation_count','manual_close_count'):
        if f in d:
            sets.append(f'{f} = ?')
            v = d[f]
            # 빈 문자열은 NULL로 저장 (자동 카운트로 복귀)
            params.append(None if v == '' else v)
    if not sets:
        return jsonify({'error': '수정할 필드 없음'}), 400
    sets.append("updated_at = datetime('now','localtime')")
    params.append(sid)
    execute(f'UPDATE cs_surveys SET {", ".join(sets)} WHERE id = ?', params)
    return jsonify({'id': sid})


@bp.route('/api/cs/surveys/<int:sid>', methods=['DELETE'])
@login_required
def api_cs_survey_delete(sid):
    execute('DELETE FROM cs_surveys WHERE id=?', (sid,))
    return jsonify({'ok': True})


# ----- Findings (세부 항목) -----

def _next_finding_no(survey_id, category):
    r = query(
        'SELECT COALESCE(MAX(no), 0) + 1 AS n FROM cs_findings WHERE survey_id=? AND category=?',
        (survey_id, category), one=True,
    )
    return r['n']


@bp.route('/api/cs/surveys/<int:sid>/findings', methods=['POST'])
@login_required
def api_cs_finding_create(sid):
    """단건 또는 배치(엑셀 붙여넣기) 추가.
    body: { category: 'Defect'|'Observation', items: [{description,remark,status},...] }
    또는 단건: { category, description, remark, status }
    """
    if not query('SELECT id FROM cs_surveys WHERE id=?', (sid,), one=True):
        abort(404)
    d = request.get_json(silent=True) or {}
    cat = d.get('category')
    if cat not in ('Defect','Observation'):
        return jsonify({'error': "category는 Defect 또는 Observation"}), 400

    items = d.get('items')
    if items is None:
        items = [{
            'item':        d.get('item'),
            'description': d.get('description'),
            'remark':      d.get('remark'),
            'status':      d.get('status') or 'Open',
        }]

    next_no = _next_finding_no(sid, cat)
    created_ids = []
    for it in items:
        st = it.get('status') or 'Open'
        if st not in ('Open','Closed'): st = 'Open'
        fid = execute("""
            INSERT INTO cs_findings (survey_id, category, no, item, description, remark, status)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (sid, cat, next_no,
              it.get('item') or '',
              it.get('description') or '',
              it.get('remark') or '',
              st))
        created_ids.append(fid)
        next_no += 1
    return jsonify({'ids': created_ids, 'count': len(created_ids)}), 201


@bp.route('/api/cs/findings/<int:fid>', methods=['PUT'])
@login_required
def api_cs_finding_update(fid):
    cur = query('SELECT survey_id, status FROM cs_findings WHERE id=?', (fid,), one=True)
    if not cur:
        abort(404)
    d = request.get_json(silent=True) or {}
    sets, params = [], []
    for f in ('item','description','remark','status'):
        if f in d:
            sets.append(f'{f} = ?')
            params.append(d[f])
    if not sets:
        return jsonify({'error': '수정할 필드 없음'}), 400
    sets.append("updated_at = datetime('now','localtime')")
    params.append(fid)
    execute(f'UPDATE cs_findings SET {", ".join(sets)} WHERE id = ?', params)

    # status 변경 시 cs_surveys.updated_at 갱신 (선박 헤더의 Last update에 반영)
    if 'status' in d and d['status'] != cur['status']:
        execute(
            "UPDATE cs_surveys SET updated_at = datetime('now','localtime') WHERE id=?",
            (cur['survey_id'],),
        )
    return jsonify({'id': fid})


@bp.route('/api/cs/findings/<int:fid>', methods=['DELETE'])
@login_required
def api_cs_finding_delete(fid):
    f = query('SELECT survey_id, category, no FROM cs_findings WHERE id=?', (fid,), one=True)
    if not f: abort(404)
    execute('DELETE FROM cs_findings WHERE id=?', (fid,))
    # No 재정렬: 같은 survey + category 내에서
    rows = query(
        'SELECT id FROM cs_findings WHERE survey_id=? AND category=? ORDER BY no, id',
        (f['survey_id'], f['category']),
    )
    for idx, r in enumerate(rows, 1):
        execute('UPDATE cs_findings SET no=? WHERE id=?', (idx, r['id']))
    return jsonify({'ok': True})
