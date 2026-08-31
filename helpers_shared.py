"""Shared helpers used by more than one boundary (loaded first).

Every symbol here is consumed by two or more boundaries, or by ``app.py``
itself.  Extracting them breaks the boundary-to-boundary dependency cycle:
after this file, each ``routes_*``/``ai_gemini`` boundary may depend only on
``app.py`` and this module — never on a sibling boundary.  That layering is
enforced by ``tests/test_boundary_dependency_graph.py``.

Sections preserve the original per-file order, and the file order matches the
old loader order, so every load-time reference that worked before the
extraction still resolves.  Since 2026-08-12 this file is a real imported
module: it takes its primitives from ``app_core`` (config, the Flask app and
the DB helpers), which is what removed the ``app.py`` ↔ helpers cycle that
forced the old ``exec`` loading.
"""

import json
import math
import os
import re
import secrets
import sqlite3
import threading
import uuid
from datetime import date, datetime, timedelta
from functools import wraps

from flask import abort, g, jsonify, redirect, request, session, url_for
from werkzeug.utils import secure_filename

from app_core import (
    ALLOWED_EXT, BASE_DIR, DATABASE, DOCKATT_FILE_DIR, INSTANCE_DIR,
    SOA_REVIEW_PDF_DIR, UPLOAD_DIR, _NON_STT_UPLOAD_MAX, app,
    execute, get_db, query,
)


# ══════════════════════════════════════════════════════════════════
#  From routes_core.py (8 symbols)
# ══════════════════════════════════════════════════════════════════
def _session_account():
    """쿠키 세션의 계정을 요청마다 DB 로 재확인한다. (row 또는 None)

    🔴 Bearer 경로(app.py)는 요청마다 `active=1` + 비번지문을 검사하는데 쿠키 경로만
       그 검사가 없었다. 그래서 `active=0` 으로 비활성화한 계정이 이미 열려 있던 브라우저
       세션으로는 PERMANENT_SESSION_LIFETIME(7일) 동안 계속 통과했다 — 퇴사/사고 대응으로
       계정을 껐는데 실제로는 안 꺼지는 구멍이라 여기서 막는다.
    요청당 1회만 조회한다(users.id 는 PK).
    """
    uid = session.get('user_id')
    if not uid:
        return None
    if not hasattr(g, '_sess_account'):
        try:
            g._sess_account = query(
                'SELECT id, role, active FROM users WHERE id=?', (uid,), one=True)
        except Exception:
            # DB 조회 자체가 실패하면 인증을 통과시키지 않는다(fail-closed).
            app.logger.exception('session account 재확인 실패 uid=%s', uid)
            g._sess_account = None
    return g._sess_account


def _session_account_ok():
    u = _session_account()
    return bool(u and u['active'] == 1)


def login_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if 'user_id' not in session or not _session_account_ok():
            session.clear()             # 비활성화된 계정의 잔여 쿠키를 여기서 끊는다
            if request.path.startswith('/api/'):
                return jsonify({'error': 'unauthorized'}), 401
            return redirect(url_for('routes_core.login', next=request.path))
        return f(*args, **kwargs)
    return wrapped
def admin_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if 'user_id' not in session or not _session_account_ok():
            session.clear()
            return jsonify({'error': 'unauthorized'}), 401
        # 🔴 권한은 세션에 굳은 값이 아니라 DB 를 본다 — admin 을 member 로 강등해도
        #    기존 세션이 살아 있는 동안 admin API 가 계속 열려 있던 문제(권한 상승 잔존).
        acct = _session_account()       # 위 검사를 통과했으므로 None 이 아니다
        if session.get('role') != 'admin' or acct['role'] != 'admin':
            return jsonify({'error': 'forbidden'}), 403
        return f(*args, **kwargs)
    return wrapped
def _dashboard_ctx():
    """대시보드 집계 컨텍스트(stats/events/scope) — Fleet Map 상단 KPI 스트립과
    구 카드형(/dashboard/classic) 양쪽에서 공유."""
    today   = date.today().isoformat()
    horizon = (date.today() + timedelta(days=30)).isoformat()
    cal_end = (date.today() + timedelta(days=7)).isoformat()
    is_admin = (session.get('role') == 'admin')

    sup_id = session.get('supervisor_id')
    scoped = bool(sup_id)
    sup_name = None
    vessel_ids = []
    if scoped:
        srow = query("SELECT name FROM supervisors WHERE id=?", (sup_id,), one=True)
        sup_name = srow['name'] if srow else None
        vessel_ids = [r['vessel_id'] for r in
                      query("SELECT vessel_id FROM supervisor_vessels WHERE supervisor_id=?", (sup_id,))]

    def vin(col):
        """담당선박 IN 절. 미연결=전체(1=1), 연결+선박없음=0건(0=1)."""
        if not scoped:
            return ("1=1", [])
        if not vessel_ids:
            return ("0=1", [])
        return (f"{col} IN ({','.join('?' * len(vessel_ids))})", list(vessel_ids))

    # 1) 현안 요약 — 감독 연결 시 그 감독 이슈만(issues.supervisor_id)
    iss_where = "WHERE supervisor_id=?" if scoped else ""
    iss_params = (sup_id,) if scoped else ()
    iss = query(
        "SELECT "
        "SUM(CASE WHEN status!='Closed' THEN 1 ELSE 0 END) open_cnt, "
        "SUM(CASE WHEN status!='Closed' AND priority='Urgent' THEN 1 ELSE 0 END) urgent_cnt, "
        "SUM(CASE WHEN status!='Closed' AND priority='COC & Flag' THEN 1 ELSE 0 END) coc_cnt, "
        "SUM(CASE WHEN status!='Closed' AND priority='Next DD' THEN 1 ELSE 0 END) dd_cnt "
        f"FROM issues {iss_where}", iss_params, one=True)

    # 2) Class 만기 임박 (due_date D-30, 담당선박)
    cvf, cvp = vin("cs.vessel_id")
    class_due = query(
        "SELECT COUNT(*) c FROM class_status_items i JOIN class_status cs ON cs.id=i.cs_id "
        "WHERE i.due_date IS NOT NULL AND i.due_date != '' "
        f"AND i.due_date >= ? AND i.due_date <= ? AND {cvf}",
        (today, horizon, *cvp), one=True)['c']

    # 3) Vetting 미해결 (Open observation, 담당선박)
    vvf, vvp = vin("vt.vessel_id")
    vrow = query(
        "SELECT "
        "SUM(CASE WHEN f.status='Open' THEN 1 ELSE 0 END) open_cnt, "
        "SUM(CASE WHEN f.status='Open' AND f.priority=1 THEN 1 ELSE 0 END) pri_cnt "
        "FROM vt_findings f JOIN vettings vt ON vt.id=f.vetting_id "
        f"WHERE {vvf}", (*vvp,), one=True)

    # 4) 다가오는 일정 (7일) — 담당선박/본인/공용
    if scoped:
        evf, evp = vin("vessel_id")
        events = query(
            "SELECT title, start_date, category, color, completed FROM calendar_events "
            "WHERE start_date >= ? AND start_date <= ? "
            f"AND (supervisor_id=? OR supervisor_id IS NULL OR {evf}) "
            "ORDER BY start_date ASC, COALESCE(start_time,'') ASC LIMIT 8",
            (today, cal_end, sup_id, *evp))
    else:
        events = query(
            "SELECT title, start_date, category, color, completed FROM calendar_events "
            "WHERE start_date >= ? AND start_date <= ? "
            "ORDER BY start_date ASC, COALESCE(start_time,'') ASC LIMIT 8",
            (today, cal_end))

    # 7일 일정 총건수(KPI 스트립용) — events 는 LIMIT 8 미리보기라 카운트와 분리.
    if scoped:
        evf2, evp2 = vin("vessel_id")
        events_count = query(
            "SELECT COUNT(*) c FROM calendar_events WHERE start_date >= ? AND start_date <= ? "
            f"AND (supervisor_id=? OR supervisor_id IS NULL OR {evf2})",
            (today, cal_end, sup_id, *evp2), one=True)['c']
    else:
        events_count = query(
            "SELECT COUNT(*) c FROM calendar_events WHERE start_date >= ? AND start_date <= ?",
            (today, cal_end), one=True)['c']

    # 오늘 일정(KPI 스트립 = 당일 요약, 손유석 지시 2026-06-29). start_date=오늘만.
    if scoped:
        evf3, evp3 = vin("vessel_id")
        today_events = query(
            "SELECT title, category, start_time, completed FROM calendar_events WHERE start_date = ? "
            f"AND (supervisor_id=? OR supervisor_id IS NULL OR {evf3}) "
            "ORDER BY COALESCE(start_time,'') ASC", (today, sup_id, *evp3))
    else:
        today_events = query(
            "SELECT title, category, start_time, completed FROM calendar_events WHERE start_date = ? "
            "ORDER BY COALESCE(start_time,'') ASC", (today,))
    today_count = len(today_events)

    stats = {
        'issues_open':   (iss['open_cnt']   or 0) if iss else 0,
        'issues_urgent': (iss['urgent_cnt'] or 0) if iss else 0,
        'issues_coc':    (iss['coc_cnt']    or 0) if iss else 0,
        'issues_dd':     (iss['dd_cnt']     or 0) if iss else 0,
        'class_due':     class_due,
        'vetting_open':  (vrow['open_cnt'] or 0) if vrow else 0,
        'vetting_pri':   (vrow['pri_cnt']  or 0) if vrow else 0,
        'aor_pending':   0,
        'aor_crew_submitted': 0,
    }
    # 자동화 위젯은 admin 만 (탭 자체가 admin 전용) — 전사 큐라 감독 스코프 무관
    if is_admin:
        ap = query("SELECT COUNT(*) c FROM aor_draft WHERE status='pending'", one=True)
        stats['aor_pending'] = ap['c'] if ap else 0
        try:
            r = query("SELECT v FROM api_settings WHERE k='aor_crew_submitted'", one=True)
            stats['aor_crew_submitted'] = int(r['v'] or 0) if r else 0
        except sqlite3.Error:
            pass

    vlcc_last_push = None
    if is_admin:
        try:
            r = query("SELECT v FROM api_settings WHERE k='vlcc_last_push_at'", one=True)
            vlcc_last_push = r['v'] if r else None
        except sqlite3.Error:
            pass

    return dict(stats=stats, events=events, events_count=events_count,
                today_events=today_events, today_count=today_count, is_admin=is_admin,
                scoped=scoped, sup_name=sup_name, vlcc_last_push=vlcc_last_push)
def _issue_to_dict(row):
    d = dict(row)
    try:
        d['actions'] = json.loads(d['actions']) if d.get('actions') else []
    except Exception as e:
        app.logger.warning('issue-to-dict: %s', e)
        d['actions'] = []
    return d
def _gen_summary_rows(supervisor_id=None):
    """해당 스코프의 현재 이슈 행을 만든다.

    Gemini 입력은 상태가 정확히 Open/InProgress 인 행으로 제한한다. Closed 행은
    ``_run_summary_generate`` 가 저장된 이전 요약을 재사용하고, 이전 요약이 없는
    첫 실행만 아래 결정적 원문 fallback 을 쓴다.
    """
    conds, params = ['1=1'], []
    if supervisor_id:
        conds.append('i.supervisor_id = ?'); params.append(supervisor_id)
    sql = f'''
        SELECT i.*, s.display_order AS sv_order, v.name AS vessel_name,
               v.vessel_type AS vessel_type
          FROM issues i
          JOIN supervisors s ON s.id = i.supervisor_id
          JOIN vessels     v ON v.id = i.vessel_id
         WHERE {' AND '.join(conds)}
         ORDER BY s.display_order ASC, s.id ASC, i.issue_date ASC, i.id ASC
    '''
    rows = [_issue_to_dict(r) for r in query(sql, params)]
    payload = [{'i': idx,
                'description': r.get('description') or '',
                'action': _latest_action_progress(r.get('actions'))}
               for idx, r in enumerate(rows)
               if r.get('status') in ('Open', 'InProgress')]
    summaries = _gen_issue_summaries(payload)
    STAT = {'Open': 'Open', 'InProgress': '진행중', 'Closed': 'Closed'}
    out = []
    for idx, r in enumerate(rows):
        s = summaries.get(idx, {})
        desc = s.get('desc') or (r.get('description') or '').strip().split('\n')[0]
        ad, araw = _latest_action(r.get('actions'))
        action = s.get('action') or araw
        head = f"{_md_label(r.get('issue_date') or '')} {r.get('item_topic') or ''}".strip()
        lines = [head]
        if desc:
            lines.append(f'1) {desc}')
        if action:
            md = _md_label(ad)
            lines.append(f'2) {md} {action}'.strip() if md else f'2) {action}')
        out.append({'no': idx + 1,
                    'issue_id': r.get('id'),
                    'item': r.get('item_topic') or '',
                    'supervisor_id': r.get('supervisor_id'),
                    'vessel_id': r.get('vessel_id'),
                    'vessel_name': r.get('vessel_name') or '',
                    'vessel_type': r.get('vessel_type') or '',
                    'issue': '\n'.join(lines),
                    'priority': r.get('priority') or '',
                    'status_raw': r.get('status') or '',
                    'status': STAT.get(r.get('status'), r.get('status') or '')})
    return out
def _ensure_summary_table():
    execute("""CREATE TABLE IF NOT EXISTS issue_summaries (
                 scope TEXT PRIMARY KEY, data TEXT, generated_at TEXT )""")


def _preserve_closed_summary(rows, previous_rows, fallback_rows=None):
    """Closed 행의 AI 요약 본문을 같은 ``issue_id`` 의 저장본에서 복원한다.

    상태·우선순위·선박명 같은 표시 메타데이터는 현재 DB 값을 유지한다. 따라서
    직전 생성 뒤 종결된 항목도 마지막 진행중 요약 본문은 그대로 두되 완료 탭으로
    정확히 이동한다. ID 없는 legacy 저장행이나 깨진 저장값은 추측 매칭하지 않는다.
    """
    previous_by_id = {}
    # 전체 scope는 감독 재배정 뒤 새 감독 scope에 저장본이 없을 때의 보조 저장본이다.
    # 같은 ID가 둘 다 있으면 더 구체적인 현재 scope 저장본이 마지막에 덮어쓴다.
    for source in (fallback_rows, previous_rows):
        for old in source if isinstance(source, list) else []:
            if isinstance(old, dict) and old.get('issue_id') is not None:
                previous_by_id[str(old['issue_id'])] = old

    merged = []
    for row in rows:
        current = dict(row)
        if current.get('status_raw') == 'Closed':
            old = previous_by_id.get(str(current.get('issue_id')))
            if old and isinstance(old.get('issue'), str):
                current['issue'] = old['issue']
        merged.append(current)
    return merged


def _run_summary_generate(sid=None):
    """업무요약 생성+저장 코어 (UI 버튼·API키 스케줄러 공용).

    Open/InProgress 만 새로 요약하고 Closed 본문은 각 저장 scope의 직전 값을
    유지한다. 반환값은 ``(rows, gen_at, counts)``.
    """
    from datetime import datetime
    _ensure_summary_table()
    rows = _gen_summary_rows(sid)
    gen_at = datetime.now().strftime('%Y-%m-%d %H:%M')

    def _load(scope):
        old = query('SELECT data FROM issue_summaries WHERE scope=?', (scope,), one=True)
        try:
            return json.loads(old['data']) if old and old['data'] else []
        except (TypeError, ValueError, json.JSONDecodeError):
            return []

    def _save(scope, scope_rows):
        previous = _load(scope)
        fallback = _load('all') if scope != 'all' else []
        scope_rows = _preserve_closed_summary(scope_rows, previous, fallback)
        # scope 내에서 No. 재넘버링
        renum = []
        for i, r in enumerate(scope_rows, start=1):
            rr = dict(r); rr['no'] = i; renum.append(rr)
        execute("INSERT OR REPLACE INTO issue_summaries (scope, data, generated_at) VALUES (?, ?, ?)",
                (scope, json.dumps(renum, ensure_ascii=False), gen_at))
        return renum

    counts = {}
    if sid:
        rows = _save(str(sid), rows)
        counts[str(sid)] = len(rows)
    else:
        all_rows = _save('all', rows)
        counts['all'] = len(all_rows)
        # 감독별로 분리 저장 (각 감독 탭의 요약도 동시 갱신)
        by_sv = {}
        for r in rows:
            by_sv.setdefault(r.get('supervisor_id'), []).append(r)
        all_sv = [s['id'] for s in query('SELECT id FROM supervisors')]
        for sv_id in all_sv:
            saved = _save(str(sv_id), by_sv.get(sv_id, []))
            counts[str(sv_id)] = len(saved)
        rows = all_rows
    return rows, gen_at, counts
def _issue_write_scope(iid=None, payload=None):
    """Return a scoped issue row or raise 403 for non-admin cross-supervisor writes."""
    if session.get('role') == 'admin':
        if not iid:
            return None
        row = query('SELECT id, supervisor_id, vessel_id FROM issues WHERE id=?', (iid,), one=True)
        if not row:
            abort(404)
        return row
    sup_id = session.get('supervisor_id')
    if not sup_id:
        abort(403)
    if iid:
        row = query('SELECT id, supervisor_id, vessel_id FROM issues WHERE id=?', (iid,), one=True)
        if not row:
            abort(404)
        if row['supervisor_id'] != sup_id:
            abort(403)
        return row
    if payload is not None:
        # The browser cannot choose another supervisor, and the vessel must belong to it.
        if int(payload.get('supervisor_id') or 0) != sup_id:
            abort(403)
        vessel_id = int(payload.get('vessel_id') or 0)
        if not query('SELECT 1 FROM supervisor_vessels WHERE supervisor_id=? AND vessel_id=?',
                     (sup_id, vessel_id), one=True):
            abort(403)
    return None


# ══════════════════════════════════════════════════════════════════
#  From ai_gemini.py (16 symbols)
# ══════════════════════════════════════════════════════════════════
# ─── 보고서 → 항목 자동 추출 (Gemini + 엑셀 파서) ─────────────
def _findings_workbook(title, subtitle, headers, rows, wrap_cols, widths):
    """검사 findings → 스타일된 1시트 워크북 BytesIO 반환."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook(); ws = wb.active; ws.title = 'List'
    F = 'Malgun Gothic'
    N = len(headers)
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    title_fill = PatternFill('solid', start_color='1F3A5F')
    sub_fill   = PatternFill('solid', start_color='2C5282')
    hdr_fill   = PatternFill('solid', start_color='34495E')
    def_fill   = PatternFill('solid', start_color='FCE8E6')   # Defect 행 연한 적색
    thin = Side(style='thin', color='BBBBBB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name=F, size=14, bold=True, color='FFFFFF'); c.fill = title_fill
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N)
    c = ws.cell(row=2, column=1, value=subtitle)
    c.font = Font(name=F, size=10, italic=True, color='ECF0F1'); c.fill = sub_fill
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6

    HDR = 4
    for ci, h in enumerate(headers, start=1):
        cc = ws.cell(row=HDR, column=ci, value=h)
        cc.font = Font(name=F, size=11, bold=True, color='FFFFFF'); cc.fill = hdr_fill
        cc.alignment = Alignment(horizontal='center', vertical='center'); cc.border = border
    ws.row_dimensions[HDR].height = 24

    body = Font(name=F, size=10)
    top_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
    center = Alignment(horizontal='center', vertical='top')
    r_idx = HDR + 1
    for row in rows:
        max_len = 1
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=ci, value=val)
            cell.font = body; cell.border = border
            cell.alignment = top_wrap if ci in wrap_cols else center
            if ci in wrap_cols and val:
                w = widths[ci - 1]
                max_len = max(max_len, sum((len(ln) // max(int(w / 1.6), 1)) + 1
                                           for ln in str(val).split('\n')))
        # Defect 행 살짝 음영
        if 'Category' in headers:
            cat_col = headers.index('Category') + 1
            if ws.cell(row=r_idx, column=cat_col).value == 'Defect':
                for ci in range(1, N + 1):
                    ws.cell(row=r_idx, column=ci).fill = def_fill
        ws.row_dimensions[r_idx].height = max(20, min(120, 15 * max_len + 4))
        r_idx += 1

    ws.freeze_panes = f'A{HDR + 1}'
    if r_idx - 1 > HDR:
        ws.auto_filter.ref = f'A{HDR}:{get_column_letter(N)}{r_idx - 1}'
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f'{HDR}:{HDR}'

    bio = BytesIO(); wb.save(bio); bio.seek(0)
    return bio
def _gemini_call_json(parts, model=None):
    """parts(list) → Gemini generateContent → 파싱된 JSON dict 또는 {'error':...}."""
    if not GEMINI_API_KEY:
        return {'error': 'NO_API_KEY'}
    import urllib.request, urllib.error
    mdl = model or GEMINI_MODEL
    body = {'contents': [{'parts': parts}],
            'generationConfig': {'response_mime_type': 'application/json'}}
    url = (f'https://generativelanguage.googleapis.com/v1beta/models/'
           f'{mdl}:generateContent')
    req = urllib.request.Request(
        url, data=json.dumps(body).encode('utf-8'),
        headers={'content-type': 'application/json', 'x-goog-api-key': GEMINI_API_KEY},
        method='POST')
    try:
        with urllib.request.urlopen(req, timeout=90) as resp:
            data = json.loads(resp.read().decode('utf-8'))
    except urllib.error.HTTPError as he:
        try:
            detail = he.read().decode('utf-8')[:300]
        except Exception:
            app.logger.exception('gemini-call-json')
            detail = str(he)
        return {'error': 'API_CALL_FAILED', 'detail': detail}
    except Exception as e:
        app.logger.exception('gemini-call-json')
        return {'error': 'API_CALL_FAILED', 'detail': str(e)}
    text = ''
    try:
        cands = data.get('candidates') or []
        if not cands:
            return {'error': 'API_CALL_FAILED', 'detail': json.dumps(data)[:300]}
        for part in (cands[0].get('content', {}).get('parts') or []):
            if isinstance(part.get('text'), str):
                text += part['text']
    except Exception as e:
        app.logger.exception('gemini-call-json')
        return {'error': 'PARSE_FAILED', 'raw': str(e)}
    text = text.strip()
    if text.startswith('```'):
        text = text.strip('`')
        if text[:4].lower() == 'json':
            text = text[4:]
        text = text.strip()
    try:
        return json.loads(text)
    except Exception:
        app.logger.exception('gemini-call-json')
        return {'error': 'PARSE_FAILED', 'raw': text[:300]}
def _coerce_translation_items(res):
    """Gemini 응답을 [{'i':int,'en':str}] 리스트로 정규화. list/dict/다양한 키 모두 수용."""
    if isinstance(res, dict):
        if res.get('error'):
            return None  # 호출 자체 실패
        arr = (res.get('translations') or res.get('items')
               or res.get('results') or res.get('data'))
        if arr is None:
            # 단일 객체이거나 {i:en} 매핑일 수 있음
            if 'i' in res and ('en' in res or 'text' in res):
                arr = [res]
            else:
                arr = []
    elif isinstance(res, list):
        arr = res
    else:
        arr = []
    return arr if isinstance(arr, list) else []
def _translate_batch_en(texts, group):
    """group(인덱스 리스트) 한 묶음 번역 → {원본인덱스: 영문}. 실패 시 None."""
    payload = json.dumps([{'i': i, 'text': texts[i]} for i in group], ensure_ascii=False)
    prompt = (
        "너는 선박 기술 감독(ship superintendent)이다. 아래 JSON 배열의 각 한국어(또는 한영 혼용) "
        "텍스트를 선박 관리 현업에서 자연스럽게 쓰는 영어로 번역하라.\n"
        "- 장비명·약어·단위·수치(예: BRG, RPM, S/W pump, LT cooler, EGCS, °C, kts)는 그대로 둔다.\n"
        "- 줄바꿈과 번호 매김(1. 2. ...) 구조를 그대로 보존한다.\n"
        "- 이미 영어인 부분은 그대로 둔다. 의미를 바꾸거나 내용을 덧붙이지 마라.\n"
        "반드시 {\"translations\":[...]} 형태의 JSON 객체로만 답하라. 입력의 i를 그대로 사용하라.\n"
        '형식: {"translations":[{"i":0,"en":"..."}]}\n\n[입력]\n' + payload)
    res = _gemini_call_json([{'text': prompt}], model=_model_for('translate'))
    arr = _coerce_translation_items(res)
    if arr is None:
        return None  # API 호출 실패 → 상위에서 분할 재시도
    out = {}
    for tr in arr:
        if not isinstance(tr, dict):
            continue
        try:
            i = int(tr.get('i'))
        except (TypeError, ValueError):
            continue
        en = tr.get('en') if isinstance(tr.get('en'), str) else tr.get('text')
        if isinstance(en, str) and en.strip():
            out[i] = en
    return out
def _gen_issue_summaries(payload_items):
    """payload_items: [{'i':int,'description':str,'action':str}] →
    {i: {'desc':str, 'action':str}} (한국어 요약). 키 없음/실패 시 빈 dict 부분 반환."""
    result = {}
    if not GEMINI_API_KEY or not payload_items:
        return result

    def run(group, depth=0):
        if not group:
            return
        sub = [payload_items[k] for k in group]
        prompt = (
            "너는 선박 기술 감독(ship superintendent)이다. 아래 JSON 배열의 각 업무 항목에 대해 "
            "두 가지를 한국어로 작성하라.\n"
            "- desc: description의 핵심 문제를 1문장(최대 2문장)으로 짧게 요약\n"
            "- action: action(최신 조치내용)을 한 줄로 짧게 요약 (내용 없으면 빈 문자열)\n"
            "■ 매우 중요: 요약은 원문(description/action)에 실제로 쓰인 단어와 표현을 그대로 사용해 "
            "압축하라. 동의어로 바꾸거나 새 표현을 지어내지 말고, 불필요한 부분만 덜어내라. "
            "원문에 있는 장비명·기술용어·약어·표현(예: EGCS, Pump, Auto mode, Maker Trouble Shooting, BRG, RPM, LT cooler)은 "
            "그대로 보존한다. 과장/추측/내용 추가 금지.\n"
            "입력의 i를 그대로 사용해 JSON 객체로만 답하라.\n"
            '형식: {"items":[{"i":0,"desc":"...","action":"..."}]}\n\n[입력]\n'
            + json.dumps(sub, ensure_ascii=False))
        res = _gemini_call_json([{'text': prompt}], model=_model_for('summary'))
        arr = _coerce_translation_items(res)  # translations/items/results/data 모두 수용
        if arr is None:
            if len(group) > 1 and depth < 6:
                mid = len(group) // 2
                run(group[:mid], depth + 1); run(group[mid:], depth + 1)
            return
        got = set()
        for o in arr:
            if not isinstance(o, dict):
                continue
            try:
                i = int(o.get('i'))
            except (TypeError, ValueError):
                continue
            result[i] = {
                'desc':   (o.get('desc') or o.get('desc_summary') or '').strip(),
                'action': (o.get('action') or o.get('action_summary') or '').strip(),
            }
            got.add(i)
        missing = [k for k in group if k not in got]
        if missing and len(group) > 1 and depth < 6:
            mid = max(1, len(missing) // 2)
            run(missing[:mid], depth + 1); run(missing[mid:], depth + 1)

    CHUNK = 12
    idxs = list(range(len(payload_items)))
    for s in range(0, len(idxs), CHUNK):
        run(idxs[s:s + CHUNK])
    return result
def _latest_action_progress(acts):
    if not acts:
        return ''
    try:
        best = sorted(acts, key=lambda a: (a.get('date') or ''))[-1]
    except Exception as e:
        app.logger.warning('latest-action-progress: %s', e)
        best = acts[-1]
    return (best.get('progress') or '').strip()
def _latest_action(acts):
    """최신 action(날짜 최댓값)의 (date, progress) 반환."""
    if not acts:
        return '', ''
    try:
        best = sorted(acts, key=lambda a: (a.get('date') or ''))[-1]
    except Exception as e:
        app.logger.warning('latest-action: %s', e)
        best = acts[-1]
    return (best.get('date') or '').strip(), (best.get('progress') or '').strip()
def _md_label(d):
    try:
        y, m, dd = d.split('-')
        return f'[{int(m)}/{int(dd)}]'
    except Exception as e:
        app.logger.warning('md-label: %s', e)
        return f'[{d}]' if d else ''
def _translate_texts_en(texts):
    """한국어(한영 혼용) 문자열 리스트 → 선박 감독 현업 영어. 키 없음/실패 시 원문 유지.
    묶음 실패 시 절반→1:1로 분할 재시도하여 '일부 누락'을 방지."""
    if not GEMINI_API_KEY:
        return list(texts)
    out = list(texts)
    idxs = [i for i, t in enumerate(texts) if t and str(t).strip()]

    def run(group, depth=0):
        if not group:
            return
        res = _translate_batch_en(texts, group)
        if res is None:
            # 호출 실패 → 분할 재시도
            if len(group) > 1 and depth < 6:
                mid = len(group) // 2
                run(group[:mid], depth + 1)
                run(group[mid:], depth + 1)
            return
        missing = [i for i in group if i not in res]
        for i, en in res.items():
            out[i] = en
        # 일부만 응답에 빠진 경우도 분할 재시도
        if missing and len(group) > 1 and depth < 6:
            mid = max(1, len(missing) // 2)
            run(missing[:mid], depth + 1)
            run(missing[mid:], depth + 1)

    CHUNK = 12
    for s in range(0, len(idxs), CHUNK):
        run(idxs[s:s + CHUNK])
    return out
def _dd_ko_prompt(payload):
    """감독 DD report 행(영문) → 형의 일일보고 어조(한국어) 번역 프롬프트.

    어조 정본은 라이브 `issues` 229행에서 관찰한 형의 작성 패턴이다: 한국어 서술 +
    장비명·약어는 영문 그대로 + `~함/~됨/~예정/~필요` 음슴체.  🔴 실제 지적 문구를
    few-shot 으로 매번 외부 API 에 보내지 않는다(선박 실데이터 유출) — 패턴만 규칙으로
    적고 예시는 식별정보 없는 합성문으로 쓴다.
    """
    return (
        "너는 선박 기술 감독(ship superintendent)이다. 아래 JSON 배열은 조선소 감독이 영문으로 "
        "쓴 입거(dry dock) 일일 작업 항목이다. 각 항목을 한국어 일일보고 문장으로 옮겨라.\n"
        "- 문체는 '~함/~됨/~예정/~필요' 음슴체 개조식. 1문장(길면 최대 2문장)으로 짧게.\n"
        "- 장비명·부위명·약어·모델명·단위·수치는 **영문 그대로** 둔다 "
        "(예: Hatch Cover, Rope Guard, Stern Tube, Windlass, Cooler Plate, T/C, ME, AE, "
        "S/W pump, LT cooler, EGCS, BWTS, UTM, RPM, °C, No.1).\n"
        "- 원문에 없는 내용·원인·평가를 덧붙이지 마라. 요약은 하되 사실을 만들지 마라.\n"
        "- 시각(예: 11:42 LT)·수량·호기 번호는 반드시 보존한다.\n"
        "- 존댓말·'-습니다' 금지. 제목처럼 명사만 나열하지 말고 서술로 쓴다.\n"
        + _MARITIME_TERMS +
        "예: 'Hatch Cover No.1 dismantle hydraulic jacks and pins' → "
        "'Hatch Cover No.1 hydraulic jack 및 pin 분해함'\n"
        "입력의 i를 그대로 사용해 JSON 객체로만 답하라.\n"
        '형식: {"translations":[{"i":0,"ko":"..."}]}\n\n[입력]\n' + payload)


def _translate_batch_ko(texts, group):
    """group(인덱스 리스트) 한 묶음 EN→KO → {원본인덱스: 한국어}. 실패 시 None.

    `_translate_batch_en` 의 거울.  🔴 번역 함수를 이 한 곳으로 몰아 두는 이유는
    나중에 덴키(로컬 Qwen)로 갈아탈 때 이 함수만 바꾸면 되게 하려는 것이다
    (지금 서버는 tailscale 이 없어 맥의 `127.0.0.1:8000` 에 닿지 못한다).
    """
    payload = json.dumps([{'i': i, 'text': texts[i]} for i in group], ensure_ascii=False)
    res = _gemini_call_json([{'text': _dd_ko_prompt(payload)}], model=_model_for('translate'))
    arr = _coerce_translation_items(res)
    if arr is None:
        return None  # API 호출 실패 → 상위에서 분할 재시도
    out = {}
    for tr in arr:
        if not isinstance(tr, dict):
            continue
        try:
            i = int(tr.get('i'))
        except (TypeError, ValueError):
            continue
        ko = tr.get('ko') if isinstance(tr.get('ko'), str) else tr.get('text')
        if isinstance(ko, str) and ko.strip():
            out[i] = ko.strip()
    return out


def translate_texts_ko(texts):
    """영문 문자열 리스트 → 한국어. 키 없음/실패 시 **원문 유지**.

    🔴 실패를 예외로 올리지 않는다. 번역이 안 되면 영문 그대로 카드에 들어가는 게
    맞다 — 형은 영문 원문도 읽을 수 있고, 여기서 터뜨리면 파일 읽기 전체가 실패한다.
    """
    if not GEMINI_API_KEY:
        return list(texts)
    out = list(texts)
    idxs = [i for i, t in enumerate(texts) if t and str(t).strip()]

    def run(group, depth=0):
        if not group:
            return
        res = _translate_batch_ko(texts, group)
        if res is None:
            if len(group) > 1 and depth < 6:
                mid = len(group) // 2
                run(group[:mid], depth + 1)
                run(group[mid:], depth + 1)
            return
        missing = [i for i in group if i not in res]
        for i, ko in res.items():
            out[i] = ko
        if missing and len(group) > 1 and depth < 6:
            mid = max(1, len(missing) // 2)
            run(missing[:mid], depth + 1)
            run(missing[mid:], depth + 1)

    CHUNK = 12
    for s in range(0, len(idxs), CHUNK):
        run(idxs[s:s + CHUNK])
    return out


def _translate_rows_en(rows):
    """이슈 행들의 item_topic/description/actions[].progress 를 영문으로 치환(제자리)."""
    bucket, texts = [], []
    for r in rows:
        if r.get('item_topic'):
            bucket.append((r, 'item_topic', None)); texts.append(r['item_topic'])
        if r.get('description'):
            bucket.append((r, 'description', None)); texts.append(r['description'])
        for ai, a in enumerate(r.get('actions') or []):
            if a.get('progress'):
                bucket.append((r, 'actions', ai)); texts.append(a['progress'])
    if not texts:
        return
    tr = _translate_texts_en(texts)
    for (r, field, ai), en in zip(bucket, tr):
        if field == 'actions':
            r['actions'][ai]['progress'] = en
        else:
            r[field] = en
_MARITIME_TERMS = (
    " 요약은 선박 현업(감독/기관부) 용어로 옮긴다. 일반어 → 현업어 매핑: "
    "repair=수리(※'보수'로 쓰지 말 것), cleaning/clean=소제, replace/renew/renewal=신환, "
    "install/fitting=설치, overhaul=O/H(분해점검), inspection/survey=수검, maintenance=정비, "
    "check/verify=확인, adjust/adjustment=조정, calibration=교정, test=시험, crack=균열, "
    "corrosion/rust=부식, leak/leakage=누설(누유/누수), wear/weardown=마모, deformation=변형, "
    "spare parts=예비품, weld/welding=용접, coating/painting=도장, submit=제출, "
    "place onboard=본선 비치. "
    "목록에 없어도 선박에서 통용되는 자연스러운 표현을 우선 사용한다. "
)
# ═════════════════════════════════════════════════════════════════
#  API — Vetting Status (비정기, 선박당 0~N건, CNTR 제외)
# ═════════════════════════════════════════════════════════════════
VETTING_TYPES = ('VLCC', 'AFRAMAX', 'LR', 'MR')


def _clean_vetting_overall_remark(value):
    """과거 종합소견에 잘못 포함된 Full report 내부 블록만 표시 전에 제거한다."""
    begin = '[SIRE Full Report 자동반영]'
    end = '[/SIRE Full Report 자동반영]'
    text = value or ''
    text = re.sub(
        r'\s*-\s*' + re.escape(begin) + r'\n.*?^' + re.escape(end) + r'\s*',
        '\n', text, flags=re.S | re.M,
    )
    text = re.sub(
        r'^' + re.escape(begin) + r'\n.*?^' + re.escape(end) + r'\s*',
        '', text, flags=re.S | re.M,
    )
    return '\n'.join(line.rstrip() for line in text.splitlines() if line.strip()).strip()


def _vetting_with_counts(v):
    """vetting dict에 카운트 추가. manual override 적용."""
    vid = v['id']
    rows = query("""
        SELECT status, COUNT(*) AS n
          FROM vt_findings
         WHERE vetting_id = ?
         GROUP BY status
    """, (vid,))
    auto_open = auto_closed = 0
    for r in rows:
        if r['status'] == 'Closed': auto_closed = r['n']
        else: auto_open = r['n']
    auto_total = auto_open + auto_closed

    d = dict(v)
    d['overall_remark'] = _clean_vetting_overall_remark(d.get('overall_remark'))
    d['observation_count'] = v['manual_observation_count'] if v['manual_observation_count'] is not None else auto_total
    d['close_count']       = v['manual_close_count']       if v['manual_close_count']       is not None else auto_closed
    d['open_count']        = v['manual_open_count']        if v['manual_open_count']        is not None else max(0, d['observation_count'] - d['close_count'])
    d['observation_manual'] = v['manual_observation_count'] is not None
    d['open_manual']        = v['manual_open_count']        is not None
    d['close_manual']       = v['manual_close_count']       is not None
    # 첨부 카운트
    ar = query('SELECT COUNT(*) AS n FROM vt_attachments WHERE vetting_id=? AND inactive_at IS NULL',
               (vid,), one=True)
    d['attach_count'] = ar['n'] if ar else 0
    return d
def _vetting_display_order(rows):
    """선박 1척의 vetting 행 표시 순서 정본 (맨 앞 = 상단표시 기준 행).

    🔴 검사일 내림차순만 쓰면 **날짜 미입력 행이 항상 맨 밑**으로 밀린다. 새 Vetting 을
       추가해 'Next Plan'(계획된 다음 검사)으로 지정해도 검사일을 아직 모르면 목록 끝에
       숨어버려서, 정작 제일 먼저 봐야 할 계획을 못 본다(손유석 지시 2026-07-31).
       그래서 상태를 1순위로 두고 날짜는 그 안에서만 본다:
       ① 'Next Plan' 을 항상 위. 여러 개면 새로 만든 것(id 최신) 우선
          — `_vetting_pick` 의 latest 선정과 같은 규칙이라 rows[0] == latest 가 성립한다.
       ② 나머지(Report)는 기존대로 검사일 내림차순, 같은 날짜면 id 내림차순.
    """
    nexts  = [r for r in rows if (r.get('valid') or '') == 'Next Plan']
    others = [r for r in rows if (r.get('valid') or '') != 'Next Plan']
    nexts.sort(key=lambda r: r.get('id') or 0, reverse=True)
    others.sort(key=lambda r: ((r.get('inspection_date') or ''), r.get('id') or 0),
                reverse=True)
    return nexts + others
def _vetting_pick(vessel_id):
    """선박 1척의 vetting 중 (상단표시 기준 행, 전체) 를 고른다.

    🔴 이 선정 규칙은 **정본이 1곳이어야 한다** — 웹 프론트 `vt.js vettingDigest`,
       `/api/ext/vetting-digests`, 위젯이 서로 다른 숫자를 보여주면 형이 못 믿는다.
       'Next Plan'(계획된 다음 검사)이 있으면 검사일 미입력이어도 그것을 상단으로.
       여러 개면 새로 만든 것(id 최신) 우선.

    🔴 요약행의 모든 값(OBS/OPEN 포함)은 **상단행 그 자체**에서 나온다 — 손유석 지시
       2026-08-11("Next Plan 일 경우 해당 Next Plan 의 OBS 및 OPEN 숫자가 표시되게").
       상단이 Next Plan 일 때 OBS 만 직전 Report 에서 끌어오던 `obs_src` 폴백은 이때
       폐기했다. 상태는 계획인데 숫자는 지난 수검 것이라 한 줄 안에서 출처가 갈렸고,
       형이 화면에서 그걸 오독으로 지목했다. **되살리지 말 것** — 살리려면 요약행을
       두 줄로 나누는 설계부터 다시 받아야 한다.
    반환: (latest, enr). vetting 이 없으면 (None, []).
    """
    vts = query("SELECT * FROM vettings WHERE vessel_id=? "
                "ORDER BY inspection_date DESC, id DESC", (vessel_id,))
    if not vts:
        return None, []
    enr = _vetting_display_order([_vetting_with_counts(v) for v in vts])
    return enr[0], enr

def _vetting_summary_counts(latest):
    """요약/미러용 OBS 수치. Next Plan 은 아직 수검 전이므로 미입력(null)으로 둔다."""
    if not latest or (latest.get('valid') or '') == 'Next Plan':
        return None, None, None
    # Last Result 의 기존 계약은 결측도 0으로 표시하는 것. Next Plan 외 동작은 바꾸지 않는다.
    return (latest.get('observation_count') or 0,
            latest.get('open_count') or 0,
            latest.get('close_count') or 0)
# ═════════════════════════════════════════════════════════════════
#  API — Calendar Events (일정 모듈)
# ═════════════════════════════════════════════════════════════════
CAL_VALID_COLORS = ('gray','red','amber','yellow','green','blue','purple','pink')


# ══════════════════════════════════════════════════════════════════
#  From routes_calendar_dock.py (81 symbols)
# ══════════════════════════════════════════════════════════════════
def _safe_filename(s):
    """파일명에서 OS 비호환 문자 제거"""
    import re
    s = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', s)
    s = s.strip().strip('.')
    return s[:80] or 'report'
# ═════════════════════════════════════════════════════════════════
#  API — attachments
# ═════════════════════════════════════════════════════════════════
def _ext_allowed(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXT
GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY', '')
GEMINI_MODEL   = os.environ.get('GEMINI_MODEL', 'gemini-3.1-flash-lite')
# 용도별 모델 — /etc/trmt.env 에서 지정 (미지정 시 GEMINI_MODEL 사용)
#   MODEL_SUMMARY  : 요약        (텍스트)
#   MODEL_TRANSLATE: 영문 번역    (텍스트)
#   MODEL_FINDINGS : 리포트 추출  (멀티모달 필요)
#   MODEL_REMARK   : 리마크 요약  (텍스트)
#   MODEL_RECEIPT  : 영수증 비전  (멀티모달 필수)
_MODEL_ENV = {
    'summary':   'MODEL_SUMMARY',
    'translate': 'MODEL_TRANSLATE',
    'findings':  'MODEL_FINDINGS',
    'remark':    'MODEL_REMARK',
    'receipt':   'MODEL_RECEIPT',
    'krcon':     'MODEL_KRCON',
}
def _model_for(purpose):
    """용도별 모델 ID 반환 (환경변수 우선, 없으면 기본 GEMINI_MODEL)."""
    env = _MODEL_ENV.get(purpose)
    return (os.environ.get(env) if env else None) or GEMINI_MODEL
# ═════════════════════════════════════════════════════════════════
#  외부 연동용 데이터 API (읽기 전용, API 키 보호)
#  · 출장 경비(biz_*) 제외 — 그 외 전체 탭 공개
# ═════════════════════════════════════════════════════════════════
#: `_ensure_api_table()` 를 이미 끝낸 DB 파일 경로들(프로세스 로컬 캐시).
#: 경로를 키로 쓰는 이유는 테스트가 `app.config['DATABASE']` 를 임시 파일로 바꿔치기해서다
#: — 단순 bool 플래그면 새 임시 DB 에서 테이블 없이 진행해 조용히 깨진다.
_API_TABLE_READY = {}  # path -> (device, inode): DB 파일 교체 감지(정상 DB 쓰기는 cache 무효화 안 함)
_API_TABLE_LOCK = threading.Lock()
def _ensure_api_table():
    """`api_settings` 존재 보장 — **프로세스당 DB 경로별 1회만**.

    예전엔 API 키가 붙은 **모든 요청**마다 `CREATE TABLE IF NOT EXISTS` + `commit()` 이 돌았다
    (`_check_api_key` → `_get_api_key` → 여기). no-op DDL 이라도 `execute()` 가 커밋을 하므로
    요청마다 쓰기 트랜잭션이 열렸다 — 순수 낭비이고, 읽기전용 API 가 writer lock 을 잡는
    부작용까지 있었다. 테이블 자체는 `init_db()` 가 이미 만든다(위 api_settings 블록).
    """
    path = app.config.get('DATABASE', DATABASE)
    with _API_TABLE_LOCK:
        try:
            st = os.stat(path)
            sig = (st.st_dev, st.st_ino)
        except OSError:
            sig = None
        if _API_TABLE_READY.get(path) == sig and sig is not None:
            return
        execute("""CREATE TABLE IF NOT EXISTS api_settings (
                     k TEXT PRIMARY KEY, v TEXT )""")
        st = os.stat(path)
        _API_TABLE_READY[path] = (st.st_dev, st.st_ino)
def _get_api_key(create=True):
    _ensure_api_table()
    row = query("SELECT v FROM api_settings WHERE k='api_key'", one=True)
    if row and row['v']:
        return row['v']
    if not create:
        return None
    key = secrets.token_hex(24)
    execute("INSERT OR REPLACE INTO api_settings (k, v) VALUES ('api_key', ?)", (key,))
    return key
def _check_api_key():
    provided = (request.headers.get('X-API-Key')
                or request.args.get('key') or '').strip()
    if not provided:
        return False
    real = _get_api_key(create=False)
    if not real:
        return False
    return secrets.compare_digest(provided, real)
def _vkey(name):
    return (name or '').strip().lower()
def api_key_required(fn):
    @wraps(fn)
    def wrapper(*a, **k):
        if not _check_api_key():
            return jsonify({'error': 'unauthorized',
                            'message': 'valid API key required (X-API-Key header or ?key=)'}), 401
        return fn(*a, **k)
    return wrapper
# ═════════════════════════════════════════════════════════════════
#  자동화 헬스 보드 (하트비트) — 맥측 health_push.py 가 POST, admin 이 /health 로 조회
# ═════════════════════════════════════════════════════════════════
# 러너 기술키 → (한글 표시명, 돈경로 여부). 미등록 키는 raw key 그대로 표시.
AUTOMATION_LABELS = {
    'fundreq-auto':    ('비용청구 자동상신',      True),
    'jeonja-auto':     ('전자결재 자동상신',      True),
    'soa-approve':     ('SOA 주말 자동승인',      True),
    'invoice-auto':    ('인보이스 자동처리',      True),
    'aor-prep':        ('AOR 준비 카드',          False),
    'dock-sync':       ('입거 발주 SVMS 동기화',  False),
    'fleet-map':       ('선박 위치지도 갱신',      False),
    'fleet-map-crawl': ('선위 AIS 수집',          False),
    'cls-push':        ('선급 검사현황 동기화',    False),
    'shipwiki-ingest': ('선박 위키 수집',          False),
    'trmt-summary':    ('현안 요약 생성',          False),
    'money-watch':     ('돈경로 감시견',          False),
    'git-backup':      ('작업 백업',              False),
    'jeonja-verify':   ('전자결재 검증',          False),
    'wfmail':          ('메일→현안 카드 수집',     False),
    'logrotate':       ('로그 정리',              False),
}
# 은퇴한 러너 — 보드에서 숨긴다. 맥측 health_push 에서 빼도 automation_health 에 남은
# 과거 행 때문에 카드가 계속 보이므로(그리고 갱신이 끊겨 unknown/fail 로 굳으므로) 여기서 차단한다.
# 운영 DB 를 손으로 지우는 대신 코드 필터로 처리 — prune 규약(러너당 30행)상 잔존 행은 무해하다.
RETIRED_RUNNER_KEYS = {
    'mail-brief',   # 아침 메일 브리핑, 2026-07-31 폐기(형 지시)
}
# status 정렬 우선순위(fail 먼저, 그다음 warn, ok, unknown)
_HEALTH_ORDER = {'fail': 0, 'warn': 1, 'ok': 2, 'unknown': 3}
def _automation_health_summary():
    """러너별 최신 관측 + 최근 14개 히스토리(oldest→newest)를 조립.
    반환: (runners[list], counts[dict]). Feature1 read 와 Feature2 cockpit 이 공유."""
    rows = query("SELECT id, runner_key, status, note, ran_at, next_run, reported_at "
                 "FROM automation_health ORDER BY runner_key, reported_at, id")
    by_key = {}
    for r in rows:
        if r['runner_key'] in RETIRED_RUNNER_KEYS:
            continue
        by_key.setdefault(r['runner_key'], []).append(r)

    runners = []
    counts = {'ok': 0, 'warn': 0, 'fail': 0, 'unknown': 0, 'total': 0}
    for key, obs in by_key.items():
        latest = obs[-1]
        status = latest['status'] if latest['status'] in _HEALTH_ORDER else 'unknown'
        label, money = AUTOMATION_LABELS.get(key, (key, False))
        history = [(o['status'] if o['status'] in _HEALTH_ORDER else 'unknown')
                   for o in obs[-14:]]
        runners.append({
            'key': key, 'label': label, 'money': money,
            'status': status, 'note': latest['note'],
            'ran_at': latest['ran_at'], 'next_run': latest['next_run'],
            'reported_at': latest['reported_at'], 'history': history,
        })
        counts[status] = counts.get(status, 0) + 1
        counts['total'] += 1

    # fail → warn → ok → unknown, 동급이면 돈경로 먼저, 그다음 라벨
    runners.sort(key=lambda x: (_HEALTH_ORDER.get(x['status'], 3),
                                0 if x['money'] else 1, x['label']))
    return runners, counts
# category → SVMS OW_COMP_ID. 돈 분기(Slip 출금상신·검증강도)의 근거라
# 코드 상수로 격리 — DB·UI 어디서도 편집 불가.
SOA_CATEGORY_OWNER = {'silver': '037', 'skrt': '001'}
def _soa_groups_load(active_only=True, db=None):
    """soa_group + membership → dict 리스트. 그룹 수와 무관하게 SQL 한 번."""
    # 실행 경계의 첫 인자를 완전한 literal로 둔다. 값 injection 위험뿐 아니라 동적 SQL
    # site 자체를 늘리지 않는 게 repository SQL-construction 계약이다.
    if db is not None:
        if active_only:
            rows = db.execute(
                'SELECT g.id,g.key,g.label,g.category,g.mode,g.sort_order,g.active,v.vsl_cd '
                'FROM soa_group g LEFT JOIN soa_group_vessel v ON v.group_id=g.id '
                'WHERE g.active=1 ORDER BY g.sort_order,g.key,v.vsl_cd').fetchall()
        else:
            rows = db.execute(
                'SELECT g.id,g.key,g.label,g.category,g.mode,g.sort_order,g.active,v.vsl_cd '
                'FROM soa_group g LEFT JOIN soa_group_vessel v ON v.group_id=g.id '
                'ORDER BY g.sort_order,g.key,v.vsl_cd').fetchall()
    elif active_only:
        rows = query('SELECT g.id,g.key,g.label,g.category,g.mode,g.sort_order,g.active,v.vsl_cd '
                     'FROM soa_group g LEFT JOIN soa_group_vessel v ON v.group_id=g.id '
                     'WHERE g.active=1 ORDER BY g.sort_order,g.key,v.vsl_cd')
    else:
        rows = query('SELECT g.id,g.key,g.label,g.category,g.mode,g.sort_order,g.active,v.vsl_cd '
                     'FROM soa_group g LEFT JOIN soa_group_vessel v ON v.group_id=g.id '
                     'ORDER BY g.sort_order,g.key,v.vsl_cd')
    out = []
    by_id = {}
    for r in rows:
        group = by_id.get(r['id'])
        if group is None:
            group = {
                'key': r['key'], 'label': r['label'], 'category': r['category'],
                'owner_comp_id': SOA_CATEGORY_OWNER.get(r['category']),
                'mode': r['mode'], 'sort_order': r['sort_order'], 'active': r['active'],
                'vessels': [],
            }
            by_id[r['id']] = group
            out.append(group)
        if r['vsl_cd'] is not None:
            group['vessels'].append(r['vsl_cd'])
    return out
def _soa_owner_map():
    """SVMS My Vessel owner 스냅샷(러너가 push). 표시 전용 — 실행 대상 판정 근거 아님."""
    return {r['vsl_cd']: r['owner_comp_id']
            for r in query('SELECT vsl_cd, owner_comp_id FROM soa_vessel_owner')}
def _soa_group_members(g, owner_map=None):
    """이 그룹에 지금 리스트업된 선박. explicit=명시 선박 ∩ owner, dynamic=owner 전체.
    owner 스냅샷이 없으면 explicit 은 명시 선박 그대로, dynamic 은 빈 리스트."""
    if owner_map is None:
        owner_map = _soa_owner_map()
    oc = SOA_CATEGORY_OWNER.get(g['category'])
    pool = {v for v, o in owner_map.items() if o == oc}
    if g['mode'] == 'dynamic_owner':
        return sorted(pool)
    if not pool:
        return sorted(g['vessels'])
    return sorted(v for v in g['vessels'] if v in pool)
# ===== dock_procure 수동 SVMS 발주 새로고침(dock_sync 온디맨드 트리거) — roster-sync 패턴 =====
def _dock_sync_flag_bump():
    """dock_sync 온디맨드 트리거 flag 를 세우고 **실효 flag** 를 반환한다 — 단일 writer.

    호출자 = ①'SVMS 발주 새로고침' 버튼 ②견적요청 성공 콜백(`api_ext_dock_inquiry_result`).
    맥 watcher(launchd `ai.openclaw.dock-sync-watch`, 60s)가 `/api/ext/dock_procure/sync/pending`
    (**flag>done 일 때만** flag 반환)을 폴링해 `dock_sync.sh --live` 를 1회 돌리고
    `/api/ext/dock_procure/sync/done` 으로 flag+결과를 기록(=clear)한다. dock_sync 는 선박 전체를
    다시 읽는 멱등 read-only 폴러라 flag 가 몇 번 세워져도 결과가 달라지지 않는다.

    🔴 flag 는 **항상 전진**하고, `done` 보다도 **엄격히 크게** 만든다. 이유는 두 개의 조용한 유실:
      · 같은 초에 done 이 이미 찍혀 있으면(버튼→sync→done 직후 견적요청 성공) `now == done` 이라
        pending 이 false → 우리 요청이 아무 sync 도 못 일으키고 사라진다.
      · watcher 가 flag=T1 을 들고 sync 를 **이미 시작한 뒤** 우리 write 가 들어오면, 그 sync 는
        우리 변경 이전의 SVMS 를 읽고 done=T1 로 flag 를 닫는다. 그래서 pending 인 flag 를 보고
        "이미 예약됨"으로 재사용하지 않고 새 시각으로 밀어 **우리 write 이후에 시작하는 sync** 를
        보장한다. 최악의 대가 = 멱등 폴러 1회 더 실행(read-only).

    🔴 write 는 **단일 SQL 의 조건부 upsert**(`WHERE excluded.v > api_settings.v`)로 한다 — 읽고
       계산한 뒤 무조건 덮으면, 두 호출이 겹쳤을 때 늦게 도착한 쪽이 **더 작은 값으로 후퇴**시켜
       앞선 요청이 조용히 유실될 수 있다(올마이트 지적 수용). 반환값은 upsert 후 실제 저장값이다.
    🔴 파싱 불가한 쓰레기값(수동 DB 편집 등)은 **비교 대상에서 지운다.** 남겨두면
       ①`max()` 가 'zzz' 같은 lexical high 값을 floor 로 골라 `+1 second` 가 NULL → now 폴백 →
         `flag > done` 보장이 깨지고, ②조건부 upsert 도 `'2026-…' > 'zzz'` 가 false 라 영구 잠긴다.
       특히 `dock_sync_done` 이 쓰레기면 pending 판정(`done < flag`)이 영구 false 라 **버튼·자동
       반영 둘 다 죽는다** — 그래서 경고 로그를 남기고 지운다(done 은 기계가 쓰는 북마크일 뿐,
       watcher 는 자기 `.state/sync_last_flag` 로 중복을 막으므로 지워도 재실행 1회가 최대 대가).

    ⚠️ 반영 지연 가능성은 남는다: 폴러가 읽는 `SP_GET_INQ_LIST` 가 write 직후 즉시 최신인지는
       미실측이다(워커의 성공 판정 readback 은 `SP_GET_REP_INFO`/부품그리드로 **다른 프로시저**).
       여기서 "미반영이면 다시 flag" 같은 재시도 루프는 **일부러 넣지 않았다** — 종료조건이
       SVMS 응답에 달려 무한 재트리거가 될 수 있다. 늦으면 기존 1시간 폴러가 결국 채운다
       (= 최악이 이 수정 전 상태와 동일, 회귀 없음).
    """
    _ensure_api_table()
    now = query("SELECT datetime('now','localtime') t", one=True)['t']
    vals = {}
    for k in ('dock_sync_flag', 'dock_sync_done'):
        r = query("SELECT v FROM api_settings WHERE k=?", (k,), one=True)
        v = (r['v'] if r else None) or ''
        if v:
            ok = query("SELECT datetime(?) t", (v,), one=True)
            if not (ok and ok['t']):                       # sqlite 가 시각으로 못 읽는 값 = 쓰레기
                app.logger.warning('api_settings.%s 가 시각이 아님(%r) — 비교 대상에서 제거', k, v[:40])
                execute("DELETE FROM api_settings WHERE k=?", (k,))
                v = ''
        vals[k] = v
    # 같은 'YYYY-MM-DD HH:MM:SS' 포맷이라 문자열 비교 = 시각 비교(기존 pending 판정과 동일 방식)
    floor = max(vals['dock_sync_flag'], vals['dock_sync_done'])
    flag = now
    if floor >= now:
        nxt = query("SELECT datetime(?, '+1 second') t", (floor,), one=True)
        flag = (nxt['t'] if nxt else None) or now
    execute("INSERT INTO api_settings (k, v) VALUES ('dock_sync_flag', ?) "
            "ON CONFLICT(k) DO UPDATE SET v=excluded.v WHERE excluded.v > api_settings.v", (flag,))
    eff = query("SELECT v FROM api_settings WHERE k='dock_sync_flag'", one=True)
    return ((eff['v'] if eff else None) or flag)          # 경합 시 더 큰 쪽이 이긴 실효 flag
def _soa_review_attachment_path(stored_name):
    if not stored_name:
        return None
    return os.path.join(SOA_REVIEW_PDF_DIR, os.path.basename(stored_name))
def _soa_review_case_unlock(run_id, *, result=None):
    rows = query('SELECT id FROM soa_review_case WHERE queued_run_id=?', (run_id,))
    for r in rows:
        execute(
            "UPDATE soa_review_case SET queued_action=NULL, queued_run_id=NULL, queued_at=NULL, "
            "last_action_at=datetime('now','localtime'), last_action_result=?, updated_at=datetime('now','localtime') "
            "WHERE id=? AND queued_run_id=?",
            (result, r['id'], run_id),
        )
#: **행 단위** absorbing — 이 상태의 행은 status/aor_cd 를 바꿀 수 없다(DB trigger 로 강제).
#: 러너 skip 근거는 아니지만(위 참조), "상신 이력을 사후 편집하지 않는다"는 불변식 자체는 유효해
#: 그대로 유지한다. 새 SVMS 사이클은 이 행을 고치지 않고 **새 행을 INSERT** 해서 표현한다.
#: ⚠️ 여기에 상태를 추가하려면 **그 상태에서 나가는 전이가 하나도 없음**을 먼저 증명할 것.
#:    회귀 가드 = tests/test_aor_statuses.py 의 absorbing 전이 소스 스캔.
AOR_ROW_ABSORBING_STATUSES = ('submitted',)
def _aor_status_list_sql(statuses):
    """상태 튜플 → `'a','b'` SQL 리터럴 목록. 상태값을 SQL 에 박는 유일한 경로."""
    return ','.join("'%s'" % s.replace("'", "''") for s in statuses)
def _aor_absorbing_trigger_sql():
    """absorbing 상태에서 **나가는 UPDATE 자체를 DB 가 거부**하게 하는 trigger 문.

    소스 정적 스캔만으로는 부족하다(올마이트 R17): f-string SQL·소문자·`OR` 섞인 WHERE·
    다른 모듈·마이그레이션·수동 SQL 을 못 잡는다. 상신 완료 이력을 되살려 재상신하는 경로가
    생기면 SVMS 이중상신 위험이므로, 경로가 몇 개든 **DB 층에서 한 번** 막는다.
    ⚠️ `IF NOT EXISTS` 라 상수를 나중에 넓혀도 기존 trigger 는 안 바뀐다 — 그래서
       `_aor_absorbing_trigger_install()` 이 매 부팅 DROP 후 재생성하고,
       `_aor_absorbing_trigger_ok()` 가 런타임에 실물과 대조한다.
    ⚠️ 기준 상수는 `AOR_ROW_ABSORBING_STATUSES`(행 단위 불변식)다 —
       `AOR_REINGEST_TERMINAL_STATUSES`(러너 skip 계약)와 갈라졌다(2026-07-30).
    """
    lst = _aor_status_list_sql(AOR_ROW_ABSORBING_STATUSES)
    # 🔴 불변식의 단위는 **행**이다(2026-07-30 올마이트 지적으로 교정). 예전엔 "같은 canonical
    #    key 를 대표하는 다른 absorbing 행이 남아 있으면 허용"하는 `NOT EXISTS` 예외가 있었다.
    #    그건 불변식의 단위가 key 였을 때(= 러너 skip 집합의 key 가 aor_cd 였을 때) 성립하던
    #    타협이고, init_db 중복정리가 'submitted' loser 를 'duplicate' 로 강등할 수 있게 하려고
    #    뚫어둔 구멍이었다. 지금은 둘 다 사라졌다:
    #      · skip 최적화 철회(`AOR_REINGEST_TERMINAL_STATUSES = ()`) → key 단위로 볼 이유 없음
    #      · 'submitted' 가 활성군에서 빠져 중복정리 UPDATE 가 애초에 그 행을 건드리지 않음
    #    반면 같은 aor_cd 의 'submitted' 이력행이 2개 이상 쌓이는 건 이제 **정상**(SVMS 리젝→
    #    재상신 사이클 2회)이라, 예외를 남겨두면 그 순간 두 행 다 자유롭게 변경 가능해져
    #    불변식이 통째로 무력화된다. 그래서 예외를 제거했다.
    # ⚠️ init_db 의 `SET aor_cd=upper(trim(aor_cd))` 정규화는 canonical key 를 바꾸지 않으므로
    #    아래 WHEN 조건에 걸리지 않는다(= 부팅이 ABORT 되지 않는다). 비교를 raw 가 아니라
    #    canonical 로 하는 이유가 그것.
    # `aor_cd` 도 감시 대상이다(올마이트 R18 blocker): absorbing 행의 canonical key 를 바꿔
    # 다른 문서번호의 이력으로 이식하는 것도 막는다.
    # `IS NOT` 사용: `<>` 는 한쪽이 NULL 이면 NULL 로 평가돼 trigger 가 안 뜬다(status 는 NOT NULL
    # 이라 어차피 거부되지만, 방어를 제약 하나에 의존시키지 않는다).
    return ("CREATE TRIGGER IF NOT EXISTS trg_aor_draft_absorbing "
            "BEFORE UPDATE OF status, aor_cd ON aor_draft FOR EACH ROW "
            "WHEN OLD.status IN ({lst}) "
            "AND (NEW.status IS NOT OLD.status "
            "OR upper(trim(NEW.aor_cd)) IS NOT upper(trim(OLD.aor_cd))) "
            "BEGIN SELECT RAISE(ABORT, 'aor_draft: absorbing status transition denied'); "
            "END".format(lst=lst))
def _aor_absorbing_trigger_install(conn):
    """trigger 를 **항상 현재 상수와 일치하게** 심는다.

    `CREATE TRIGGER IF NOT EXISTS` 만 쓰면 정의를 바꿔도 옛 trigger 가 그대로 남고,
    `_aor_absorbing_trigger_ok()` 가 불일치로 판단해 skip 이 영구 비활성된다
    (안전하지만 최적화가 조용히 죽는다 — 올마이트 R19). DROP 후 재생성해 그 표류를 없앤다.
    """
    conn.execute("DROP TRIGGER IF EXISTS trg_aor_draft_absorbing")
    conn.execute(_aor_absorbing_trigger_sql())
#: `uq_aor_draft_active_cd` predicate 와 같은 상태군(= 동시에 하나만 존재할 수 있는 "활성" 상태).
#: = "아직 처리가 끝나지 않아 사람/러너의 다음 액션을 기다리는" 카드. 이 안에서만 aor_cd 유일.
#: ⛔ 'submitted' 는 활성이 아니다(2026-07-30) — SVMS 로 상신을 끝낸 **이력행**이다. SVMS 는
#:    리젝→수정→재상신으로 같은 aor_cd 를 다시 결재대기(STATUS='S')로 되돌리므로, 이력행이
#:    활성으로 남아 있으면 새 사이클의 카드 적재를 영구히 막는다(실측 ATGRCA2607220002 —
#:    사유·이중상신 방어 근거는 `AOR_REINGEST_TERMINAL_STATUSES` 주석 참조).
#: ⛔ 'rejected'/'failed'/'reject_failed'/'duplicate' 도 같은 이유로 활성이 아니다(원래부터).
_AOR_ACTIVE_STATUSES = ('pending', 'hold', 'approved', 'submitting',
                        'rejecting', 'reject_submitting')
def _aor_active_index_sql(name='uq_aor_draft_active_cd'):
    """활성행 유일성 index 문 — key 는 **canonical(`upper(trim(aor_cd))`)**.

    예전엔 raw 컬럼 `aor_draft(aor_cd)` 에 걸려 있었다. 실제로 중복이 안 생긴 건 모든 writer 가
    미리 `strip().upper()` 를 하기 때문인데, 그건 **관례**라 정규화를 빠뜨린 writer 가 하나만
    생겨도 `'ABC'` 와 `' abc '` 가 동시에 활성으로 앉는다. 그러면 러너 skip 의 key 인 canonical
    aor_cd 당 활성행이 2개가 되어 skip 판정 근거가 무너진다(런타임 가드가 그걸 잡아 skip 을 통째로
    끄므로 결과는 안전하지만 최적화가 죽는다). 표현식 index 로 바꾸면 그 가정을 DB 가 강제한다.
    """
    return ("CREATE UNIQUE INDEX IF NOT EXISTS %s ON aor_draft(upper(trim(aor_cd))) "
            "WHERE status IN (%s)" % (name, _aor_status_list_sql(_AOR_ACTIVE_STATUSES)))
def _aor_active_index_install(conn):
    """index 를 **항상 현재 정의와 일치하게** 심는다(raw 컬럼 구버전 포함 교체).

    `IF NOT EXISTS` 만으로는 기존 배포에 남은 옛 raw-컬럼 index 가 영영 안 바뀐다.
    ⚠️ DROP 을 먼저 하면 CREATE 가 실패했을 때 **index 가 아예 없는 상태**로 남는다 — 그건
       지금보다 나쁘다. 그래서 교체를 깨뜨릴 유일한 원인(canonical 중복 활성행)을 **먼저**
       확인하고, 있으면 손대지 않고 예외로 올린다(호출부가 판단).
    """
    want = ' '.join(_aor_active_index_sql().replace('IF NOT EXISTS ', '').split())
    row = conn.execute("SELECT sql FROM sqlite_master WHERE type='index' AND name=?",
                       ('uq_aor_draft_active_cd',)).fetchone()
    got = ' '.join((row[0] or '').split()) if row and row[0] else None
    if got == want:
        return False
    # 중복 점검·DROP·CREATE는 하나의 savepoint로 묶는다. CREATE 실패/동시 writer가
    # 있어도 ROLLBACK TO로 기존 index를 보존한다.
    conn.execute('SAVEPOINT aor_active_index_upgrade')
    try:
        dup = conn.execute(
            "SELECT upper(trim(aor_cd)) k, COUNT(*) n FROM aor_draft "
            "WHERE status IN (%s) GROUP BY k HAVING n > 1 LIMIT 1"
            % _aor_status_list_sql(_AOR_ACTIVE_STATUSES)).fetchone()
        if dup:
            raise RuntimeError(
                'aor_draft: canonical key %r 활성행이 %d 개 — 중복 정리 전에는 '
                'uq_aor_draft_active_cd 를 교체할 수 없음' % (dup[0], dup[1]))
        conn.execute("DROP INDEX IF EXISTS uq_aor_draft_active_cd")
        conn.execute(_aor_active_index_sql())
    except Exception:
        conn.execute('ROLLBACK TO aor_active_index_upgrade')
        conn.execute('RELEASE aor_active_index_upgrade')
        raise
    conn.execute('RELEASE aor_active_index_upgrade')
    return True
# ═════════════════════════════════════════════════════════════════
#  자동화 모음 (SOA/전자결재 온디맨드 버튼 → 맥미니 launchd 폴링 실행)
# ═════════════════════════════════════════════════════════════════
# task = 실행단위. mode: 'verify'(읽기전용 DRY) | 'live'(자동 승인/상신).
# 맥미니가 task+mode를 스크립트+env로 매핑(서버는 명령어를 모름 — 안전).
# ===== 비용청구(Fund Request) 2단게이트 =====
#   · review 엔진(맥)이 장금 Technical Submitted 검토결과를 POST /api/ext/fundreq/drafts (카드 적재, [검증] 버튼)
#   · 사람이 /fundreq 탭서 카드마다 승인(approved) / 리젝(rejecting, 사유) 결정
#   · [자동상신] 버튼 → 맥 fundreq_exec 가 approved=SP_SET_OPEX 상신(STATUS=U) / rejecting=STATUS=R+통보메일
# ---- SVMS 첨부(인보이스·증빙) 미리보기 cache ----
#   맥 fundreq_review 가 SP_GET_FILE 로 받은 첨부를 카드 적재 직후 idx 순서대로 업로드.
#   idx = attach_files JSON 배열의 위치(=웹/앱 목록 순서)라 이름·순서가 항상 1:1 로 맞는다.
#   읽기는 admin 세션(웹)·Bearer(앱) 전용. 파일명은 did/idx/확장자만으로 만들어 경로주입 불가.
_FUNDREQ_ATT_MIME = {
    'pdf':  'application/pdf',
    'jpg':  'image/jpeg', 'jpeg': 'image/jpeg', 'png': 'image/png', 'gif': 'image/gif',
    'heic': 'image/heic', 'heif': 'image/heif', 'webp': 'image/webp', 'bmp': 'image/bmp',
    'tif':  'image/tiff', 'tiff': 'image/tiff',
    'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'xls':  'application/vnd.ms-excel',
    'docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    'doc':  'application/msword',
    'pptx': 'application/vnd.openxmlformats-officedocument.presentationml.presentation',
    'ppt':  'application/vnd.ms-powerpoint',
}
# 브라우저가 렌더할 수 있는 것만 inline. Office 는 어차피 못 그리니 다운로드로 넘긴다
# (앱 QuickLook 은 Content-Disposition 과 무관하게 바이트만 받아 확장자로 판별).
_FUNDREQ_ATT_INLINE = {'pdf', 'jpg', 'jpeg', 'png', 'gif', 'webp', 'bmp'}
_FUNDREQ_ATT_MAX = _NON_STT_UPLOAD_MAX   # before_request 가 비-STT 업로드를 조이는 값과 일치(초과분은 거기서 413)
def _fundreq_att_ext(name):
    """확장자 정규화 — 허용목록에 없으면 None(= 업로드/서빙 거부)."""
    ext = str(name or '').rsplit('.', 1)[-1].strip().lower()
    return ext if ext in _FUNDREQ_ATT_MIME else None
def _fundreq_att_sniff_ok(ext, data):
    """확장자와 실제 바이트가 같은 계열인지 확인 — 위장 업로드로 inline 서빙 되는 걸 막는다."""
    if ext == 'pdf':
        return data[:5] == b'%PDF-'
    if ext in ('jpg', 'jpeg'):
        return data[:3] == b'\xff\xd8\xff'
    if ext == 'png':
        return data[:8] == b'\x89PNG\r\n\x1a\n'
    if ext == 'gif':
        return data[:6] in (b'GIF87a', b'GIF89a')
    if ext == 'bmp':
        return data[:2] == b'BM'
    if ext == 'webp':
        return data[:4] == b'RIFF' and data[8:12] == b'WEBP'
    if ext in ('heic', 'heif'):
        return data[4:8] == b'ftyp'
    if ext in ('tif', 'tiff'):
        return data[:4] in (b'II*\x00', b'MM\x00*')
    if ext in ('xlsx', 'docx', 'pptx'):
        return data[:4] == b'PK\x03\x04'          # OOXML = zip
    if ext in ('xls', 'doc', 'ppt'):
        return data[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'   # 레거시 OLE
    return False
def _reqgen_vsl_prefix(vtype):
    """선종 텍스트 → 선명 접두어. 컨테이너=M/V, 그 외(VLCC·탱커)=M/T(기본)."""
    t = (vtype or '').upper()
    if 'CONT' in t or 'BOX' in t:
        return 'M/V'
    return 'M/T'
def _reqgen_build_subj(vsl_cd, sheet, vnm, prefix, subject):
    """SVMS 제목 = [DOCK][<VSL_CD> <sheet>]<M/T> <선명> - <제목>. 수리(R)와 동일 규칙.
    선명에 이미 M/T·MT 등 접두어가 박혀있으면 제거 후 재부착(중복 방지)."""
    import re as _re
    nm = _re.sub(r'^(M/?[TV])\s+', '', vnm.strip(), flags=_re.I) if vnm else None
    tag = f"[{vsl_cd} {sheet}]" if vsl_cd else f"[{sheet}]"
    core = tag + (f"{prefix} {nm}" if nm else prefix)
    if subject:
        core += f" - {subject}"
    return f"[DOCK]{core}"
AUTOMATION_TASKS_BASE = {
    'jeonja':   '전자결재 자동상신',
    'fundreq':  '비용청구(Fund Request) 자동상신 — 장금·Technical·Submitted',
    'invoice_confirm': '인보이스 자동컨펌 — PIC/SUP/Remit 교정 + SVMS 컨펌 (승인 건만 처리)',
    'remittance_sync': '송금요청 후보 동기화 — SVMS Fund Request + Invoice 읽기전용',
    'soa_resend': '리젝 통보메일 재발송 (실패분)',
    'aor_prep':   'AOR(Technical) prep — Submitted AOR 카드화 (/aor 큐 적재)',
    'aor_submit': 'AOR 상신 — 승인된 건 SVMS 제출 (approve 시 자동큐)',
    'aor_reject': 'AOR 리젝 — STATUS=R + 관리사 통보메일 (reject 시 자동큐)',
    'reqgen_save': '구매청구 DRAFT 저장 — 승인된 입거 requisition 시트 SVMS 저장 (approve 시 자동큐)',
    'shipwiki_ingest': '선박 위키 신규수집 — 범주 메일 최근 7일 크롤·분류·적재 (외부 발송·승인 0)',
    'soa_vessel': '선박별 SOA 검증 — 선박코드 입력 (검증단계까지만: 체크박스+리젝리마크, 승인·출금·제출·메일 안 함)',
}
# verify=읽기전용 / live=자동승인·상신 / reject_dry=리젝후보표시 / reject_mark=리젝라인체크 / reject_submit=리젝제출+메일 / remark_cleanup=컨펌된 라인 잔존 RJT_RMK 삭제(SVMS UI버그 보정)
AUTOMATION_MODES = ('verify', 'live', 'reject_dry', 'reject_mark', 'reject_submit', 'remark_cleanup')
def soa_task_key(group_key):
    """그룹키(G1/SKRT) → 자동화 task 키(soa_g1/soa_skrt). 러너도 같은 규칙으로 역변환."""
    return 'soa_' + str(group_key).lower()
def _soa_task_label(g, owner_map=None):
    """허브 버튼에 뜰 문구. dynamic_owner 는 owner 스냅샷 기준 현재 편입 선박을 노출."""
    if g['mode'] == 'dynamic_owner':
        mem = _soa_group_members(g, owner_map)
        body = ('·'.join(mem) + ' · 신규선 자동편입') if mem else '전체·신규선 자동편입(현재 편입 미확인)'
    else:
        body = '·'.join(g['vessels']) if g['vessels'] else '선박 미지정'
    tail = ' +출금상신' if g['category'] == 'skrt' else ''
    return f"{g['label']} ({body}){tail}"
def automation_tasks():
    """정적 task + DB soa_group 파생 task 병합(SOA 그룹이 앞). 화면·검증 공용 SSOT."""
    out = {}
    try:
        owner_map = _soa_owner_map()
        for g in _soa_groups_load(active_only=True):
            out[soa_task_key(g['key'])] = _soa_task_label(g, owner_map)
    except sqlite3.Error:
        pass          # DB 미초기화 등 → 정적 task 만. 그룹 버튼은 안 뜨고, 실행도 거부(fail-closed)
    out.update(AUTOMATION_TASKS_BASE)
    return out
def _automation_enabled():
    row = query("SELECT v FROM api_settings WHERE k='automation_enabled'", one=True)
    return (row['v'] if row else '1') != '0'
def _soa_vessel_codes_from_params(p):
    raw = p.get('vsl_cds')
    if raw is None:
        raw = p.get('vsl_cd')
    if isinstance(raw, str):
        candidates = re.split(r'[\s,;/]+', raw.strip().upper())
    elif isinstance(raw, list):
        candidates = [str(x or '').strip().upper() for x in raw]
    else:
        candidates = []
    out = []
    seen = set()
    for code in candidates:
        if not code:
            continue
        if not re.match(r'^[A-Z]{4}$', code):
            raise ValueError('선박코드(VSL_CD 4자 영문)를 정확히 입력하세요.')
        if code not in seen:
            out.append(code)
            seen.add(code)
    if not out:
        raise ValueError('선박코드(VSL_CD 4자 영문)를 정확히 입력하세요.')
    if len(out) > 5:
        raise ValueError('선박별 SOA 검증은 한 번에 최대 5척까지 실행합니다.')
    return out
def _soa_vessel_params(p, vsl):
    fm, to, sl = (str(p.get(k) or '').strip() for k in ('fm_dm', 'to_dm', 'sl_tp'))
    def _ym(v): return bool(re.match(r'^[0-9]{6}$', v)) and '01' <= v[4:] <= '12'
    if (fm and not _ym(fm)) or (to and not _ym(to)) or (fm and to and fm > to):
        raise ValueError('기간(YYYYMM, 시작<=끝)을 확인하세요.')
    if sl and sl not in ('04', '05'):
        raise ValueError('부서는 05(Technical)/04(Crew)만.')
    review_model = str(p.get('review_model') or 'auto').strip()
    if review_model not in ('auto', 'claude-haiku-4-5', 'openai/gpt-5.4-mini'):
        raise ValueError('검증모델 선택값이 올바르지 않습니다.')
    pp = {'vsl_cd': vsl}
    if fm: pp['fm_dm'] = fm
    if to: pp['to_dm'] = to
    if sl: pp['sl_tp'] = sl
    pp['review_model'] = review_model
    return pp
# ===================== Dock Procurement (입거 발주현황 트래커) =====================
_DOCKPROC_CAT_NM = {'R': 'SHORE REPAIR', 'S': 'SPARE', 'ST': 'STORE',
                    'P': 'PAINT', 'SY': 'SHIPYARD'}
def _dockproc_cat_code(req_no):
    import re as _re
    m = _re.match(r'^(SY|ST|R|S|P)\d+$', (req_no or '').strip().upper())
    return m.group(1) if m else None
def _dockproc_source(code, prepared_by):
    """견적출처 결정: 페인트P·조선소SY=MAIL(메일견적) / MANAGER=AOR / OWNER(R·S·ST)=SVMS."""
    if code in ('P', 'SY'):
        return 'MAIL'
    if (prepared_by or '').strip().upper() == 'MANAGER':
        return 'AOR'
    return 'SVMS'
# Phase 2 역동기화: SVMS Status → 진행단계 rank(누적). HQ Canceled=무시(맵 없음→0).
_DOCKPROC_STATUS_RANK = {
    # 1=견적작성 / 2=벤더제출 / 3=벤더컨펌·결재상신 / 4=발주완료 (누적).
    # HQ Canceled·미등재=무시(rank0).
    'HQ CONFIRMED': 1,          # 견적작성 (수리·구매 공통)
    'QUOTATION INQUIRY': 2,     # 벤더제출(견적의뢰)
    # 구매 결재 반려는 견적/업체선택 데이터가 살아 있는 재상신 가능 단계다.
    # 2026-08-03 BGBB S10 실측: 상세 STATUS=X, STATUS_NM='HQ Rejected'.
    'HQ REJECTED': 2,           # 벤더제출로 복귀(수정 후 다시 업체선택·상신 가능)
    # 'Submit'은 발주가 아니라 **업체 선택 후 결재 상신** 단계. 벤더 견적제출(VNDR_STATS='Submitted')과
    # 구분되지만 발주서 미발행(ODR_YN='N')이라 발주완료로 올리면 안 된다.
    # 반면 실발주건(SAPS)은 헤더 'HQ Ordered' + 벤더 'Ordered' + ODR_YN='Y'.
    # 4로 두면 "발주완료인데 금액 —" 이 구조적으로 발생함(dock_sync._repair_order 는 'Ordered'만 읽음).
    'SUBMIT': 3,                # 벤더컨펌 (수리 — 업체 선택 후 결재 상신)
    'HQ ORDERED': 4,            # 발주완료 (수리 — 실발주)
    'ORDERED': 4,               # 발주완료 (구매 발주)
    'VENDOR CONFIRMED': 4,      # 발주완료 (구매 — 업체확정 완료)
    # 'Approval(Procssing)' 도 발주완료가 아니다 — 업체 선택 후 결재 진행 중인 rank3.
    # 근거: BELGIUM B S10(BGBBES2607B11) PC_PRO 행이 스스로 밝힘 —
    #   ODR_STEP = "[Order] Order is Progressing (Not Approved)" · ODR_STATUS_CD='A'
    #   ODR_NO 는 이미 발급('BGBBES2607B11A')인데 SP_GET_ODR_LIST(BGBB)=0행 → 금액 소스에 아직 없음.
    # 4로 두면 "발주완료인데 금액 —"이 또 구조적으로 발생함(실제로 이 행 1건이 그렇게 떴다).
    # ⚠️ODR_NO 존재만으로는 구매 발주근거가 약하다는 반례이기도 함(승인 전에도 번호가 붙음).
    'APPROVAL(PROCSSING)': 3,   # 벤더컨펌 (구매 — 업체 선택 후 발주 승인 진행 중)
    # 같은 BGBB S10을 HQ Rejected에서 재상신한 직후 라이브 상태.
    'HQ PROGRESSING': 3,        # 벤더컨펌 (구매 — HQ 결재 진행 중)
}
def _dockproc_status_rank(status):
    return _DOCKPROC_STATUS_RANK.get((status or '').strip().upper(), 0)
# 🔴 rank 0 중에서도 **되돌림을 허용하는 '견적의뢰 이전' 라벨 allowlist**(2026-08-03 실사고로 신설).
#   rank 0 은 원래 통째로 `link_only`(단계 미변경)라, 상신/견적요청이 나간 건이 SVMS 에서 **회수**돼
#   헤더가 rank 0 라벨로 되돌아오면 `stg_*` 가 영구히 켜진 채 남아 재요청 게이트가 영구 잠겼다.
#   실사고: BGBBME26073116([BGBB R22]) 견적요청 LIVE 성공(18:05) → 형이 SVMS 에서 회수 →
#     SVMS 는 목록·상세 모두 STATUS='AP'(HQ Received)·벤더그리드 0행인데 DB 는 `stg_vendor=1` +
#     `svms_status='Quotation Inquiry'`(`api_ext_dock_inquiry_result` 의 낙관적 표시)로 고착 →
#     `_dock_inq_blocked` 가 '이미 벤더제출 이후 단계'로 409 → 큐 재적재 불가.
#   ⚠️기존 주석은 "반려는 헤더가 'RE'(rank 2)로 돌아오니 갱신 경로에 걸린다"고 전제했는데 **틀렸다** —
#     회수는 'AP'(HQ Received, rank 0)로 돌아온다. 그래서 그 안전장치가 발동하지 않았다.
#   allowlist 로 좁히는 이유: 미지의 rank 0 라벨까지 되돌리면 처음 보는 상태 하나로 단계가 조용히
#     꺼진다. 여기 등재된 **확인된 pre-inquiry 라벨만** 되돌리고 나머지는 종전대로 link_only 다.
#   ⚠️빈 라벨('')은 절대 넣지 않는다 — SVMS 미연결 수동관리 행이고(2026-08-03 라이브 73행 중 50행이
#     사람이 켠 단계 보유) 넣으면 그 수동 체크를 sync 가 지운다.
#   🔴 **실측된 라벨만** 넣는다(올마이트 지적 수용). 'Approved'(라이브 1행, 단계 0)는 구매/수리 의미
#     충돌이 확인 안 돼 여전히 넣지 않는다 — 효과 0 · 위험 잠재.
#   🔴 2026-08-04 추가: **'VSL Approved' 도 회수 경로로 실측됐다.** 구매 INQ `BGBBES2607B41` 을
#     `PKG_PC_INQ.SP_SET_INQ_RTN` 으로 회수한 직후 REQ `BGBBES2607B4` 헤더가 `STATUS='C'(HQ Confirmed)`
#     → **`STATUS='N'(VSL Approved)`·CFM_DT=null** 로 돌아갔다(화면 회수 안내문 "청구서 컨펌이전
#     상태로 됩니다" 그대로). 즉 구매 회수의 도착지는 상황에 따라 'HQ Confirmed'(2026-08-03 실측)
#     **또는 'VSL Approved'** 다. 넣지 않으면 그 회수건이 `link_only` 로 빠져 위 BGBBME26073116 과
#     **똑같이 영구 잠긴다**. 라이브 영향은 0 — 이 라벨 10행 전원 `stg_*` 0 이라 no-op 이고,
#     되돌림 대상은 방금 회수된 행뿐이다. (구 주석의 "회수 경로로 실측된 라벨은 'HQ Received' 하나"는
#     이 실측으로 폐기.)
#   ⚠️ stale·순서역전 sync 가 이 라벨을 실어와 진행 중 행을 되돌릴 위험은 **'HQ RECEIVED' 와 동일한
#     기존 위험**이고 새로 생긴 게 아니다(올마이트 지적). freshness 로 걸러내고 싶어도 dock_procure
#     sync payload 에는 시각·리비전 필드가 아예 없다(실측: `status`/`inq_no`/`amt`/`quotes`/`files`/
#     `ordered_evidence` 뿐) ⇒ 방어선은 아래 fail-closed 가드(발주완료·발주금액·제출수>0)와 워커
#     pre-read 게이트 두 겹뿐이다. 라벨을 하나 늘릴 때마다 이 가드가 유일한 안전선임을 기억할 것.
_DOCKPROC_PRE_INQUIRY = {
    'HQ RECEIVED',              # 본선 요청이 HQ 에 접수된 상태 = 견적의뢰 전. 회수 시 여기로 돌아온다.
    'VSL APPROVED',             # 본선 승인·본사확인 전. 구매 회수가 컨펌을 풀면 여기로 돌아온다(실측).
}
def _dockproc_submit_has_quotes(raw):
    """`svms_submit`(SVMS `SUBMIT` = "제출수/요청수") 가 **실제 견적 제출 흔적**인가 — 순수함수.

    🔴 왜 필요한가(2026-08-04 실측): 되돌림 fail-closed 가드가 `svms_submit` 을 **존재하기만 하면**
       발주흔적으로 봤다. 그런데 견적요청이 나가면 이 칸은 곧바로 `"0/0"`·`"0/1"` 처럼 **제출 0** 값으로
       채워진다 ⇒ 회수된 행은 언제나 이 가드에 걸려 `link_only` 로 빠지고, allowlist 를 고쳐도 단계가
       영구히 켜진 채 남는다(S14 `BGBBES2607B4`=`"0/0"`, 정상 회수건도 `"0/1"` 이라 같은 결과).
    ⇒ 판정을 **분자(제출수) > 0** 으로 좁힌다. 제출된 견적이 있으면 하류 데이터(금액·첨부)가 존재할 수
       있으니 종전대로 보호하고, 제출 0 이면 지울 이력이 없으므로 회수 되돌림을 허용한다.
    ⚠️**파싱 실패는 흔적 있음(True)** — 처음 보는 형식 하나로 이력을 조용히 지우지 않는다(닫힘 쪽 실패).
       괄호도 **쌍으로만** 인정한다(`"(0/1"` 같은 반쪽은 미지 형식 → True). 올마이트 지적 수용.
    🔴 `sub_quotes`/`att_files` 는 **여기서 보지 않는다**(올마이트 대안 반대 — 근거 실측):
       그 두 칸은 폴러 payload 의 3상태 계약이 정본이라 **키 없음=기존 유지 / `[]`=0건 확정 → clear** 로
       이미 보호된다. 즉 저장된 값은 "지금 SVMS 에 견적이 있다"는 증거가 아니라 지난 sync 의 스냅샷이다.
       이걸 되돌림 가드에 넣으면 회수건이 옛 스냅샷 때문에 다시 영구 잠기는 **똑같은 부류의 버그**가
       다른 칸에서 재발한다(#12 R39 = `quotes:[]` 로 회수와 정합하게 clear 되는 경로가 그 반례).
       라이브 실측(2026-08-04, 161행): 제출0/NULL + 하류데이터 보유 + 발주흔적 없음 조합 = **0행**.
    """
    s = (raw or '').strip()
    if not s:
        return False                                   # 값 없음 = 흔적 없음
    if len(s) >= 2 and s[0] == '(' and s[-1] == ')':   # SVMS 목록 `SUBMIT` 은 "(0/4)" 꼴, DB 는 "0/4" 꼴
        s = s[1:-1].strip()
    m = re.fullmatch(r'(\d+)\s*/\s*\d+', s)
    if not m:
        return True                                    # 미지 형식 = fail-closed
    return int(m.group(1)) > 0
def _dockproc_submit_pair(raw):
    """`svms_submit`("3/5"·"(3/5)") → (제출수, 요청수). 미지 형식은 None.

    ⚠️`_dockproc_submit_has_quotes` 와 **판정 목적이 다르다** — 저건 되돌림 가드용이라 미지 형식을
      True(흔적 있음)로 fail-closed 하지만, 여기는 푸시 트리거라 미지 형식이면 **아무 것도 안 한다**
      (모르는 값으로 알림을 만들면 형에게 틀린 숫자가 간다).
    """
    s = (raw or '').strip()
    if len(s) >= 2 and s[0] == '(' and s[-1] == ')':
        s = s[1:-1].strip()
    m = re.fullmatch(r'(\d+)\s*/\s*(\d+)', s)
    return (int(m.group(1)), int(m.group(2))) if m else None
# 결재반려 라벨 — 웹 `SBM_REJ_LABELS`·iOS `rejectedLabels` 와 1:1. 부분일치 금지(정확일치).
_DOCKPROC_REJ_LABELS = {'HQ REJECTED'}
def _dockproc_push_events(row, status, ordered, submit, vendor, amt, cur):
    """Dock 단계전이 → 푸시 이벤트 목록. **판정만 하고 발송은 안 한다**(테스트 가능하게 분리).

    🔴 `event_key` 설계 = "중복 1건 < 미탐 1건"(BV 감시 교훈)이되 폭주는 막는다.
       키에 **직전 `updated_at`(전이 직전 행 상태의 지문)** 을 넣는다 — SVMS 에는 회차 번호가 없고,
       상태값만으로 키를 만들면 **되풀이되는 전이가 첫 번째 것과 충돌해 두 번째부터 영구히 묻힌다**
       (올마이트 지적: 제출수 3→2→3 재증가, 같은 날 2차 반려). `updated_at` 은 전이가 일어날 때마다
       sync 가 새로 찍으므로 회차 대용으로 쓸 수 있고, **시계가 아니라 데이터에서 나오므로**
       fast/full 이 같은 변화를 동시에 봐도 두 프로세스가 같은 키를 만든다(= 중복발송 안 남).
       · `dock_quote`  : 제출수까지 넣어 제출이 늘 때마다 1번.
       · `dock_ordered`: 발주완료 1번(되돌렸다 다시 완료되면 그건 새 전이라 다시 1번).
       · `dock_reject` : 반려 1번. 재상신→재반려는 그 사이에 `updated_at` 이 바뀌므로 또 온다.
    🔴 제출수 증가는 **목록 라벨이 그대로여도 일어난다**('Quotation Inquiry' 유지). 그래서 라벨 전이로는
       절대 못 잡고, 이 판정에는 상세조회로 채운 `submit` 이 반드시 있어야 한다(fast 폴러가 견적진행
       중인 행을 항상 상세조회하는 이유).
    """
    ev = []
    _g = (lambda k: (row[k] if k in row.keys() else None))
    rid = _g('id')
    seq = (_g('updated_at') or '').strip() or '0'     # 전이 회차 대용(위 docstring)
    head = ('[%s] %s' % ((_g('vsl_nm') or '').strip(), (_g('req_no') or '').strip())).strip()
    subj = (_g('subject') or '').strip()
    tail = (' · ' + subj[:60]) if subj else ''
    # ① 벤더 견적제출 = 분자 증가.
    was = _dockproc_submit_pair(_g('svms_submit'))
    now = _dockproc_submit_pair(submit)
    if now and now[0] > (was[0] if was else 0):
        ev.append({'kind': 'dock_quote', 'collapse': 'dq%s' % rid,
                   'event_key': 'dock_quote:%s:%s:%s/%s' % (rid, seq, now[0], now[1]),
                   'title': '%s 견적 제출' % head,
                   'body': '%s/%s개 업체 제출%s' % (now[0], now[1], tail)})
    # ② 발주완료. 분할발주는 **전 업체 승인**돼야 여기 온다(상류 부분완료 게이트가 이미 rank 를 내림).
    #   ⚠️`not stg_order` 때문에 **사람이 먼저 발주완료를 켜둔 행은 SVMS 추인 때 푸시가 안 간다**
    #     (올마이트 2026-08-07 지적 → 실측 판정: 의도된 억제, `stg_manual` floor 도입 전과 동일 동작).
    #     사람이 직접 켠 상태를 SVMS 가 뒤늦게 따라오는 건 형에게 새 정보가 아니다.
    if ordered and not _g('stg_order'):
        money = ''
        if isinstance(amt, (int, float)):
            money = ' · %s %s' % ((cur or '').strip(), format(amt, ',.0f'))
        ev.append({'kind': 'dock_ordered', 'collapse': 'do%s' % rid,
                   'event_key': 'dock_ordered:%s:%s' % (rid, seq),
                   'title': '%s 발주완료' % head,
                   'body': '%s%s%s' % ((vendor or '업체 미상'), money, tail)})
    # ③ 결재 반려 — 라벨 정확일치. 이미 반려 상태였으면 재발송 안 함.
    cu = (status or '').strip().upper()
    pv = (_g('svms_status') or '').strip().upper()
    if cu in _DOCKPROC_REJ_LABELS and pv not in _DOCKPROC_REJ_LABELS:
        ev.append({'kind': 'dock_reject', 'collapse': 'dr%s' % rid,
                   'event_key': 'dock_reject:%s:%s' % (rid, seq),
                   'title': '%s 결재 반려' % head,
                   'body': '%s%s' % ((status or '').strip(), tail)})
    return ev
# 견적요청(벤더 Submit)이 쓰는 SVMS 문서종류 — cat_code 기준. 페인트(P)/기타(SY)는 봉투 자체가 없다.
_DOCK_INQ_DOC = {'R': 'MARP', 'S': 'PCRQ', 'ST': 'PCRQ'}
def _dockproc_inq_target(row):
    """그 행의 견적요청 대상 = (문서종류, SVMS 키). 키가 빈 문자열이면 '연결 안 됨'.

    🔴 키가 컬럼마다 다르다 — 수리는 `svms_req_no`(=REP_CD)지만 **구매는 `svms_pc_req_no`(=REQ_NO)** 다.
       구매의 `svms_req_no` 는 견적요청이 나간 **뒤에** 발급되는 INQ_NO 라서 요청 전에는 항상 NULL 이고
       (실측 2026-08-03 라이브 87행), 그 칸을 REQ_NO 로 덮으면 Phase ③ 상신이 INQ_NO 를 잃는다.
    """
    g = (row.get if isinstance(row, dict)
         else (lambda k, d=None: (row[k] if k in row.keys() else d)))
    doc = _DOCK_INQ_DOC.get((g('cat_code') or '').strip().upper())
    if not doc:
        return '', ''
    return doc, ((g('svms_req_no') if doc == 'MARP' else g('svms_pc_req_no')) or '').strip()
# 견적요청이 열리는 SVMS 헤더 라벨 — 워커 pre-read(`inquiry_watch.py`)의 코드 게이트와 1:1 대응이다.
#   수리 `OPEN_STATUS=('AP','RC')`                    = 'HQ Received' / 'HQ Confirmed'
#   구매 `PC_OPEN_STATUS=('C',)` + `PC_CONFIRM_STATUS=('N',)` = 'HQ Confirmed' / 'VSL Approved'
# 🔴 'VSL APPROVED'(=구매 `STATUS='N'`) 추가 근거(2026-08-03 형 요청 "컨펌 버튼 누르는 기능까지 포함"):
#    워커가 견적요청 **직전에** SVMS Confirm(`SP_SET_REQ_INFO`+STATUS='C')을 대신 눌러 'C' 로 올린 뒤
#    요청한다. 그래서 이 라벨은 더 이상 '헛클릭'이 아니다. 수리(MARP)에는 추가하지 않는다 —
#    수리 쪽 대응 코드가 미실측이고, 워커도 수리 Confirm 은 `AP`/`RC` 에서만 한다.
_DOCK_INQ_STAGE_OK = {'MARP': ('HQ RECEIVED', 'HQ CONFIRMED'),
                      'PCRQ': ('HQ CONFIRMED', 'VSL APPROVED')}
# 그 단계가 **아님이 실측된** 라벨만 미리 회색처리한다(2026-08-03 라이브 전수 + 폴러 관측값).
# 🔴 allowlist 반전(=OK 아니면 전부 차단)을 쓰지 않는 이유: 아직 못 본 라벨과 폴러가 못 채운 빈
#    라벨까지 닫혀 **실제로 가능한 건이 영구히 막힌다**. 최종 안전선은 워커 pre-read 이고(SVMS
#    실시간 재조회 · 불일치면 write 0 으로 거부), 이 게이트는 헛클릭을 줄이는 표시층이다.
_DOCK_INQ_STAGE_BLOCK = {
    # 🔴 구매(PCRQ)는 이 라벨을 **더 이상 차단하지 않는다** — `_DOCK_INQ_STAGE_OK['PCRQ']` 가 먼저
    #    통과시키고, 워커가 견적요청 직전에 SVMS Confirm 을 대신 누른다(2026-08-03).
    #    여기 남겨두는 건 **수리(MARP) 등 다른 문서종류**에서 같은 라벨이 나올 때를 위한 사유다.
    'VSL APPROVED':      '본선 승인 단계 — 이 문서종류는 HQ 확인 후에만 견적요청 가능',
    'HQ REJECTED':       'HQ 반려됨',
    'QUOTATION INQUIRY': '이미 견적요청된 건',
    'VENDOR CONFIRMED':  '이미 벤더 확정 이후 단계',
    'ORDERED':           '이미 발주 이후 단계',
    'HQ ORDERED':        '이미 발주 이후 단계',
}
def _dockproc_inq_stage_block(doc, svms_status):
    """SVMS 단계 라벨만 보는 견적요청 게이트 — 사유(문자열) 또는 None(=열어둠).

    쿼리 0 인 순수 함수라 목록 API 가 행마다 불러도 비용이 없다. 웹·iOS 가 이 결과(`inq_block`)만
    보고 버튼을 회색처리하므로 두 클라이언트가 각자 라벨을 해석하다 어긋나는 일이 없다."""
    lbl = str(svms_status or '').strip()
    if not lbl or not doc:
        return None                      # 라벨 미관측 = 판단 보류(워커 pre-read 가 최종 판정)
    if lbl.upper() in _DOCK_INQ_STAGE_OK.get(doc, ()):
        return None
    why = _DOCK_INQ_STAGE_BLOCK.get(lbl.upper())
    if not why:
        return None                      # 처음 보는 라벨 — 추측으로 닫지 않는다
    # 라벨은 폴러가 채운 스냅샷이라 SVMS 가 방금 움직였으면 낡을 수 있다 — 영구 차단으로 오해하지
    # 않도록 해소 경로를 사유에 같이 적는다(`_dock_inq_prior` 의 '동기화 후 다시 열림'과 같은 규약).
    return 'SVMS 단계가 %s — %s (동기화 후 단계가 바뀌면 다시 열림)' % (lbl, why)
_DOCKPROC_QUOTE_MAX = 20            # 한 건에 붙는 벤더 수 상한(표시전용 스냅샷이라 넉넉하되 무한 아님)
_DOCKPROC_GAP_MAX = 5               # 업체당 보여줄 결함 품목 줄 수 상한(전체 건수는 gap_n 이 따로 말한다)
def _dockproc_hard_n(q, gap_n=None):
    """업체 견적 스냅샷 한 건의 **상신을 막는 결함 수**. 읽기 규칙을 여기 한 곳만 둔다.

    정규화(적재)와 상신 preview(조회)가 각자 계산하면 같은 행을 놓고 화면이 서로 다른 말을 한다.
    🔴 `hard_n` 도 `gaps[].hard` 도 없는 스냅샷(hard/soft 분리 이전 폴러가 적재한 행)은 차단 여부를
       **모른다**. 그때는 0 이 아니라 `gap_n` 으로 본다 — 모르는 걸 "상신 가능" 이라고 말하면 형이
       버튼을 누른 뒤 실패하고, 반대 방향은 한 번 더 확인하는 것뿐이다. 다음 sync 에서 새 폴러
       값이 들어오면 자동 해소된다(올마이트 2026-08-04 지적 반영).
    """
    def _c(v):
        try:
            return int(v or 0)
        except (TypeError, ValueError, OverflowError):    # int(float('inf'))=OverflowError
            return 0
    gaps = [g for g in (q.get('gaps') if isinstance(q.get('gaps'), list) else []) if isinstance(g, dict)]
    n = _c(q.get('gap_n')) if gap_n is None else gap_n
    n = max(0, min(999, n), len(gaps))
    if q.get('hard_n') is None and not any('hard' in g for g in gaps):
        return n
    return min(n, max(0, _c(q.get('hard_n')), sum(1 for g in gaps if g.get('hard') is True)))
def _dockproc_norm_quotes(raw):
    """폴러가 보낸 **벤더 제출견적** 목록 → canonical JSON 문자열(쓸 값이 없으면 None).
    발주금액(quote_amt)과 다른 값이다 — 제출견적은 아직 발주가 아니므로 절대 섞지 않는다.
    `cd`(SVMS VNDR_CD)는 Phase ③ 상신에서 `SELETED_VDR` 로 쓰는 업체코드다. 표시용 `nm` 과 달리
      **정본 식별자**이므로 형식검증(대문자·숫자 1~20)을 통과하지 못하면 None 으로 떨군다.
    표시전용이라 값 신뢰보다 형태 방어가 우선: 개수 캡·타입 강제·통화 3글자 검증.

    canonical 두 겹(멱등 목적):
      · 원소 키 정렬(sort_keys) + 고정 separators — dict 순서·공백 흔들림 흡수
      · **리스트 자체를 정렬** — 벤더 배열 순서는 의미가 없는데 SVMS 가 순서를 바꿔 주면 같은 견적
        집합이 '변경'으로 잡혀 매 폴링마다 UPDATE 가 돈다(올마이트 지적).

    '최저' 판정도 여기서 한다(`best:1` 플래그). 프런트에서 하면 JS 테스트 target 이 없어 검증 공백이
    생기고 통화 혼재 비교 버그가 조용히 살아난다 — 그래서 테스트되는 층으로 끌어내렸다.
    비교 규칙: 금액 있는 견적 **전원이 usd 를 가질 때만** usd 로 비교. 하나라도 없으면 통화가 전부
      같을 때만 원표시금액으로 비교하고, 통화가 섞였으면 비교를 포기한다(best 없음 → 화면에 '최저' 안 씀).
      부분집합만 비교하면 usd 없는 견적이 조용히 후보에서 빠져 오답이 된다."""
    if not isinstance(raw, list):
        return None
    out = []
    def _num(v):
        try:
            n = None if v in (None, '') else float(str(v).replace(',', ''))
        except (TypeError, ValueError):
            return None
        return n if (n is None or math.isfinite(n)) else None   # inf/nan 은 JSON 직렬화도 못 함

    for q in raw[:_DOCKPROC_QUOTE_MAX]:
        if not isinstance(q, dict):
            continue
        try:
            att = int(q.get('att') or 0)
        except (TypeError, ValueError, OverflowError):    # int(float('inf'))=OverflowError (올마이트 지적)
            att = 0
        cur = str(q.get('cur') or '').strip().upper()
        cd = str(q.get('cd') or '').strip().upper()[:20]
        # 품목 견적 결함(견적 미제출 · 단가 0 …). 🔴 판정은 맥 워커가 **상신 게이트와 같은 함수**
        #   (`dock_items.item_gaps`)로 이미 했다 — 여기서는 형태만 방어한다(개수·길이 캡). 서버가
        #   다시 판정하면 두 판정이 갈려서 "화면은 조용한데 상신만 실패" 가 재발한다.
        #   구버전 폴러는 이 키를 안 보내 gap_n=0 이 되고, 화면은 경고를 안 띄운다(하위호환).
        #   ⚠️ 키가 항상 있으므로 배포 후 첫 sync 에서 기존 S/ST 행은 canonical 문자열이 달라져
        #      한 번 UPDATE 된다(값 변화 없음, 멱등은 그 다음 sync 부터 복귀).
        #   `hard_n` = 그 중 상신을 실제로 막는 건수. 단가 0 은 hard 가 아니다(형 지시 = 인폼만).
        def _cnt(v):
            try:
                return int(v or 0)
            except (TypeError, ValueError, OverflowError):    # int(float('inf'))=OverflowError
                return 0
        gap_n = _cnt(q.get('gap_n'))
        gaps = []
        for g in (q.get('gaps') if isinstance(q.get('gaps'), list) else [])[:_DOCKPROC_GAP_MAX]:
            if not isinstance(g, dict):
                continue
            gaps.append({'seq': str(g.get('seq') or '').strip()[:20],
                         'why': str(g.get('why') or '').strip()[:20],
                         'hard': g.get('hard') is True,
                         'label': str(g.get('label') or '').strip()[:200]})
        # 라벨이 건수보다 많으면(캡·이상값) 건수를 라벨 수로 올린다 — "외 −1건" 같은 표시 방지.
        gap_n = max(0, min(999, gap_n), len(gaps))
        # hard 는 전체를 넘을 수 없다. 폴러가 hard_n 을 안 보내도(구버전) 라벨의 hard 플래그로 복원하고,
        # 그것마저 없으면 gap_n 으로 본다 — 규칙 정본은 `_dockproc_hard_n` 한 곳(조회 경로와 공유).
        hard_n = _dockproc_hard_n(q, gap_n)
        out.append({'nm': str(q.get('nm') or '').strip()[:120],
                    'gap_n': gap_n,
                    'hard_n': hard_n if gap_n else 0,
                    'gaps': gaps if gap_n else [],
                    # cd = SVMS VNDR_CD. Phase ③ 상신 봉투의 SELETED_VDR 가 이 값이다.
                    # 구버전 폴러는 안 보내므로 None 가능 — 그 경우 상신 대상에서 제외(fail-closed).
                    'cd': cd if re.fullmatch(r'[A-Z0-9]{1,20}', cd) else None,
                    'amt': _num(q.get('amt')),
                    'usd': _num(q.get('usd')),           # 달러환산액 — 통화 혼재 시 '최저' 비교는 이걸로만
                    # 구매(S/ST) 표시용: amt/gross=P_RS_VNDR.TAMT(승인 스냅샷 정본),
                    # dc_rate=P_RS_VNDR.DIS_RATE, final=P_RS_ODR.TAMT/USD_TAMT.
                    # 수리(R)·구버전 폴러는 이 키가 없어 모두 None 이며 기존 UI/승인 계약 유지.
                    'gross_amt': _num(q.get('gross_amt')),
                    'dc_rate': _num(q.get('dc_rate')),
                    'final_amt': _num(q.get('final_amt')),
                    'final_usd': _num(q.get('final_usd')),
                    'cur': cur if re.fullmatch(r'[A-Z]{3}', cur) else None,
                    'att': max(0, min(99, att)),
                    'st': str(q.get('st') or '').strip()[:20]})
    if not out:
        return None
    out.sort(key=lambda q: (q['nm'], q['cd'] or '', q['cur'] or '',
                            q['amt'] is None, q['amt'] or 0.0, q['st']))
    priced = [q for q in out if q['amt'] is not None]
    best = None
    if priced:
        if all(q['usd'] is not None for q in priced):
            best = min(priced, key=lambda q: q['usd'])
        elif len({q['cur'] for q in priced}) == 1:
            best = min(priced, key=lambda q: q['amt'])
    if best is not None:
        best['best'] = 1
    return json.dumps(out, ensure_ascii=False, sort_keys=True, separators=(',', ':'))
_DOCKPROC_ATT_MAX = 20              # 한 건에 붙는 견적서 파일 수 상한(실측 BGBB 최대 2 — 넉넉하되 무한 아님)
def _dockproc_norm_files(raw):
    """폴러가 보낸 **벤더 견적서 첨부 목록** → canonical JSON 문자열(쓸 값이 없으면 None).
    원소 = {nm 파일명, kb 크기, vndr 업체코드, vnm 업체명, dt 업로드일, sv SVMS 저장명}.

    ⚠️`kb` 는 SVMS `FILE_SIZE` 원값이고 **단위가 KB 지 bytes 가 아니다**(실측: 362 → 실제
      370,998 bytes). bytes 로 읽어서 화면에 쓰면 371KB 파일이 '362B' 로 보인다.

    **배열 위치(idx)가 preview cache 파일명**이 되므로 정렬을 서버가 못박는다 — SVMS 응답 순서가
    흔들리면 같은 파일이 다른 idx 로 옮겨가 캐시된 PDF 와 목록의 이름이 어긋난다(= 형이 A업체
    견적서를 열었는데 B업체 파일이 뜨는 사고). 정렬키 1순위는 SVMS 저장명(`sv`)으로, 이름이
    같은 두 파일도 구분된다.
    canonical JSON(키 정렬·고정 separators)은 sub_quotes 와 같은 이유 — 멱등 비교용."""
    if not isinstance(raw, list):
        return None
    out = []
    for f in raw[:_DOCKPROC_ATT_MAX]:
        if not isinstance(f, dict):
            continue
        nm = str(f.get('nm') or '').strip()[:160]
        if not nm:
            continue                                     # 이름 없는 첨부는 열 수도 표시할 수도 없다
        try:
            kb = int(float(str(f.get('kb') or 0).replace(',', '')))
        except (TypeError, ValueError, OverflowError):    # int(float('inf'))=OverflowError
            kb = 0
        out.append({'nm': nm,
                    'kb': max(0, min(99_999_999, kb)),
                    'vndr': str(f.get('vndr') or '').strip()[:20],
                    'vnm': str(f.get('vnm') or '').strip()[:120],
                    'dt': str(f.get('dt') or '').strip()[:20],
                    'sv': str(f.get('sv') or '').strip()[:160]})
    if not out:
        return None
    out.sort(key=lambda f: (f['sv'], f['nm'], f['vndr']))
    return json.dumps(out, ensure_ascii=False, sort_keys=True, separators=(',', ':'))
_DOCKPROC_ORDER_MAX = 10            # 한 청구를 나눠 발주할 수 있는 업체 수 상한(실측 2 — 넉넉하되 무한 아님)
def _dockproc_norm_orders(raw):
    """폴러가 보낸 **발주서(ODR_NO)별 업체·금액** 목록 → canonical JSON 문자열(쓸 값 없으면 None).

    왜 별도 칸(`ord_vendors`)인가: `quote_amt` 는 값이 **하나**뿐이고 `remark` 도 업체명 한 칸이다.
      자재(S)·스토어(ST)는 한 청구를 업체 2곳으로 나눠 발주할 수 있어서(실측 2026-08-05 [BGBB S1]
      = 딘텍 KRW 14,700,100 + 에버런스 USD 42,523.32) 단일 칸으로는 한쪽이 통째로 사라진다.
      게다가 통화가 섞이면 합계 자체가 성립하지 않는다 ⇒ **합치지 않고 업체별로 그대로 보관**한다.
    `sub_quotes`(제출견적) 와 섞지 말 것 — 그건 '누가 얼마에 제출했나'고 이건 '누구에게 실제로
      발주가 나갔나'다. 섞으면 발주 안 된 업체가 발주완료로 보인다.
    표시전용이라 값 신뢰보다 형태 방어가 우선: 개수 캡·타입 강제·통화 3글자 검증(`sub_quotes` 동일).
    🔴 `amt=None` 은 **0 이 아니라 '아직 확정 안 됨'** 이다(결재 중인 발주서는 ODR_LIST 에 없다).
      0·음수·inf 도 None 으로 떨군다 — 화면이 '0원 발주'로 그리면 형이 무료 발주로 읽는다.
    canonical 두 겹(원소 키 정렬 + 리스트 정렬)은 `sub_quotes` 와 같은 이유 — SVMS 응답 순서가
      흔들려도 같은 집합이 '변경'으로 잡혀 매 폴링 UPDATE 가 돌지 않게. 정렬 1순위는 유일 식별자
      `odr_no` 다.
    """
    if not isinstance(raw, list):
        return None
    out = []
    seen = set()
    for o in raw[:_DOCKPROC_ORDER_MAX]:
        if not isinstance(o, dict):
            continue
        odr = str(o.get('odr_no') or '').strip().upper()[:40]
        if not odr or odr in seen:
            continue                     # 번호 없음 = '발주'라 말할 수 없음 / 중복 = 같은 발주서 두 줄 표시 방지
        seen.add(odr)
        try:
            amt = None if o.get('amt') in (None, '') else float(str(o.get('amt')).replace(',', ''))
        except (TypeError, ValueError):
            amt = None
        if amt is not None and (not math.isfinite(amt) or amt <= 0):
            amt = None                                   # 금액 미확정으로 본다(0원 발주 표시 방지)
        cur = str(o.get('cur') or '').strip().upper()
        cd = str(o.get('cd') or '').strip().upper()[:20]
        out.append({'odr_no': odr,
                    'nm': str(o.get('nm') or '').strip()[:120],
                    'cd': cd if re.fullmatch(r'[A-Z0-9]{1,20}', cd) else None,
                    'st': str(o.get('st') or '').strip()[:40],
                    'amt': amt,
                    'cur': cur if re.fullmatch(r'[A-Z]{3}', cur) else None,
                    # 🔴 `is True` — 폴러가 문자열 'false'/1 을 보내도 발주완료로 읽지 않는다(닫힘 쪽 실패).
                    'ordered': o.get('ordered') is True})
    if not out:
        return None
    out.sort(key=lambda o: (o['odr_no'], o['nm']))
    return json.dumps(out, ensure_ascii=False, sort_keys=True, separators=(',', ':'))
def _dockproc_orders_of(raw):
    """저장된 `ord_vendors` JSON → 리스트(깨진 값은 빈 목록). 서버 내부 판정·테스트용."""
    if isinstance(raw, str):
        try:
            v = json.loads(raw or '[]')
        except (ValueError, TypeError):
            return []
        return [o for o in v if isinstance(o, dict)] if isinstance(v, list) else []
    return []
def _dockproc_files_of(raw):
    """저장된 att_files JSON → 리스트(깨진 값은 빈 목록). 서버 내부 비교·검증용."""
    if isinstance(raw, str):
        try:
            raw = json.loads(raw)
        except (ValueError, TypeError):
            return []
    if not isinstance(raw, list):
        return []
    return [f for f in raw if isinstance(f, dict)][:_DOCKPROC_ATT_MAX]
# ---- 벤더 견적서 preview cache (fundreq 첨부 cache 와 같은 규약: 확장자 allowlist + magic-byte) ----
#   경로는 row id/idx/지문/확장자만으로 만들어 경로주입 불가. 읽기는 세션(웹)·Bearer(앱).
#
# 🔴 왜 캐시 파일명에 **지문**을 박는가 (올마이트 2026-07-31 지적 반영):
#   처음엔 `{rid}_{idx}.{ext}` 였다. 그런데 idx 는 '목록의 몇 번째'일 뿐이라서, 목록이 바뀌면
#   (앞 첨부가 SVMS 에서 삭제되면 뒤가 앞으로 밀린다) **같은 idx 가 다른 파일을 가리킨다.**
#   그 상태에서 옛 캐시가 남아 있거나(무효화 실패·프로세스 중단), 폴러가 pending 을 받은 뒤 목록이
#   바뀐 다음 업로드하면 → 형이 A업체 견적서 자리에서 **B업체 파일**을 열게 된다.
#   경로에 지문이 있으면 현재 목록과 안 맞는 파일은 **애초에 찾아지지 않는다**(fail-closed).
#   덕분에 무효화(GC)는 '정확성'이 아니라 '용량'만 담당하게 되어, 실패해도 오열람이 없다.
def _dockatt_fp(f):
    """첨부 신원 지문 = (SVMS 저장명, 파일명, 크기). 서버가 이 공식의 **단일 정본**이다 —
    폴러는 pending 으로 받은 지문을 그대로 되돌려주기만 한다(공식 중복구현 금지).
    크기를 넣는 이유: SVMS 가 같은 저장명으로 내용을 바꿔치면 지문이 달라져 다시 받는다."""
    import hashlib as _hl
    s = '%s|%s|%s' % (f.get('sv') or '', f.get('nm') or '', f.get('kb') or 0)
    return _hl.sha1(s.encode('utf-8')).hexdigest()[:12]
def _dockatt_path(rid, idx, fp, ext):
    return os.path.join(DOCKATT_FILE_DIR, '%d_%d_%s.%s' % (int(rid), int(idx), str(fp)[:12], ext))
def _dockatt_find(rid, idx, fp):
    """(경로, 확장자) — 지문까지 일치하는 캐시만. 없으면 (None, None)."""
    for ext in _FUNDREQ_ATT_MIME:
        p = _dockatt_path(rid, idx, fp, ext)
        if os.path.exists(p):
            return p, ext
    return None, None
def _dockatt_disk_map():
    """디스크 1회 스캔 → {row_id: {(idx, fp): ext}}.
    행마다 listdir 하면 목록 API 가 O(행수) 로 느려진다."""
    out = {}
    try:
        names = os.listdir(DOCKATT_FILE_DIR)
    except OSError:
        return out
    for name in names:
        stem, _, ext = name.rpartition('.')
        if ext.lower() not in _FUNDREQ_ATT_MIME:
            continue
        parts = stem.split('_')
        if len(parts) != 3 or not parts[0].isdigit() or not parts[1].isdigit():
            continue
        out.setdefault(int(parts[0]), {})[(int(parts[1]), parts[2])] = ext.lower()
    return out
def _dockatt_cached_idx(files, disk_row):
    """현재 목록 기준으로 **실제로 열 수 있는** idx 목록(지문 일치분만)."""
    return [i for i, f in enumerate(files) if (i, _dockatt_fp(f)) in (disk_row or {})]
def _dockatt_gc(rid, files, disk_row=None):
    """현재 목록이 참조하지 않는 캐시 파일 정리. **용량 회수용이고 정확성 담보가 아니다** —
    실패하거나 아예 안 돌아도 서빙은 지문 불일치로 이미 막힌다(그래서 예외를 삼켜도 안전)."""
    if disk_row is None:
        disk_row = _dockatt_disk_map().get(int(rid), {})
    live = {(i, _dockatt_fp(f)) for i, f in enumerate(files)}
    dropped = 0
    for (idx, fp), ext in list(disk_row.items()):
        if (idx, fp) in live:
            continue
        try:
            os.remove(_dockatt_path(rid, idx, fp, ext)); dropped += 1
        except OSError:
            app.logger.exception('dockatt-gc')     # 남아도 오열람 없음 — 다음 GC 에서 다시 시도
    return dropped
def _dockproc_hash(equipment, subject):
    import hashlib as _hl
    s = f"{(equipment or '').strip().upper()}|{(subject or '').strip().upper()}"
    return _hl.md5(s.encode('utf-8')).hexdigest()[:16]
def _dockproc_subject_from_svms(subj):
    """SVMS 제목 → 트래커 `subject`. `_reqgen_build_subj` 의 역함수.

    '[DOCK][BGBB S33]M/T BELGIUM B - AUXILIARY BOILER SPARE PARTS' → 'AUXILIARY BOILER SPARE PARTS'.
    ' - ' 가 없는 수동작성 제목은 태그만 떼고 남은 전체를 쓴다(비우는 것보다 낫다).
    """
    import re as _re
    s = (subj or '').strip()
    s = _re.sub(r'^\s*\[DOCK\]\s*', '', s, flags=_re.I)
    s = _re.sub(r'^\s*\[[^\]]*\]\s*', '', s)             # [BGBB S33]
    head, sep, tail = s.partition(' - ')
    return ((tail if sep else head).strip() or None)
def _dockproc_fill_key(row, doc, svms_key):
    """SVMS 문서번호를 **문서종류에 맞는 칸**에 채운다(비어 있을 때만).

    🔴 키 분리 규약: 수리=`svms_req_no`(REP_CD) / 구매=`svms_pc_req_no`(REQ_NO).
       섞으면 Phase ③ 상신이 INQ_NO 를 잃는다(`_dockproc_inq_target` 주석 참조).
    """
    key = (svms_key or '').strip()
    if not key or row is None:
        return 'none'
    col = 'svms_req_no' if (doc or '').strip().upper() in ('MA', 'MARP') else 'svms_pc_req_no'
    keys = row.keys() if hasattr(row, 'keys') else ()
    if col not in keys:                                  # 마이그레이션 전 DB — 닫힘 쪽 실패(키만 안 채움)
        return 'none'
    if ((row[col] or '').strip()):
        return 'same' if row[col].strip() == key else 'conflict'
    execute(f"UPDATE dock_procure SET {col}=?, updated_at=datetime('now','localtime') WHERE id=?",
            (key, row['id']))
    return 'filled'
def _dockproc_adopt_svms(vsl_nm, vsl_cd, req_no, subject, equipment, doc, svms_key):
    """SVMS 에 **이미 존재하는** 입거 청구 1건을 발주현황 행으로 적재. 반환 (row_id, created, key_state).

    발주현황 행은 종전에 INDEX 엑셀 업로드로만 생겼고 역동기화는 update-only 였다 →
    엑셀에 없는 시트번호로 청구가 나가면 붙을 행이 없어 `unmatched` 로 조용히 버려졌다
    (2026-08-05 실측: BGBB S2·S18·S21~S24·S33 = 7건이 SVMS 에만 있고 화면엔 없었음).
    ⚠️청구 존재가 확인된 건에만 쓴다 — 추측으로 행을 만들지 않는다.
    ⚠️P(페인트)·SY(조선소)는 메일견적이라 SVMS 청구가 아니다 → 자동적재 대상에서 제외.
    """
    code = _dockproc_cat_code(req_no)
    if code not in ('R', 'S', 'ST'):
        # 🔴 반드시 3-tuple 로 반환한다 — 호출부가 `lid, created, key_state` 로 언패킹한다.
        #    2-tuple 로 나가면 `ValueError` 가 `api_ext_reqgen_result` 의 트랜잭션을 rollback 시켜
        #    **SVMS 에는 저장됐는데 카드는 `saving` 에 갇히고 6h 뒤 failed** 가 된다(형이 실패로 봄).
        #    도달 경로: 시트명이 `R\d+/S\d+/ST\d+` 형식이 아닐 때(P·SY 시트, `S33A` 같은 변형).
        return None, False, 'none'
    rq = (req_no or '').strip().upper()
    ex = query("SELECT * FROM dock_procure WHERE vsl_nm=? AND UPPER(req_no)=?", (vsl_nm, rq), one=True)
    if ex:                                               # 이미 있으면 만들지 않고 키만 채운다(멱등)
        key_state = _dockproc_fill_key(ex, doc, svms_key)
        return ex['id'], False, key_state
    nxt = query("SELECT COALESCE(MAX(sort_no),0)+1 n FROM dock_procure WHERE vsl_nm=?",
                (vsl_nm,), one=True)['n']                # 목록 맨 뒤(NULL 이면 정렬 맨 앞으로 튄다)
    lid = execute(
        "INSERT INTO dock_procure (vsl_nm, vsl_cd, req_no, cat_code, category, equipment, subject, "
        "prepared_by, source, content_hash, sort_no, remark) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
        (vsl_nm, vsl_cd, rq, code, _DOCKPROC_CAT_NM.get(code), equipment, subject,
         'OWNER', _dockproc_source(code, 'OWNER'), _dockproc_hash(equipment, subject), nxt,
         'SVMS 청구 자동적재'))
    key_state = _dockproc_fill_key(query("SELECT * FROM dock_procure WHERE id=?", (lid,), one=True), doc, svms_key)
    return lid, True, key_state


# ══════════════════════════════════════════════════════════════════
#  From routes_tail.py (41 symbols)
# ══════════════════════════════════════════════════════════════════
# ───────────────────────── Fleet Map (대시보드) ─────────────────────────
FLEET_MAP_FILE = os.path.join(INSTANCE_DIR, 'fleet_map.json')
FLEET_MAP_PACKAGED_DIR = os.path.join(BASE_DIR, 'data', 'fleet_map')
FLEET_MAP_AUTOMATION_DIR = os.path.abspath(os.path.join(BASE_DIR, '..', '..', 'automation', 'fleet-map'))
FLEET_LOCODE_FILES = (
    os.path.join(FLEET_MAP_PACKAGED_DIR, 'locode.json'),
    os.path.join(FLEET_MAP_AUTOMATION_DIR, 'locode.json'),
)
FLEET_LOCODE_NAME_FILES = (
    os.path.join(FLEET_MAP_PACKAGED_DIR, 'locode_name.json'),
    os.path.join(FLEET_MAP_AUTOMATION_DIR, 'locode_name.json'),
)
FLEET_COUNTRY_MAP_FILES = (
    os.path.join(FLEET_MAP_PACKAGED_DIR, 'country_map.json'),
    os.path.join(FLEET_MAP_AUTOMATION_DIR, 'country_map.json'),
)
FLEET_LOCODE_LABEL_FILES = (
    os.path.join(FLEET_MAP_PACKAGED_DIR, 'locode_labels.json'),
)
_fleet_port_catalog_cache = None
def _norm_port_text(s):
    return re.sub(r'[^A-Z0-9]+', '', str(s or '').upper())
def _norm_locode(s):
    code = _norm_port_text(s)
    return code if re.fullmatch(r'[A-Z]{2}[A-Z0-9]{3}', code or '') else ''
def _valid_latlng_pair(xy):
    if not (isinstance(xy, (list, tuple)) and len(xy) == 2):
        return None
    try:
        lat, lng = float(xy[0]), float(xy[1])
    except (TypeError, ValueError):
        return None
    if not (math.isfinite(lat) and math.isfinite(lng)):
        return None
    if not (-90 <= lat <= 90 and -180 <= lng <= 180):
        return None
    return [lat, lng]
def _load_json_first(paths):
    for path in paths:
        try:
            with open(path, encoding='utf-8') as f:
                return json.load(f)
        except (FileNotFoundError, ValueError, OSError, json.JSONDecodeError):
            continue
    return None
def _fleet_port_catalog():
    """Read-only deterministic port catalog used by Fleet Map correction/overrides."""
    global _fleet_port_catalog_cache
    if _fleet_port_catalog_cache is not None:
        return _fleet_port_catalog_cache
    locodes, by_name, countries, labels = {}, {}, {}, {}
    raw = _load_json_first(FLEET_LOCODE_FILES)
    if isinstance(raw, dict):
        for code, xy in raw.items():
            point = _valid_latlng_pair(xy)
            if isinstance(code, str) and point:
                locodes[_norm_port_text(code)] = point
    raw_countries = _load_json_first(FLEET_COUNTRY_MAP_FILES)
    if isinstance(raw_countries, dict):
        countries = {str(k).upper(): str(v).upper() for k, v in raw_countries.items()}
    raw_labels = _load_json_first(FLEET_LOCODE_LABEL_FILES)
    if isinstance(raw_labels, dict):
        labels = {_norm_locode(k): str(v).strip() for k, v in raw_labels.items()
                  if _norm_locode(k) and str(v).strip()}
    idx = _load_json_first(FLEET_LOCODE_NAME_FILES)
    if isinstance(idx, dict):
        for key, xy in (idx.get('by') or {}).items():
            if not (isinstance(key, str) and '|' in key and isinstance(xy, list) and len(xy) == 2):
                continue
            point = _valid_latlng_pair(xy)
            if not point:
                continue
            point = (round(point[0], 6), round(point[1], 6))
            iso, name_key = key.split('|', 1)
            by_name.setdefault(name_key, set()).add(point)
            by_name.setdefault(iso + '|' + name_key, set()).add(point)
    for code, label in labels.items():
        xy = locodes.get(code)
        if not xy:
            continue
        point = (round(float(xy[0]), 6), round(float(xy[1]), 6))
        name_key = _norm_port_text(label)
        by_name.setdefault(name_key, set()).add(point)
        by_name.setdefault(code[:2] + '|' + name_key, set()).add(point)
    _fleet_port_catalog_cache = {
        'locodes': locodes,
        'by_name': by_name,
        'countries': countries,
        'labels': labels,
    }
    return _fleet_port_catalog_cache
def _fleet_extract_next_port_code(v):
    port = v.get('next_port') if isinstance(v, dict) else None
    candidates = []
    if isinstance(port, dict):
        candidates.extend(port.get(key) for key in ('cd', 'code', 'locode', 'unlocode'))
    if not isinstance(v, dict):
        return ''
    candidates.extend((v.get('next_port_cd'), v.get('dest_cd')))
    for candidate in candidates:
        code = _norm_locode(candidate)
        if code:
            return code
    return ''
def _fleet_apply_code_first_next_port(v):
    """Correct automatic Next Port coordinates when an explicit code is present."""
    if not isinstance(v, dict):
        return
    code = _fleet_extract_next_port_code(v)
    if not code:
        return
    xy = _fleet_port_catalog()['locodes'].get(code)
    if not xy:
        return
    port = v.get('next_port')
    if not isinstance(port, dict):
        port = {}
    port['cd'] = code
    port['xy'] = [float(xy[0]), float(xy[1])]
    if not port.get('name'):
        port['name'] = _fleet_port_catalog()['labels'].get(code) or code
    v['next_port'] = port
    v['dest_xy'] = port['xy']
    v['dest_port'] = port.get('name') or v.get('dest_port')
    v['next_port_source'] = 'code'
    v['route_legs'] = _fleet_route_to_destination(v, port['xy'])
def _fleet_route_to_destination(v, dest_xy):
    dest = _valid_latlng_pair(dest_xy)
    if not dest:
        return []
    legs = v.get('route_legs') if isinstance(v, dict) else None
    if isinstance(legs, list):
        valid_legs = []
        for leg in legs:
            if not isinstance(leg, list):
                continue
            pts = []
            for point in leg:
                pt = _valid_latlng_pair(point)
                if pt:
                    pts.append(pt)
            if len(pts) >= 2:
                valid_legs.append(pts)
        if valid_legs:
            valid_legs[-1][-1] = dest
            return valid_legs
    here = _valid_latlng_pair([v.get('lat'), v.get('lng')]) if isinstance(v, dict) else None
    return [[here, dest]] if here else []
def _fleet_visible_auto_vessels():
    """Fleet items visible to the current UI user, with corrected automatic next port only."""
    try:
        with open(FLEET_MAP_FILE, encoding='utf-8') as f:
            data = json.load(f)
    except (FileNotFoundError, ValueError, OSError, json.JSONDecodeError):
        return []
    fleet = data.get('fleet') or []
    for v in fleet:
        _fleet_apply_code_first_next_port(v)
    vsup = {_vkey(r['vname']): r['sname'] for r in
            query("SELECT v.name AS vname, s.name AS sname FROM supervisor_vessels sv "
                  "JOIN vessels v ON v.id=sv.vessel_id JOIN supervisors s ON s.id=sv.supervisor_id")}
    for v in fleet:
        v['supervisor'] = vsup.get(_vkey(v.get('name')))
    fleet = [v for v in fleet if v.get('supervisor')]
    is_admin = (session.get('role') == 'admin')
    sup_id = session.get('supervisor_id')
    if sup_id and not is_admin:
        srow = query("SELECT name FROM supervisors WHERE id=?", (sup_id,), one=True)
        sup_name = srow['name'] if srow else None
        allowed = {(_vkey(r['name'])) for r in
                   query("SELECT v.name FROM supervisor_vessels sv "
                         "JOIN vessels v ON v.id=sv.vessel_id WHERE sv.supervisor_id=?", (sup_id,))}
        if allowed:
            fleet = [v for v in fleet if _vkey(v.get('name')) in allowed]
        elif sup_name:
            fleet = [v for v in fleet if v.get('supervisor') == sup_name]
        else:
            fleet = []
    return fleet
# ═════════════════════════════════════════════════════════════════
#  CLASS STATUS (선급 Class Status Report 업로드/추출/매칭)
# ═════════════════════════════════════════════════════════════════
import re as _re_cls
def _norm_vessel_name(name):
    """선명 정규화: 대문자, M/T·M/V 접두 제거, 공백 단일화."""
    if not name:
        return ''
    s = str(name).upper().strip()
    s = _re_cls.sub(r'^(M[\./]?\s*[TV][\./]?|MT|MV)\s+', '', s)  # M/T, M.V., MT, MV ...
    s = _re_cls.sub(r'[^A-Z0-9 ]+', ' ', s)
    s = _re_cls.sub(r'\s+', ' ', s).strip()
    return s
def _match_vessel_by_name(name):
    """보고서 선명 → vessels 행 매칭. 정확 일치 우선, 없으면 부분포함. 실패 시 None."""
    target = _norm_vessel_name(name)
    if not target:
        return None
    rows = query('SELECT * FROM vessels WHERE active=1')
    norm = [(v, _norm_vessel_name(v['name'])) for v in rows]
    for v, n in norm:
        if n == target:
            return v
    # 부분 포함 (한쪽이 다른 쪽을 포함)
    for v, n in norm:
        if n and (n in target or target in n):
            return v
    return None
def _annotate_drafts_with_vessel(drafts):
    """P4 표시전용(read-only): 각 draft 행에 matched_vessel:{id,name,in_my_roster} 부가.

    돈 파이프라인·draft 원본·status·금액 무변경. money 테이블 write 없음(읽기시점 계산).
    매칭 순서: vessels.vsl_cd 정확일치 우선 → 없으면 선명 정규화(_match_vessel_by_name).
    in_my_roster = 매칭 선박이 현재 세션 감독의 supervisor_vessels 에 포함되는지
      (supervisor_id 미설정 admin은 전체 로스터로 간주 → 매칭되면 True).
    각 draft dict 에 'matched_vessel' 키만 추가(없으면 None). 리스트 그대로 반환.
    """
    if not drafts:
        return drafts
    try:
        vrows = query('SELECT id, name, vsl_cd FROM vessels WHERE active=1')
    except Exception:
        # 조회 실패 시 표시기능만 조용히 생략 — 목록 응답 자체는 절대 깨지 않는다.
        for d in drafts:
            d.setdefault('matched_vessel', None)
        return drafts
    # 매칭 블록 전체를 방어적으로 감싼다 — supervisor_vessels 조회나 선명매칭이
    # 어떤 이유로 예외를 던져도 목록 API(500)를 깨지 않고 표시기능만 조용히 생략.
    try:
        by_cd = {}
        for v in vrows:
            cd = (v['vsl_cd'] or '').strip().upper()
            if cd:
                by_cd.setdefault(cd, v)
        # 내 로스터(현재 세션 감독) 선박 id 집합. 감독 미설정이면 None(=전체 로스터).
        sup_id = session.get('supervisor_id')
        my_ids = None
        if sup_id:
            my_ids = {r['vessel_id'] for r in
                      query('SELECT vessel_id FROM supervisor_vessels WHERE supervisor_id=?', (sup_id,))}
        for d in drafts:
            mv = None
            cd = (d.get('vsl_cd') or '').strip().upper()
            v = by_cd.get(cd) if cd else None
            if v is None:
                v = _match_vessel_by_name(d.get('vsl_nm') or d.get('vsl_cd'))
            if v is not None:
                in_roster = True if my_ids is None else (v['id'] in my_ids)
                mv = {'id': v['id'], 'name': v['name'], 'in_my_roster': bool(in_roster)}
            d['matched_vessel'] = mv
    except Exception:
        for d in drafts:
            d.setdefault('matched_vessel', None)
    return drafts
def _class_status_prompt():
    return (
        "다음은 선박 선급(Classification Society)의 'Class Status Report' 또는 "
        "'Survey Status Report'다. (선급 예: DNV, BV, KR, ABS, LR, NK 등 — 포맷이 다를 수 있다.)\n"
        "아래 정보를 추출해 지정한 JSON으로만 답하라.\n"
        "■ 공통 정보\n"
        "- vessel_name: 보고서의 선명(Name of vessel / Ship name). 대문자 원문.\n"
        "- class_society: 발행 선급 약어 (DNV / BV / KR / ABS / LR / NK 중 하나, 식별 가능하면).\n"
        "- report_date: 보고서 발행일/생성일 (Date of issue / Generated on). 가능하면 YYYY-MM-DD.\n"
        "■ 추출 대상 — 'Open(미해소)' 상태인 항목만:\n"
        "  (1) coc  = Condition of Class / 선급지적. 선급별 명칭 예:\n"
        "      DNV 'Conditions related to class', BV 'Conditions of Class', "
        "ABS 'Conditions of Class / Outstanding', LR 'Conditions of Class(COC)', "
        "또한 BV 'Planned Inspection Items'의 Recommendation(R)/Condition of Class 도 포함.\n"
        "  (2) statutory = Condition of Statutory / 기국(법정)사항. 예:\n"
        "      DNV 'Conditions related to statutory certificates', "
        "BV 'Statutory Recommendations' 및 'Planned Inspection Items'의 Statutory Condition/Recommendation 항목. "
        "⚠️ 단, Type이 'Observation'(Obs)인 행은 statutory에도 coc에도 절대 넣지 마라(아래 제외 규칙).\n"
        "■ 제외(절대 추출 금지): 단순 Survey 예정표(1-Year Planner/Surveys 목록), 인증서 목록, "
        "**모든 Observation 항목 — Type/VS 칸이 'Obs' 또는 'Observation'인 행(특히 'Planned Inspection Items'의 Obs 행, "
        "예: 'STS plan to be approved and placed on board', 'BWMP to be approved …')은 due date가 있어도 절대 추출하지 마라**, "
        "그리고 **Memoranda 섹션 전체**. 제목에 'Memoranda(메모란다)'가 들어간 표·섹션 — "
        "'Class Memoranda', 'Statutory Memoranda', 'Description of (Class/Statutory) Memoranda' 등 — 의 항목은 "
        "내용이 지적·기국처럼 보여도(예: 'Engine Power Limitation (SHaPoLi) approved, limiting … kW') 절대 추출하지 마라. "
        "⚠️ 'Statutory Memoranda'는 'Statutory Recommendations'와 전혀 다른 별개 섹션이다 — 'Statutory' 단어가 같다고 혼동 금지. "
        "메모란다는 단순 정보성 기록(approved/완료 통보 등)이라 미해소 조치사항이 아니다. "
        "이미 Closed/Cleared/Deleted 되었거나 조치 확인 완료된 항목도 제외. 'None'이면 빈 배열.\n"
        "■ 각 항목 필드:\n"
        "- issued_date: 발행/기재일 (가능하면 YYYY-MM-DD, 없으면 빈 문자열)\n"
        "- description: 지적/기국 본문을 원문 그대로 복사(영문이면 영문 그대로). 요약·변형 금지.\n"
        "- due_date: 마감/처리기한 (Due/Limit date, 가능하면 YYYY-MM-DD, 없으면 빈 문자열). "
        "⚠️ **연장(postpone/extend)된 경우 반드시 최종(연장된) 날짜를 due_date로 한다.** "
        "보고서에 원래 기한과 연장 기한이 함께 있거나(예: 'Original due 2025-04-26, postponed to 2026-04-26', "
        "'Limit date revised/extended to …', 'New limit date …', 'Postponed until …'), "
        "여러 날짜가 보이면 **가장 나중(최신) 유효 기한**을 due_date로 쓴다. 원래(이른) 날짜를 쓰지 마라.\n"
        "- remark: description의 핵심을 한국어 1~2문장으로 간결히 요약(전체 직역 금지). "
        "문장은 '~함/~됨/~음' 음슴체(개조식). 기술 명칭·장비명·약어·인증명(예: COC, SEEMP, IHM, "
        "BNWAS, Load Line, Plimsoll Mark, EGCS, BWTS)은 영문 그대로 둔다." + _MARITIME_TERMS + "\n"
        "없는 내용을 지어내지 말 것.\n"
        '형식: {"vessel_name":"","class_society":"","report_date":"",'
        '"coc":[{"issued_date":"","description":"","due_date":"","remark":""}],'
        '"statutory":[{"issued_date":"","description":"","due_date":"","remark":""}]}'
    )
def _cls_item(it):
    if not isinstance(it, dict):
        return None
    rec = {
        'issued_date': (it.get('issued_date') or '').strip(),
        'description': (it.get('description') or '').strip(),
        'due_date':    (it.get('due_date') or '').strip(),
        'remark':      (it.get('remark') or '').strip(),
    }
    return rec if rec['description'] else None
def _normalize_class_status(parsed):
    if not isinstance(parsed, dict):
        return None
    def lst(key):
        out = []
        for it in (parsed.get(key) or []):
            r = _cls_item(it)
            if r:
                out.append(r)
        return out
    return {
        'vessel_name':   (parsed.get('vessel_name') or '').strip(),
        'class_society': (parsed.get('class_society') or '').strip().upper(),
        'report_date':   (parsed.get('report_date') or '').strip(),
        'coc':           lst('coc'),
        'statutory':     lst('statutory'),
    }
def _xlsx_to_text(raw_bytes):
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    lines = []
    for ws in wb.worksheets:
        for r in ws.iter_rows(values_only=True):
            cells = ['' if c is None else str(c).strip() for c in r]
            if any(cells):
                lines.append('\t'.join(cells))
            if len(lines) > 600:
                break
    return '\n'.join(lines)
def _extract_class_status_from_upload(f):
    """업로드 FileStorage → (data, err). data = _normalize_class_status 결과."""
    name = (f.filename or '').lower()
    ext = name.rsplit('.', 1)[-1] if '.' in name else ''
    raw = f.read()
    size_mb = len(raw) / (1024 * 1024)
    prompt = _class_status_prompt()

    if ext == 'pdf':
        if size_mb > 15:
            return None, {'reason': 'TOO_LARGE',
                          'message': f'PDF가 너무 큽니다({size_mb:.1f}MB). 15MB 이하로 줄여주세요.'}
        b64 = __import__('base64').standard_b64encode(raw).decode()
        parsed = _gemini_call_json([
            {'inline_data': {'mime_type': 'application/pdf', 'data': b64}},
            {'text': prompt},
        ], model=_model_for('findings'))
    elif ext in ('png', 'jpg', 'jpeg', 'webp', 'gif', 'bmp'):
        if size_mb > 15:
            return None, {'reason': 'TOO_LARGE', 'message': f'이미지가 너무 큽니다({size_mb:.1f}MB).'}
        import mimetypes
        media = mimetypes.guess_type(name)[0] or 'image/jpeg'
        b64 = __import__('base64').standard_b64encode(raw).decode()
        parsed = _gemini_call_json([
            {'inline_data': {'mime_type': media, 'data': b64}},
            {'text': prompt},
        ], model=_model_for('findings'))
    elif ext in ('xlsx', 'xls'):
        try:
            txt = _xlsx_to_text(raw)
        except Exception as e:
            app.logger.exception('extract-class-status-from-upload')
            return None, {'reason': 'XLSX_PARSE_FAILED', 'message': f'엑셀을 읽지 못했습니다: {e}'}
        parsed = _gemini_call_json([{'text': prompt + '\n\n[보고서 표 내용]\n' + txt}],
                                   model=_model_for('findings'))
    else:
        return None, {'reason': 'BAD_TYPE', 'message': 'PDF · 이미지 · 엑셀(xlsx) 파일만 지원합니다.'}

    if isinstance(parsed, dict) and parsed.get('error') == 'NO_API_KEY':
        return None, {'reason': 'no_api_key', 'message': 'AI 자동추출이 설정되지 않았습니다(키 미설정).'}
    if isinstance(parsed, dict) and parsed.get('error'):
        return None, {'reason': parsed['error'], 'message': '자동 추출에 실패했습니다.',
                      'detail': parsed.get('detail') or parsed.get('raw')}
    data = _normalize_class_status(parsed)
    if data is None:
        return None, {'reason': 'PARSE_FAILED', 'message': '추출 결과를 해석하지 못했습니다.'}
    return data, None
def _cls_delete_file(path):
    """보관 파일 삭제(교체 시 이전 파일 자동삭제). 경로가 업로드 폴더 내일 때만."""
    if not path:
        return
    try:
        full = os.path.join(BASE_DIR, path) if not os.path.isabs(path) else path
        if os.path.commonpath([os.path.realpath(full), os.path.realpath(UPLOAD_DIR)]) == os.path.realpath(UPLOAD_DIR) \
                and os.path.isfile(full):
            os.remove(full)
    except Exception as e:
        print(f'[cls] old file remove skip: {e}')
def _cls_save_snapshot(vessel_id, vessel_name_raw, data, filename, source_path=None):
    """선박 스냅샷 교체(최신만 유지). 이전 스냅샷의 보관파일도 자동삭제.
    vessel_id None 이면 미매칭으로 저장(같은 정규화 선명의 기존 미매칭 제거 후 삽입)."""
    conn = get_db()
    user = session.get('username')
    _ndesc = lambda s: ' '.join((s or '').strip().lower().split())
    preserved = {}   # (category, 정규화 description) -> action_taken — 스냅샷 교체에도 손유석 조치사항 유지
    if vessel_id is not None:
        try:
            for r in conn.execute(
                "SELECT i.category, i.description, i.action_taken "
                "FROM class_status_items i JOIN class_status c ON c.id = i.cs_id "
                "WHERE c.vessel_id = ? AND IFNULL(i.action_taken,'') <> ''", (vessel_id,)).fetchall():
                preserved[(r['category'], _ndesc(r['description']))] = r['action_taken']
        except Exception:
            app.logger.exception('cls-save-snapshot')
            preserved = {}
        for r in conn.execute('SELECT source_path FROM class_status WHERE vessel_id=?', (vessel_id,)).fetchall():
            _cls_delete_file(r['source_path'])
        conn.execute('DELETE FROM class_status WHERE vessel_id=?', (vessel_id,))
    else:
        # 같은 (정규화) 선명의 기존 미매칭 스냅샷 제거
        tgt = _norm_vessel_name(vessel_name_raw)
        for r in conn.execute('SELECT id, vessel_name_raw, source_path FROM class_status WHERE vessel_id IS NULL').fetchall():
            if _norm_vessel_name(r['vessel_name_raw']) == tgt:
                _cls_delete_file(r['source_path'])
                conn.execute('DELETE FROM class_status WHERE id=?', (r['id'],))
    cur = conn.execute(
        '''INSERT INTO class_status
             (vessel_id, vessel_name_raw, class_society, report_date, source_filename, source_path, uploaded_by)
           VALUES (?,?,?,?,?,?,?)''',
        (vessel_id, vessel_name_raw, data.get('class_society'),
         data.get('report_date'), filename, source_path, user))
    cs_id = cur.lastrowid
    for cat, key in (('COC', 'coc'), ('STATUTORY', 'statutory')):
        for n, it in enumerate(data.get(key) or [], start=1):
            act = preserved.get((cat, _ndesc(it.get('description'))), '')
            conn.execute(
                '''INSERT INTO class_status_items
                     (cs_id, category, no, issued_date, description, due_date, remark, action_taken)
                   VALUES (?,?,?,?,?,?,?,?)''',
                (cs_id, cat, n, it.get('issued_date'), it.get('description'),
                 it.get('due_date'), it.get('remark'), act))
    conn.commit()
    return cs_id
def _cls_handle_files(files):
    """업로드 파일들 → AI추출 → 선박매칭 → 저장. 원본파일도 선박별 최신만 보관. (UI·BV Pushing 공용)"""
    cls_dir = os.path.join(UPLOAD_DIR, 'class_status')
    os.makedirs(cls_dir, exist_ok=True)
    results = []
    for f in [x for x in files if x and x.filename]:
        fname = f.filename
        # 원본 바이트 보관(추출이 스트림을 소비하므로 추출 전에 읽고 seek 리셋)
        raw = None
        try:
            f.stream.seek(0); raw = f.read(); f.stream.seek(0)
        except Exception as _e:
            app.logger.warning('cls-handle-files: %s', _e)
            raw = None
        data, err = _extract_class_status_from_upload(f)
        if err:
            results.append({'filename': fname, 'ok': False, **err})
            continue
        vname = data.get('vessel_name') or ''
        v = _match_vessel_by_name(vname)
        vessel_id = v['id'] if v else None
        src_rel = None
        if raw:
            uniq = uuid.uuid4().hex[:8] + '_' + datetime.now().strftime('%Y%m%d%H%M%S%f') + '_' + (secure_filename(fname) or 'report')
            try:
                with open(os.path.join(cls_dir, uniq), 'wb') as out:
                    out.write(raw)
                src_rel = os.path.join('static', 'uploads', 'class_status', uniq)
            except Exception as e:
                print(f'[cls] file save skip: {e}')
        _cls_save_snapshot(vessel_id, vname, data, fname, src_rel)
        results.append({
            'filename': fname, 'ok': True,
            'vessel_name': vname,
            'matched': bool(v),
            'vessel_id': vessel_id,
            'matched_name': v['name'] if v else None,
            'class_society': data.get('class_society'),
            'report_date': data.get('report_date'),
            'coc_count': len(data.get('coc') or []),
            'statutory_count': len(data.get('statutory') or []),
        })
    return results
# ═════════════════════════════════════════════════════════════════
#  iOS 푸시알림 (APNs) — 디바이스 등록 · 발송 · 종류별 on/off
# ═════════════════════════════════════════════════════════════════
# 발송 자체는 apns_push.py(curl --http2 + ES256 JWT). 키가 없으면 라우트는 살아 있고
# 'not_configured' 를 정직하게 반환한다 — 조용히 성공으로 위장하지 않는다.
#
# 알림 종류. 앱 설정화면이 이 목록을 그대로 그리므로 키를 바꾸면 기존 저장값이 끊긴다.
PUSH_KINDS = [
    ('dock_quote',       'Dock 견적 제출',      '업체가 견적을 올렸을 때'),
    ('dock_ordered',     'Dock 발주완료',       '발주가 전부 완료됐을 때'),
    ('dock_reject',      'Dock 결재반려',       '상신이 반려돼 단계가 되돌아갔을 때'),
    ('dock_submit_fail', 'Dock 상신 실패',      'SVMS 상신이 실패로 끝났을 때'),
    ('auto_fail',        '자동화 실패',         '러너 실패·killswitch·차단'),
    ('approval_new',     '승인 대기 신규',      '전자결재/Fund Request 신규 대기'),
    ('issue_urgent',     '긴급 현안',           '긴급으로 등록된 현안'),
    ('class_due',        '선급·증서 만료',      'D-30/D-7 요약'),
    ('calendar_daily',   '오늘 일정 요약',      '매일 10시·14시, 미완료 일정만'),
    ('test',             '테스트 알림',         '설정 확인용'),
]
PUSH_KIND_KEYS = {k for k, _l, _d in PUSH_KINDS}
_PUSH_DEVICE_CAP = 20        # 단일 요청에서 발송할 디바이스 상한(폭주 방지)
def _push_module():
    """apns_push 지연 import — 모듈이 없어도 앱 전체가 죽지 않게."""
    try:
        import apns_push
        return apns_push
    except Exception as e:
        print(f'[push] apns_push import 실패: {e}')
        return None
def _push_prefs(row):
    try:
        p = json.loads(row['prefs'] or '{}')
    except Exception:
        p = {}
    return p if isinstance(p, dict) else {}
def _push_kind_enabled(row, kind):
    """미설정 = 켜짐(기본 on). 형이 명시적으로 끈 것만 뺀다."""
    return bool(_push_prefs(row).get(kind, 1))
def _push_devices(user_ids=None, kind=None):
    """발송 대상 기기. `user_ids=None` = 전체 브로드캐스트(호출측이 의도적으로 None 을 줄 때만).

    🔴 cap 은 **prefs 필터 뒤에** 적용한다(올마이트 지적). SQL LIMIT 으로 먼저 자르면
       상위 N 대가 그 종류를 껐을 때 켜둔 기기가 영구히 미발송된다 — 조용한 미탐.
    """
    sql = "SELECT * FROM ios_device WHERE active=1"
    params = []
    if user_ids:
        sql += " AND user_id IN (%s)" % ','.join('?' * len(user_ids))
        params += list(user_ids)
    sql += " ORDER BY updated_at DESC"
    rows = query(sql, tuple(params))
    if kind:
        rows = [r for r in rows if _push_kind_enabled(r, kind)]
    return rows[:_PUSH_DEVICE_CAP]
def _push_dispatch(kind, event_key, title, body, link=None,
                   user_ids=None, collapse_id=None):
    """알림 1건 발송. 반환 dict(ok/sent/failed/reason).

    🔴 2단 커밋(BV 감시 교훈): event_key 를 먼저 claim 해 중복을 막되, **디바이스가 있었는데
       전부 발송 실패**면 claim 을 풀어 다음 폴링에 재시도되게 한다. 안 풀면 일시적 네트워크
       오류 1회가 그 이벤트를 영구 미탐으로 만든다.
       디바이스가 0대면 claim 을 유지한다(배달할 곳이 없는 과거 이벤트가 나중에 폭주하는 것 방지).
    """
    if kind not in PUSH_KIND_KEYS:
        return {'ok': False, 'reason': 'unknown_kind', 'sent': 0, 'failed': 0}
    ap = _push_module()
    if ap is None:
        return {'ok': False, 'reason': 'module_missing', 'sent': 0, 'failed': 0}
    try:
        conf = ap.load_conf()
    except Exception as e:
        return {'ok': False, 'reason': 'not_configured', 'detail': str(e),
                'sent': 0, 'failed': 0}

    try:                                     # ── claim (중복발송 차단)
        execute("INSERT INTO push_log (event_key, kind, title, body, link) "
                "VALUES (?,?,?,?,?)", (event_key, kind, title, body, link))
    except sqlite3.IntegrityError:
        return {'ok': True, 'dup': True, 'reason': 'already_sent',
                'sent': 0, 'failed': 0}

    targets = []
    sent, failed, detail = 0, 0, []
    try:
        targets = _push_devices(user_ids, kind=kind)
        for d in targets:
            payload = ap.alert_payload(title, body, link=link, kind=kind)
            ok, st, reason = ap.send(d['token'], payload, env=(d['env'] or 'production'),
                                     conf=conf, collapse_id=collapse_id)
            if ok:
                sent += 1
                execute("UPDATE ios_device SET last_push_at=datetime('now','localtime') "
                        "WHERE id=?", (d['id'],))
            else:
                failed += 1
                detail.append(f"dev{d['id']}:{st}:{reason}")
                if ap.is_dead(st, reason):   # 영구 사망 사유만 비활성(일시실패로는 안 끔)
                    execute("UPDATE ios_device SET active=0, dead_reason=?, "
                            "updated_at=datetime('now','localtime') WHERE id=?",
                            (f'{st}:{reason}', d['id']))
    except Exception as e:
        # 🔴 claim 만 남기고 발송 루프가 예외로 죽으면, 다음 시도는 IntegrityError 를 만나
        #    `dup/already_sent` (ok=True) 로 응답한다 → 대기함(outbox) 행이 "성공"으로 지워져
        #    그 알림이 **영구 유실**된다. 성공분이 0일 때만 claim 을 풀어 재시도를 살린다.
        app.logger.exception('push 발송 예외 key=%s', event_key)
        if sent == 0:
            try:
                execute("DELETE FROM push_log WHERE event_key=? AND sent_n=0", (event_key,))
            except Exception:
                app.logger.exception('push claim 해제 실패 key=%s', event_key)
            return {'ok': False, 'reason': 'exception', 'detail': str(e)[:200],
                    'sent': 0, 'failed': failed}
        # 일부는 이미 나갔다 — claim 을 유지해 재발송(중복 알림)을 막고 실적만 기록한다.
        detail.append(f'exception:{str(e)[:120]}')

    if targets and sent == 0:                # 전부 실패 → claim 해제(재시도 가능하게)
        execute("DELETE FROM push_log WHERE event_key=? AND sent_n=0", (event_key,))
        return {'ok': False, 'reason': 'all_failed', 'sent': 0,
                'failed': failed, 'detail': '; '.join(detail)[:500]}

    execute("UPDATE push_log SET sent_n=?, fail_n=?, detail=? WHERE event_key=?",
            (sent, failed, '; '.join(detail)[:500] or None, event_key))
    return {'ok': True, 'sent': sent, 'failed': failed,
            'devices': len(targets), 'detail': '; '.join(detail)[:500]}
_PUSH_OUTBOX_MAX_TRIES = 6      # 15분 폴 기준 약 1.5시간 재시도
_PUSH_OUTBOX_MAX_AGE_H = 24     # 그보다 늙은 알림은 형에게 가치가 없다(늦은 알림 = 오보에 가깝다)
def _push_outbox_add(kind, event_key, title, body, link=None, collapse_id=None):
    """발송 예정 이벤트를 durable 하게 적재. 이미 있으면(=같은 전이 재관측) 조용히 통과.

    🔴 반드시 **상태 UPDATE 보다 먼저** 부른다 — 순서가 뒤집히면 그 사이에 죽었을 때
       "행은 갱신됐는데 알림은 없는" 영구 미탐이 남는다(schema.sql push_outbox 주석 참조).
    """
    execute("INSERT OR IGNORE INTO push_outbox (event_key, kind, title, body, link, collapse_id) "
            "VALUES (?,?,?,?,?,?)",
            (event_key, kind, title, body, link, collapse_id))
def _push_outbox_drain(limit=20):
    """대기함 발송. 성공·중복이면 삭제, 실패면 tries+1 하고 남긴다. 결과 목록 반환.

    ⚠️`sent=0` 은 실패가 아니다 — 등록된 기기가 0대면 보낼 곳이 없을 뿐이고 `_push_dispatch` 가
      claim 을 유지하므로 여기서도 지운다(안 지우면 기기 등록 순간 과거 알림이 폭주한다).
    """
    rows = query("SELECT * FROM push_outbox ORDER BY created_at LIMIT ?", (limit,))
    out = []
    for r in rows:
        key = r['event_key']
        old = query("SELECT CAST((julianday('now','localtime') - julianday(?)) * 24 AS REAL) h",
                    (r['created_at'],), one=True)
        if r['tries'] >= _PUSH_OUTBOX_MAX_TRIES or (old and (old['h'] or 0) > _PUSH_OUTBOX_MAX_AGE_H):
            execute("DELETE FROM push_outbox WHERE event_key=?", (key,))
            app.logger.warning('push outbox 포기 key=%s tries=%s', key, r['tries'])
            out.append({'kind': r['kind'], 'key': key, 'ok': False, 'sent': 0, 'reason': 'dropped'})
            continue
        try:
            res = _push_dispatch(r['kind'], key, r['title'] or '', r['body'] or '',
                                 link=r['link'], collapse_id=r['collapse_id'])
        except Exception as e:
            app.logger.exception('push outbox 발송 예외 key=%s', key)
            res = {'ok': False, 'reason': 'exception', 'detail': str(e)[:200]}
        if res.get('ok'):
            execute("DELETE FROM push_outbox WHERE event_key=?", (key,))
        else:
            execute("UPDATE push_outbox SET tries=tries+1, last_error=? WHERE event_key=?",
                    ((res.get('reason') or '')[:200], key))
        out.append({'kind': r['kind'], 'key': key, 'ok': bool(res.get('ok')),
                    'sent': res.get('sent', 0), 'reason': res.get('reason')})
    return out
