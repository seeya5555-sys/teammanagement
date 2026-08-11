# ---- ext (맥 push_cards.py / apply_decisions.py) ----
@app.route('/api/ext/shipwiki/push', methods=['POST'])
@api_key_required
def api_ext_shipwiki_push():
    """맥이 pending/wiki 노트를 적재(upsert by slug+fname). 사람 결정(decision/card_status)이
    이미 걸린 카드는 내용만 갱신하고 결정은 보존 — 재push해도 사람 판단 안 풀림."""
    d = request.get_json(silent=True) or {}
    cards = d.get('cards') or []
    slug = (d.get('slug') or '').strip()
    purge = bool(d.get('purge'))                        # 해당 slug 의 open 미결정 카드 중 이번에 없는 건 정리
    db = get_db()
    n_ins = n_upd = 0
    seen = set()
    try:
        for c in cards:
            cslug = (c.get('slug') or slug or '').strip()
            fname = (c.get('fname') or '').strip()
            if not cslug or not fname:
                continue
            seen.add((cslug, fname))
            ex = db.execute("SELECT id, card_status FROM shipwiki_card WHERE slug=? AND fname=?",
                            (cslug, fname)).fetchone()
            vals = (cslug, c.get('ship_nm'), fname, (c.get('tier') or 'pending'), c.get('title'),
                    c.get('category'), c.get('confidence'), c.get('llm_conf'),
                    1 if c.get('multi') else 0, c.get('msg_count'),
                    json.dumps(c.get('needs_human') or [], ensure_ascii=False),
                    c.get('judgment'), c.get('evidence'), c.get('raw_links'),
                    json.dumps(c.get('source_msgids') or [], ensure_ascii=False),
                    json.dumps(c.get('equipment') or [], ensure_ascii=False),
                    json.dumps(c.get('vendors') or [], ensure_ascii=False),
                    json.dumps(c.get('ref_numbers') or [], ensure_ascii=False),
                    c.get('date_first'), c.get('date_last'), c.get('wiki_thread_id'))
            if ex:
                # 내용만 갱신(결정/상태 보존)
                db.execute(
                    "UPDATE shipwiki_card SET ship_nm=?, tier=?, title=?, category=?, confidence=?, "
                    "llm_conf=?, multi=?, msg_count=?, needs_human=?, judgment=?, evidence=?, raw_links=?, "
                    "source_msgids=?, equipment=?, vendors=?, ref_numbers=?, date_first=?, date_last=?, "
                    "wiki_thread_id=?, pushed_at=datetime('now','localtime') WHERE id=?",
                    vals[1:2] + vals[3:] + (ex['id'],))   # slug(0)·fname(2) 제외
                n_upd += 1
            else:
                db.execute(
                    "INSERT INTO shipwiki_card (slug, ship_nm, fname, tier, title, category, confidence, "
                    "llm_conf, multi, msg_count, needs_human, judgment, evidence, raw_links, source_msgids, "
                    "equipment, vendors, ref_numbers, date_first, date_last, wiki_thread_id) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", vals)
                n_ins += 1
        purged = 0
        if purge and slug:
            for r in db.execute("SELECT id, slug, fname FROM shipwiki_card "
                                "WHERE slug=? AND card_status='open' AND decision IS NULL",
                                (slug,)).fetchall():
                if (r['slug'], r['fname']) not in seen:
                    db.execute("DELETE FROM shipwiki_card WHERE id=?", (r['id'],))
                    purged += 1
        db.commit()
    except Exception:
        db.rollback()
        raise
    return jsonify({'ok': True, 'inserted': n_ins, 'updated': n_upd,
                    'purged': (purged if purge and slug else 0)})


# ───────────────────────── Fleet Map (대시보드) ─────────────────────────
FLEET_MAP_FILE = os.path.join(INSTANCE_DIR, 'fleet_map.json')
FLEET_MAP_PACKAGED_DIR = os.path.join(BASE_DIR, 'data', 'fleet_map')
FLEET_MAP_AUTOMATION_DIR = os.path.abspath(os.path.join(BASE_DIR, '..', '..', 'automation', 'fleet-map'))
FLEET_LOCODE_FILES = (
    os.path.join(FLEET_MAP_PACKAGED_DIR, 'locode.json'),
    os.path.join(FLEET_MAP_AUTOMATION_DIR, 'locode.json'),
)
FLEET_LOCODE_NAME_FILES = (
    os.path.join(FLEET_MAP_PACKAGED_DIR, 'locode_name.json'),
    os.path.join(FLEET_MAP_AUTOMATION_DIR, 'locode_name.json'),
)
FLEET_COUNTRY_MAP_FILES = (
    os.path.join(FLEET_MAP_PACKAGED_DIR, 'country_map.json'),
    os.path.join(FLEET_MAP_AUTOMATION_DIR, 'country_map.json'),
)
FLEET_LOCODE_LABEL_FILES = (
    os.path.join(FLEET_MAP_PACKAGED_DIR, 'locode_labels.json'),
)
_fleet_port_catalog_cache = None

# Fleet Map 위치는 기존 SVMS/VesselTracker 적재본을 fallback으로 유지하되, 화면 조회 시
# TRMT DB의 최신 ship-position으로 덮어쓴다. 키는 반드시 systemd EnvironmentFile에만 둔다.
TRMTDB_SHIP_POSITION_URL = os.getenv(
    'TRMTDB_SHIP_POSITION_URL',
    'https://trmtdb.duckdns.org/api/ship-position?platform=ALL',
)
# upstream(`?platform=ALL`)은 실측 33.5MB·3.4~3.8초인데, `latest_event_at`이 **정시 단위 배치**로만
# 갱신된다(2026-07-29 실측: 13:00에 172척 / 08:00에 138척). 45초 TTL은 신선도 이득 없이 하루 약
# 1,900회 × 33.5MB 를 왕복하던 순수 낭비여서 10분으로 올렸다.
TRMTDB_POSITION_CACHE_TTL = 600
# 실패한 시도까지 10분 묶어두면 upstream 일시 장애가 10분짜리 빈 화면이 된다(올마이트 지적).
# 실패 뒤에는 짧게 다시 시도한다.
TRMTDB_POSITION_ERROR_TTL = 60
_trmtdb_position_cache = {'at': 0.0, 'loaded': False, 'vessels': [], 'fetched_at': None, 'error': None}
_trmtdb_position_lock = threading.Lock()
_trmtdb_position_refreshing = False
_fleet_next_port_lock = threading.RLock()
_fleet_eta_lock = threading.RLock()


def _norm_port_text(s):
    return re.sub(r'[^A-Z0-9]+', '', str(s or '').upper())


def _norm_locode(s):
    code = _norm_port_text(s)
    return code if re.fullmatch(r'[A-Z]{2}[A-Z0-9]{3}', code or '') else ''


def _valid_latlng_pair(xy):
    if not (isinstance(xy, (list, tuple)) and len(xy) == 2):
        return None
    try:
        lat, lng = float(xy[0]), float(xy[1])
    except (TypeError, ValueError):
        return None
    if not (math.isfinite(lat) and math.isfinite(lng)):
        return None
    if not (-90 <= lat <= 90 and -180 <= lng <= 180):
        return None
    return [lat, lng]


def _load_json_first(paths):
    for path in paths:
        try:
            with open(path, encoding='utf-8') as f:
                return json.load(f)
        except (FileNotFoundError, ValueError, OSError, json.JSONDecodeError):
            continue
    return None


def _fleet_port_catalog():
    """Read-only deterministic port catalog used by Fleet Map correction/overrides."""
    global _fleet_port_catalog_cache
    if _fleet_port_catalog_cache is not None:
        return _fleet_port_catalog_cache
    locodes, by_name, countries, labels = {}, {}, {}, {}
    raw = _load_json_first(FLEET_LOCODE_FILES)
    if isinstance(raw, dict):
        for code, xy in raw.items():
            point = _valid_latlng_pair(xy)
            if isinstance(code, str) and point:
                locodes[_norm_port_text(code)] = point
    raw_countries = _load_json_first(FLEET_COUNTRY_MAP_FILES)
    if isinstance(raw_countries, dict):
        countries = {str(k).upper(): str(v).upper() for k, v in raw_countries.items()}
    raw_labels = _load_json_first(FLEET_LOCODE_LABEL_FILES)
    if isinstance(raw_labels, dict):
        labels = {_norm_locode(k): str(v).strip() for k, v in raw_labels.items()
                  if _norm_locode(k) and str(v).strip()}
    idx = _load_json_first(FLEET_LOCODE_NAME_FILES)
    if isinstance(idx, dict):
        for key, xy in (idx.get('by') or {}).items():
            if not (isinstance(key, str) and '|' in key and isinstance(xy, list) and len(xy) == 2):
                continue
            point = _valid_latlng_pair(xy)
            if not point:
                continue
            point = (round(point[0], 6), round(point[1], 6))
            iso, name_key = key.split('|', 1)
            by_name.setdefault(name_key, set()).add(point)
            by_name.setdefault(iso + '|' + name_key, set()).add(point)
    for code, label in labels.items():
        xy = locodes.get(code)
        if not xy:
            continue
        point = (round(float(xy[0]), 6), round(float(xy[1]), 6))
        name_key = _norm_port_text(label)
        by_name.setdefault(name_key, set()).add(point)
        by_name.setdefault(code[:2] + '|' + name_key, set()).add(point)
    _fleet_port_catalog_cache = {
        'locodes': locodes,
        'by_name': by_name,
        'countries': countries,
        'labels': labels,
    }
    return _fleet_port_catalog_cache


def _fleet_extract_next_port_code(v):
    port = v.get('next_port') if isinstance(v, dict) else None
    candidates = []
    if isinstance(port, dict):
        candidates.extend(port.get(key) for key in ('cd', 'code', 'locode', 'unlocode'))
    if not isinstance(v, dict):
        return ''
    candidates.extend((v.get('next_port_cd'), v.get('dest_cd')))
    for candidate in candidates:
        code = _norm_locode(candidate)
        if code:
            return code
    return ''


def _fleet_auto_next_port_identity(v):
    """Normalized automatic Next Port identity. Prefer explicit code over display text."""
    if not isinstance(v, dict):
        return None
    code = _fleet_extract_next_port_code(v)
    if code:
        return 'CODE:' + code
    port = v.get('next_port') if isinstance(v.get('next_port'), dict) else {}
    text = port.get('name') or v.get('dest_port') or v.get('next_port')
    norm = _norm_port_text(text)
    return ('TEXT:' + norm) if norm else None


def _fleet_resolve_port_input(value):
    """Resolve user-entered UN/LOCODE or unambiguous catalog name to name/code/xy."""
    if not isinstance(value, str):
        return None, 'port must be text'
    raw = value.strip()
    if not raw:
        return None, 'port required'
    if len(raw) > 120:
        return None, 'port too long'
    cat = _fleet_port_catalog()
    code = _norm_locode(raw)
    if code:
        xy = cat['locodes'].get(code)
        if xy:
            return {'label': cat['labels'].get(code) or code, 'code': code, 'xy': xy}, None

    parts = [p.strip() for p in raw.split(',') if p.strip()]
    name_raw = parts[0] if parts else raw
    name_key = _norm_port_text(re.sub(r'\(.*?\)|=.*', ' ', name_raw).split('/')[0])
    if not name_key:
        return None, 'port required'
    lookup_key = name_key
    if len(parts) > 1:
        iso = cat['countries'].get(parts[-1].upper())
        if not iso:
            return None, 'unknown country'
        lookup_key = iso + '|' + name_key
    matches = cat['by_name'].get(lookup_key) or set()
    if len(matches) != 1:
        return None, 'unknown or ambiguous port'
    lat, lng = next(iter(matches))
    label = re.sub(r'\s+', ' ', raw).strip()
    return {'label': label, 'code': None, 'xy': [float(lat), float(lng)]}, None


def _fleet_apply_code_first_next_port(v):
    """Correct automatic Next Port coordinates when an explicit code is present."""
    if not isinstance(v, dict):
        return
    code = _fleet_extract_next_port_code(v)
    if not code:
        return
    xy = _fleet_port_catalog()['locodes'].get(code)
    if not xy:
        return
    port = v.get('next_port')
    if not isinstance(port, dict):
        port = {}
    port['cd'] = code
    port['xy'] = [float(xy[0]), float(xy[1])]
    if not port.get('name'):
        port['name'] = _fleet_port_catalog()['labels'].get(code) or code
    v['next_port'] = port
    v['dest_xy'] = port['xy']
    v['dest_port'] = port.get('name') or v.get('dest_port')
    v['next_port_source'] = 'code'
    v['route_legs'] = _fleet_route_to_destination(v, port['xy'])


def _fleet_route_to_destination(v, dest_xy):
    dest = _valid_latlng_pair(dest_xy)
    if not dest:
        return []
    legs = v.get('route_legs') if isinstance(v, dict) else None
    if isinstance(legs, list):
        valid_legs = []
        for leg in legs:
            if not isinstance(leg, list):
                continue
            pts = []
            for point in leg:
                pt = _valid_latlng_pair(point)
                if pt:
                    pts.append(pt)
            if len(pts) >= 2:
                valid_legs.append(pts)
        if valid_legs:
            valid_legs[-1][-1] = dest
            return valid_legs
    here = _valid_latlng_pair([v.get('lat'), v.get('lng')]) if isinstance(v, dict) else None
    return [[here, dest]] if here else []


def _ensure_fleet_next_port_override_table():
    execute("""
        CREATE TABLE IF NOT EXISTS fleet_next_port_override (
            vessel_key     TEXT PRIMARY KEY,
            vessel_name    TEXT NOT NULL,
            manual_label   TEXT NOT NULL,
            manual_code    TEXT,
            manual_lat     REAL NOT NULL,
            manual_lng     REAL NOT NULL,
            auto_snapshot  TEXT NOT NULL,
            active         INTEGER NOT NULL DEFAULT 1,
            inactivated_at TEXT,
            inactivated_reason TEXT,
            created_by     TEXT,
            created_at     TEXT NOT NULL DEFAULT (datetime('now','localtime')),
            updated_by     TEXT,
            updated_at     TEXT NOT NULL DEFAULT (datetime('now','localtime'))
        )
    """)
    cols = {r['name'] for r in query("PRAGMA table_info(fleet_next_port_override)")}
    for col, ddl in (
            ('active', "ALTER TABLE fleet_next_port_override ADD COLUMN active INTEGER NOT NULL DEFAULT 1"),
            ('inactivated_at', "ALTER TABLE fleet_next_port_override ADD COLUMN inactivated_at TEXT"),
            ('inactivated_reason', "ALTER TABLE fleet_next_port_override ADD COLUMN inactivated_reason TEXT"),
            ('updated_by', "ALTER TABLE fleet_next_port_override ADD COLUMN updated_by TEXT")):
        if col not in cols:
            execute(ddl)


def _fleet_load_manual_overrides(ensure_schema=True):
    if ensure_schema:
        _ensure_fleet_next_port_override_table()
    elif not query("SELECT name FROM sqlite_master WHERE type='table' AND name='fleet_next_port_override'", one=True):
        return {}
    return {
        r['vessel_key']: dict(r)
        for r in query("SELECT * FROM fleet_next_port_override WHERE active=1")
    }


def _fleet_apply_manual_next_port_overrides(fleet, ensure_schema=True):
    overrides = _fleet_load_manual_overrides(ensure_schema=ensure_schema)
    if not overrides:
        return
    for v in fleet:
        key = _vkey(v.get('name'))
        row = overrides.get(key)
        if not row:
            continue
        auto_id = _fleet_auto_next_port_identity(v)
        if not auto_id or row['auto_snapshot'] != auto_id:
            continue
        xy = [float(row['manual_lat']), float(row['manual_lng'])]
        auto_port = v.get('next_port') if isinstance(v.get('next_port'), dict) else {}
        v['next_port_auto'] = {'name': auto_port.get('name'), 'cd': auto_port.get('cd') or auto_port.get('code')}
        v['next_port'] = {
            'name': row['manual_label'],
            'cd': row['manual_code'],
            'xy': xy,
            'manual': True,
            'source': 'manual',
        }
        v['dest_port'] = row['manual_label']
        v['dest_xy'] = xy
        v['next_port_manual'] = {'active': True, 'label': row['manual_label'], 'code': row['manual_code']}
        v['route_legs'] = _fleet_route_to_destination(v, xy)


# ── Fleet Map 수동 ETA 기입 (noon report ETA 누락 시 사람이 직접 입력) ──────────
# next_port override와 달리 fallback-only: noon ETA가 있으면 항상 auto 우선(fresh),
# 없을 때만 수동값 표시. lat/lng·snapshot 무효화 불필요(표시 문자열뿐)이라 단순.
_ETA_MANUAL_RE = re.compile(r'^(\d{1,2})/(\d{1,2})(?:\s+(\d{1,2}):(\d{2}))?$')


def _fleet_normalize_manual_eta(s):
    """사용자 입력 ETA → 'MM/DD' 또는 'MM/DD HH:MM'(LT). (정규화값, None) 또는 (None, 에러)."""
    if not isinstance(s, str):
        return None, 'eta required'
    s = s.strip()
    if not s:
        return None, 'eta required'
    m = _ETA_MANUAL_RE.match(s)
    if not m:
        return None, 'format: MM/DD or MM/DD HH:MM'
    mo, da, hh, mi = m.group(1), m.group(2), m.group(3), m.group(4)
    mo_i, da_i = int(mo), int(da)
    if not (1 <= mo_i <= 12 and 1 <= da_i <= 31):
        return None, 'invalid date (MM 1-12, DD 1-31)'
    if hh is not None:
        hh_i, mi_i = int(hh), int(mi)
        if not (0 <= hh_i <= 23 and 0 <= mi_i <= 59):
            return None, 'invalid time (HH 0-23, MM 0-59)'
        return f'{mo_i:02d}/{da_i:02d} {hh_i:02d}:{mi_i:02d}', None
    return f'{mo_i:02d}/{da_i:02d}', None


def _ensure_fleet_eta_override_table():
    execute("""
        CREATE TABLE IF NOT EXISTS fleet_eta_override (
            vessel_key    TEXT PRIMARY KEY,
            vessel_name   TEXT NOT NULL,
            manual_eta    TEXT NOT NULL,
            next_port_key TEXT,
            created_by    TEXT,
            created_at    TEXT NOT NULL DEFAULT (datetime('now','localtime')),
            updated_by    TEXT,
            updated_at    TEXT NOT NULL DEFAULT (datetime('now','localtime'))
        )
    """)
    cols = {r['name'] for r in query("PRAGMA table_info(fleet_eta_override)")}
    if 'next_port_key' not in cols:
        execute("ALTER TABLE fleet_eta_override ADD COLUMN next_port_key TEXT")


def _fleet_load_manual_eta_overrides(ensure_schema=True):
    if ensure_schema:
        _ensure_fleet_eta_override_table()
    elif not query("SELECT name FROM sqlite_master WHERE type='table' AND name='fleet_eta_override'", one=True):
        return {}
    return {r['vessel_key']: dict(r)
            for r in query("SELECT * FROM fleet_eta_override")}


def _fleet_apply_manual_eta_overrides(fleet, ensure_schema=True):
    overrides = _fleet_load_manual_eta_overrides(ensure_schema=ensure_schema)
    if not overrides:
        return
    for v in fleet:
        row = overrides.get(_vkey(v.get('name')))
        if not row:
            continue
        # 목적지(next_port) 바뀌면 = voyage 변경 → 이전 voyage용 수동 ETA는 stale. 표시 안 함.
        npk = row.get('next_port_key')
        if npk and npk != _fleet_auto_next_port_identity(v):
            continue
        # 수동값 자체는 노출(패널 입력칸 prefill/Reset 렌더용). 실제 표시는 noon ETA 없을 때만(auto 우선).
        v['eta_manual_value'] = row['manual_eta']
        if not v.get('eta'):
            v['eta'] = row['manual_eta']
            v['eta_zd'] = None            # 사람 입력=목적지 LT, 숫자 offset 없음
            v['eta_manual'] = True


def _fleet_consume_eta_overrides_on_noon(fleet):
    """Push 시 noon report ETA가 관측된 선박의 수동 ETA는 소비(삭제) — 진짜 갭필러(one-shot)로
    만들어, 이후 noon ETA가 다시 누락돼도 과거 수동값이 stale하게 재노출되지 않게 한다."""
    if not query("SELECT name FROM sqlite_master WHERE type='table' AND name='fleet_eta_override'", one=True):
        return 0
    have_noon = {_vkey(v.get('name')) for v in (fleet or []) if v.get('eta')}
    if not have_noon:
        return 0
    with _fleet_eta_lock:
        existing = {r['vessel_key'] for r in query("SELECT vessel_key FROM fleet_eta_override")}
        targets = have_noon & existing
        for k in targets:
            execute("DELETE FROM fleet_eta_override WHERE vessel_key=?", (k,))
    return len(targets)


def _fleet_invalidate_next_port_overrides_from_push(fleet, actor='fleet-push'):
    """One-way invalidate active manual overrides whose automatic source changed/missing."""
    _ensure_fleet_next_port_override_table()
    active = query("SELECT vessel_key, auto_snapshot FROM fleet_next_port_override WHERE active=1")
    if not active:
        return 0
    current = {}
    for v in fleet or []:
        if not isinstance(v, dict):
            continue
        _fleet_apply_code_first_next_port(v)
        current[_vkey(v.get('name'))] = _fleet_auto_next_port_identity(v)
    updates = []
    for row in active:
        if row['vessel_key'] not in current:
            continue
        auto_id = current.get(row['vessel_key'])
        if not auto_id:
            updates.append(('auto identity missing', row['vessel_key']))
        elif auto_id != row['auto_snapshot']:
            updates.append(('auto identity changed', row['vessel_key']))
    if not updates:
        return 0
    db = get_db()
    for reason, key in updates:
        db.execute("""
            UPDATE fleet_next_port_override
               SET active=0,
                   inactivated_at=datetime('now','localtime'),
                   inactivated_reason=?,
                   updated_by=?,
                   updated_at=datetime('now','localtime')
             WHERE vessel_key=? AND active=1
        """, (reason, actor, key))
    db.commit()
    return len(updates)


def _fleet_visible_auto_vessels():
    """Fleet items visible to the current UI user, with corrected automatic next port only."""
    try:
        with open(FLEET_MAP_FILE, encoding='utf-8') as f:
            data = json.load(f)
    except (FileNotFoundError, ValueError, OSError, json.JSONDecodeError):
        return []
    fleet = data.get('fleet') or []
    for v in fleet:
        _fleet_apply_code_first_next_port(v)
    vsup = {_vkey(r['vname']): r['sname'] for r in
            query("SELECT v.name AS vname, s.name AS sname FROM supervisor_vessels sv "
                  "JOIN vessels v ON v.id=sv.vessel_id JOIN supervisors s ON s.id=sv.supervisor_id")}
    for v in fleet:
        v['supervisor'] = vsup.get(_vkey(v.get('name')))
    fleet = [v for v in fleet if v.get('supervisor')]
    is_admin = (session.get('role') == 'admin')
    sup_id = session.get('supervisor_id')
    if sup_id and not is_admin:
        srow = query("SELECT name FROM supervisors WHERE id=?", (sup_id,), one=True)
        sup_name = srow['name'] if srow else None
        allowed = {(_vkey(r['name'])) for r in
                   query("SELECT v.name FROM supervisor_vessels sv "
                         "JOIN vessels v ON v.id=sv.vessel_id WHERE sv.supervisor_id=?", (sup_id,))}
        if allowed:
            fleet = [v for v in fleet if _vkey(v.get('name')) in allowed]
        elif sup_name:
            fleet = [v for v in fleet if v.get('supervisor') == sup_name]
        else:
            fleet = []
    return fleet


def _trmtdb_positions_refresh(api_key):
    """upstream 1회 갱신 — **백그라운드 스레드 전용**. 예외는 캐시 error 로만 남긴다
    (스레드에서 raise 하면 아무도 못 받고 삼켜지므로 여기서 끝낸다)."""
    global _trmtdb_position_refreshing
    try:
        req = urllib.request.Request(
            TRMTDB_SHIP_POSITION_URL,
            headers={'x-api-key': api_key, 'Accept': 'application/json'},
        )
        with urllib.request.urlopen(req, timeout=20) as res:
            payload = json.loads(res.read().decode('utf-8'))
        vessels = payload.get('vessels') if isinstance(payload, dict) else None
        if not isinstance(vessels, list):
            raise ValueError('TRMT DB ship-position payload missing vessels[]')
        with _trmtdb_position_lock:
            _trmtdb_position_cache.update(
                {'at': time.monotonic(), 'loaded': True, 'vessels': vessels,
                 'fetched_at': datetime.utcnow().isoformat(timespec='seconds'), 'error': None})
    except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError, OSError,
            http.client.HTTPException, ValueError, json.JSONDecodeError) as exc:
        # 오류 문자열은 사용자 API 응답에 내보내지 않는다(상세 upstream 정보/키 누출 방지).
        with _trmtdb_position_lock:
            _trmtdb_position_cache.update({'at': time.monotonic(), 'error': type(exc).__name__})
    finally:
        with _trmtdb_position_lock:
            _trmtdb_position_refreshing = False


def _trmtdb_positions():
    """TRMT DB 위치 API를 서버에서만 조회한다. upstream 장애 시 마지막 정상본/SVMS fallback.

    🔴 **요청 경로에서 upstream 을 기다리지 않는다(stale-while-revalidate).**
    실측 2026-07-29: `?platform=ALL` 응답이 33.5MB·3.4~3.8초다(선박 316척 × history·by_platform
    각 26KB — overlay 는 선박당 `latest` 431B 만 쓴다). 옛 구조는 TTL 만료 뒤 첫 요청이 이 왕복+파싱을
    통째로 뒤집어써서 `/api/fleet-map/data` 가 콜드 4초였고, gunicorn `-w 1` 이라 그 사이 다른 요청까지
    밀렸다(위젯 페이지 전환이 안 넘어가 보인 원인 중 하나).
    → 만료되면 **마지막 정상본을 즉시 반환**하고 갱신은 백그라운드 스레드가 한다.
    ⚠️ upstream 파라미터로 이력을 줄이는 길은 없었다(history=0·latest_only·include=latest 전부 무시,
       응답 33,484,019 bytes 동일 / gzip 요청도 무압축). 그건 upstream 쪽 과제로 남김.
    """
    global _trmtdb_position_refreshing
    now = time.monotonic()
    api_key = os.getenv('TRMTDB_API_KEY')
    start = False
    with _trmtdb_position_lock:
        cached = _trmtdb_position_cache
        # 마지막 시도가 실패였으면 짧은 TTL 로 곧 재시도한다.
        ttl = TRMTDB_POSITION_ERROR_TTL if cached['error'] else TRMTDB_POSITION_CACHE_TTL
        fresh = (cached['loaded'] or cached['error']) and now - cached['at'] < ttl
        if not fresh and api_key and not _trmtdb_position_refreshing:
            _trmtdb_position_refreshing = True
            start = True
        vessels, fetched_at = cached['vessels'], cached['fetched_at']
        error, loaded = cached['error'], cached['loaded']
    if start:
        # daemon = 워커 종료를 막지 않는다. 실패해도 다음 요청이 다시 건다.
        try:
            threading.Thread(target=_trmtdb_positions_refresh, args=(api_key,),
                             name='trmtdb-pos-refresh', daemon=True).start()
        except RuntimeError:
            # 스레드 생성 실패 시 플래그를 되돌린다 — 안 그러면 True 로 굳어 갱신이 영구히 멈춘다
            # (올마이트 지적). 다음 요청이 다시 시도하게 둔다.
            with _trmtdb_position_lock:
                _trmtdb_position_refreshing = False
    if not api_key:
        error = 'TRMT DB API key not configured'
    # 4번째 값 = "캐시본을 내줬는가". 이제 요청 경로는 **항상** 캐시본을 내주므로 적재 여부와 같다.
    return vessels, fetched_at, error, loaded


def _trmtdb_track_points(row):
    """TRMT DB AIS 응답의 과거 선위를 지도용 최소 필드로 정규화한다.

    ship-position API 배포본별 배열 키(history/track/positions)를 모두 받아들이되,
    원본의 MMSI·provider 메타데이터는 브라우저에 전달하지 않는다.
    """
    if not isinstance(row, dict):
        return []
    raw_points = []
    for key in ('history', 'track', 'positions'):
        values = row.get(key)
        if isinstance(values, list):
            raw_points.extend(values)
    points, seen = [], set()
    for point in raw_points:
        if not isinstance(point, dict):
            continue
        try:
            raw_lat = point.get('latitude')
            if raw_lat is None:
                raw_lat = point.get('lat')
            raw_lng = point.get('longitude')
            if raw_lng is None:
                raw_lng = point.get('lng', point.get('lon'))
            lat = float(raw_lat)
            lng = float(raw_lng)
        except (TypeError, ValueError):
            continue
        if not (-90 <= lat <= 90 and -180 <= lng <= 180):
            continue
        event_at = point.get('event_at') or point.get('timestamp') or point.get('reported_at')
        event_at = str(event_at) if event_at is not None else ''
        # 동일 측위 중복은 polyline의 불필요한 정점을 만들지 않는다.
        identity = (lat, lng, event_at)
        if identity in seen:
            continue
        seen.add(identity)
        points.append({'lat': lat, 'lng': lng, 'event_at': event_at})
    # ISO-8601 timestamps sort lexically. timestamp 없는 레코드는 마지막에 보낸다.
    points.sort(key=lambda item: (not bool(item['event_at']), item['event_at'] or ''))
    # API가 고해상도 이력을 주는 경우에도 지도 렌더 비용은 제한하되 가장 최신 구간을 남긴다.
    return points[-2000:]


def _trmtdb_track_row_for_vessel(rows, vessel):
    """IMO가 있으면 IMO exact-match만 허용하고, 없을 때만 정규화 선명 fallback한다."""
    wanted_imo = str(vessel.get('imo') or '').strip()
    if wanted_imo:
        return next((row for row in rows if isinstance(row, dict)
                     and str(row.get('imo') or '').strip() == wanted_imo), None)
    wanted_name = _vkey(vessel.get('name'))
    return next((row for row in rows if isinstance(row, dict)
                 and _vkey(row.get('vessel_name') or row.get('name')) == wanted_name), None)


def _overlay_trmtdb_positions(fleet, override_keys):
    """TRMT DB의 latest 위치를 fleet_map 항목에 병합. 이메일 수동 override가 최우선이다."""
    upstream, fetched_at, error, cached = _trmtdb_positions()
    by_name, by_imo = {}, {}
    for row in upstream:
        if not isinstance(row, dict):
            continue
        latest = row.get('latest') if isinstance(row.get('latest'), dict) else {}
        try:
            lat, lng = float(latest.get('latitude')), float(latest.get('longitude'))
        except (TypeError, ValueError):
            continue
        if not (-90 <= lat <= 90 and -180 <= lng <= 180):
            continue
        row = {**row, '_latest': latest, '_lat': lat, '_lng': lng}
        if row.get('vessel_name'):
            by_name[_vkey(row['vessel_name'])] = row
        if row.get('imo') not in (None, ''):
            by_imo[str(row['imo']).strip()] = row
    matched = 0
    for v in fleet:
        if _vkey(v.get('name')) in override_keys:
            continue
        src = by_imo.get(str(v.get('imo') or '').strip()) or by_name.get(_vkey(v.get('name')))
        if not src:
            continue
        latest = src['_latest']
        v['lat'], v['lng'] = src['_lat'], src['_lng']
        for source_key, target_key in (('heading', 'course'), ('speed', 'speed')):
            if latest.get(source_key) is not None:
                try:
                    v[target_key] = float(latest[source_key])
                except (TypeError, ValueError):
                    pass
        # 상태는 기존 SVMS 상태 체계를 유지한다. 이 API는 실시간 위치 전용이다.
        v['position_source'] = 'TRMT DB ' + str(latest.get('platform') or '')
        v['position_ts'] = latest.get('event_at') or src.get('latest_event_at')
        v['pos_source'] = 'trmtdb'
        v['pos_reported_at'] = v['position_ts']
        event_date = str(v['position_ts'] or '')[:10].replace('-', '')
        if len(event_date) == 8 and event_date.isdigit():
            v['rpt_dt'] = event_date
        matched += 1
    return {'source': 'TRMT DB', 'fetched_at': fetched_at, 'matched': matched,
            'upstream_vessels': len(upstream), 'cached': cached, 'error': error}


@app.route('/api/ext/fleet-map/push', methods=['POST'])
@api_key_required
def api_ext_fleet_map_push():
    """맥 스케줄러(run.sh)가 SVMS noon+TRMT 조인한 fleet_enriched.json 적재.
    파일 저장만(스키마 무관). 대시보드가 /api/fleet-map/data 로 읽음."""
    if request.content_length and request.content_length > 8 * 1024 * 1024:
        return jsonify({'ok': False, 'error': 'payload too large'}), 413
    d = request.get_json(silent=True)
    if not isinstance(d, dict) or not isinstance(d.get('fleet'), list):
        return jsonify({'ok': False, 'error': 'invalid payload (fleet[] required)'}), 400
    if len(d['fleet']) > 500:
        return jsonify({'ok': False, 'error': 'too many vessels'}), 400
    # 각 선박 최소 필드/타입 검증(오염 데이터 저장 차단)
    for v in d['fleet']:
        if (not isinstance(v, dict) or not v.get('name')
                or not isinstance(v.get('lat'), (int, float))
                or not isinstance(v.get('lng'), (int, float))):
            return jsonify({'ok': False, 'error': 'invalid fleet item (name/lat/lng required)'}), 400
    with _fleet_next_port_lock:
        for v in d['fleet']:
            _fleet_apply_code_first_next_port(v)
        d['_received_at'] = datetime.now().isoformat(timespec='seconds')
        tmp = FLEET_MAP_FILE + '.tmp'
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump(d, f, ensure_ascii=False)
        os.replace(tmp, FLEET_MAP_FILE)
        invalidated = _fleet_invalidate_next_port_overrides_from_push(d['fleet'])
        eta_consumed = _fleet_consume_eta_overrides_on_noon(d['fleet'])
    return jsonify({'ok': True, 'count': len(d.get('fleet') or []),
                    'generated_at': d.get('generated_at'),
                    'next_port_overrides_invalidated': invalidated,
                    'eta_overrides_consumed': eta_consumed})


FLEET_OVERRIDE_FILE = os.path.join(INSTANCE_DIR, 'fleet_map_overrides.json')


@app.route('/api/ext/fleet-map/override', methods=['POST'])
@api_key_required
def api_ext_fleet_map_override():
    """특정 선박 선위를 외부 소스(예: Master 이메일 보고)로 임시 override.
    payload: {vessel, lat, lng, course?, speed?, source?, reported_at?, clear?}
    clear=true 면 해당 선박 override 제거(=SVMS noon 위치로 복귀)."""
    d = request.get_json(silent=True)
    if not isinstance(d, dict) or not d.get('vessel'):
        return jsonify({'ok': False, 'error': 'vessel required'}), 400
    try:
        with open(FLEET_OVERRIDE_FILE, encoding='utf-8') as f:
            ov = json.load(f)
    except (FileNotFoundError, ValueError):
        ov = {}
    key = _vkey(d['vessel'])
    if d.get('clear'):
        ov.pop(key, None)
    else:
        if not isinstance(d.get('lat'), (int, float)) or not isinstance(d.get('lng'), (int, float)):
            return jsonify({'ok': False, 'error': 'lat/lng (number) required'}), 400
        ov[key] = {
            'vessel': d['vessel'], 'lat': d['lat'], 'lng': d['lng'],
            'course': d.get('course'), 'speed': d.get('speed'),
            'source': d.get('source') or 'email',
            'reported_at': d.get('reported_at'),
            'until': d.get('until'),   # 이 시각(KST ISO) 이후엔 hard override→fallback 전환
            'stored_at': datetime.now().isoformat(timespec='seconds'),
        }
    tmp = FLEET_OVERRIDE_FILE + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(ov, f, ensure_ascii=False)
    os.replace(tmp, FLEET_OVERRIDE_FILE)
    return jsonify({'ok': True, 'count': len(ov), 'key': key})


FLEET_WIND_FILE = os.path.join(INSTANCE_DIR, 'fleet_wind.json')


@app.route('/api/ext/fleet-map/wind', methods=['POST'])
@api_key_required
def api_ext_fleet_map_wind_push():
    """맥 wind_gfs.py 가 NOAA GFS 10m 바람을 leaflet-velocity 포맷으로 적재.
    payload: {grid:[{header,data},{header,data}], generated_at}. 대시보드 '바람' 토글이 GET으로 읽음."""
    if request.content_length and request.content_length > 4 * 1024 * 1024:
        return jsonify({'ok': False, 'error': 'payload too large'}), 413
    d = request.get_json(silent=True)
    grid = d.get('grid') if isinstance(d, dict) else None
    if (not isinstance(grid, list) or len(grid) != 2
            or not all(isinstance(g, dict) and isinstance(g.get('data'), list)
                       and isinstance(g.get('header'), dict) for g in grid)):
        return jsonify({'ok': False, 'error': 'invalid wind grid (2 entries with header/data[])'}), 400
    # 스키마 고정 — nx*ny=data길이, U/V 동일 길이, parameterNumber 2(U)/3(V) 확인(오염 차단)
    h0 = grid[0]['header']
    nx, ny = h0.get('nx'), h0.get('ny')
    if (not isinstance(nx, int) or not isinstance(ny, int)
            or len(grid[0]['data']) != nx * ny
            or len(grid[1]['data']) != len(grid[0]['data'])
            or {grid[0]['header'].get('parameterNumber'), grid[1]['header'].get('parameterNumber')} != {2, 3}):
        return jsonify({'ok': False, 'error': 'wind grid schema mismatch (nx*ny/len/paramNumber)'}), 400
    out = {'grid': grid, 'generated_at': d.get('generated_at'),
           '_received_at': datetime.now().isoformat(timespec='seconds')}
    tmp = FLEET_WIND_FILE + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, separators=(',', ':'))
    os.replace(tmp, FLEET_WIND_FILE)
    return jsonify({'ok': True, 'points': len(grid[0]['data']), 'generated_at': out['generated_at']})


@app.route('/api/fleet-map/wind')
@login_required
def api_fleet_map_wind():
    """대시보드 '바람' 토글용 — leaflet-velocity 그리드(GFS 10m)."""
    try:
        with open(FLEET_WIND_FILE, encoding='utf-8') as f:
            d = json.load(f)
    except (FileNotFoundError, ValueError):
        return jsonify({'grid': None, 'empty': True})
    return jsonify({'grid': d.get('grid'), 'generated_at': d.get('generated_at')})


FLEET_EMAIL_WATCH_FILE = os.path.join(INSTANCE_DIR, 'fleet_map_email_watch.json')
AIS_STALE_HOURS = 6   # AIS lastSeen이 이보다 오래면 '끊김' 자동표시(이메일 선위 후보)


def _load_email_watch():
    try:
        with open(FLEET_EMAIL_WATCH_FILE, encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, ValueError):
        return {}


# 수동 'SVMS 운항데이터 고정'(AIS off) 선박 — trmtdb/vesseltracker 오버레이를 건너뛰고 SVMS noon 위치 사용.
FLEET_AIS_OFF_FILE = os.path.join(INSTANCE_DIR, 'fleet_map_ais_off.json')


def _load_ais_off():
    try:
        with open(FLEET_AIS_OFF_FILE, encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, ValueError):
        return {}


@app.route('/api/fleet-map/email-watch', methods=['POST'])
@login_required
def api_fleet_map_email_watch_set():
    """대시보드 토글 — 선박을 '이메일 선위' watch에 등록/해제(AIS off 대응).
    payload: {vessel, enabled}. 워처(맥)가 GET /api/ext/fleet-map/email-watch 로 읽음."""
    d = request.get_json(silent=True)
    if not isinstance(d, dict) or not d.get('vessel'):
        return jsonify({'ok': False, 'error': 'vessel required'}), 400
    w = _load_email_watch()
    key = _vkey(d['vessel'])
    if d.get('enabled'):
        w[key] = {'vessel': d['vessel'],
                  'since': datetime.now().isoformat(timespec='seconds'),
                  'by': session.get('username') or session.get('supervisor_id')}
        # 상호배타 — email 켜면 수동 SVMS 고정(ais-off) 해제(구 endpoint 우회로 두 모드 공존 차단).
        _off = _load_ais_off()
        if _off.pop(key, None) is not None:
            _t = FLEET_AIS_OFF_FILE + '.tmp'
            with open(_t, 'w', encoding='utf-8') as f:
                json.dump(_off, f, ensure_ascii=False)
            os.replace(_t, FLEET_AIS_OFF_FILE)
    else:
        w.pop(key, None)
        # watch 해제 시 이메일 override도 제거 → 즉시 AIS/SVMS 위치로 복귀
        try:
            with open(FLEET_OVERRIDE_FILE, encoding='utf-8') as f:
                ov = json.load(f)
            if ov.pop(key, None) is not None:
                t2 = FLEET_OVERRIDE_FILE + '.tmp'
                with open(t2, 'w', encoding='utf-8') as f:
                    json.dump(ov, f, ensure_ascii=False)
                os.replace(t2, FLEET_OVERRIDE_FILE)
        except (FileNotFoundError, ValueError):
            pass
    tmp = FLEET_EMAIL_WATCH_FILE + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(w, f, ensure_ascii=False)
    os.replace(tmp, FLEET_EMAIL_WATCH_FILE)
    return jsonify({'ok': True, 'enabled': bool(d.get('enabled')), 'count': len(w)})


@app.route('/api/ext/fleet-map/email-watch')
@api_key_required
def api_ext_fleet_map_email_watch_get():
    """워처(맥)용 — 현재 이메일 선위 watch 켜진 선박 목록."""
    w = _load_email_watch()
    return jsonify({'ok': True, 'vessels': list(w.values()), 'keys': list(w.keys())})


@app.route('/api/fleet-map/pos-source', methods=['POST'])
@login_required
def api_fleet_map_pos_source_set():
    """대시보드 선박별 선위 소스 토글(상호배타 3택):
      source='ais'   → 자동(TRMT DB 실시간/AIS 우선). 이메일·SVMS 고정 해제.
      source='svms'  → AIS off, SVMS 운항데이터(noon)로 고정.
      source='email' → Master 이메일 선위 override watch(기존).
    payload: {vessel, source}."""
    d = request.get_json(silent=True)
    if not isinstance(d, dict) or not d.get('vessel'):
        return jsonify({'ok': False, 'error': 'vessel required'}), 400
    source = str(d.get('source') or '').strip().lower()
    if source not in ('ais', 'svms', 'email'):
        return jsonify({'ok': False, 'error': 'source required (ais|svms|email)'}), 400
    vessel = d.get('vessel')
    if not isinstance(vessel, str) or not vessel.strip() or len(vessel) > 120:
        return jsonify({'ok': False, 'error': 'vessel required'}), 400
    vessel = vessel.strip()
    key = _vkey(vessel)
    w = _load_email_watch()
    off = _load_ais_off()
    # 상호배타 — 먼저 두 모드 다 해제한 뒤 선택 모드만 설정.
    was_email = w.pop(key, None) is not None
    off.pop(key, None)
    meta = {'vessel': vessel, 'since': datetime.now().isoformat(timespec='seconds'),
            'by': session.get('username') or session.get('supervisor_id')}
    if source == 'email':
        w[key] = meta
    elif source == 'svms':
        off[key] = meta
    # 이메일 모드가 아니면 이메일 override(위치)도 항상 제거 → 즉시 AIS/SVMS 복귀(orphan override 방지).
    if source != 'email':
        try:
            with open(FLEET_OVERRIDE_FILE, encoding='utf-8') as f:
                ov = json.load(f)
            if ov.pop(key, None) is not None:
                t2 = FLEET_OVERRIDE_FILE + '.tmp'
                with open(t2, 'w', encoding='utf-8') as f:
                    json.dump(ov, f, ensure_ascii=False)
                os.replace(t2, FLEET_OVERRIDE_FILE)
        except (FileNotFoundError, ValueError):
            pass
    for path, obj in ((FLEET_EMAIL_WATCH_FILE, w), (FLEET_AIS_OFF_FILE, off)):
        tmp = path + '.tmp'
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump(obj, f, ensure_ascii=False)
        os.replace(tmp, path)
    return jsonify({'ok': True, 'source': source, 'vessel': vessel})


@app.route('/api/ext/fleet-map/ais-off')
@api_key_required
def api_ext_fleet_map_ais_off_get():
    """워처(맥 vt_overlay)용 — 수동 'SVMS 고정'(AIS off) 선박 목록. 키=선명 strip+lower."""
    off = _load_ais_off()
    return jsonify({'ok': True, 'vessels': list(off.values()), 'keys': list(off.keys())})


@app.route('/api/fleet-map/next-port-override', methods=['POST'])
@login_required
def api_fleet_map_next_port_override_set():
    """Dashboard write endpoint: save per-vessel manual Next Port override."""
    d = request.get_json(silent=True)
    if not isinstance(d, dict):
        return jsonify({'ok': False, 'error': 'invalid json'}), 400
    if session.get('role') != 'admin' and not session.get('supervisor_id'):
        return jsonify({'ok': False, 'error': 'forbidden'}), 403
    vessel = d.get('vessel')
    port_input = d.get('port')
    if not isinstance(vessel, str) or not vessel.strip() or len(vessel) > 120:
        return jsonify({'ok': False, 'error': 'vessel required'}), 400
    resolved, err = _fleet_resolve_port_input(port_input)
    if err:
        return jsonify({'ok': False, 'error': err}), 400
    key = _vkey(vessel)
    with _fleet_next_port_lock:
        visible = _fleet_visible_auto_vessels()
        v = next((x for x in visible if _vkey(x.get('name')) == key), None)
        if not v:
            return jsonify({'ok': False, 'error': 'vessel not found'}), 400
        auto_id = _fleet_auto_next_port_identity(v)
        if not auto_id:
            return jsonify({'ok': False, 'error': 'automatic next port unavailable'}), 400
        _ensure_fleet_next_port_override_table()
        xy = resolved['xy']
        user = session.get('username') or str(session.get('supervisor_id') or '')
        db = get_db()
        db.execute("""
            INSERT INTO fleet_next_port_override
                (vessel_key, vessel_name, manual_label, manual_code, manual_lat, manual_lng,
                 auto_snapshot, active, inactivated_at, inactivated_reason, created_by, updated_by, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, 1, NULL, NULL, ?, ?, datetime('now','localtime'))
            ON CONFLICT(vessel_key) DO UPDATE SET
                vessel_name=excluded.vessel_name,
                manual_label=excluded.manual_label,
                manual_code=excluded.manual_code,
                manual_lat=excluded.manual_lat,
                manual_lng=excluded.manual_lng,
                auto_snapshot=excluded.auto_snapshot,
                active=1,
                inactivated_at=NULL,
                inactivated_reason=NULL,
                updated_by=excluded.updated_by,
                updated_at=datetime('now','localtime')
        """, (key, v.get('name') or vessel.strip(), resolved['label'], resolved.get('code'),
              float(xy[0]), float(xy[1]), auto_id, user, user))
        db.commit()
    return jsonify({'ok': True, 'vessel': v.get('name'), 'next_port': {
        'name': resolved['label'], 'cd': resolved.get('code'), 'xy': [float(xy[0]), float(xy[1])],
        'manual': True,
    }})


@app.route('/api/fleet-map/next-port-override', methods=['DELETE'])
@login_required
def api_fleet_map_next_port_override_delete():
    """Dashboard write endpoint: clear per-vessel manual Next Port override."""
    d = request.get_json(silent=True)
    if not isinstance(d, dict):
        return jsonify({'ok': False, 'error': 'invalid json'}), 400
    if session.get('role') != 'admin' and not session.get('supervisor_id'):
        return jsonify({'ok': False, 'error': 'forbidden'}), 403
    vessel = d.get('vessel')
    if not isinstance(vessel, str) or not vessel.strip() or len(vessel) > 120:
        return jsonify({'ok': False, 'error': 'vessel required'}), 400
    key = _vkey(vessel)
    visible_keys = {_vkey(v.get('name')) for v in _fleet_visible_auto_vessels()}
    if key not in visible_keys:
        return jsonify({'ok': False, 'error': 'vessel not found'}), 400
    with _fleet_next_port_lock:
        _ensure_fleet_next_port_override_table()
        execute("DELETE FROM fleet_next_port_override WHERE vessel_key=?", (key,))
    return jsonify({'ok': True, 'vessel': vessel.strip()})


@app.route('/api/fleet-map/eta-override', methods=['POST'])
@login_required
def api_fleet_map_eta_override_set():
    """Dashboard write endpoint: noon ETA 누락 선박에 수동 ETA 기입."""
    d = request.get_json(silent=True)
    if not isinstance(d, dict):
        return jsonify({'ok': False, 'error': 'invalid json'}), 400
    if session.get('role') != 'admin' and not session.get('supervisor_id'):
        return jsonify({'ok': False, 'error': 'forbidden'}), 403
    vessel = d.get('vessel')
    if not isinstance(vessel, str) or not vessel.strip() or len(vessel) > 120:
        return jsonify({'ok': False, 'error': 'vessel required'}), 400
    norm, err = _fleet_normalize_manual_eta(d.get('eta'))
    if err:
        return jsonify({'ok': False, 'error': err}), 400
    key = _vkey(vessel)
    with _fleet_eta_lock:
        # 선박 존재 + scope 검증(_fleet_visible_auto_vessels = 현재 사용자 담당선만).
        v = next((x for x in _fleet_visible_auto_vessels() if _vkey(x.get('name')) == key), None)
        if not v:
            return jsonify({'ok': False, 'error': 'vessel not found'}), 400
        # 요청 취지="ETA 기입 안되어있을 경우". noon report ETA가 이미 있으면 수동 기입 불필요 → 거부
        # (auto 우선이므로 저장해도 shadow만 될 뿐, stale 재노출 소지 차단).
        if v.get('eta'):
            return jsonify({'ok': False, 'error': 'noon report ETA가 이미 있음 — 수동 기입 불필요'}), 400
        # 목적지 identity 바인딩 → voyage 바뀌면 apply에서 자동 만료(stale 방지).
        npk = _fleet_auto_next_port_identity(v)
        user = session.get('username') or str(session.get('supervisor_id') or '')
        _ensure_fleet_eta_override_table()
        db = get_db()
        db.execute("""
            INSERT INTO fleet_eta_override
                (vessel_key, vessel_name, manual_eta, next_port_key, created_by, updated_by, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, datetime('now','localtime'))
            ON CONFLICT(vessel_key) DO UPDATE SET
                vessel_name=excluded.vessel_name,
                manual_eta=excluded.manual_eta,
                next_port_key=excluded.next_port_key,
                updated_by=excluded.updated_by,
                updated_at=datetime('now','localtime')
        """, (key, v.get('name') or vessel.strip(), norm, npk, user, user))
        db.commit()
    return jsonify({'ok': True, 'vessel': v.get('name') or vessel.strip(), 'eta': norm, 'manual': True})


@app.route('/api/fleet-map/eta-override', methods=['DELETE'])
@login_required
def api_fleet_map_eta_override_delete():
    """Dashboard write endpoint: 수동 ETA 기입 삭제(noon 자동값으로 복귀)."""
    d = request.get_json(silent=True)
    if not isinstance(d, dict):
        return jsonify({'ok': False, 'error': 'invalid json'}), 400
    if session.get('role') != 'admin' and not session.get('supervisor_id'):
        return jsonify({'ok': False, 'error': 'forbidden'}), 403
    vessel = d.get('vessel')
    if not isinstance(vessel, str) or not vessel.strip() or len(vessel) > 120:
        return jsonify({'ok': False, 'error': 'vessel required'}), 400
    key = _vkey(vessel)
    if key not in {_vkey(v.get('name')) for v in _fleet_visible_auto_vessels()}:
        return jsonify({'ok': False, 'error': 'vessel not found'}), 400
    with _fleet_eta_lock:
        _ensure_fleet_eta_override_table()
        execute("DELETE FROM fleet_eta_override WHERE vessel_key=?", (key,))
    return jsonify({'ok': True, 'vessel': vessel.strip()})


@app.route('/api/fleet-map/data')
@login_required
def api_fleet_map_data():
    """대시보드 맵 데이터. 감독 연결 사용자는 본인 담당선박만(admin/미연결=전체)."""
    try:
        with open(FLEET_MAP_FILE, encoding='utf-8') as f:
            data = json.load(f)
    except (FileNotFoundError, ValueError):
        return jsonify({'fleet': [], 'supervisors': [], 'generated_at': None,
                        'empty': True})
    fleet = data.get('fleet') or []
    for v in fleet:
        _fleet_apply_code_first_next_port(v)
    # 선위 override(이메일 등 외부 소스) 병합 — 특정 선박만 임시로 다른 소스 위치 사용.
    try:
        with open(FLEET_OVERRIDE_FILE, encoding='utf-8') as f:
            overrides = json.load(f)
    except (FileNotFoundError, ValueError):
        overrides = {}
    override_keys = set()
    if overrides:
        now_k = datetime.utcnow() + timedelta(hours=9)
        for v in fleet:
            o = overrides.get(_vkey(v.get('name')))
            if not o:
                continue
            ov_date = str(o.get('reported_at') or '')[:10].replace('-', '')
            # until 지나면 hard override → fallback: SVMS가 override 보고일 이후 데이터 있으면 SVMS 사용,
            # SVMS 미갱신이면 마지막 override(이메일) 위치 유지.
            until = o.get('until')
            if until:
                try:
                    udt = datetime.strptime(str(until)[:16], '%Y-%m-%dT%H:%M')
                    if now_k >= udt:
                        svms_rpt = str(v.get('rpt_dt') or '')
                        if (len(svms_rpt) == 8 and svms_rpt.isdigit()
                                and len(ov_date) == 8 and svms_rpt >= ov_date):
                            continue   # SVMS 최신 → override 끔(SVMS 위치 사용)
                        # else: SVMS 미갱신 → 아래로 진행(override를 fallback으로 유지)
                except ValueError:
                    pass
            v['lat'] = o['lat']; v['lng'] = o['lng']
            if o.get('course') is not None: v['course'] = o['course']
            if o.get('speed') is not None: v['speed'] = o['speed']
            v['pos_source'] = o.get('source') or 'email'
            v['pos_reported_at'] = o.get('reported_at') or o.get('stored_at')
            override_keys.add(_vkey(v.get('name')))
            # 신선도 ALERT 오탐 방지: override 보고일을 rpt_dt로
            if len(ov_date) == 8 and ov_date.isdigit():
                v['rpt_dt'] = ov_date
    # 실시간 위치는 TRMT DB를 우선 사용. 이메일 override + 수동 SVMS 고정(ais-off) 선박은 덮어쓰기 제외.
    _ais_off = _load_ais_off()
    _ais_off_keys = set(_ais_off.keys())
    data['position_feed'] = _overlay_trmtdb_positions(fleet, override_keys | _ais_off_keys)
    # 감독 = TRMT supervisor_vessels(권위)로 채움 — 이슈 없는 선박도 올바른 감독/필터 표시.
    vsup = {_vkey(r['vname']): r['sname'] for r in
            query("SELECT v.name AS vname, s.name AS sname FROM supervisor_vessels sv "
                  "JOIN vessels v ON v.id=sv.vessel_id JOIN supervisors s ON s.id=sv.supervisor_id")}
    # supervisor = supervisor_vessels(TRMT DB) 권위값으로 '완전 대체'. build.py가 이슈기반으로 붙인 라벨은 무시
    # (안 그러면 매핑 삭제해도 이슈기반 라벨이 남아 필터에 뜸 — 손유석 정리 후 김흥민/이창주 잔존 버그).
    for v in fleet:
        v['supervisor'] = vsup.get(_vkey(v.get('name')))
    # 대시보드 = supervisor_vessels 배정된 선박만 표시(미배정·타팀 제외). 손유석 정리 후 손유석 담당선만 남음(손유석 지시 2026-06-29).
    # ⚠️ admin/비admin 공통 정책 — 배정 없는 감독(예 김흥민/이창주 멤버계정)은 빈 대시보드(의도). 빈 fleet은 프론트가 "표시할 선박 없음"으로 처리.
    fleet = [v for v in fleet if v.get('supervisor')]
    data['fleet'] = fleet
    data['supervisors'] = sorted({v['supervisor'] for v in fleet if v.get('supervisor')})
    # SIRE 검사일 +3주(21일) 초과인데 Observation All-close 안 됨(open>0) → 아이콘 노란 펄스
    overdue_vkeys = {
        _vkey(r['vname']) for r in query("""
            SELECT v2.name AS vname
              FROM vettings vt
              JOIN vessels v2 ON v2.id = vt.vessel_id
              LEFT JOIN (
                  SELECT vetting_id,
                         SUM(CASE WHEN status='Closed' THEN 1 ELSE 0 END) AS closed_n,
                         COUNT(*) AS total_n
                    FROM vt_findings GROUP BY vetting_id
              ) fc ON fc.vetting_id = vt.id
             WHERE vt.inspection_date IS NOT NULL AND vt.inspection_date != ''
               AND date(vt.inspection_date, '+21 days') < date('now','localtime')
               AND COALESCE(vt.manual_open_count,
                            MAX(0, COALESCE(vt.manual_observation_count, COALESCE(fc.total_n,0))
                                   - COALESCE(vt.manual_close_count, COALESCE(fc.closed_n,0)))) > 0
        """)
    }
    for v in fleet:
        v['sire_obs_overdue'] = _vkey(v.get('name')) in overdue_vkeys
    # 이메일 선위 watch 상태 + AIS 끊김 자동표시(이메일모드 후보)
    _watch = _load_email_watch()
    _now_epoch = (datetime.utcnow() - datetime(1970, 1, 1)).total_seconds()
    for v in fleet:
        _k = _vkey(v.get('name'))
        v['email_watch'] = _k in _watch
        v['ais_off'] = _k in _ais_off_keys              # 수동 SVMS 고정
        v['pos_mode'] = 'email' if v['email_watch'] else ('svms' if v['ais_off'] else 'ais')
        # 이메일 모드인데 아직 override 가 안 꽂힌 구간(watch 등록 직후 ~ 다음 watcher 실행).
        # 이 때 화면은 TRMT DB/AIS 를 보여주므로 "메일대로 안 바뀐다"로 오인됨 → 대기중임을 표면화.
        v['email_pending'] = bool(v['email_watch'] and _k not in override_keys)
        ep = v.get('position_ts_epoch')
        src = str(v.get('position_source') or '')
        # AIS 소스인데 마지막 측위가 AIS_STALE_HOURS 초과 → 끊김(이메일/SVMS 수동모드면 표시 안 함).
        # 단 email_pending 구간은 실제로 AIS 를 보여주는 중이므로 끊김 경고를 억제하면 안 됨
        # (억제하면 낡은 AIS 좌표가 아무 경고 없이 '현재 표시'로 나감 — 올마이트 지적, 2026-08-08).
        _mail_active = v['email_watch'] and not v['email_pending']
        v['ais_stale'] = bool(
            ep and 'AIS' in src and not _mail_active and not v['ais_off']
            and (_now_epoch - float(ep)) > AIS_STALE_HOURS * 3600)
    is_admin = (session.get('role') == 'admin')
    sup_id = session.get('supervisor_id')
    if sup_id and not is_admin:
        srow = query("SELECT name FROM supervisors WHERE id=?", (sup_id,), one=True)
        sup_name = srow['name'] if srow else None
        allowed = {(_vkey(r['name'])) for r in
                   query("SELECT v.name FROM supervisor_vessels sv "
                         "JOIN vessels v ON v.id=sv.vessel_id WHERE sv.supervisor_id=?", (sup_id,))}
        # 담당선박(supervisor_vessels, TRMT DB 권위) 매칭. 매핑이 비었을 때만 supervisor명 폴백.
        if allowed:
            fleet = [v for v in fleet if _vkey(v.get('name')) in allowed]
        elif sup_name:
            fleet = [v for v in fleet if v.get('supervisor') == sup_name]
        else:
            fleet = []
        data = {**data, 'fleet': fleet, 'scoped_to': sup_name}
    _fleet_apply_manual_next_port_overrides(fleet, ensure_schema=False)
    _fleet_apply_manual_eta_overrides(fleet, ensure_schema=False)
    # ── 데이터 신선도 ALERT (사이트 내 표시) ─────────────────────────────
    # KST = UTC+9 (서버 TZ 무관하게 utcnow 기준). 6h 스케줄 → 파이프라인/선박별 누락 산출.
    now_k = datetime.utcnow() + timedelta(hours=9)
    stale = {'pipeline': None, 'vessels': []}
    # 1) 파이프라인(push) 미갱신: 6h 주기 2회분(13h) 넘게 없으면 경보 + 며칠/몇시부터
    ga = data.get('generated_at')
    if ga:
        try:
            gdt = datetime.strptime(str(ga)[:16], '%Y-%m-%d %H:%M')
            age_h = (now_k - gdt).total_seconds() / 3600
            if age_h >= 13:
                stale['pipeline'] = {'last': str(ga)[:16], 'at': gdt.strftime('%-m/%-d %H:%M'),
                                     'days': int(age_h // 24), 'hours': int(age_h)}
        except ValueError:
            pass
    # 2) 선박별 noon 보고 누락: 어제(전날)도 보고 안 된 선박만 = miss>=2 (오늘 6/23이면 6/22까지 미보고).
    #    어제 보고는 정상으로 봄(손유석 2026-06-23). 며칠부터 끊겼는지 함께 표기.
    today = now_k.date()
    miss_threshold = 2
    for v in (data.get('fleet') or []):
        # SVMS noon 보고 대상이 아닌 선박(stub, 타 관리사 등)은 '누락' 집계 제외 — AIS로 추적 중.
        if v.get('no_noon'):
            continue
        sup = v.get('supervisor')
        rd = str(v.get('rpt_dt') or '')
        if len(rd) == 8 and rd.isdigit():
            try:
                d0 = datetime.strptime(rd, '%Y%m%d').date()
            except ValueError:
                continue
            miss = (today - d0).days
            if miss >= miss_threshold:
                nxt = d0 + timedelta(days=1)
                stale['vessels'].append({'name': v.get('name'), 'last_rpt': d0.strftime('%-m/%-d'),
                                         'since': nxt.strftime('%-m/%-d'), 'days': miss, 'sup': sup})
        else:
            stale['vessels'].append({'name': v.get('name'), 'last_rpt': None,
                                     'since': None, 'days': None, 'sup': sup})
    stale['vessels'].sort(key=lambda x: (x['days'] or 9999), reverse=True)
    data['staleness'] = stale
    # 로그인 사용자의 감독명(admin 포함) — 대시보드 기본필터를 본인 감독으로.
    my_sup = None
    _sid = session.get('supervisor_id')
    if _sid:
        _r = query("SELECT name FROM supervisors WHERE id=?", (_sid,), one=True)
        my_sup = _r['name'] if _r else None
    data['my_supervisor'] = my_sup
    return jsonify(data)


@app.route('/api/fleet-map/track')
@login_required
def api_fleet_map_track():
    """선택·권한범위 내 선박의 TRMT DB AIS 이전 항적만 반환한다."""
    vessel_name = str(request.args.get('vessel') or '').strip()
    if not vessel_name or len(vessel_name) > 120:
        return jsonify({'ok': False, 'error': 'vessel required'}), 400

    # 동일 로그인/담당선박 스코프를 Fleet Map 본문과 공유한다. 이름만 아는 사용자가
    # 타 담당선박 AIS 이력을 조회하는 것을 막는다.
    visible_response = api_fleet_map_data()
    visible = visible_response.get_json(silent=True) or {}
    vessel = next((v for v in visible.get('fleet') or []
                   if _vkey(v.get('name')) == _vkey(vessel_name)), None)
    if vessel is None:
        return jsonify({'ok': False, 'error': 'vessel not available'}), 404

    upstream, fetched_at, upstream_error, cached = _trmtdb_positions()
    source = _trmtdb_track_row_for_vessel(upstream, vessel)
    points = _trmtdb_track_points(source)
    return jsonify({
        'ok': True,
        'vessel': vessel.get('name'),
        'points': points,
        'available': len(points) >= 2,
        'source': 'TRMT DB AIS',
        'fetched_at': fetched_at,
        'cached': cached,
        # upstream의 상세 오류·endpoint/key는 브라우저에 노출하지 않는다.
        'error': ('unavailable' if upstream_error else None),
    })


@app.route('/dashboard/classic')
@login_required
def dashboard_classic():
    """구 대시보드(카드형) — Fleet Map 도입 후 백업 경로."""
    return render_template('dashboard_classic.html', **_dashboard_ctx())


@app.route('/api/ext/shipwiki/decided')
@api_key_required
def api_ext_shipwiki_decided():
    """맥 apply_decisions.py 가 적용할 결정건 → card_status='applying' 락(조건부).
    ?peek=1 이면 락 없이 미리보기."""
    cols = ("id, slug, fname, tier, decision, merge_group, new_title, new_category, new_conf, "
            "decided_judgment, source_msgids")
    if request.args.get('peek'):
        rows = query(f"SELECT {cols} FROM shipwiki_card WHERE card_status='decided' ORDER BY merge_group, id")
        return jsonify({'count': len(rows), 'cards': [dict(r) for r in rows], 'peek': True})
    out = [dict(r) for r in query(f"SELECT {cols} FROM shipwiki_card WHERE card_status='applying' ORDER BY merge_group, id")]
    for r in query(f"SELECT {cols} FROM shipwiki_card WHERE card_status='decided' ORDER BY merge_group, id"):
        if execute_rc("UPDATE shipwiki_card SET card_status='applying' WHERE id=? AND card_status='decided'", (r['id'],)):
            out.append(dict(r))
    return jsonify({'count': len(out), 'cards': out})


@app.route('/api/ext/shipwiki/<int:cid>/result', methods=['POST'])
@api_key_required
def api_ext_shipwiki_result(cid):
    """적용 결과: ok=True → applied(+result 파일경로), else failed(사람 재검토)."""
    d = request.get_json(silent=True) or {}
    ok = bool(d.get('ok'))
    rc = execute_rc("UPDATE shipwiki_card SET card_status=?, done_at=datetime('now','localtime'), "
                    "result=? WHERE id=? AND card_status='applying'",
                    ('applied' if ok else 'failed', (d.get('result') or '')[:2000], cid))
    return jsonify({'id': cid, 'ok': ok, 'applied': bool(rc)})


# ═════════════════════════════════════════════════════════════════
#  CLASS STATUS (선급 Class Status Report 업로드/추출/매칭)
# ═════════════════════════════════════════════════════════════════
import re as _re_cls


def _norm_vessel_name(name):
    """선명 정규화: 대문자, M/T·M/V 접두 제거, 공백 단일화."""
    if not name:
        return ''
    s = str(name).upper().strip()
    s = _re_cls.sub(r'^(M[\./]?\s*[TV][\./]?|MT|MV)\s+', '', s)  # M/T, M.V., MT, MV ...
    s = _re_cls.sub(r'[^A-Z0-9 ]+', ' ', s)
    s = _re_cls.sub(r'\s+', ' ', s).strip()
    return s


def _match_vessel_by_name(name):
    """보고서 선명 → vessels 행 매칭. 정확 일치 우선, 없으면 부분포함. 실패 시 None."""
    target = _norm_vessel_name(name)
    if not target:
        return None
    rows = query('SELECT * FROM vessels WHERE active=1')
    norm = [(v, _norm_vessel_name(v['name'])) for v in rows]
    for v, n in norm:
        if n == target:
            return v
    # 부분 포함 (한쪽이 다른 쪽을 포함)
    for v, n in norm:
        if n and (n in target or target in n):
            return v
    return None


def _annotate_drafts_with_vessel(drafts):
    """P4 표시전용(read-only): 각 draft 행에 matched_vessel:{id,name,in_my_roster} 부가.

    돈 파이프라인·draft 원본·status·금액 무변경. money 테이블 write 없음(읽기시점 계산).
    매칭 순서: vessels.vsl_cd 정확일치 우선 → 없으면 선명 정규화(_match_vessel_by_name).
    in_my_roster = 매칭 선박이 현재 세션 감독의 supervisor_vessels 에 포함되는지
      (supervisor_id 미설정 admin은 전체 로스터로 간주 → 매칭되면 True).
    각 draft dict 에 'matched_vessel' 키만 추가(없으면 None). 리스트 그대로 반환.
    """
    if not drafts:
        return drafts
    try:
        vrows = query('SELECT id, name, vsl_cd FROM vessels WHERE active=1')
    except Exception:
        # 조회 실패 시 표시기능만 조용히 생략 — 목록 응답 자체는 절대 깨지 않는다.
        for d in drafts:
            d.setdefault('matched_vessel', None)
        return drafts
    # 매칭 블록 전체를 방어적으로 감싼다 — supervisor_vessels 조회나 선명매칭이
    # 어떤 이유로 예외를 던져도 목록 API(500)를 깨지 않고 표시기능만 조용히 생략.
    try:
        by_cd = {}
        for v in vrows:
            cd = (v['vsl_cd'] or '').strip().upper()
            if cd:
                by_cd.setdefault(cd, v)
        # 내 로스터(현재 세션 감독) 선박 id 집합. 감독 미설정이면 None(=전체 로스터).
        sup_id = session.get('supervisor_id')
        my_ids = None
        if sup_id:
            my_ids = {r['vessel_id'] for r in
                      query('SELECT vessel_id FROM supervisor_vessels WHERE supervisor_id=?', (sup_id,))}
        for d in drafts:
            mv = None
            cd = (d.get('vsl_cd') or '').strip().upper()
            v = by_cd.get(cd) if cd else None
            if v is None:
                v = _match_vessel_by_name(d.get('vsl_nm') or d.get('vsl_cd'))
            if v is not None:
                in_roster = True if my_ids is None else (v['id'] in my_ids)
                mv = {'id': v['id'], 'name': v['name'], 'in_my_roster': bool(in_roster)}
            d['matched_vessel'] = mv
    except Exception:
        for d in drafts:
            d.setdefault('matched_vessel', None)
    return drafts


def _class_status_prompt():
    return (
        "다음은 선박 선급(Classification Society)의 'Class Status Report' 또는 "
        "'Survey Status Report'다. (선급 예: DNV, BV, KR, ABS, LR, NK 등 — 포맷이 다를 수 있다.)\n"
        "아래 정보를 추출해 지정한 JSON으로만 답하라.\n"
        "■ 공통 정보\n"
        "- vessel_name: 보고서의 선명(Name of vessel / Ship name). 대문자 원문.\n"
        "- class_society: 발행 선급 약어 (DNV / BV / KR / ABS / LR / NK 중 하나, 식별 가능하면).\n"
        "- report_date: 보고서 발행일/생성일 (Date of issue / Generated on). 가능하면 YYYY-MM-DD.\n"
        "■ 추출 대상 — 'Open(미해소)' 상태인 항목만:\n"
        "  (1) coc  = Condition of Class / 선급지적. 선급별 명칭 예:\n"
        "      DNV 'Conditions related to class', BV 'Conditions of Class', "
        "ABS 'Conditions of Class / Outstanding', LR 'Conditions of Class(COC)', "
        "또한 BV 'Planned Inspection Items'의 Recommendation(R)/Condition of Class 도 포함.\n"
        "  (2) statutory = Condition of Statutory / 기국(법정)사항. 예:\n"
        "      DNV 'Conditions related to statutory certificates', "
        "BV 'Statutory Recommendations' 및 'Planned Inspection Items'의 Statutory Condition/Recommendation 항목. "
        "⚠️ 단, Type이 'Observation'(Obs)인 행은 statutory에도 coc에도 절대 넣지 마라(아래 제외 규칙).\n"
        "■ 제외(절대 추출 금지): 단순 Survey 예정표(1-Year Planner/Surveys 목록), 인증서 목록, "
        "**모든 Observation 항목 — Type/VS 칸이 'Obs' 또는 'Observation'인 행(특히 'Planned Inspection Items'의 Obs 행, "
        "예: 'STS plan to be approved and placed on board', 'BWMP to be approved …')은 due date가 있어도 절대 추출하지 마라**, "
        "그리고 **Memoranda 섹션 전체**. 제목에 'Memoranda(메모란다)'가 들어간 표·섹션 — "
        "'Class Memoranda', 'Statutory Memoranda', 'Description of (Class/Statutory) Memoranda' 등 — 의 항목은 "
        "내용이 지적·기국처럼 보여도(예: 'Engine Power Limitation (SHaPoLi) approved, limiting … kW') 절대 추출하지 마라. "
        "⚠️ 'Statutory Memoranda'는 'Statutory Recommendations'와 전혀 다른 별개 섹션이다 — 'Statutory' 단어가 같다고 혼동 금지. "
        "메모란다는 단순 정보성 기록(approved/완료 통보 등)이라 미해소 조치사항이 아니다. "
        "이미 Closed/Cleared/Deleted 되었거나 조치 확인 완료된 항목도 제외. 'None'이면 빈 배열.\n"
        "■ 각 항목 필드:\n"
        "- issued_date: 발행/기재일 (가능하면 YYYY-MM-DD, 없으면 빈 문자열)\n"
        "- description: 지적/기국 본문을 원문 그대로 복사(영문이면 영문 그대로). 요약·변형 금지.\n"
        "- due_date: 마감/처리기한 (Due/Limit date, 가능하면 YYYY-MM-DD, 없으면 빈 문자열). "
        "⚠️ **연장(postpone/extend)된 경우 반드시 최종(연장된) 날짜를 due_date로 한다.** "
        "보고서에 원래 기한과 연장 기한이 함께 있거나(예: 'Original due 2025-04-26, postponed to 2026-04-26', "
        "'Limit date revised/extended to …', 'New limit date …', 'Postponed until …'), "
        "여러 날짜가 보이면 **가장 나중(최신) 유효 기한**을 due_date로 쓴다. 원래(이른) 날짜를 쓰지 마라.\n"
        "- remark: description의 핵심을 한국어 1~2문장으로 간결히 요약(전체 직역 금지). "
        "문장은 '~함/~됨/~음' 음슴체(개조식). 기술 명칭·장비명·약어·인증명(예: COC, SEEMP, IHM, "
        "BNWAS, Load Line, Plimsoll Mark, EGCS, BWTS)은 영문 그대로 둔다." + _MARITIME_TERMS + "\n"
        "없는 내용을 지어내지 말 것.\n"
        '형식: {"vessel_name":"","class_society":"","report_date":"",'
        '"coc":[{"issued_date":"","description":"","due_date":"","remark":""}],'
        '"statutory":[{"issued_date":"","description":"","due_date":"","remark":""}]}'
    )


def _cls_item(it):
    if not isinstance(it, dict):
        return None
    rec = {
        'issued_date': (it.get('issued_date') or '').strip(),
        'description': (it.get('description') or '').strip(),
        'due_date':    (it.get('due_date') or '').strip(),
        'remark':      (it.get('remark') or '').strip(),
    }
    return rec if rec['description'] else None


def _normalize_class_status(parsed):
    if not isinstance(parsed, dict):
        return None
    def lst(key):
        out = []
        for it in (parsed.get(key) or []):
            r = _cls_item(it)
            if r:
                out.append(r)
        return out
    return {
        'vessel_name':   (parsed.get('vessel_name') or '').strip(),
        'class_society': (parsed.get('class_society') or '').strip().upper(),
        'report_date':   (parsed.get('report_date') or '').strip(),
        'coc':           lst('coc'),
        'statutory':     lst('statutory'),
    }


def _xlsx_to_text(raw_bytes):
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    lines = []
    for ws in wb.worksheets:
        for r in ws.iter_rows(values_only=True):
            cells = ['' if c is None else str(c).strip() for c in r]
            if any(cells):
                lines.append('\t'.join(cells))
            if len(lines) > 600:
                break
    return '\n'.join(lines)


def _extract_class_status_from_upload(f):
    """업로드 FileStorage → (data, err). data = _normalize_class_status 결과."""
    name = (f.filename or '').lower()
    ext = name.rsplit('.', 1)[-1] if '.' in name else ''
    raw = f.read()
    size_mb = len(raw) / (1024 * 1024)
    prompt = _class_status_prompt()

    if ext == 'pdf':
        if size_mb > 15:
            return None, {'reason': 'TOO_LARGE',
                          'message': f'PDF가 너무 큽니다({size_mb:.1f}MB). 15MB 이하로 줄여주세요.'}
        b64 = __import__('base64').standard_b64encode(raw).decode()
        parsed = _gemini_call_json([
            {'inline_data': {'mime_type': 'application/pdf', 'data': b64}},
            {'text': prompt},
        ], model=_model_for('findings'))
    elif ext in ('png', 'jpg', 'jpeg', 'webp', 'gif', 'bmp'):
        if size_mb > 15:
            return None, {'reason': 'TOO_LARGE', 'message': f'이미지가 너무 큽니다({size_mb:.1f}MB).'}
        import mimetypes
        media = mimetypes.guess_type(name)[0] or 'image/jpeg'
        b64 = __import__('base64').standard_b64encode(raw).decode()
        parsed = _gemini_call_json([
            {'inline_data': {'mime_type': media, 'data': b64}},
            {'text': prompt},
        ], model=_model_for('findings'))
    elif ext in ('xlsx', 'xls'):
        try:
            txt = _xlsx_to_text(raw)
        except Exception as e:
            app.logger.exception('extract-class-status-from-upload')
            return None, {'reason': 'XLSX_PARSE_FAILED', 'message': f'엑셀을 읽지 못했습니다: {e}'}
        parsed = _gemini_call_json([{'text': prompt + '\n\n[보고서 표 내용]\n' + txt}],
                                   model=_model_for('findings'))
    else:
        return None, {'reason': 'BAD_TYPE', 'message': 'PDF · 이미지 · 엑셀(xlsx) 파일만 지원합니다.'}

    if isinstance(parsed, dict) and parsed.get('error') == 'NO_API_KEY':
        return None, {'reason': 'no_api_key', 'message': 'AI 자동추출이 설정되지 않았습니다(키 미설정).'}
    if isinstance(parsed, dict) and parsed.get('error'):
        return None, {'reason': parsed['error'], 'message': '자동 추출에 실패했습니다.',
                      'detail': parsed.get('detail') or parsed.get('raw')}
    data = _normalize_class_status(parsed)
    if data is None:
        return None, {'reason': 'PARSE_FAILED', 'message': '추출 결과를 해석하지 못했습니다.'}
    return data, None


def _cls_snapshot_dict(cs_row, items_by_cs):
    items = items_by_cs.get(cs_row['id'], [])
    coc = [dict(i) for i in items if i['category'] == 'COC']
    stat = [dict(i) for i in items if i['category'] == 'STATUTORY']
    return {
        'id':              cs_row['id'],
        'vessel_id':       cs_row['vessel_id'],
        'vessel_name_raw': cs_row['vessel_name_raw'],
        'class_society':   cs_row['class_society'],
        'report_date':     cs_row['report_date'],
        'source_filename': cs_row['source_filename'],
        'has_file':        bool(cs_row['source_path']) if 'source_path' in cs_row.keys() else False,
        'updated_at':      cs_row['updated_at'],
        'coc':             coc,
        'statutory':       stat,
    }


def _cls_delete_file(path):
    """보관 파일 삭제(교체 시 이전 파일 자동삭제). 경로가 업로드 폴더 내일 때만."""
    if not path:
        return
    try:
        full = os.path.join(BASE_DIR, path) if not os.path.isabs(path) else path
        if os.path.commonpath([os.path.realpath(full), os.path.realpath(UPLOAD_DIR)]) == os.path.realpath(UPLOAD_DIR) \
                and os.path.isfile(full):
            os.remove(full)
    except Exception as e:
        print(f'[cls] old file remove skip: {e}')


def _cls_save_snapshot(vessel_id, vessel_name_raw, data, filename, source_path=None):
    """선박 스냅샷 교체(최신만 유지). 이전 스냅샷의 보관파일도 자동삭제.
    vessel_id None 이면 미매칭으로 저장(같은 정규화 선명의 기존 미매칭 제거 후 삽입)."""
    conn = get_db()
    user = session.get('username')
    _ndesc = lambda s: ' '.join((s or '').strip().lower().split())
    preserved = {}   # (category, 정규화 description) -> action_taken — 스냅샷 교체에도 손유석 조치사항 유지
    if vessel_id is not None:
        try:
            for r in conn.execute(
                "SELECT i.category, i.description, i.action_taken "
                "FROM class_status_items i JOIN class_status c ON c.id = i.cs_id "
                "WHERE c.vessel_id = ? AND IFNULL(i.action_taken,'') <> ''", (vessel_id,)).fetchall():
                preserved[(r['category'], _ndesc(r['description']))] = r['action_taken']
        except Exception:
            app.logger.exception('cls-save-snapshot')
            preserved = {}
        for r in conn.execute('SELECT source_path FROM class_status WHERE vessel_id=?', (vessel_id,)).fetchall():
            _cls_delete_file(r['source_path'])
        conn.execute('DELETE FROM class_status WHERE vessel_id=?', (vessel_id,))
    else:
        # 같은 (정규화) 선명의 기존 미매칭 스냅샷 제거
        tgt = _norm_vessel_name(vessel_name_raw)
        for r in conn.execute('SELECT id, vessel_name_raw, source_path FROM class_status WHERE vessel_id IS NULL').fetchall():
            if _norm_vessel_name(r['vessel_name_raw']) == tgt:
                _cls_delete_file(r['source_path'])
                conn.execute('DELETE FROM class_status WHERE id=?', (r['id'],))
    cur = conn.execute(
        '''INSERT INTO class_status
             (vessel_id, vessel_name_raw, class_society, report_date, source_filename, source_path, uploaded_by)
           VALUES (?,?,?,?,?,?,?)''',
        (vessel_id, vessel_name_raw, data.get('class_society'),
         data.get('report_date'), filename, source_path, user))
    cs_id = cur.lastrowid
    for cat, key in (('COC', 'coc'), ('STATUTORY', 'statutory')):
        for n, it in enumerate(data.get(key) or [], start=1):
            act = preserved.get((cat, _ndesc(it.get('description'))), '')
            conn.execute(
                '''INSERT INTO class_status_items
                     (cs_id, category, no, issued_date, description, due_date, remark, action_taken)
                   VALUES (?,?,?,?,?,?,?,?)''',
                (cs_id, cat, n, it.get('issued_date'), it.get('description'),
                 it.get('due_date'), it.get('remark'), act))
    conn.commit()
    return cs_id


@app.route('/api/class-status', methods=['GET'])
@login_required
def api_class_status_list():
    """매칭 선박별 스냅샷 + 미매칭 버킷.
    Query: ?supervisor_id=N (지정 시 해당 감독 담당선박만, 미매칭은 미포함)"""
    sup_id = request.args.get('supervisor_id', type=int)

    all_cs = query('SELECT * FROM class_status ORDER BY updated_at DESC')
    cs_ids = [r['id'] for r in all_cs]
    items_by_cs = {cid: [] for cid in cs_ids}
    if cs_ids:
        ph = ','.join('?' * len(cs_ids))
        for it in query(f'SELECT * FROM class_status_items WHERE cs_id IN ({ph}) '
                        f'ORDER BY cs_id, category, no', tuple(cs_ids)):
            items_by_cs[it['cs_id']].append(it)

    snap_by_vessel = {r['vessel_id']: r for r in all_cs if r['vessel_id'] is not None}

    # 대상 선박: 스냅샷 보유 선박만 (감독 필터 적용)
    vessel_ids = list(snap_by_vessel.keys())
    vessels = []
    if vessel_ids:
        ph = ','.join('?' * len(vessel_ids))
        sql = f'SELECT * FROM vessels WHERE id IN ({ph})'
        params = list(vessel_ids)
        if sup_id:
            sql += (' AND EXISTS (SELECT 1 FROM supervisor_vessels sv '
                    'WHERE sv.vessel_id=vessels.id AND sv.supervisor_id=?)')
            params.append(sup_id)
        sql += ' ORDER BY name'
        vessels = query(sql, tuple(params))

    sv_map = {}
    if vessels:
        vids = [v['id'] for v in vessels]
        ph2 = ','.join('?' * len(vids))
        for r in query(f'SELECT vessel_id, supervisor_id FROM supervisor_vessels '
                       f'WHERE vessel_id IN ({ph2})', tuple(vids)):
            sv_map.setdefault(r['vessel_id'], []).append(r['supervisor_id'])

    vessel_out = []
    for v in vessels:
        vd = dict(v)
        vd['supervisor_ids'] = sv_map.get(v['id'], [])
        vessel_out.append({
            'vessel': vd,
            'snapshot': _cls_snapshot_dict(snap_by_vessel[v['id']], items_by_cs),
        })

    unmatched = []
    if not sup_id:
        for r in all_cs:
            if r['vessel_id'] is None:
                unmatched.append(_cls_snapshot_dict(r, items_by_cs))

    return jsonify({'vessels': vessel_out, 'unmatched': unmatched})


def _cls_handle_files(files):
    """업로드 파일들 → AI추출 → 선박매칭 → 저장. 원본파일도 선박별 최신만 보관. (UI·BV Pushing 공용)"""
    cls_dir = os.path.join(UPLOAD_DIR, 'class_status')
    os.makedirs(cls_dir, exist_ok=True)
    results = []
    for f in [x for x in files if x and x.filename]:
        fname = f.filename
        # 원본 바이트 보관(추출이 스트림을 소비하므로 추출 전에 읽고 seek 리셋)
        raw = None
        try:
            f.stream.seek(0); raw = f.read(); f.stream.seek(0)
        except Exception as _e:
            app.logger.warning('cls-handle-files: %s', _e)
            raw = None
        data, err = _extract_class_status_from_upload(f)
        if err:
            results.append({'filename': fname, 'ok': False, **err})
            continue
        vname = data.get('vessel_name') or ''
        v = _match_vessel_by_name(vname)
        vessel_id = v['id'] if v else None
        src_rel = None
        if raw:
            uniq = uuid.uuid4().hex[:8] + '_' + datetime.now().strftime('%Y%m%d%H%M%S%f') + '_' + (secure_filename(fname) or 'report')
            try:
                with open(os.path.join(cls_dir, uniq), 'wb') as out:
                    out.write(raw)
                src_rel = os.path.join('static', 'uploads', 'class_status', uniq)
            except Exception as e:
                print(f'[cls] file save skip: {e}')
        _cls_save_snapshot(vessel_id, vname, data, fname, src_rel)
        results.append({
            'filename': fname, 'ok': True,
            'vessel_name': vname,
            'matched': bool(v),
            'vessel_id': vessel_id,
            'matched_name': v['name'] if v else None,
            'class_society': data.get('class_society'),
            'report_date': data.get('report_date'),
            'coc_count': len(data.get('coc') or []),
            'statutory_count': len(data.get('statutory') or []),
        })
    return results


@app.route('/api/class-status/upload', methods=['POST'])
@login_required
def api_class_status_upload():
    files = request.files.getlist('files') or (
        [request.files['file']] if 'file' in request.files else [])
    if not [f for f in files if f and f.filename]:
        return jsonify({'ok': False, 'message': '파일이 없습니다.'}), 400
    results = _cls_handle_files(files)
    return jsonify({'ok': any(r.get('ok') for r in results), 'results': results})


@app.route('/api/class-status/push', methods=['POST'])
@admin_required
def api_class_status_push():
    """'BV에서 Pushing' 버튼 — 맥 러너가 폴링해서 BV→Class Status 동기화하도록 플래그."""
    _ensure_api_table()
    now = query("SELECT datetime('now','localtime') t", one=True)['t']
    execute("INSERT OR REPLACE INTO api_settings (k, v) VALUES ('cls_push_flag', ?)", (now,))
    return jsonify({'ok': True, 'flagged_at': now})


@app.route('/api/class-status/items/<int:iid>', methods=['PUT'])
@login_required
def api_class_status_item_update(iid):
    row = query('SELECT * FROM class_status_items WHERE id=?', (iid,), one=True)
    if not row:
        abort(404)
    d = request.get_json(silent=True) or {}
    fields, params = [], []
    for col in ('importance', 'remark', 'description', 'issued_date', 'due_date', 'action_taken'):
        if col in d:
            val = d[col]
            if col == 'importance' and val not in ('', 'Urgent'):
                val = 'Urgent' if val else ''
            fields.append(f'{col}=?'); params.append(val)
    if not fields:
        return jsonify({'ok': True})
    fields.append("updated_at=datetime('now','localtime')")
    params.append(iid)
    execute(f'UPDATE class_status_items SET {", ".join(fields)} WHERE id=?', tuple(params))
    return jsonify({'ok': True})


@app.route('/api/class-status/<int:cs_id>', methods=['DELETE'])
@login_required
def api_class_status_delete(cs_id):
    if not query('SELECT id FROM class_status WHERE id=?', (cs_id,), one=True):
        abort(404)
    execute('DELETE FROM class_status WHERE id=?', (cs_id,))
    return jsonify({'ok': True})


@app.route('/api/class-status/<int:cs_id>/assign', methods=['POST'])
@login_required
def api_class_status_assign(cs_id):
    """미매칭 스냅샷을 특정 선박에 수동 배정(기존 선박 스냅샷은 교체)."""
    snap = query('SELECT * FROM class_status WHERE id=?', (cs_id,), one=True)
    if not snap:
        abort(404)
    d = request.get_json(silent=True) or {}
    vessel_id = d.get('vessel_id')
    if not vessel_id or not query('SELECT id FROM vessels WHERE id=?', (vessel_id,), one=True):
        return jsonify({'ok': False, 'message': '유효한 선박을 선택하세요.'}), 400
    conn = get_db()
    # 대상 선박의 기존 스냅샷 제거 후 배정
    conn.execute('DELETE FROM class_status WHERE vessel_id=? AND id<>?', (vessel_id, cs_id))
    conn.execute("UPDATE class_status SET vessel_id=?, updated_at=datetime('now','localtime') "
                 "WHERE id=?", (vessel_id, cs_id))
    conn.commit()
    return jsonify({'ok': True})


@app.route('/api/class-status/<int:cs_id>/export')
@login_required
def api_class_status_export(cs_id):
    from flask import send_file
    snap = query('SELECT * FROM class_status WHERE id=?', (cs_id,), one=True)
    if not snap:
        abort(404)
    vname = snap['vessel_name_raw'] or ''
    if snap['vessel_id']:
        vrow = query('SELECT name FROM vessels WHERE id=?', (snap['vessel_id'],), one=True)
        if vrow:
            vname = vrow['name']
    items = query('SELECT * FROM class_status_items WHERE cs_id=? ORDER BY category, no', (cs_id,))
    cat_ko = {'COC': '선급지적(COC)', 'STATUTORY': '기국(Statutory)'}
    rows = []
    for it in items:
        rows.append([
            cat_ko.get(it['category'], it['category']),
            it['no'],
            it['issued_date'] or '',
            it['description'] or '',
            it['due_date'] or '',
            it['remark'] or '',
            it['action_taken'] or '',
            it['importance'] or '',
        ])
    headers = ['Category', 'No', 'Issued', 'Description', 'Due', '한글 요약', '조치사항', 'Urgent']
    subtitle = f"{snap['class_society'] or ''}  ·  발행 {snap['report_date'] or '-'}"
    bio = _findings_workbook(
        f'{vname} Class Status', subtitle, headers, rows,
        wrap_cols={4, 6, 7}, widths=[16, 5, 13, 60, 13, 40, 40, 8])
    safe = _re_cls.sub(r'[^A-Za-z0-9가-힣 _-]', '', vname).strip() or 'class_status'
    return send_file(bio, as_attachment=True,
                     download_name=f'{safe}_ClassStatus.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/class-status/<int:cs_id>/file')
@login_required
def api_class_status_file(cs_id):
    """선박별 보관된 최신 Class Status 원본 파일. 기본 inline(브라우저 미리보기), ?dl=1 이면 다운로드."""
    import mimetypes
    from flask import send_file
    snap = query('SELECT source_path, source_filename FROM class_status WHERE id=?', (cs_id,), one=True)
    if not snap or not snap['source_path']:
        abort(404)
    full = os.path.join(BASE_DIR, snap['source_path'])
    if not os.path.isfile(full):
        abort(404)
    dl = request.args.get('dl') == '1'
    name = snap['source_filename'] or os.path.basename(full)
    mime = mimetypes.guess_type(name)[0] or mimetypes.guess_type(full)[0] or 'application/octet-stream'
    return send_file(full, mimetype=mime, as_attachment=dl, download_name=name)


@app.route('/api/class-status/export-all')
@login_required
def api_class_status_export_all():
    """전체 선박 Class Status 엑셀 (선박별 COC/기국 지적 전부, 1시트). 감독 필터 지원."""
    from flask import send_file
    sup_id = request.args.get('supervisor_id', type=int)
    snaps = query('SELECT * FROM class_status WHERE vessel_id IS NOT NULL')
    name_by_v = {r['id']: r['name'] for r in query('SELECT id, name FROM vessels')}
    allowed = None
    if sup_id:
        allowed = {r['vessel_id'] for r in
                   query('SELECT vessel_id FROM supervisor_vessels WHERE supervisor_id=?', (sup_id,))}
    # 선박명 정렬
    snaps = sorted(snaps, key=lambda s: (name_by_v.get(s['vessel_id']) or s['vessel_name_raw'] or '').lower())
    cat_ko = {'COC': '선급지적(COC)', 'STATUTORY': '기국(Statutory)'}
    rows = []
    for s in snaps:
        if allowed is not None and s['vessel_id'] not in allowed:
            continue
        vname = name_by_v.get(s['vessel_id']) or s['vessel_name_raw'] or ''
        items = query('SELECT * FROM class_status_items WHERE cs_id=? ORDER BY category, no', (s['id'],))
        if not items:
            rows.append([vname, s['class_society'] or '', '', '', '지적 없음', '', '', '', ''])
            continue
        for it in items:
            rows.append([
                vname, s['class_society'] or '',
                cat_ko.get(it['category'], it['category']),
                it['issued_date'] or '', it['description'] or '',
                it['due_date'] or '', it['remark'] or '', it['action_taken'] or '', it['importance'] or '',
            ])
    headers = ['Vessel', 'Class', 'Category', 'Issued', 'Description', 'Due', '한글 요약', '조치사항', 'Urgent']
    today = query("SELECT date('now','localtime') d", one=True)['d']
    bio = _findings_workbook(
        '전체 선박 Class Status', f'생성 {today}', headers, rows,
        wrap_cols={5, 7, 8}, widths=[20, 7, 16, 13, 58, 13, 38, 38, 8])
    return send_file(bio, as_attachment=True,
                     download_name=f'ClassStatus_All_{today}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


UNASSIGNED_MGR = '(Unassigned)'


def _class_export_vessels(sup_id=None):
    """관리사별 추출 대상: active 선박 중 **최신 class_status에 지적(item)이 1개 이상**인 선박만
    (지적 없는 선박 자동 제외). sup_id 주면 그 담당 감독 선박으로 한정.
    반환 [{id, name, class_society, manager, items[]}]."""
    if sup_id:
        vrows = query("""SELECT v.id, v.name, v.class_society, v.manager
                           FROM vessels v
                           JOIN supervisor_vessels sv ON sv.vessel_id = v.id
                          WHERE v.active = 1 AND sv.supervisor_id = ?
                          ORDER BY v.name COLLATE NOCASE""", (sup_id,))
    else:
        vrows = query("""SELECT id, name, class_society, manager FROM vessels
                          WHERE active = 1 ORDER BY name COLLATE NOCASE""")
    out = []
    for v in vrows:
        snap = query('SELECT id FROM class_status WHERE vessel_id=? ORDER BY updated_at DESC LIMIT 1',
                     (v['id'],), one=True)
        if not snap:
            continue
        items = query('SELECT * FROM class_status_items WHERE cs_id=? ORDER BY category, no', (snap['id'],))
        if not items:
            continue   # 지적 없는 선박 제외
        out.append({'id': v['id'], 'name': v['name'],
                    'class_society': v['class_society'] or '',
                    'manager': (v['manager'] or '').strip(),
                    'items': items})
    return out


@app.route('/api/class-status/managers')
@login_required
def api_class_status_managers():
    """관리사 목록 + 선박수(지적 있는 선박만). supervisor_id 주면 그 감독 담당선박만 집계."""
    sup_id = request.args.get('supervisor_id', type=int)
    counts = {}
    for v in _class_export_vessels(sup_id):
        key = v['manager'] or UNASSIGNED_MGR
        counts[key] = counts.get(key, 0) + 1
    managers = [{'manager': k, 'vessels': n} for k, n in counts.items()]
    managers.sort(key=lambda m: (m['manager'] == UNASSIGNED_MGR, m['manager'].lower()))
    return jsonify({'managers': managers})


@app.route('/api/class-status/export-by-manager')
@login_required
def api_class_status_export_by_manager():
    """관리사 선택 → 그 관리사 선박 Class Status 지적 엑셀 일괄 추출 (영문, 지적없는선박 제외).
    supervisor_id 주면 그 담당 감독 선박만. 컬럼: Vessel/Class/Category/Issued/Description/Due/
    Management Action Plan & Progress(blank)."""
    from flask import send_file
    mgr = (request.args.get('manager') or '').strip()
    sup_id = request.args.get('supervisor_id', type=int)
    if not mgr:
        return jsonify({'error': 'manager required'}), 400
    cat_en = {'COC': 'Condition of Class (COC)', 'STATUTORY': 'Statutory (Flag)'}
    rows = []
    for v in _class_export_vessels(sup_id):
        if (v['manager'] or UNASSIGNED_MGR) != mgr:
            continue
        for it in v['items']:
            rows.append([
                v['name'], v['class_society'],
                cat_en.get(it['category'], it['category']),
                it['issued_date'] or '', it['description'] or '',
                it['due_date'] or '', '',   # Management Action Plan & Progress = blank
            ])
    headers = ['Vessel', 'Class', 'Category', 'Issued', 'Description', 'Due',
               'Management Action Plan & Progress']
    today = query("SELECT date('now','localtime') d", one=True)['d']
    safe_mgr = re.sub(r'[^\w\-]+', '_', mgr) or 'manager'
    bio = _findings_workbook(
        f'Class Status - {mgr}', f'Generated {today}', headers, rows,
        wrap_cols={5, 7}, widths=[20, 7, 20, 13, 58, 13, 40])
    return send_file(bio, as_attachment=True,
                     download_name=f'ClassStatus_{safe_mgr}_{today}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ═════════════════════════════════════════════════════════════════
#  iOS 푸시알림 (APNs) — 디바이스 등록 · 발송 · 종류별 on/off
# ═════════════════════════════════════════════════════════════════
# 발송 자체는 apns_push.py(curl --http2 + ES256 JWT). 키가 없으면 라우트는 살아 있고
# 'not_configured' 를 정직하게 반환한다 — 조용히 성공으로 위장하지 않는다.
#
# 알림 종류. 앱 설정화면이 이 목록을 그대로 그리므로 키를 바꾸면 기존 저장값이 끊긴다.
PUSH_KINDS = [
    ('dock_quote',       'Dock 견적 제출',      '업체가 견적을 올렸을 때'),
    ('dock_ordered',     'Dock 발주완료',       '발주가 전부 완료됐을 때'),
    ('dock_reject',      'Dock 결재반려',       '상신이 반려돼 단계가 되돌아갔을 때'),
    ('dock_submit_fail', 'Dock 상신 실패',      'SVMS 상신이 실패로 끝났을 때'),
    ('auto_fail',        '자동화 실패',         '러너 실패·killswitch·차단'),
    ('approval_new',     '승인 대기 신규',      '전자결재/Fund Request 신규 대기'),
    ('issue_urgent',     '긴급 현안',           '긴급으로 등록된 현안'),
    ('class_due',        '선급·증서 만료',      'D-30/D-7 요약'),
    ('calendar_daily',   '오늘 일정 요약',      '매일 10시·14시, 미완료 일정만'),
    ('test',             '테스트 알림',         '설정 확인용'),
]
PUSH_KIND_KEYS = {k for k, _l, _d in PUSH_KINDS}
_PUSH_DEVICE_CAP = 20        # 단일 요청에서 발송할 디바이스 상한(폭주 방지)


def _push_module():
    """apns_push 지연 import — 모듈이 없어도 앱 전체가 죽지 않게."""
    try:
        import apns_push
        return apns_push
    except Exception as e:
        print(f'[push] apns_push import 실패: {e}')
        return None


def _push_prefs(row):
    try:
        p = json.loads(row['prefs'] or '{}')
    except Exception:
        p = {}
    return p if isinstance(p, dict) else {}


def _push_kind_enabled(row, kind):
    """미설정 = 켜짐(기본 on). 형이 명시적으로 끈 것만 뺀다."""
    return bool(_push_prefs(row).get(kind, 1))


def _push_devices(user_ids=None, kind=None):
    """발송 대상 기기. `user_ids=None` = 전체 브로드캐스트(호출측이 의도적으로 None 을 줄 때만).

    🔴 cap 은 **prefs 필터 뒤에** 적용한다(올마이트 지적). SQL LIMIT 으로 먼저 자르면
       상위 N 대가 그 종류를 껐을 때 켜둔 기기가 영구히 미발송된다 — 조용한 미탐.
    """
    sql = "SELECT * FROM ios_device WHERE active=1"
    params = []
    if user_ids:
        sql += " AND user_id IN (%s)" % ','.join('?' * len(user_ids))
        params += list(user_ids)
    sql += " ORDER BY updated_at DESC"
    rows = query(sql, tuple(params))
    if kind:
        rows = [r for r in rows if _push_kind_enabled(r, kind)]
    return rows[:_PUSH_DEVICE_CAP]


def _push_dispatch(kind, event_key, title, body, link=None,
                   user_ids=None, collapse_id=None):
    """알림 1건 발송. 반환 dict(ok/sent/failed/reason).

    🔴 2단 커밋(BV 감시 교훈): event_key 를 먼저 claim 해 중복을 막되, **디바이스가 있었는데
       전부 발송 실패**면 claim 을 풀어 다음 폴링에 재시도되게 한다. 안 풀면 일시적 네트워크
       오류 1회가 그 이벤트를 영구 미탐으로 만든다.
       디바이스가 0대면 claim 을 유지한다(배달할 곳이 없는 과거 이벤트가 나중에 폭주하는 것 방지).
    """
    if kind not in PUSH_KIND_KEYS:
        return {'ok': False, 'reason': 'unknown_kind', 'sent': 0, 'failed': 0}
    ap = _push_module()
    if ap is None:
        return {'ok': False, 'reason': 'module_missing', 'sent': 0, 'failed': 0}
    try:
        conf = ap.load_conf()
    except Exception as e:
        return {'ok': False, 'reason': 'not_configured', 'detail': str(e),
                'sent': 0, 'failed': 0}

    try:                                     # ── claim (중복발송 차단)
        execute("INSERT INTO push_log (event_key, kind, title, body, link) "
                "VALUES (?,?,?,?,?)", (event_key, kind, title, body, link))
    except sqlite3.IntegrityError:
        return {'ok': True, 'dup': True, 'reason': 'already_sent',
                'sent': 0, 'failed': 0}

    targets = _push_devices(user_ids, kind=kind)
    sent, failed, detail = 0, 0, []
    for d in targets:
        payload = ap.alert_payload(title, body, link=link, kind=kind)
        ok, st, reason = ap.send(d['token'], payload, env=(d['env'] or 'production'),
                                 conf=conf, collapse_id=collapse_id)
        if ok:
            sent += 1
            execute("UPDATE ios_device SET last_push_at=datetime('now','localtime') "
                    "WHERE id=?", (d['id'],))
        else:
            failed += 1
            detail.append(f"dev{d['id']}:{st}:{reason}")
            if ap.is_dead(st, reason):       # 영구 사망 사유만 비활성(일시실패로는 안 끔)
                execute("UPDATE ios_device SET active=0, dead_reason=?, "
                        "updated_at=datetime('now','localtime') WHERE id=?",
                        (f'{st}:{reason}', d['id']))

    if targets and sent == 0:                # 전부 실패 → claim 해제(재시도 가능하게)
        execute("DELETE FROM push_log WHERE event_key=? AND sent_n=0", (event_key,))
        return {'ok': False, 'reason': 'all_failed', 'sent': 0,
                'failed': failed, 'detail': '; '.join(detail)[:500]}

    execute("UPDATE push_log SET sent_n=?, fail_n=?, detail=? WHERE event_key=?",
            (sent, failed, '; '.join(detail)[:500] or None, event_key))
    return {'ok': True, 'sent': sent, 'failed': failed,
            'devices': len(targets), 'detail': '; '.join(detail)[:500]}


_PUSH_OUTBOX_MAX_TRIES = 6      # 15분 폴 기준 약 1.5시간 재시도
_PUSH_OUTBOX_MAX_AGE_H = 24     # 그보다 늙은 알림은 형에게 가치가 없다(늦은 알림 = 오보에 가깝다)


def _push_outbox_add(kind, event_key, title, body, link=None, collapse_id=None):
    """발송 예정 이벤트를 durable 하게 적재. 이미 있으면(=같은 전이 재관측) 조용히 통과.

    🔴 반드시 **상태 UPDATE 보다 먼저** 부른다 — 순서가 뒤집히면 그 사이에 죽었을 때
       "행은 갱신됐는데 알림은 없는" 영구 미탐이 남는다(schema.sql push_outbox 주석 참조).
    """
    execute("INSERT OR IGNORE INTO push_outbox (event_key, kind, title, body, link, collapse_id) "
            "VALUES (?,?,?,?,?,?)",
            (event_key, kind, title, body, link, collapse_id))


def _push_outbox_drain(limit=20):
    """대기함 발송. 성공·중복이면 삭제, 실패면 tries+1 하고 남긴다. 결과 목록 반환.

    ⚠️`sent=0` 은 실패가 아니다 — 등록된 기기가 0대면 보낼 곳이 없을 뿐이고 `_push_dispatch` 가
      claim 을 유지하므로 여기서도 지운다(안 지우면 기기 등록 순간 과거 알림이 폭주한다).
    """
    rows = query("SELECT * FROM push_outbox ORDER BY created_at LIMIT ?", (limit,))
    out = []
    for r in rows:
        key = r['event_key']
        old = query("SELECT CAST((julianday('now','localtime') - julianday(?)) * 24 AS REAL) h",
                    (r['created_at'],), one=True)
        if r['tries'] >= _PUSH_OUTBOX_MAX_TRIES or (old and (old['h'] or 0) > _PUSH_OUTBOX_MAX_AGE_H):
            execute("DELETE FROM push_outbox WHERE event_key=?", (key,))
            app.logger.warning('push outbox 포기 key=%s tries=%s', key, r['tries'])
            out.append({'kind': r['kind'], 'key': key, 'ok': False, 'sent': 0, 'reason': 'dropped'})
            continue
        try:
            res = _push_dispatch(r['kind'], key, r['title'] or '', r['body'] or '',
                                 link=r['link'], collapse_id=r['collapse_id'])
        except Exception as e:
            app.logger.exception('push outbox 발송 예외 key=%s', key)
            res = {'ok': False, 'reason': 'exception', 'detail': str(e)[:200]}
        if res.get('ok'):
            execute("DELETE FROM push_outbox WHERE event_key=?", (key,))
        else:
            execute("UPDATE push_outbox SET tries=tries+1, last_error=? WHERE event_key=?",
                    ((res.get('reason') or '')[:200], key))
        out.append({'kind': r['kind'], 'key': key, 'ok': bool(res.get('ok')),
                    'sent': res.get('sent', 0), 'reason': res.get('reason')})
    return out


# ---- 앱(Bearer) : 디바이스 등록/해제 ----
@app.route('/api/ios/device', methods=['POST'])
@login_required
def api_ios_device_register():
    d = request.get_json(silent=True) or {}
    tok = (d.get('token') or '').strip()
    # APNs 토큰은 hex 문자열(보통 64자, 향후 확장 여지로 상한만 둔다)
    if not tok or not re.fullmatch(r'[0-9a-fA-F]{40,200}', tok):
        return jsonify({'error': 'bad_token'}), 400
    env = (d.get('env') or 'production').strip().lower()
    if env not in ('production', 'sandbox'):
        env = 'production'
    # 🔴 신규 기기(재설치로 토큰이 바뀐 경우 포함)는 같은 계정의 기존 prefs 를 물려받는다(올마이트 지적).
    #    안 물려받으면 형이 껐던 종류가 재설치만으로 조용히 다시 켜진다 = 형 의도를 뒤집는 동작.
    prior = query("SELECT prefs FROM ios_device WHERE user_id=? AND prefs IS NOT NULL "
                  "ORDER BY updated_at DESC LIMIT 1", (session['user_id'],))
    inherit = prior[0]['prefs'] if prior else None
    execute("""
        INSERT INTO ios_device (token, user_id, env, app_ver, device_name, active,
                                dead_reason, prefs, updated_at)
        VALUES (?,?,?,?,?,1,NULL,?,datetime('now','localtime'))
        ON CONFLICT(token) DO UPDATE SET
            user_id=excluded.user_id, env=excluded.env, app_ver=excluded.app_ver,
            device_name=excluded.device_name, active=1, dead_reason=NULL,
            -- 소유자가 바뀌면 이전 사용자의 설정을 남기지 않는다(계정 전환 잔상 차단).
            prefs=CASE WHEN ios_device.user_id = excluded.user_id
                       THEN COALESCE(ios_device.prefs, excluded.prefs)
                       ELSE excluded.prefs END,
            updated_at=datetime('now','localtime')
    """, (tok.lower(), session['user_id'], env,
          (d.get('app_ver') or '')[:32] or None, (d.get('device_name') or '')[:64] or None,
          inherit))
    return jsonify({'ok': True})


@app.route('/api/ios/device', methods=['DELETE'])
@login_required
def api_ios_device_unregister():
    d = request.get_json(silent=True) or {}
    tok = (d.get('token') or '').strip().lower()
    if not tok:
        return jsonify({'error': 'bad_token'}), 400
    execute("UPDATE ios_device SET active=0, dead_reason='unregistered_by_app', "
            "updated_at=datetime('now','localtime') WHERE token=? AND user_id=?",
            (tok, session['user_id']))
    return jsonify({'ok': True})


# ---- 앱(Bearer) : 알림 종류 on/off ----
@app.route('/api/ios/notify-prefs', methods=['GET'])
@login_required
def api_ios_prefs_get():
    rows = query("SELECT * FROM ios_device WHERE user_id=? AND active=1 "
                 "ORDER BY updated_at DESC", (session['user_id'],))
    prefs = _push_prefs(rows[0]) if rows else {}
    ap = _push_module()
    return jsonify({
        'ok': True,
        'kinds': [{'key': k, 'label': l, 'desc': dsc,
                   'enabled': bool(prefs.get(k, 1))} for k, l, dsc in PUSH_KINDS],
        'devices': len(rows),
        'configured': bool(ap and ap.configured()),
    })


@app.route('/api/ios/notify-prefs', methods=['PUT'])
@login_required
def api_ios_prefs_put():
    d = request.get_json(silent=True) or {}
    incoming = d.get('prefs')
    if not isinstance(incoming, dict):
        return jsonify({'error': 'bad_request'}), 400
    clean = {k: (1 if incoming.get(k) else 0) for k in incoming if k in PUSH_KIND_KEYS}
    # 형 폰이 여러 대여도 설정은 계정 단위로 같게 유지(기기별로 달라 헷갈리는 일 방지)
    execute("UPDATE ios_device SET prefs=?, updated_at=datetime('now','localtime') "
            "WHERE user_id=?", (json.dumps(clean, ensure_ascii=False), session['user_id']))
    return jsonify({'ok': True, 'prefs': clean})


# ---- 앱(Bearer) : 상태 확인 + 테스트 발송 ----
@app.route('/api/ios/push/status', methods=['GET'])
@login_required
def api_ios_push_status():
    ap = _push_module()
    rows = query("SELECT id, env, app_ver, device_name, active, dead_reason, "
                 "last_push_at, updated_at FROM ios_device WHERE user_id=? "
                 "ORDER BY updated_at DESC LIMIT 10", (session['user_id'],))
    # push_log 는 계정 컬럼이 없어(이벤트 단위 dedup 용) user 로 좁힐 수 없다.
    # 🔴 그래서 admin 이 아니면 title 을 지운다(올마이트 지적) — 다른 감독의 알림 내용이 새면 안 된다.
    #    진단에 필요한 kind/성패 카운트는 그대로 남긴다.
    is_admin = session.get('role') == 'admin'
    last = query("SELECT kind, title, sent_n, fail_n, created_at FROM push_log "
                 "WHERE hidden_at IS NULL ORDER BY id DESC LIMIT 5")
    recent = []
    for r in last:
        item = dict(r)
        if not is_admin:
            item['title'] = None
        recent.append(item)
    return jsonify({
        'ok': True,
        'configured': bool(ap and ap.configured()),
        'devices': [dict(r) for r in rows],
        'recent': recent,
        # 발송기록은 계정 컬럼이 없는 전사 공용 자원이라 지우면 남의 화면에서도 사라진다 →
        # admin 만. 화면은 이 값만 보고 버튼을 그린다(자체판정 금지).
        'can_clear': is_admin,
    })


@app.route('/api/ios/push/log/clear', methods=['POST'])
@login_required
def api_ios_push_log_clear():
    """최근 발송 기록 감추기.

    🔴 **행을 지우지 않는다.** `push_log.event_key` 는 화면 이력이자 **중복발송 차단 claim** 이다.
       하드 DELETE 하면 지운 직후 같은 이벤트가 다시 나간다(캘린더 슬롯 재발송, outbox 재시도가
       이미 보낸 알림을 다시 발송). 그래서 `hidden_at` 만 찍어 화면에서만 감춘다.
    """
    if session.get('role') != 'admin':
        return jsonify({'error': 'forbidden',
                        'message': '발송기록은 공용 자원이라 admin 만 정리할 수 있음'}), 403
    # 🔴 COUNT 후 UPDATE 로 나누지 않는다(올마이트 지적) — 그 사이에 들어온 알림까지 같이
    #    감춰지면서 건수만 틀린다("2건 감춤"인데 방금 온 알림도 사라짐 = 미탐으로 보임).
    #    한 문장으로 감추고 실제 영향 행수를 그대로 돌려준다.
    n = execute_rc("UPDATE push_log SET hidden_at=datetime('now','localtime') "
                   "WHERE hidden_at IS NULL")
    return jsonify({'ok': True, 'hidden': n})


@app.route('/api/ios/push/test', methods=['POST'])
@login_required
def api_ios_push_test():
    """형이 앱에서 눌러 실제로 알림이 뜨는지 확인 — 카나리 검증 수단.
    🔴 event_key 에 난수를 붙인다(올마이트 지적): 초 단위 시각만 쓰면 같은 초에 두 번 누른 두 번째가
       dup 으로 흘러 "발송 0건"으로 보이고, 형은 그걸 실패로 읽는다."""
    stamp = datetime.now().strftime('%Y%m%d%H%M%S') + '-' + uuid.uuid4().hex[:8]
    res = _push_dispatch(
        'test', f"test:{session['user_id']}:{stamp}",
        'TRMT 알림 테스트',
        f"푸시 경로 정상 — {datetime.now().strftime('%m/%d %H:%M:%S')}",
        link='trmt://automation', user_ids=[session['user_id']])
    return jsonify(res), (200 if res.get('ok') else 502)


# ---- 자동화(X-API-Key) : 이벤트 → 알림 ----
@app.route('/api/ext/push', methods=['POST'])
@api_key_required
def api_ext_push():
    """맥 워커/자동화가 상태변화를 알릴 때 호출.
    event_key 는 호출측이 '이 이벤트 1회'를 규정하는 안정키여야 한다
    (예: dock_ordered:R1:ATBG). 같은 키로 두 번 오면 두 번째는 dup 으로 조용히 흘린다."""
    d = request.get_json(silent=True) or {}
    kind = (d.get('kind') or '').strip()
    ekey = (d.get('event_key') or '').strip()
    title = (d.get('title') or '').strip()
    body = (d.get('body') or '').strip()
    if kind not in PUSH_KIND_KEYS or not ekey or not title:
        return jsonify({'error': 'bad_request',
                        'message': 'kind(등록된 값)/event_key/title 필수',
                        'kinds': sorted(PUSH_KIND_KEYS)}), 400
    # 🔴 user_ids 는 3상태다(올마이트 지적): 키 없음/null = 전체 브로드캐스트, 정상 리스트 = 그 사용자,
    #    **빈 리스트 = 대상 없음** — 이걸 None 으로 접으면 "아무에게도"가 "전원에게"로 뒤집힌다.
    #    숫자로 못 바꾸는 값이 섞이면 500 대신 400 으로 명확히 거절한다.
    uids = d.get('user_ids')
    if uids is None:
        pass
    elif not isinstance(uids, list):
        return jsonify({'error': 'bad_request', 'message': 'user_ids 는 배열이어야 함'}), 400
    elif not uids:
        return jsonify({'error': 'bad_request',
                        'message': 'user_ids 가 빈 배열 — 대상이 없으면 호출하지 말 것'}), 400
    else:
        try:
            uids = [int(x) for x in uids]
        except (TypeError, ValueError):
            return jsonify({'error': 'bad_request',
                            'message': 'user_ids 에 정수 아닌 값이 있음'}), 400
    res = _push_dispatch(kind, ekey[:200], title[:120], body[:300],
                         link=(d.get('link') or None),
                         user_ids=uids, collapse_id=d.get('collapse_id'))
    return jsonify(res), (200 if res.get('ok') else 502)


# ---- 자동화(X-API-Key) : 오늘 일정 요약(하루 2회) ----
# 형 지시(2026-08-06): "하루에 2번(10시·14시) 캘린더 일정 푸시알람(완료 제외)".
# 14시 판은 그 사이에 체크한 일정이 빠지므로 자연스럽게 '아직 안 한 것' 리마인더가 된다.
CAL_PUSH_SLOTS = {'10': ('오늘 일정', 10), '14': ('오늘 남은 일정', 14)}
_CAL_PUSH_MAX_ITEMS = 6
_CAL_PUSH_LATE_H = 3       # 슬롯 + 이 시간 이상 늦은 실행은 버린다(늦은 알림 = 오보에 가깝다)
_CAL_DAY_RE = re.compile(r'^\d{4}-\d{2}-\d{2}$')


def _cal_slot_window_ok(slot, now_hour):
    """이 실행이 해당 슬롯의 유효 시간대(hour ~ hour+3, 끝 제외) 안인가.

    🔴 **이른 실행도 막는다**(올마이트 지적). launchd `StartCalendarInterval` 은 놓친 잡을 깨어날 때
       즉시 돌린다 — 맥이 꺼져 있다 09시에 부팅하면 10시판이 09시에 나가고, 그 event_key 가
       오늘치를 선점해 **진짜 10시 실행이 dedup 으로 묻힌다**(조기발송 + 미탐 동시 발생).
    """
    h = CAL_PUSH_SLOTS[slot][1]
    return h <= now_hour < h + _CAL_PUSH_LATE_H


def _calendar_daily_items(sup_id, day):
    """`day` 에 걸치는 **미완료** 일정. 멀티데이는 기간에 걸치면 매일 포함(입거 같은 진행형).

    스코프는 캘린더 화면과 같은 규칙(해당 감독 + 공용 supervisor_id IS NULL)이다 —
    화면과 알림이 다르면 형이 둘 다 못 믿는다.
    🔴 `sup_id` 가 없는 계정은 **공용만** 본다. 화면은 전체를 보여주지만, 푸시는 능동 발송이라
       남의 감독 개인일정을 폰으로 밀어내지 않는다(fail-closed).
    """
    sql = ("SELECT title, all_day, start_time FROM calendar_events "
           "WHERE completed=0 AND start_date<=? AND COALESCE(end_date, start_date)>=?")
    params = [day, day]
    if sup_id is None:
        sql += " AND supervisor_id IS NULL"
    else:
        sql += " AND (supervisor_id=? OR supervisor_id IS NULL)"
        params.append(sup_id)
    sql += " ORDER BY all_day, COALESCE(start_time,'99:99'), id"   # 시각 있는 일정 먼저
    return query(sql, tuple(params))


def _calendar_daily_body(rows, limit=300):
    """본문 조립. 🔴 길이 초과는 뒤를 자르는 게 아니라 **항목을 줄인다**(올마이트 지적) —
    그냥 자르면 '외 N건' 꼬리가 날아가 형이 나머지가 있는 줄도 모른다."""
    def line(r):
        when = (r['start_time'] or '')[:5] if (not r['all_day'] and r['start_time']) else '종일'
        t = (r['title'] or '일정').strip()
        return '%s %s' % (when, t if len(t) <= 60 else t[:59] + '…')

    n = min(len(rows), _CAL_PUSH_MAX_ITEMS)
    while n >= 0:
        parts = [line(r) for r in rows[:n]]
        rest = len(rows) - n
        if rest > 0:
            parts.append('외 %d건' % rest)
        body = ' · '.join(parts)
        if len(body) <= limit or n == 0:
            return body[:limit]
        n -= 1


@app.route('/api/ext/push/calendar-daily', methods=['POST'])
@api_key_required
def api_ext_push_calendar_daily():
    """맥 launchd 가 10시·14시에 호출. 일정 계산은 전부 서버가 한다(러너는 트리거일 뿐).

    body: {"slot":"10"|"14", "dry":0|1, "date":"YYYY-MM-DD"(dry 전용)}
    · 0건이면 **보내지 않는다**(`skipped_empty`) — 빈 알림이 매일 2번 오면 형이 알림을 끈다.
    · 대기함(push_outbox)은 쓰지 않는다. 그 경로엔 user_ids 가 없어 재발송이 전체 브로드캐스트가
      된다(개인일정 유출). 실패는 claim 이 풀리므로 러너가 재시도한다.
    """
    d = request.get_json(silent=True) or {}
    slot = str(d.get('slot') or '').strip()
    if slot not in CAL_PUSH_SLOTS:
        return jsonify({'error': 'bad_request', 'message': 'slot 은 10 또는 14',
                        'slots': sorted(CAL_PUSH_SLOTS)}), 400
    # 🔴 dry 는 문자열 "0"/"false" 를 참으로 보면 안 된다(올마이트 지적) — 실발송 의도가
    #    조용히 dry 로 접히면 그 슬롯은 통째로 미탐이다.
    _dry = d.get('dry')
    dry = bool(_dry) and str(_dry).strip().lower() not in ('0', 'false', 'no', '')
    now = datetime.now()
    today = now.strftime('%Y-%m-%d')
    day = (str(d.get('date') or '').strip() or today)
    if not _CAL_DAY_RE.match(day):
        return jsonify({'error': 'bad_request', 'message': 'date 는 YYYY-MM-DD'}), 400
    if day != today and not dry:
        # 과거/미래 날짜 실발송은 막는다 — 백필 한 번이 알림 폭주가 된다.
        return jsonify({'error': 'bad_request',
                        'message': '오늘이 아닌 날짜는 dry 에서만 허용'}), 400

    label, _hour = CAL_PUSH_SLOTS[slot]
    if not dry and not _cal_slot_window_ok(slot, now.hour):
        return jsonify({'ok': True, 'skipped': 'out_of_window', 'slot': slot, 'day': day,
                        'now_hour': now.hour}), 200

    rows = query("SELECT DISTINCT d.user_id, u.supervisor_id FROM ios_device d "
                 "JOIN users u ON u.id=d.user_id "
                 "WHERE d.active=1 AND u.active=1")
    results = []
    for r in rows:
        uid = r['user_id']
        items = _calendar_daily_items(r['supervisor_id'], day)
        if not items:
            results.append({'user_id': uid, 'n': 0, 'ok': True, 'reason': 'skipped_empty'})
            continue
        title = '%s %d건' % (label, len(items))
        body = _calendar_daily_body(items)
        ekey = 'calendar_daily:%s:%s:%s' % (uid, day, slot)
        if dry:
            results.append({'user_id': uid, 'n': len(items), 'ok': True, 'dry': True,
                            'event_key': ekey, 'title': title, 'body': body})
            continue
        res = _push_dispatch('calendar_daily', ekey, title, body,
                             link='trmt://calendar', user_ids=[uid],
                             collapse_id='cal-%s-%s' % (day, slot))
        results.append({'user_id': uid, 'n': len(items), 'ok': bool(res.get('ok')),
                        'sent': res.get('sent', 0), 'reason': res.get('reason')})

    failed = [x for x in results if not x['ok']]
    return jsonify({'ok': not failed, 'day': day, 'slot': slot, 'dry': dry,
                    'users': len(rows), 'results': results}), (200 if not failed else 502)


# ═════════════════════════════════════════════════════════════════
#  CLI entry
# ═════════════════════════════════════════════════════════════════
