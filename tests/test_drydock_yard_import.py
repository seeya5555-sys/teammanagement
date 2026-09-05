import io
import sqlite3
import unittest

from openpyxl import Workbook

import drydock_integration as integration


def workbook_bytes():
    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    cover.append(["Final Discount :", 0.10])
    quote = wb.create_sheet("Quotation")
    quote.append(["Item No.", None, "Work Description", "Qty", "Net Total"])
    quote.append([2, None, "GENERAL SERVICE", None, 0])
    quote.append([2.1, None, "Circulating Water", None, 0])
    quote.append([None, None, "Connect", 1, 100])
    quote.append(["2.1.1", None, "Supply", 2, 25])
    quote.append([2.2, None, "Ballast Water", 1, 10])
    quote.append(["2.2.1", None, "Please provide quotation for temporary staging", None, 0])
    quote.append(["2.2.2", None, "Following tank cleaning to be quoted.", None, 0])
    quote.append(["2.2.3", None, "For quotation only", None, 0])
    quote.append([3, None, "DOCKING", None, 0])
    quote.append([24, None, "Fire Wire Reel", 1, 8])
    quote.append([None, None, "Normal Total Price/USD", None, 143])
    quote.append([None, None, "Final discount", None, 0.10])
    quote.append([None, None, "Total Price after dicount/USD Net", None, 128.70])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def make_db():
    db = sqlite3.connect(":memory:")
    db.row_factory = sqlite3.Row
    db.executescript("""
        CREATE TABLE vessels (id TEXT PRIMARY KEY, dc_rate REAL DEFAULT 0);
        CREATE TABLE jobs (
            id INTEGER PRIMARY KEY AUTOINCREMENT, vessel_id TEXT, number TEXT,
            section TEXT, category TEXT, description TEXT, vendor TEXT,
            budget REAL, consumption REAL, start_date TEXT, end_date TEXT,
            completion REAL, remarks TEXT, updated_at TEXT
        );
        INSERT INTO vessels(id,dc_rate) VALUES('v_test',0);
    """)
    return db


def make_queue_db():
    db = sqlite3.connect(":memory:")
    db.row_factory = sqlite3.Row
    db.executescript("""
        CREATE TABLE trmt_svms_job_import (
            token TEXT PRIMARY KEY, vessel_id TEXT NOT NULL, vsl_cd TEXT NOT NULL,
            dk_cd TEXT NOT NULL, status TEXT NOT NULL, payload_json TEXT, error TEXT,
            requested_at TEXT NOT NULL, completed_at TEXT, applied_at TEXT
        );
    """)
    return db


class DockYardImportTests(unittest.TestCase):
    def test_svms_request_coalesces_pending_but_never_reuses_ready_preview(self):
        db = make_queue_db()
        db.execute("""INSERT INTO trmt_svms_job_import
            (token,vessel_id,vsl_cd,dk_cd,status,requested_at,completed_at)
            VALUES('old-ready','v_test','V1','D1','ready',datetime('now'),datetime('now'))""")
        reusable = integration._pending_svms_import_request(db, 'v_test')
        self.assertIsNone(reusable)

        db.execute("""INSERT INTO trmt_svms_job_import
            (token,vessel_id,vsl_cd,dk_cd,status,requested_at)
            VALUES('in-flight','v_test','V1','D1','pending',datetime('now'))""")
        reusable = integration._pending_svms_import_request(db, 'v_test')
        self.assertEqual(('in-flight', 'pending'), (reusable['token'], reusable['status']))

    def test_parser_matches_quote_total_and_kuwait_rollup_shape(self):
        parsed = integration._parse_yard_job_workbook(workbook_bytes())
        jobs = {job["number"]: job for job in parsed["jobs"]}
        self.assertEqual(6, parsed["job_count"])
        self.assertEqual(143.0, parsed["gross_total"])
        self.assertEqual(10.0, parsed["discount_rate"])
        self.assertEqual(128.7, parsed["after_discount"])
        self.assertEqual(125.0, jobs["2.1"]["budget"])
        self.assertEqual(0.0, jobs["2.1.1"]["budget"])
        self.assertEqual(10.0, jobs["2.2"]["budget"])
        self.assertEqual(8.0, jobs["24"]["budget"])
        self.assertEqual("GENERAL", jobs["2.1"]["section"])
        self.assertEqual("DECK", jobs["24"]["section"])
        self.assertEqual([], parsed["warnings"])
        self.assertNotIn("2.2.1", jobs)
        self.assertNotIn("2.2.2", jobs)
        self.assertNotIn("2.2.3", jobs)

    def test_quote_request_filter_is_narrow(self):
        self.assertTrue(integration._yard_is_quote_request(
            "Please provide quotation for removal of staging"))
        self.assertTrue(integration._yard_is_quote_request(
            "Following tank cleaning to be quoted."))
        self.assertFalse(integration._yard_is_quote_request(
            "Cargo and Slop Tank Cleaning (Bottom Plate and Bulkhead)"))
        self.assertFalse(integration._yard_is_quote_request(
            "Prepare quotation comparison and report"))
        self.assertFalse(integration._yard_is_quote_request(
            "Work to be quoted after inspection and then completed"))

    def test_priced_request_wording_is_not_silently_dropped(self):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(workbook_bytes()))
        ws = wb["Quotation"]
        ws.insert_rows(9)
        ws.cell(9, 1).value = "2.3"
        ws.cell(9, 3).value = "Please provide quotation for confirmed repair"
        ws.cell(9, 5).value = 33
        out = io.BytesIO(); wb.save(out)
        parsed = integration._parse_yard_job_workbook(out.getvalue())
        jobs = {item["number"]: item for item in parsed["jobs"]}
        self.assertIn("2.3", jobs)
        self.assertEqual(33, jobs["2.3"]["budget"])

    def test_real_child_survives_filtered_quote_request_parent(self):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(workbook_bytes()))
        ws = wb["Quotation"]
        ws.insert_rows(9, 2)
        ws.cell(9, 1).value = "2.3"
        ws.cell(9, 3).value = "Please provide quotation for optional work"
        ws.cell(10, 1).value = "2.3.1"
        ws.cell(10, 3).value = "Confirmed steel repair"
        ws.cell(10, 5).value = 44
        out = io.BytesIO(); wb.save(out)
        parsed = integration._parse_yard_job_workbook(out.getvalue())
        jobs = {item["number"]: item for item in parsed["jobs"]}
        self.assertNotIn("2.3", jobs)
        self.assertEqual(44, jobs["2.3.1"]["budget"])

    def test_apply_preserves_live_progress_and_manual_classification(self):
        db = make_db()
        db.execute("""INSERT INTO jobs(vessel_id,number,section,category,description,vendor,budget,
            consumption,start_date,end_date,completion,remarks)
            VALUES('v_test','2.1','CANCEL','Shipyard','Owner wording','Cancelled',1,55,
                   '2026-01-01','2026-01-02',80,'[{"date":"2026-01-01"}]')""")
        db.execute("""INSERT INTO jobs(vessel_id,number,section,category,description,vendor,budget,
            consumption,completion,remarks) VALUES('v_test','2.2','GENERAL','Crew','Manual Crew','',7,0,0,'[]')""")
        parsed = integration._parse_yard_job_workbook(workbook_bytes())
        result = integration._apply_yard_job_import(db, "v_test", parsed)
        db.commit()
        kept = db.execute("SELECT * FROM jobs WHERE vessel_id='v_test' AND number='2.1'").fetchone()
        crew = db.execute("SELECT * FROM jobs WHERE vessel_id='v_test' AND number='2.2'").fetchone()
        self.assertEqual("CANCEL", kept["section"])
        self.assertEqual("Owner wording", kept["description"])
        self.assertEqual(125.0, kept["budget"])
        self.assertEqual(55.0, kept["consumption"])
        self.assertEqual(80.0, kept["completion"])
        self.assertEqual(7.0, crew["budget"])
        self.assertEqual(1, result["preserved_manual"])
        self.assertEqual(10.0, db.execute("SELECT dc_rate FROM vessels WHERE id='v_test'").fetchone()[0])

    def test_svms_dock_rows_classify_subject_vendor_and_svms_usd(self):
        parsed = integration._normalize_svms_dock_jobs({
            "P_RS_DP": [{"SUBJ": "[P1] DOCK PAINT", "VNDR_NM": "Paint Co",
                          "AMT": 100, "AMT_USD": 100, "CUR_CD": "USD"}],
            "P_RS_SS": [
                {"SUBJ": "[S1-A] M/E SPARE", "VNDR_NM": "Maker",
                 "AMT": 1350000, "AMT_USD": 1000, "CUR_CD": "KRW"},
                {"SUBJ": "[ST2] CABIN STORE", "AMT": 200, "AMT_USD": 200, "CUR_CD": "USD"},
                {"SUBJ": "[S3] LOCAL SUPPLY - STORE", "AMT": 16000,
                 "AMT_USD": 100, "CUR_CD": "JPY"},
                {"SUBJ": "[X4] UNCLASSIFIED ITEM", "R_AMT": 50,
                 "R_AMT_USD": 50, "R_CUR_CD": "USD"},
            ],
            "P_RS_SR": [{"SUBJ": "[R1] M/E OVHL", "VNDR_NM": "Repair Co",
                          "AMT": 45000, "AMT_USD": 45000, "CUR_CD": "USD"}],
        })
        jobs = {job["number"]: job for job in parsed["jobs"]}
        self.assertEqual(("Paint", "PAINT", "DOCK PAINT"),
                         (jobs["P1"]["category"], jobs["P1"]["section"], jobs["P1"]["description"]))
        self.assertEqual("Spare", jobs["S1-A"]["category"])
        self.assertEqual(1000, jobs["S1-A"]["budget"])
        self.assertEqual("Maker", jobs["S1-A"]["vendor"])
        self.assertEqual("Store", jobs["ST2"]["category"])
        self.assertEqual("Store", jobs["S3"]["category"])
        self.assertEqual("Spare", jobs["X4"]["category"])
        self.assertEqual("Shore Repair", jobs["R1"]["category"])
        self.assertEqual({"Paint": 1, "Spare": 2, "Store": 2, "Shore Repair": 1}, parsed["counts"])

    def test_svms_apply_is_idempotent_and_preserves_live_progress(self):
        db = make_db()
        db.execute("""INSERT INTO jobs(vessel_id,number,section,category,description,vendor,budget,
            consumption,start_date,end_date,completion,remarks)
            VALUES('v_test','S1','SPARE','Spare','Old title','Old vendor',1,55,
                   '2026-01-01','2026-01-02',80,'[{"date":"2026-01-01"}]')""")
        parsed = integration._normalize_svms_dock_jobs({
            "P_RS_DP": [],
            "P_RS_SS": [{"SUBJ": "[S1] New title", "VNDR_NM": "New vendor",
                          "AMT": 1350000, "AMT_USD": 1000, "CUR_CD": "KRW"}],
            "P_RS_SR": [],
        })
        first = integration._apply_svms_job_import(db, "v_test", parsed)
        second = integration._apply_svms_job_import(db, "v_test", parsed)
        row = db.execute("SELECT * FROM jobs WHERE vessel_id='v_test' AND number='S1'").fetchone()
        self.assertEqual((0, 1, 0), (first["inserted"], first["updated"], first["unchanged"]))
        self.assertEqual((0, 0, 1), (second["inserted"], second["updated"], second["unchanged"]))
        self.assertEqual(("Old title", "Old vendor", 1000),
                         (row["description"], row["vendor"], row["budget"]))
        self.assertEqual((55, 80, '2026-01-01', '[{"date":"2026-01-01"}]'),
                         (row["consumption"], row["completion"], row["start_date"], row["remarks"]))

    def test_svms_reimport_matches_number_across_category_drift_and_updates_budget_only(self):
        db = make_db()
        db.execute("""INSERT INTO jobs(vessel_id,number,section,category,description,vendor,budget,
            consumption,start_date,end_date,completion,remarks)
            VALUES('v_test','S9','SPARE','Spare','Manual title','Manual vendor',10,55,
                   '2026-01-01','2026-01-02',80,'[{"date":"2026-01-01"}]')""")
        parsed = {"jobs": [{"number": "S9", "section": "STORE", "category": "Store",
                             "description": "SVMS changed title", "vendor": "SVMS vendor",
                             "budget": 20}],
                  "job_count": 1, "total_budget": 20}
        result = integration._apply_svms_job_import(db, "v_test", parsed)
        rows = db.execute("SELECT * FROM jobs WHERE vessel_id='v_test'").fetchall()
        self.assertEqual((0, 1, 0), (result["inserted"], result["updated"], result["unchanged"]))
        self.assertEqual(1, len(rows))
        self.assertEqual(("SPARE", "Spare", "Manual title", "Manual vendor", 20),
                         (rows[0]["section"], rows[0]["category"], rows[0]["description"],
                          rows[0]["vendor"], rows[0]["budget"]))

    def test_svms_preview_deduplicates_same_number_even_if_category_drifts(self):
        parsed = integration._normalize_svms_dock_jobs({
            "P_RS_DP": [],
            "P_RS_SS": [
                {"SUBJ": "[S9] FIRST", "AMT_USD": 10, "CUR_CD": "USD"},
                {"SUBJ": "[S9] LOCAL SUPPLY - STORE", "AMT_USD": 20, "CUR_CD": "USD"},
            ],
            "P_RS_SR": [],
        })
        self.assertEqual(1, parsed["job_count"])
        self.assertEqual(("S9", "Spare", 10),
                         (parsed["jobs"][0]["number"], parsed["jobs"][0]["category"],
                          parsed["jobs"][0]["budget"]))

    def test_svms_noncanonical_number_does_not_cross_category_boundary(self):
        parsed = integration._normalize_svms_dock_jobs({
            "P_RS_DP": [],
            "P_RS_SS": [{"SUBJ": "[X4] SPARE ITEM", "AMT_USD": 10, "CUR_CD": "USD"}],
            "P_RS_SR": [{"SUBJ": "[X4] REPAIR ITEM", "AMT_USD": 20, "CUR_CD": "USD"}],
        })
        self.assertEqual(2, parsed["job_count"])
        self.assertEqual(["Spare", "Shore Repair"], [job["category"] for job in parsed["jobs"]])

    def test_svms_numberless_reimport_updates_budget_only_when_title_is_stable(self):
        db = make_db()
        db.execute("""INSERT INTO jobs(vessel_id,number,section,category,description,vendor,budget,
            consumption,start_date,end_date,completion,remarks)
            VALUES('v_test','','SPARE','Spare','NO NUMBER','Manual vendor',10,55,
                   '2026-01-01','2026-01-02',80,'[]')""")
        parsed = {"jobs": [{"number": "", "section": "SPARE", "category": "Spare",
                             "description": "NO NUMBER", "vendor": "SVMS vendor",
                             "budget": 20}],
                  "job_count": 1, "total_budget": 20}
        result = integration._apply_svms_job_import(db, "v_test", parsed)
        row = db.execute("SELECT * FROM jobs WHERE vessel_id='v_test'").fetchone()
        self.assertEqual((0, 1, 0), (result["inserted"], result["updated"], result["unchanged"]))
        self.assertEqual(("Manual vendor", 20, 55, 80),
                         (row["vendor"], row["budget"], row["consumption"], row["completion"]))

    def test_svms_import_accepts_bare_and_numberless_subjects(self):
        parsed = integration._normalize_svms_dock_jobs({
            "P_RS_DP": [],
            "P_RS_SS": [
                {"SUBJ": "NO JOB NUMBER", "AMT": 10, "AMT_USD": 10, "CUR_CD": "USD"},
                {"SUBJ": "[S7] LOCAL SUPPLY - STORES", "AMT_USD": 25, "CUR_CD": "KRW"},
                {"SUBJ": "[S8] LOCAL SUPPLY - SPARE", "AMT": 1350,
                 "AMT_USD": 1, "CUR_CD": "KRW"},
                {"SUBJ": "ST9 BARE STORE", "AMT": 30, "AMT_USD": 30, "CUR_CD": "USD"},
            ],
            "P_RS_SR": [],
        })
        self.assertEqual(4, parsed["job_count"])
        self.assertEqual(("", "Spare", 10),
                         (parsed["jobs"][0]["number"], parsed["jobs"][0]["category"], parsed["jobs"][0]["budget"]))
        self.assertEqual(("Store", 25), (parsed["jobs"][1]["category"], parsed["jobs"][1]["budget"]))
        self.assertEqual(("Spare", 1), (parsed["jobs"][2]["category"], parsed["jobs"][2]["budget"]))
        self.assertEqual(("ST9", "Store", "BARE STORE"),
                         (parsed["jobs"][3]["number"], parsed["jobs"][3]["category"], parsed["jobs"][3]["description"]))
        self.assertIn("Job No. 없음", parsed["warnings"][0])


if __name__ == "__main__":
    unittest.main()
