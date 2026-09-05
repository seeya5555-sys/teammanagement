import datetime
import sqlite3
import unittest
from urllib.parse import unquote

from flask import Flask
from openpyxl import load_workbook

import drydock_integration as integration


def job(**overrides):
    value = {
        "number": "S1",
        "category": "Spare",
        "description": "Main engine spare",
        "vendor": "Maker",
        "budget": 1200,
        "consumption": 300,
        "start_date": "2026-09-11",
        "end_date": "2026-09-14",
        "completion": 25,
        "remarks": '[{"date":"2026-09-12","progress":"On board","important":true}]',
    }
    value.update(overrides)
    return value


class JobProgressExportTests(unittest.TestCase):
    def test_fills_selected_vessel_and_preserves_template_layout(self):
        output = integration._build_job_progress_workbook(
            {"name": "BELGIUM B", "dock_in": "2026-09-13"}, [job()]
        )
        workbook = load_workbook(output, data_only=False)
        self.assertEqual(["Job progress", "Class item"], workbook.sheetnames)
        sheet = workbook["Job progress"]
        self.assertEqual("BELGIUM B DD JOB PROGRESS", sheet["B2"].value)
        self.assertEqual("F7", sheet.freeze_panes)
        self.assertEqual(datetime.datetime(2026, 9, 13), sheet["M4"].value)
        self.assertEqual(datetime.datetime(2026, 9, 3), sheet["P6"].value)
        self.assertEqual(datetime.datetime(2026, 10, 10), sheet["BA6"].value)
        self.assertEqual("Sep", sheet["P5"].value)
        self.assertEqual("Oct", sheet["AR5"].value)
        self.assertIn("P5:AQ5", {str(value) for value in sheet.merged_cells.ranges})
        self.assertIn("AR5:BA5", {str(value) for value in sheet.merged_cells.ranges})

        self.assertEqual("S1", sheet["B7"].value)
        self.assertEqual("Spare", sheet["C7"].value)
        self.assertEqual("Main engine spare", sheet["D7"].value)
        self.assertEqual("2026-09-12 ! On board", sheet["F7"].value)
        self.assertEqual(1200, sheet["H7"].value)
        self.assertEqual(300, sheet["I7"].value)
        self.assertEqual("=H7-I7", sheet["J7"].value)
        self.assertEqual(0.25, sheet["O7"].value)
        self.assertEqual("0%", sheet["O7"].number_format)
        self.assertTrue(all(sheet.cell(row, col).value is None
                            for row in range(8, integration._JOB_PROGRESS_LAST_ROW + 1)
                            for col in (2, 3, 4, 5, 6, 7, 8, 9, 11, 12, 15)))
        self.assertEqual("=H8-I8", sheet["J8"].value)
        self.assertEqual("=SUMIF(C7:C370,G3,H7:H370)", sheet["H3"].value)
        self.assertTrue(all(sheet.cell(row, col).fill.fgColor.rgb == "FFFFFFFF"
                            for row in (7, 8, 370) for col in range(2, 54)))
        gantt_formulas = {
            formula
            for conditional_range in sheet.conditional_formatting
            if "P7" in str(conditional_range)
            for rule in sheet.conditional_formatting[conditional_range]
            for formula in (rule.formula or [])
        }
        self.assertIn('AND($K7<>"",$L7<>"",P$6>$K7,P$6<$L7)', gantt_formulas)
        self.assertIn('AND($L7<>"",P$6=$L7)', gantt_formulas)
        self.assertIn('AND($K7<>"",P$6=$K7)', gantt_formulas)

        class_sheet = workbook["Class item"]
        self.assertEqual("Class item", class_sheet["B2"].value)
        self.assertEqual("No.", class_sheet["B3"].value)
        self.assertTrue(all(class_sheet.cell(row, col).value is None
                            for row in range(4, class_sheet.max_row + 1)
                            for col in range(2, 9)))

    def test_uses_earliest_job_date_when_dock_in_is_missing(self):
        output = integration._build_job_progress_workbook(
            {"name": "PLANNED", "dock_in": None},
            [job(start_date="2026-11-20"), job(start_date="2026-11-10")],
        )
        sheet = load_workbook(output, data_only=False)["Job progress"]
        self.assertIsNone(sheet["M4"].value)
        self.assertEqual(datetime.datetime(2026, 10, 31), sheet["P6"].value)

    def test_rejects_more_jobs_than_the_template_can_hold(self):
        with self.assertRaisesRegex(ValueError, "364"):
            integration._build_job_progress_workbook(
                {"name": "TOO MANY", "dock_in": None}, [job() for _ in range(365)]
            )

    def test_dirty_numeric_values_fall_back_without_breaking_export(self):
        output = integration._build_job_progress_workbook(
            {"name": "DIRTY", "dock_in": "2026-09-13"},
            [job(budget="bad", consumption=float("nan"), completion="unknown")],
        )
        sheet = load_workbook(output, data_only=False)["Job progress"]
        self.assertEqual((0, 0, 0), (sheet["H7"].value, sheet["I7"].value, sheet["O7"].value))

    def test_translates_korean_progress_remarks_and_preserves_dates(self):
        calls = []

        def translate(values):
            calls.append(values)
            return ["2026-09-12 Attendance scheduled"]

        output = integration._build_job_progress_workbook(
            {"name": "BELGIUM B", "dock_in": "2026-09-13"},
            [job(remarks='[{"date":"2026-09-12","progress":"입회 예정"}]')],
            translator=translate,
        )
        sheet = load_workbook(output, data_only=False)["Job progress"]
        self.assertEqual([["2026-09-12 입회 예정"]], calls)
        self.assertEqual("2026-09-12 Attendance scheduled", sheet["F7"].value)

    def test_rejects_incomplete_korean_translation(self):
        with self.assertRaisesRegex(RuntimeError, "영문 번역"):
            integration._build_job_progress_workbook(
                {"name": "BELGIUM B", "dock_in": "2026-09-13"},
                [job(remarks='[{"date":"2026-09-12","progress":"입회 예정"}]')],
                translator=lambda values: values,
            )

    def test_endpoint_requires_admin_and_returns_xlsx_with_unicode_filename(self):
        db = sqlite3.connect(":memory:")
        db.row_factory = sqlite3.Row
        db.executescript("""
            CREATE TABLE vessels (id TEXT PRIMARY KEY, name TEXT, dock_in TEXT);
            CREATE TABLE jobs (
                id INTEGER PRIMARY KEY, vessel_id TEXT, number TEXT, category TEXT,
                description TEXT, vendor TEXT, budget REAL, consumption REAL,
                start_date TEXT, end_date TEXT, completion REAL, remarks TEXT
            );
            INSERT INTO vessels VALUES('v1','대한민국/호','2026-09-13');
        """)

        class FakeDD:
            admin = False

            @classmethod
            def is_admin(cls):
                return cls.admin

            @staticmethod
            def get_db():
                return db

        app = Flask(__name__)
        integration._install_job_progress_export(FakeDD, app)
        client = app.test_client()
        self.assertEqual(401, client.get("/api/vessels/v1/jobs/progress.xlsx").status_code)
        FakeDD.admin = True
        response = client.get("/api/vessels/v1/jobs/progress.xlsx")
        self.assertEqual(200, response.status_code)
        self.assertTrue(response.data.startswith(b"PK\x03\x04"))
        self.assertEqual(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response.mimetype,
        )
        self.assertIn("대한민국_호_DD_JOB_PROGRESS.xlsx", unquote(response.headers["Content-Disposition"]))


if __name__ == "__main__":
    unittest.main()
